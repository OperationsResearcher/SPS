@api_bp.route('/surec/karne/pg-veri-detay', methods=['GET'])
@login_required
def api_surec_karne_pg_veri_detay():
    """Süreç karnesinde bir PG'nin kullanıcı bazlı veri detaylarını getir"""
    try:
        surec_pg_id = request.args.get('surec_pg_id', type=int)
        ceyrek = request.args.get('ceyrek', type=str)  # String olarak al (çeyrek, ay, hafta, gun, yillik olabilir)
        yil = request.args.get('yil', type=int, default=datetime.now().year)
        
        if not surec_pg_id:
            return jsonify({'success': False, 'message': 'Eksik parametreler'}), 400

        # Bu endpoint kullanıcı bazlı veri döndürdüğü için yönetim rolleri ile sınırla
        surec_pg = SurecPerformansGostergesi.query.get(surec_pg_id)
        if not surec_pg:
            return jsonify({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
        surec = Surec.query.get(surec_pg.surec_id)
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404

        if current_user.sistem_rol == 'admin':
            pass
        elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            if surec.kurum_id != current_user.kurum_id:
                return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403
        else:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
        # Bu süreç PG'sine bağlı tüm bireysel PG'leri bul
        bireysel_pgler = BireyselPerformansGostergesi.query.filter_by(
            kaynak_surec_pg_id=surec_pg_id,
            kaynak='Süreç'
        ).all()
        
        if not bireysel_pgler:
            return jsonify({'success': True, 'veriler': []})
        
        # Her bir bireysel PG için bu periyottaki veri girişlerini topla
        tum_veriler = []
        for bireysel_pg in bireysel_pgler:
            veri_girisleri = []
            
            # Periyot tipine göre sorgu yap
            if ceyrek == 'yillik':
                # Yıllık veri
                veri_girisleri = PerformansGostergeVeri.query.filter_by(
                    bireysel_pg_id=bireysel_pg.id,
                    yil=yil,
                    ceyrek=None,
                    ay=None
                ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
            elif ceyrek and ceyrek.isdigit() and int(ceyrek) >= 1 and int(ceyrek) <= 4:
                # Çeyreklik veri
                ceyrek_no = int(ceyrek)
                # Çeyreklik kaydı ve bu çeyreğin aylarındaki veriler
                ceyrek_aylari = get_ceyrek_aylari(ceyrek_no)
                veri_girisleri = PerformansGostergeVeri.query.filter(
                    PerformansGostergeVeri.bireysel_pg_id == bireysel_pg.id,
                    PerformansGostergeVeri.yil == yil
                ).filter(
                    or_(
                        and_(PerformansGostergeVeri.ceyrek == ceyrek_no, PerformansGostergeVeri.ay.is_(None)),  # Çeyreklik kayıt
                        PerformansGostergeVeri.ay.in_(ceyrek_aylari)  # Bu çeyreğin aylarındaki veriler
                    )
                ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
            elif ceyrek and ceyrek.isdigit() and int(ceyrek) >= 1 and int(ceyrek) <= 12:
                # Aylık veri
                ay_no = int(ceyrek)
                # Aylık kayıt ve bu ayın tüm verileri (haftalık/günlük dahil)
                veri_girisleri = PerformansGostergeVeri.query.filter_by(
                    bireysel_pg_id=bireysel_pg.id,
                    yil=yil,
                    ay=ay_no
                ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
            
                # Eğer bu ay için veri yoksa, çeyreklik veriyi kontrol et
                if not veri_girisleri:
                    ilgili_ceyrek = get_ay_ceyreği(ay_no)
                    if ilgili_ceyrek:
                        ceyrek_verileri = PerformansGostergeVeri.query.filter_by(
                            bireysel_pg_id=bireysel_pg.id,
                            yil=yil,
                            ceyrek=ilgili_ceyrek,
                            ay=None
                        ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
                        if ceyrek_verileri:
                            veri_girisleri = ceyrek_verileri
                
                # Eğer hala veri yoksa, yıllık veriyi kontrol et
                if not veri_girisleri:
                    yillik_verileri = PerformansGostergeVeri.query.filter_by(
                        bireysel_pg_id=bireysel_pg.id,
                        yil=yil,
                        ceyrek=None,
                        ay=None
                    ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
                    if yillik_verileri:
                        veri_girisleri = yillik_verileri
            elif ceyrek and ceyrek.isdigit() and int(ceyrek) >= 1 and int(ceyrek) <= 5:
                # Haftalık veri - hafta numarası
                hafta_no = int(ceyrek)
                # Mevcut ayın haftalarını al
                mevcut_ay = datetime.now().month
                haftalar = get_ay_haftalari(mevcut_ay, yil)
                if 1 <= hafta_no <= len(haftalar):
                    hafta_bilgi = haftalar[hafta_no - 1]
                    hafta_baslangic = hafta_bilgi['baslangic_tarih']
                    hafta_bitis = hafta_bilgi['bitis_tarih']
                    veri_girisleri = PerformansGostergeVeri.query.filter(
                        PerformansGostergeVeri.bireysel_pg_id == bireysel_pg.id,
                        PerformansGostergeVeri.yil == yil,
                        PerformansGostergeVeri.veri_tarihi >= hafta_baslangic,
                        PerformansGostergeVeri.veri_tarihi <= hafta_bitis
                    ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
            elif ceyrek and ceyrek.isdigit() and int(ceyrek) >= 1 and int(ceyrek) <= 31:
                # Günlük veri - gün numarası
                gun_no = int(ceyrek)
                # Mevcut ayın günlerini al
                mevcut_ay = datetime.now().month
                gunler = get_ay_gunleri(mevcut_ay, yil)
                if 1 <= gun_no <= len(gunler):
                    gun_tarih = gunler[gun_no - 1]['tarih']
                    veri_girisleri = PerformansGostergeVeri.query.filter_by(
                        bireysel_pg_id=bireysel_pg.id,
                        yil=yil,
                        veri_tarihi=gun_tarih
                    ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
            else:
                # Varsayılan: çeyrek olarak kabul et
                if ceyrek and ceyrek.isdigit():
                    ceyrek_no = int(ceyrek)
                    veri_girisleri = PerformansGostergeVeri.query.filter_by(
                        bireysel_pg_id=bireysel_pg.id,
                        yil=yil,
                        ceyrek=ceyrek_no
                    ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
            
            current_app.logger.info(f"Bireysel PG {bireysel_pg.id} için {len(veri_girisleri)} veri bulundu (Periyot: {ceyrek})")
            
            for veri in veri_girisleri:
                kullanici = User.query.get(veri.user_id)
                if kullanici:
                    if kullanici.first_name and kullanici.last_name:
                        kullanici_adi = f"{kullanici.first_name} {kullanici.last_name}"
                    else:
                        kullanici_adi = kullanici.username
                else:
                    kullanici_adi = 'Bilinmiyor'
                
                current_app.logger.info(f"Veri: user_id={veri.user_id}, kullanici={kullanici_adi}, tarih={veri.veri_tarihi}, ay={veri.ay}, ceyrek={veri.ceyrek}")
                
                tum_veriler.append({
                    'id': veri.id,
                    'kullanici_id': veri.user_id,
                    'kullanici_adi': kullanici_adi,
                    'hedef_deger': veri.hedef_deger,
                    'gerceklesen_deger': veri.gerceklesen_deger,
                    'durum': veri.durum,
                    'durum_yuzdesi': veri.durum_yuzdesi,
                    'veri_tarihi': veri.veri_tarihi.isoformat() if veri.veri_tarihi else None,
                    'aciklama': veri.aciklama if veri.aciklama else None,
                    'created_at': veri.created_at.isoformat() if veri.created_at else None
                })
        
        # Tarihe göre sırala (en yeni en üstte)
        tum_veriler.sort(key=lambda x: x['veri_tarihi'] or '', reverse=True)
        
        current_app.logger.info(f"Toplam {len(tum_veriler)} veri döndürülüyor")
        
        return jsonify({
            'success': True,
            'veriler': tum_veriler
        })
        
    except Exception as e:
        current_app.logger.error(f'PG veri detay getirme hatası: {str(e)}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


################################################################################

@api_bp.route('/export/surec_karnesi/excel', methods=['GET'])
@login_required
def export_surec_karnesi_excel():
    """Süreç karnesi verilerini Excel olarak dışa aktarır."""
    surec_id = request.args.get('surec_id', type=int)
    yil = request.args.get('yil', type=int)
    
    if not surec_id or not yil:
        return jsonify({"success": False, "message": "Süreç ID ve Yıl gereklidir."}), 400
    
    surec = Surec.query.get(surec_id)
    if not surec:
        return jsonify({"success": False, "message": "Süreç bulunamadı."}), 404
    
    # Check user permission
    if current_user.sistem_rol == 'admin':
        # Admin tüm süreçleri görüntüleyebilir
        pass
    elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
        # Kurum yöneticileri kendi kurumlarındaki süreçleri görüntüleyebilir
        if surec.kurum_id != current_user.kurum_id:
            return jsonify({"success": False, "message": "Bu süreci görüntüleme yetkiniz yok."}), 403
    else:
        # Normal kullanıcı için lider/üye kontrolü
        lider_mi = db.session.query(surec_liderleri).filter(
            surec_liderleri.c.surec_id == surec_id,
            surec_liderleri.c.user_id == current_user.id
        ).first() is not None
        
        uye_mi = db.session.query(surec_uyeleri).filter(
            surec_uyeleri.c.surec_id == surec_id,
            surec_uyeleri.c.user_id == current_user.id
        ).first() is not None
        
        if not (lider_mi or uye_mi):
            return jsonify({"success": False, "message": "Bu süreci görüntüleme yetkiniz yok."}), 403
    
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception as import_err:
            current_app.logger.error(f'openpyxl import hatası: {import_err}')
            return jsonify({
                "success": False,
                "message": "Excel aktarımı için openpyxl gerekli. Lütfen bağımlılığı yükleyin."
            }), 500

        # Create workbook and sheets
        wb = Workbook()
        ws_pg = wb.active
        ws_pg.title = "Performans Göstergeleri"
        ws_faaliyet = wb.create_sheet("Faaliyetler")
        
        # --- Sheet 1: Performans Göstergeleri ---
        
        # Headers
        pg_headers = [
            "Kodu", "Performans Adı", "Hedef", "Periyot", "Ağırlık (%)",
            "I. Çeyrek Hedef", "I. Çeyrek Gerç.", "I. Çeyrek Durum",
            "II. Çeyrek Hedef", "II. Çeyrek Gerç.", "II. Çeyrek Durum",
            "III. Çeyrek Hedef", "III. Çeyrek Gerç.", "III. Çeyrek Durum",
            "IV. Çeyrek Hedef", "IV. Çeyrek Gerç.", "IV. Çeyrek Durum"
        ]
        ws_pg.append(pg_headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws_pg[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Fetch PG data
        pg_query = SurecPerformansGostergesi.query.filter_by(surec_id=surec_id).order_by(SurecPerformansGostergesi.id)
        performans_gostergeleri = pg_query.all()
        
        for pg in performans_gostergeleri:
            row_data = [
                pg.kodu if hasattr(pg, 'kodu') and pg.kodu else f"PG-{pg.id}",
                pg.ad,
                pg.hedef_deger,
                pg.periyot,
                pg.agirlik if hasattr(pg, 'agirlik') else 0,
            ]
            # Fetch data for each quarter
            for ceyrek in range(1, 5):
                # This is a simplified data fetching. A more robust implementation would query PerformansGostergeVeri
                # and aggregate the results based on the logic in the frontend.
                # For now, we'll leave these blank as a placeholder.
                row_data.extend(["", "", ""]) # Hedef, Gerç., Durum
            ws_pg.append(row_data)
        
        # --- Sheet 2: Faaliyetler ---
        faaliyet_headers = ["No", "Faaliyet Adı", "Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
        ws_faaliyet.append(faaliyet_headers)
        
        # Style headers
        for cell in ws_faaliyet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Fetch Faaliyet data
        faaliyet_query = SurecFaaliyet.query.filter_by(surec_id=surec_id).order_by(SurecFaaliyet.id)
        faaliyetler = faaliyet_query.all()
        
        for i, faaliyet in enumerate(faaliyetler):
            row_data = [i + 1, faaliyet.ad]
            # Fetch monthly completion data
            # Note: FaaliyetTakip modeli ile ilişki kontrol edilmeli
            for ay in range(1, 13):
                row_data.append("")  # Placeholder - gerçek veri için FaaliyetTakip sorgusu gerekli
            ws_faaliyet.append(row_data)
        
        # Save to a memory buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"surec_karnesi_{surec.ad.replace(' ', '_')}_{yil}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        current_app.logger.error(f'Excel export hatası: {e}')
        return jsonify({"success": False, "message": str(e)}), 500


################################################################################

@api_bp.route('/pg-veri/detay/<int:veri_id>', methods=['GET'])
@login_required
@csrf.exempt
def api_pg_veri_detay(veri_id):
    """Tek bir PG verisinin detaylarını ve audit log'unu getir"""
    try:
        # Veriyi bul
        pg_veri = PerformansGostergeVeri.query.get(veri_id)
        if not pg_veri:
            return jsonify({'success': False, 'message': 'Veri bulunamadı'}), 404
        
        # Yetki kontrolü - verinin ait olduğu bireysel PG'yi kontrol et
        bireysel_pg = BireyselPerformansGostergesi.query.get(pg_veri.bireysel_pg_id)
        if not bireysel_pg:
            return jsonify({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
        
        # Süreç bazlı PG ise süreç yetkisi kontrol et
        surec_id = None
        if bireysel_pg.kaynak == 'Süreç' and bireysel_pg.kaynak_surec_id:
            surec_id = bireysel_pg.kaynak_surec_id
        elif bireysel_pg.kaynak_surec_pg_id:
            kaynak_surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
            surec_id = kaynak_surec_pg.surec_id if kaynak_surec_pg else None

        is_surec_verisi = surec_id is not None
        if is_surec_verisi:
            # Kurum yöneticisi kontrolü
            kurum_yoneticisi_mi = current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']
            if kurum_yoneticisi_mi:
                surec = Surec.query.get(surec_id)
                if not surec:
                    return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
                    return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
            else:
                # Normal kullanıcı için lider/üye kontrolü
                lider_mi = db.session.query(surec_liderleri).filter(
                    surec_liderleri.c.surec_id == surec_id,
                    surec_liderleri.c.user_id == current_user.id
                ).first() is not None

                uye_mi = db.session.query(surec_uyeleri).filter(
                    surec_uyeleri.c.surec_id == surec_id,
                    surec_uyeleri.c.user_id == current_user.id
                ).first() is not None

                if not (lider_mi or uye_mi):
                    return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
        
        # Audit log'ları getir
        audit_logs = PerformansGostergeVeriAudit.query.filter_by(
            pg_veri_id=veri_id
        ).order_by(PerformansGostergeVeriAudit.islem_tarihi.desc()).all()
        
        # Kullanıcı bilgilerini al
        veri_sahibi = User.query.get(pg_veri.user_id)
        veri_sahibi_adi = ''
        if veri_sahibi:
            if veri_sahibi.first_name and veri_sahibi.last_name:
                veri_sahibi_adi = f"{veri_sahibi.first_name} {veri_sahibi.last_name}"
            else:
                veri_sahibi_adi = veri_sahibi.username
        
        # Audit log'ları formatla
        audit_list = []
        for audit in audit_logs:
            islem_yapan = User.query.get(audit.islem_yapan_user_id)
            islem_yapan_adi = ''
            if islem_yapan:
                if islem_yapan.first_name and islem_yapan.last_name:
                    islem_yapan_adi = f"{islem_yapan.first_name} {islem_yapan.last_name}"
                else:
                    islem_yapan_adi = islem_yapan.username
            
            audit_list.append({
                'id': audit.id,
                'islem_tipi': audit.islem_tipi,
                'eski_deger': audit.eski_deger,
                'yeni_deger': audit.yeni_deger,
                'degisiklik_aciklama': audit.degisiklik_aciklama,
                'islem_yapan': islem_yapan_adi,
                'islem_tarihi': audit.islem_tarihi.strftime('%d.%m.%Y %H:%M:%S') if audit.islem_tarihi else None
            })
        
        # Son güncelleyen bilgisini audit log'dan al
        son_guncelleyen = None
        son_guncelleme_tarihi = None
        if audit_list and len(audit_list) > 0:
            son_islem = audit_list[0]  # En son işlem
            if son_islem['islem_tipi'] == 'GUNCELLE':
                son_guncelleyen = son_islem['islem_yapan']
                son_guncelleme_tarihi = son_islem['islem_tarihi']
        
        # Veri bilgilerini formatla (frontend'in beklediği alan adlarıyla)
        veri_data = {
            'id': pg_veri.id,
            'bireysel_pg_id': pg_veri.bireysel_pg_id,
            'pg_adi': bireysel_pg.ad,
            'yil': pg_veri.yil,
            'veri_tarihi': pg_veri.veri_tarihi.strftime('%d.%m.%Y') if pg_veri.veri_tarihi else None,
            'giris_periyot_tipi': pg_veri.giris_periyot_tipi,
            'giris_periyot_no': pg_veri.giris_periyot_no,
            'giris_periyot_ay': pg_veri.giris_periyot_ay,
            'hedef_deger': pg_veri.hedef_deger,
            'gerceklesen_deger': str(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger is not None else '-',
            'durum': pg_veri.durum,
            'durum_yuzdesi': pg_veri.durum_yuzdesi,
            'aciklama': pg_veri.aciklama,
            'olusturan': veri_sahibi_adi,  # Frontend'in beklediği alan adı
            'olusturma_tarihi': pg_veri.created_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.created_at else None,  # Frontend'in beklediği alan adı
            'guncelleyen': son_guncelleyen,  # Son güncelleyen audit log'dan
            'guncelleme_tarihi': son_guncelleme_tarihi,  # Son güncelleme tarihi audit log'dan
            # Geriye dönük uyumluluk için eski alan adları da ekle
            'veri_sahibi': veri_sahibi_adi,
            'created_at': pg_veri.created_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.created_at else None,
            'updated_at': pg_veri.updated_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.updated_at else None
        }
        
        # Yetki bilgileri
        if is_surec_verisi:
            # Süreç karnesi yazma işlemleri sadece yönetim rolleri
            duzenleyebilir = current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']
            silebilir = duzenleyebilir
        else:
            # Süreç dışı (bireysel) KPI'larda mevcut davranışı koru
            duzenleyebilir = (pg_veri.user_id == current_user.id) or (current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim'])
            silebilir = duzenleyebilir

        yetki = {
            'duzenleyebilir': duzenleyebilir,
            'silebilir': silebilir
        }
        
        # Hesaplama yöntemi bilgisini ekle (bireysel PG'den)
        hesaplama_yontemi = None
        if bireysel_pg.kaynak_surec_pg_id:
            surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
            if surec_pg:
                hesaplama_yontemi = surec_pg.veri_toplama_yontemi
        
        veri_data['hesaplama_yontemi'] = hesaplama_yontemi
        veri_data['surec_pg_id'] = bireysel_pg.kaynak_surec_pg_id if bireysel_pg.kaynak == 'Süreç' else None
        
        return jsonify({
            'success': True,
            'veri': veri_data,
            'audit_log': audit_list,
            'yetki': yetki,
            'bireysel_pg_id': bireysel_pg.id,
            'surec_pg_id': bireysel_pg.kaynak_surec_pg_id if bireysel_pg.kaynak == 'Süreç' else None
        })
    except Exception as e:
        current_app.logger.error(f'PG veri detay hatası: {e}')
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


################################################################################

@api_bp.route('/pg-veri/detay/toplu', methods=['POST'])
@login_required
@csrf.exempt
def api_pg_veri_detay_toplu():
    """Birden fazla PG verisinin detaylarını toplu olarak getir"""
    try:
        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400

        data = request.get_json() or {}
        current_app.logger.info(f'Toplu veri detay isteği alındı: {data}')
        
        if not data:
            return jsonify({'success': False, 'message': 'Request body boş'}), 400
        
        veri_idleri = data.get('veri_idleri', [])
        current_app.logger.info(f'Veri ID listesi: {veri_idleri}, Tip: {type(veri_idleri)}')
        
        if not veri_idleri:
            return jsonify({'success': False, 'message': 'Veri ID listesi boş'}), 400
        
        if not isinstance(veri_idleri, list):
            return jsonify({'success': False, 'message': f'Veri ID listesi geçersiz tip: {type(veri_idleri)}'}), 400
        
        # Liste elemanlarını integer'a çevir
        try:
            veri_idleri = [int(vid) for vid in veri_idleri]
        except (ValueError, TypeError) as e:
            current_app.logger.error(f'Veri ID dönüştürme hatası: {e}')
            return jsonify({'success': False, 'message': f'Veri ID\'leri integer\'a dönüştürülemedi: {e}'}), 400
        
        # Verileri getir
        pg_veriler = PerformansGostergeVeri.query.filter(PerformansGostergeVeri.id.in_(veri_idleri)).all()

        if not pg_veriler:
            return jsonify({'success': False, 'message': 'Veriler bulunamadı'}), 404

        pg_veri_by_id = {v.id: v for v in pg_veriler}
        missing_ids = [vid for vid in veri_idleri if vid not in pg_veri_by_id]
        if missing_ids:
            return jsonify({'success': False, 'message': 'Bazı veriler bulunamadı'}), 404
        
        # Her veri için yetki kontrolü ve detayları topla (kısmi sızıntıyı engellemek için strict)
        veriler_listesi = []
        for veri_id in veri_idleri:
            pg_veri = pg_veri_by_id[veri_id]
            # Yetki kontrolü
            bireysel_pg = BireyselPerformansGostergesi.query.get(pg_veri.bireysel_pg_id)
            if not bireysel_pg:
                return jsonify({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
            
            # Süreç bazlı PG ise süreç yetkisi kontrol et
            surec_id = None
            if bireysel_pg.kaynak == 'Süreç' and bireysel_pg.kaynak_surec_id:
                surec_id = bireysel_pg.kaynak_surec_id
            elif bireysel_pg.kaynak_surec_pg_id:
                kaynak_surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
                surec_id = kaynak_surec_pg.surec_id if kaynak_surec_pg else None

            if surec_id is not None:
                surec = Surec.query.get(surec_id)
                if not surec:
                    return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403

                if current_user.sistem_rol == 'admin':
                    pass
                elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
                    if surec.kurum_id != current_user.kurum_id:
                        return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                else:
                    if surec.kurum_id != current_user.kurum_id:
                        return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403

                    lider_mi = db.session.query(surec_liderleri).filter(
                        surec_liderleri.c.surec_id == surec_id,
                        surec_liderleri.c.user_id == current_user.id
                    ).first() is not None

                    uye_mi = db.session.query(surec_uyeleri).filter(
                        surec_uyeleri.c.surec_id == surec_id,
                        surec_uyeleri.c.user_id == current_user.id
                    ).first() is not None

                    if not (lider_mi or uye_mi):
                        return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
            else:
                # Süreç dışı (bireysel) KPI verisi
                if current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
                    if current_user.sistem_rol != 'admin':
                        veri_sahibi = User.query.get(pg_veri.user_id)
                        if veri_sahibi and veri_sahibi.kurum_id != current_user.kurum_id:
                            return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                else:
                    if pg_veri.user_id != current_user.id:
                        return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
            
            # Kullanıcı bilgilerini al
            veri_sahibi = User.query.get(pg_veri.user_id)
            veri_sahibi_adi = ''
            if veri_sahibi:
                if veri_sahibi.first_name and veri_sahibi.last_name:
                    veri_sahibi_adi = f"{veri_sahibi.first_name} {veri_sahibi.last_name}"
                else:
                    veri_sahibi_adi = veri_sahibi.username
            
            # Audit log'ları getir (tüm geçmişi göster)
            audit_logs = PerformansGostergeVeriAudit.query.filter_by(
                pg_veri_id=pg_veri.id
            ).order_by(PerformansGostergeVeriAudit.islem_tarihi.desc()).all()
            
            audit_list = []
            for audit in audit_logs:
                islem_yapan = User.query.get(audit.islem_yapan_user_id)
                islem_yapan_adi = ''
                if islem_yapan:
                    if islem_yapan.first_name and islem_yapan.last_name:
                        islem_yapan_adi = f"{islem_yapan.first_name} {islem_yapan.last_name}"
                    else:
                        islem_yapan_adi = islem_yapan.username
                
                audit_list.append({
                    'id': audit.id,
                    'islem_tipi': audit.islem_tipi,
                    'eski_deger': audit.eski_deger,
                    'yeni_deger': audit.yeni_deger,
                    'degisiklik_aciklama': audit.degisiklik_aciklama,
                    'islem_yapan': islem_yapan_adi,
                    'islem_tarihi': audit.islem_tarihi.strftime('%d.%m.%Y %H:%M:%S') if audit.islem_tarihi else None
                })
            
            # Frontend'in beklediği yapıya uygun formatla
            veriler_listesi.append({
                'veri': {
                    'id': pg_veri.id,
                    'bireysel_pg_id': pg_veri.bireysel_pg_id,
                    'pg_adi': bireysel_pg.ad,
                    'yil': pg_veri.yil,
                    'veri_tarihi': pg_veri.veri_tarihi.strftime('%d.%m.%Y') if pg_veri.veri_tarihi else None,
                    'giris_periyot_tipi': pg_veri.giris_periyot_tipi,
                    'giris_periyot_no': pg_veri.giris_periyot_no,
                    'giris_periyot_ay': pg_veri.giris_periyot_ay,
                    'hedef_deger': pg_veri.hedef_deger,
                    'gerceklesen_deger': str(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger is not None else '-',
                    'durum': pg_veri.durum,
                    'durum_yuzdesi': pg_veri.durum_yuzdesi,
                    'aciklama': pg_veri.aciklama,
                    'olusturan': veri_sahibi_adi,
                    'olusturma_tarihi': pg_veri.created_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.created_at else None,
                    'guncelleyen': None,  # Son güncelleyen audit log'dan alınacak
                    'guncelleme_tarihi': pg_veri.updated_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.updated_at else None
                },
                'audit_log': audit_list,
                'yetki': {
                    'can_edit': True,
                    'can_delete': True
                }
            })
            
            # Son güncelleyen bilgisini audit log'dan al
            if audit_list and len(audit_list) > 0:
                son_islem = audit_list[0]  # En son işlem
                if son_islem['islem_tipi'] == 'GUNCELLE':
                    veriler_listesi[-1]['veri']['guncelleyen'] = son_islem['islem_yapan']
                    veriler_listesi[-1]['veri']['guncelleme_tarihi'] = son_islem['islem_tarihi']
        
        # Yetki bilgileri
        # (Bu endpoint süreç karnesi veri detayında kullanılıyor; yazma işlemleri yönetim rolleri ile sınırlandırılır)
        yetki = {
            'duzenleyebilir': current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim'],
            'silebilir': current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']
        }
        
        return jsonify({
            'success': True,
            'veriler': veriler_listesi,
            'yetki': yetki
        })
    except Exception as e:
        current_app.logger.error(f'Toplu PG veri detay hatası: {e}')
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


################################################################################

@api_bp.route('/pg-veri/guncelle/<int:veri_id>', methods=['PUT'])
@login_required
@csrf.exempt
def api_pg_veri_guncelle(veri_id):
    """PG verisini güncelle - Süreç lideri tüm kullanıcıların verilerini güncelleyebilir"""
    try:
        # Veriyi bul
        pg_veri = PerformansGostergeVeri.query.get(veri_id)
        if not pg_veri:
            return jsonify({'success': False, 'message': 'Veri bulunamadı'}), 404
        
        # Bireysel PG'yi bul
        bireysel_pg = BireyselPerformansGostergesi.query.get(pg_veri.bireysel_pg_id)
        if not bireysel_pg:
            return jsonify({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
        
        # Yetki kontrolü
        # Süreç karnesi verileri için yazma işlemleri sadece yönetim rolleri
        surec_id = None
        if bireysel_pg.kaynak == 'Süreç' and bireysel_pg.kaynak_surec_id:
            surec_id = bireysel_pg.kaynak_surec_id
        elif bireysel_pg.kaynak_surec_pg_id:
            kaynak_surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
            surec_id = kaynak_surec_pg.surec_id if kaynak_surec_pg else None

        is_surec_verisi = surec_id is not None
        kurum_yoneticisi_mi = current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']

        if is_surec_verisi:
            if not kurum_yoneticisi_mi:
                return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
            surec = Surec.query.get(surec_id)
            if not surec:
                return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
            if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
                return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
        else:
            # Süreç dışı (bireysel) KPI davranışını koru
            if kurum_yoneticisi_mi:
                # Yönetim rolleri kendi kurumundaki verileri güncelleyebilir
                veri_sahibi = User.query.get(pg_veri.user_id)
                if veri_sahibi and current_user.sistem_rol != 'admin' and veri_sahibi.kurum_id != current_user.kurum_id:
                    return jsonify({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
            else:
                if pg_veri.user_id != current_user.id:
                    return jsonify({'success': False, 'message': 'Bu veriyi güncelleme yetkiniz yok'}), 403
        
        # Güncelleme verilerini al
        data = request.get_json()
        gerceklesen_deger = data.get('gerceklesen_deger')
        aciklama = data.get('aciklama')
        
        # Eski değerleri kaydet (audit log için)
        eski_gerceklesen = pg_veri.gerceklesen_deger
        eski_aciklama = pg_veri.aciklama
        
        # Değerleri güncelle
        if gerceklesen_deger is not None:
            try:
                # String'den float'a çevir
                pg_veri.gerceklesen_deger = float(gerceklesen_deger) if gerceklesen_deger else None
            except (ValueError, TypeError):
                return jsonify({'success': False, 'message': 'Geçersiz gerçekleşen değer formatı'}), 400
        
        if aciklama is not None:
            pg_veri.aciklama = aciklama
        
        # Durumu hesapla
        if pg_veri.hedef_deger and pg_veri.gerceklesen_deger is not None:
            durum, durum_yuzdesi = hesapla_durum(
                float(pg_veri.hedef_deger) if pg_veri.hedef_deger else None,
                float(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger else None
            )
            pg_veri.durum = durum
            pg_veri.durum_yuzdesi = durum_yuzdesi
        
        pg_veri.updated_by = current_user.id
        pg_veri.updated_at = datetime.utcnow()
        
        # Audit log oluştur
        degisiklik_aciklama = []
        if eski_gerceklesen != pg_veri.gerceklesen_deger:
            degisiklik_aciklama.append(f'Gerçekleşen değer: {eski_gerceklesen} → {pg_veri.gerceklesen_deger}')
        if eski_aciklama != pg_veri.aciklama:
            degisiklik_aciklama.append(f'Açıklama güncellendi')
        
        if degisiklik_aciklama:
            audit = PerformansGostergeVeriAudit(
                pg_veri_id=pg_veri.id,
                islem_tipi='GUNCELLE',
                eski_deger=str(eski_gerceklesen) if eski_gerceklesen is not None else None,
                yeni_deger=str(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger is not None else None,
                degisiklik_aciklama='; '.join(degisiklik_aciklama),
                islem_yapan_user_id=current_user.id,
                islem_tarihi=datetime.utcnow()
            )
            db.session.add(audit)
        
        db.session.commit()
        
        # Skor Motoru: PG verisi güncellenince Vizyon puanını tüm hiyerarşide yeniden hesapla
        try:
            kurum_id = getattr(current_user, 'kurum_id', None)
            if kurum_id:
                from services.score_engine_service import recalc_on_pg_data_change
                recalc_on_pg_data_change(kurum_id)
        except Exception as ex:
            current_app.logger.debug("Skor motoru güncellemesi atlandı: %s", ex)
        
        return jsonify({
            'success': True,
            'message': 'Veri başarıyla güncellendi'
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'PG veri güncelleme hatası: %s', e)
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


################################################################################

