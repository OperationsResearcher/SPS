"""Bayi / Holding alt-tenant yönetim sayfası (Sprint C).

Yetki: can_manage_sub_tenants() — Platform Admin / Bayi admin / Holding admin

Bayi:
  - Alt tenant + ilk admin oluşturur (CRUD)
  - Alt-tenant'ın diğer kullanıcılarına / verisine erişemez

Holding:
  - Aynı işlemleri yapar
  - +  Ek olarak alt-tenant verilerini READ-ONLY görür (Sprint D)
"""
from __future__ import annotations

from datetime import datetime, timezone

from flask import render_template, jsonify, request, current_app
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash

from platform_core import app_bp
from app.extensions import csrf
from extensions import db
from app.models.core import Tenant, User, Role
from app.utils.tenant_scope import (
    can_manage_sub_tenants, is_platform_admin,
    is_dealer_user, is_holding_user, check_sub_tenant_limit,
)


def _403():
    return jsonify({"success": False, "message": "Yetkiniz yok."}), 403


# ─── Sayfa ───────────────────────────────────────────────────────────────────

@app_bp.route("/admin/sub-tenants")
@login_required
def admin_sub_tenants_page():
    if not can_manage_sub_tenants(current_user):
        return render_template("platform/errors/403.html"), 403

    # Hangi parent'ın alt-tenantlarını göstereceğiz?
    # - Platform Admin: kendi seçtiği bir parent_id ile (query param) yoksa tüm parent-capable'ları listele
    # - Bayi/Holding admin: kendi tenant'ı parent
    parent = None
    if is_platform_admin(current_user):
        parent_id = request.args.get("parent_id", type=int)
        if parent_id:
            parent = Tenant.query.get(parent_id)
    else:
        parent = current_user.tenant

    # Aktif paketler — alt-tenant açarken seçilebilir
    try:
        from app.models.saas import SubscriptionPackage
        packages = SubscriptionPackage.query.filter_by(is_active=True).order_by(SubscriptionPackage.name).all()
    except Exception:
        packages = []

    return render_template(
        "platform/admin/sub_tenants.html",
        parent=parent,
        packages=packages,
        is_platform_admin=is_platform_admin(current_user),
    )


# ─── Liste API ───────────────────────────────────────────────────────────────

@app_bp.route("/admin/api/sub-tenants")
@login_required
def admin_api_sub_tenants_list():
    if not can_manage_sub_tenants(current_user):
        return _403()

    # Parent tenant tespiti
    if is_platform_admin(current_user):
        parent_id = request.args.get("parent_id", type=int)
        if not parent_id:
            return jsonify({"success": False, "message": "Platform Admin için parent_id gerekli."}), 400
        parent = Tenant.query.get_or_404(parent_id)
    else:
        parent = current_user.tenant

    if parent.tenant_type not in ("dealer", "holding"):
        return jsonify({"success": False, "message": "Bu kurum tipi alt kurum açamaz."}), 400

    rows = Tenant.query.filter_by(parent_tenant_id=parent.id).order_by(Tenant.created_at.desc()).all()

    def _serialize(t: Tenant) -> dict:
        # İlk admin (varsa) — kayıt sırasıyla 1. tenant_admin rol
        admin = (
            db.session.query(User)
            .join(Role, User.role_id == Role.id)
            .filter(User.tenant_id == t.id, Role.name == "tenant_admin", User.is_active == True)
            .order_by(User.id)
            .first()
        )
        users_count = User.query.filter_by(tenant_id=t.id, is_active=True).count()
        return {
            "id": t.id,
            "name": t.name,
            "short_name": t.short_name,
            "is_active": t.is_active,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "admin": {
                "id": admin.id, "email": admin.email,
                "first_name": admin.first_name, "last_name": admin.last_name,
            } if admin else None,
            "users_count": users_count,
            "max_user_count": t.max_user_count,
            "package": {
                "id": t.package.id, "name": t.package.name,
            } if t.package else None,
        }

    return jsonify({
        "success": True,
        "parent": {
            "id": parent.id, "name": parent.name, "tenant_type": parent.tenant_type,
            "sub_tenant_limit": parent.sub_tenant_limit,
            "is_holding": parent.tenant_type == "holding",
        },
        "items": [_serialize(t) for t in rows],
        "active_count": sum(1 for t in rows if t.is_active),
    })


# ─── Yeni alt-tenant aç + ilk admin oluştur ──────────────────────────────────

@app_bp.route("/admin/api/sub-tenants", methods=["POST"])
@csrf.exempt
@login_required
def admin_api_sub_tenants_create():
    if not can_manage_sub_tenants(current_user):
        return _403()

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    admin_email = (data.get("admin_email") or "").strip().lower()
    admin_password = (data.get("admin_password") or "").strip()
    admin_first = (data.get("admin_first_name") or "").strip()
    admin_last = (data.get("admin_last_name") or "").strip()

    if not name or not admin_email:
        return jsonify({"success": False, "message": "Kurum adı ve admin e-posta zorunlu."}), 400

    # Parent tespit
    if is_platform_admin(current_user):
        parent_id = data.get("parent_tenant_id")
        if not parent_id:
            return jsonify({"success": False, "message": "Platform Admin için üst kurum ID gerekli."}), 400
        parent = Tenant.query.get(int(parent_id))
    else:
        parent = current_user.tenant

    if not parent or parent.tenant_type not in ("dealer", "holding"):
        return jsonify({"success": False, "message": "Geçersiz üst kurum."}), 400

    # Yetki: bayi/holding admin sadece kendi parent'ında işlem yapabilir
    if not is_platform_admin(current_user) and parent.id != current_user.tenant_id:
        return _403()

    # Kota
    ok, err = check_sub_tenant_limit(parent)
    if not ok:
        return jsonify({"success": False, "message": err}), 400

    # E-posta unique kontrolü
    if User.query.filter(db.func.lower(User.email) == admin_email).first():
        return jsonify({"success": False, "message": "Bu e-posta zaten başka bir kullanıcıya ait."}), 400

    # Geçici şifre üretimi (verilmezse)
    import secrets
    if not admin_password:
        admin_password = "Kp_" + secrets.token_urlsafe(8)
    elif len(admin_password) < 6:
        return jsonify({"success": False, "message": "Şifre en az 6 karakter olmalı."}), 400

    # tenant_admin rolünü bul
    tenant_admin_role = Role.query.filter_by(name="tenant_admin").first()
    if not tenant_admin_role:
        return jsonify({"success": False, "message": "Kurum yöneticisi rolü tanımlı değil."}), 500

    # Paket id (opsiyonel)
    package_id = data.get("package_id")
    if package_id:
        try:
            package_id = int(package_id)
            from app.models.saas import SubscriptionPackage
            pkg = SubscriptionPackage.query.get(package_id)
            if not pkg or not pkg.is_active:
                return jsonify({"success": False, "message": "Seçilen paket geçersiz veya pasif."}), 400
        except (TypeError, ValueError):
            return jsonify({"success": False, "message": "Paket ID geçersiz."}), 400
    else:
        package_id = None

    try:
        # Tenant oluştur
        t = Tenant(
            name=name,
            short_name=(data.get("short_name") or "").strip() or None,
            tenant_type="normal",
            parent_tenant_id=parent.id,
            max_user_count=int(data["max_user_count"]) if data.get("max_user_count") else 5,
            contact_email=admin_email,
            package_id=package_id,
        )
        db.session.add(t)
        db.session.flush()

        # İlk admin
        admin = User(
            email=admin_email,
            password_hash=generate_password_hash(admin_password),
            first_name=admin_first or None,
            last_name=admin_last or None,
            tenant_id=t.id,
            role_id=tenant_admin_role.id,
            is_active=True,
        )
        db.session.add(admin)
        db.session.commit()

        # Audit
        try:
            from app.utils.audit_logger import AuditLogger
            AuditLogger.log(
                action="SUB_TENANT_CREATE", resource_type="Kurum Yönetimi",
                resource_id=t.id,
                description=f"{current_user.email} bir alt-tenant açtı: '{t.name}' (parent={parent.name}, admin={admin.email})",
            )
        except Exception:
            pass

        return jsonify({
            "success": True,
            "message": f"Alt kurum '{t.name}' açıldı ve yönetici oluşturuldu.",
            "tenant_id": t.id,
            "admin_id": admin.id,
            "admin_email": admin.email,
            "admin_password": admin_password,  # plain — UI kullanıcıya gösterip kapatacak
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_api_sub_tenants_create] {e}", exc_info=True)
        return jsonify({"success": False, "message": f"Hata: {e}"}), 500


# ─── Alt-tenant admin'inin parolasını sıfırla ────────────────────────────────

@app_bp.route("/admin/api/sub-tenants/<int:sub_tenant_id>/admin/reset-password", methods=["POST"])
@csrf.exempt
@login_required
def admin_api_sub_tenant_admin_reset_password(sub_tenant_id):
    if not can_manage_sub_tenants(current_user):
        return _403()

    sub = Tenant.query.get_or_404(sub_tenant_id)
    parent = sub.parent_tenant
    if not parent:
        return jsonify({"success": False, "message": "Bu kurum bir alt kurum değil."}), 400

    # Yetki
    if not is_platform_admin(current_user) and parent.id != current_user.tenant_id:
        return _403()

    # İlk admin'i bul
    tenant_admin_role = Role.query.filter_by(name="tenant_admin").first()
    admin = (
        User.query
        .filter(User.tenant_id == sub.id, User.role_id == tenant_admin_role.id, User.is_active == True)
        .order_by(User.id)
        .first()
    )
    if not admin:
        return jsonify({"success": False, "message": "Bu alt kurumun aktif yöneticisi yok."}), 404

    import secrets
    new_password = "Kp_" + secrets.token_urlsafe(8)
    admin.password_hash = generate_password_hash(new_password)
    db.session.commit()

    try:
        from app.utils.audit_logger import AuditLogger
        AuditLogger.log(
            action="SUB_TENANT_ADMIN_PASSWORD_RESET", resource_type="GÜVENLİK",
            resource_id=admin.id,
            description=f"{current_user.email} {sub.name} alt-tenant'ının admin parolasını sıfırladı",
        )
    except Exception:
        pass

    return jsonify({
        "success": True,
        "message": f"{admin.email} kullanıcısının yeni şifresi üretildi.",
        "new_password": new_password,
        "admin_email": admin.email,
    })


# ─── Alt-tenant aktif/pasif ──────────────────────────────────────────────────

@app_bp.route("/admin/api/sub-tenants/<int:sub_tenant_id>/toggle", methods=["POST"])
@csrf.exempt
@login_required
def admin_api_sub_tenant_toggle(sub_tenant_id):
    if not can_manage_sub_tenants(current_user):
        return _403()

    sub = Tenant.query.get_or_404(sub_tenant_id)
    if not sub.parent_tenant:
        return jsonify({"success": False, "message": "Bu kurum bir alt kurum değil."}), 400
    if not is_platform_admin(current_user) and sub.parent_tenant_id != current_user.tenant_id:
        return _403()

    sub.is_active = not sub.is_active
    db.session.commit()

    try:
        from app.utils.audit_logger import AuditLogger
        AuditLogger.log(
            action="SUB_TENANT_TOGGLE", resource_type="Kurum Yönetimi",
            resource_id=sub.id,
            description=f"{current_user.email} {sub.name} alt-tenant'ını {'aktifleştirdi' if sub.is_active else 'pasifleştirdi'}",
        )
    except Exception:
        pass

    return jsonify({
        "success": True,
        "is_active": sub.is_active,
        "message": f"{sub.name} {'aktifleştirildi' if sub.is_active else 'pasifleştirildi'}.",
    })


# ─── Sprint K: Konsolide Kullanım Raporu ────────────────────────────────────

@app_bp.route("/admin/sub-tenants/usage")
@login_required
def admin_sub_tenants_usage_page():
    if not can_manage_sub_tenants(current_user):
        return render_template("platform/errors/403.html"), 403
    parent = None
    if is_platform_admin(current_user):
        parent_id = request.args.get("parent_id", type=int)
        if parent_id:
            parent = Tenant.query.get(parent_id)
    else:
        parent = current_user.tenant
    return render_template(
        "platform/admin/sub_tenants_usage.html",
        parent=parent,
        is_platform_admin=is_platform_admin(current_user),
    )


@app_bp.route("/admin/api/sub-tenants/usage")
@login_required
def admin_api_sub_tenants_usage():
    if not can_manage_sub_tenants(current_user):
        return _403()
    if is_platform_admin(current_user):
        parent_id = request.args.get("parent_id", type=int)
        if not parent_id:
            return jsonify({"success": False, "message": "Platform Admin için parent_id gerekli."}), 400
        parent = Tenant.query.get_or_404(parent_id)
    else:
        parent = current_user.tenant
    if parent.tenant_type not in ("dealer", "holding"):
        return jsonify({"success": False, "message": "Bu kurum tipi alt kurum yönetimine sahip değil."}), 400

    from app.services.sub_tenant_billing_service import build_consolidated_usage
    try:
        data = build_consolidated_usage(parent.id)
        if "error" in data:
            return jsonify({"success": False, "message": data["error"]}), 400
        return jsonify({"success": True, **data})
    except Exception as e:
        current_app.logger.error(f"[sub_tenants_usage] {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500


@app_bp.route("/admin/api/sub-tenants/usage/export.xlsx")
@login_required
def admin_api_sub_tenants_usage_export():
    """Excel dışa aktarma."""
    if not can_manage_sub_tenants(current_user):
        return _403()
    if is_platform_admin(current_user):
        parent_id = request.args.get("parent_id", type=int)
        if not parent_id:
            return jsonify({"success": False, "message": "parent_id gerekli."}), 400
        parent = Tenant.query.get_or_404(parent_id)
    else:
        parent = current_user.tenant
    if parent.tenant_type not in ("dealer", "holding"):
        return jsonify({"success": False, "message": "Bu kurum tipi alt kurum yönetimine sahip değil."}), 400

    from app.services.sub_tenant_billing_service import build_consolidated_usage
    data = build_consolidated_usage(parent.id)
    if "error" in data:
        return jsonify({"success": False, "message": data["error"]}), 400

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        return jsonify({"success": False, "message": "openpyxl kütüphanesi kurulu değil."}), 500
    from flask import send_file
    import datetime as _dt

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alt Tenant Kullanım"

    # Başlık satırı
    headers = [
        "Tenant Adı", "Kısa Ad", "Durum", "Oluşturulma",
        "Lisans Bitiş", "Paket",
        "Kullanıcı", "Maks. Kullanıcı", "Doluluk %",
        "KPI Sayısı", "Initiative Sayısı", "Plan Year Sayısı",
        "LLM Çağrı (bu ay)", "LLM Token (bu ay)", "LLM Maliyet USD (bu ay)",
    ]
    bold = Font(bold=True, color="FFFFFF")
    bg = PatternFill("solid", fgColor="4F46E5")
    center = Alignment(horizontal="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.fill = bg
        cell.alignment = center

    for row_idx, t in enumerate(data["sub_tenants"], 2):
        ws.cell(row=row_idx, column=1, value=t["name"])
        ws.cell(row=row_idx, column=2, value=t["short_name"] or "")
        ws.cell(row=row_idx, column=3, value="Aktif" if t["is_active"] else "Pasif")
        ws.cell(row=row_idx, column=4, value=t["created_at"][:10] if t["created_at"] else "")
        ws.cell(row=row_idx, column=5, value=t["license_end_date"] or "")
        ws.cell(row=row_idx, column=6, value=t["package"]["name"] if t["package"] else "—")
        ws.cell(row=row_idx, column=7, value=t["users_count"])
        ws.cell(row=row_idx, column=8, value=t["max_user_count"] or "—")
        ws.cell(row=row_idx, column=9, value=f"%{t['user_utilization_pct']}" if t["user_utilization_pct"] is not None else "—")
        ws.cell(row=row_idx, column=10, value=t["kpi_count"])
        ws.cell(row=row_idx, column=11, value=t["initiative_count"])
        ws.cell(row=row_idx, column=12, value=t["plan_year_count"])
        ws.cell(row=row_idx, column=13, value=t["llm_calls_this_month"])
        ws.cell(row=row_idx, column=14, value=t["llm_tokens_this_month"])
        ws.cell(row=row_idx, column=15, value=t["llm_cost_usd_this_month"])

    # Toplam satırı
    last_row = len(data["sub_tenants"]) + 2
    tot = data["totals"]
    tot_bg = PatternFill("solid", fgColor="EEF2FF")
    tot_font = Font(bold=True)
    ws.cell(row=last_row, column=1, value="TOPLAM").font = tot_font
    for c in range(1, 16):
        ws.cell(row=last_row, column=c).fill = tot_bg
    ws.cell(row=last_row, column=7, value=tot["users_total"]).font = tot_font
    ws.cell(row=last_row, column=8, value=tot["max_users_total"] or "—").font = tot_font
    ws.cell(row=last_row, column=10, value=tot["kpi_total"]).font = tot_font
    ws.cell(row=last_row, column=11, value=tot["initiative_total"]).font = tot_font
    ws.cell(row=last_row, column=13, value=tot["llm_calls_month_total"]).font = tot_font
    ws.cell(row=last_row, column=14, value=tot["llm_tokens_month_total"]).font = tot_font
    ws.cell(row=last_row, column=15, value=tot["llm_cost_usd_month_total"]).font = tot_font

    # Sütun genişliği
    widths = [28, 14, 10, 14, 14, 18, 10, 12, 12, 12, 16, 14, 16, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"alt-tenant-kullanim-{parent.short_name or 'parent'}-{_dt.date.today().isoformat()}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )
