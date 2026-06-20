"""Admin modülü — kullanıcı, kurum, paket, bileşen yönetimi."""

import datetime
import mimetypes
import os
import tempfile

from flask import render_template, jsonify, request, current_app, send_file, flash, redirect, url_for, abort
from flask_login import login_required, current_user
from sqlalchemy import func, and_, or_, text
from sqlalchemy.orm import selectinload
from werkzeug.security import generate_password_hash, check_password_hash

from platform_core import app_bp
from app.models import db
from app.models.core import User, Role, Tenant
from app.models.audit import AuditLog
from app.utils.audit_logger import AuditLogger
from app.utils.db_sequence import is_pk_duplicate, sync_pg_sequence_if_needed
from micro.modules.admin.constants import AKTIVITE_ETIKETLER, RESOURCE_IKONLAR

# _is_admin() = SADECE platform Admin (tüm kurumlar). tenant_admin tek-kurum
# yöneticisidir; cross-tenant kapıları açmaması için PLATFORM_ADMIN_ROLES kullanılır.
# (ADMIN_ROLES = {"Admin","tenant_admin"} idi → tenant_admin'e platform yetkisi sızdırıyordu.)
from app.constants.roles import PLATFORM_ADMIN_ROLES as _ADMIN_ROLES_SET, PRIVILEGED_ROLES as _PRIVILEGED_ROLES_SET
_ADMIN_ROLES   = tuple(_ADMIN_ROLES_SET)
_MANAGER_ROLES = tuple(_PRIVILEGED_ROLES_SET)


# Hangi rol, hangi rolleri atayabilir
ASSIGNABLE_ROLES = {
    "Admin":             ["Admin", "User", "tenant_admin", "executive_manager", "standard_user"],
    "tenant_admin":      ["executive_manager", "standard_user"],
    "executive_manager": ["executive_manager", "standard_user"],
}


def _is_admin():
    return current_user.role and current_user.role.name in _ADMIN_ROLES


def _is_manager():
    return current_user.role and current_user.role.name in _MANAGER_ROLES


def _403():
    return jsonify({"success": False, "message": "Bu işlem için yetkiniz yok."}), 403


_TENANT_LOGO_EXT = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"}
_TENANT_LOGO_MAX_BYTES = 2 * 1024 * 1024


def _tenant_logos_dir():
    d = os.path.join(current_app.instance_path, "uploads", "tenant_logos")
    os.makedirs(d, exist_ok=True)
    return d


@app_bp.route("/tenant-logo/<int:tenant_id>")
@login_required
def tenant_logo(tenant_id):
    """Oturum açan kullanıcı yalnızca kendi kurumunun logosunu veya Admin tüm kurumları görebilir."""
    if not _is_admin() and getattr(current_user, "tenant_id", None) != tenant_id:
        abort(403)
    t = Tenant.query.get_or_404(tenant_id)
    if not t.logo_path:
        abort(404)
    folder = _tenant_logos_dir()
    path = os.path.join(folder, t.logo_path)
    if not os.path.isfile(path):
        abort(404)
    mt, _ = mimetypes.guess_type(path)
    return send_file(path, mimetype=mt or "application/octet-stream", max_age=86400, conditional=True)






def _is_admin_or_tenant_admin():
    role_name = current_user.role.name if current_user.role else ""
    return role_name in ("Admin", "tenant_admin")


def get_tenant_list():
    """
    Aktif tenant listesi.
    - Admin: tüm aktif tenantlar
    - tenant_admin: sadece kendi tenantı
    """
    if not _is_admin_or_tenant_admin():
        return []
    if _is_admin():
        rows = Tenant.query.filter_by(is_active=True).order_by(Tenant.name).all()
        return [{"id": t.id, "name": t.name} for t in rows]
    if current_user.tenant_id:
        t = Tenant.query.filter_by(id=current_user.tenant_id, is_active=True).first()
        return [{"id": t.id, "name": t.name}] if t else []
    return []


def get_login_stats(tenant_id=None):
    """
    Login istatistikleri.
    Döner: online_now, active_now, last_24h, last_7d, last_30d, all_time
    """
    now_utc = datetime.datetime.now(datetime.timezone.utc)

    login_predicate = or_(
        AuditLog.action.in_(("OTURUM AÇMA", "LOGIN")),
        func.upper(AuditLog.action).like("OTURUM A%MA%"),
        func.upper(AuditLog.action).like("%LOGIN%"),
    )
    logout_predicate = or_(
        AuditLog.action.in_(("OTURUM KAPATMA", "LOGOUT")),
        func.upper(AuditLog.action).like("OTURUM KAPATMA%"),
        func.upper(AuditLog.action).like("%LOGOUT%"),
    )

    def _tenant_filter(query):
        if tenant_id is not None:
            tenant_user_ids_sq = db.session.query(User.id).filter(User.tenant_id == tenant_id)
            return query.filter(
                or_(
                    AuditLog.tenant_id == tenant_id,
                    and_(AuditLog.tenant_id.is_(None), AuditLog.user_id.in_(tenant_user_ids_sq)),
                )
            )
        return query

    def _unique_login_count(since_dt=None):
        q = db.session.query(func.count(func.distinct(AuditLog.user_id))).filter(
            login_predicate, AuditLog.user_id.isnot(None),
        )
        q = _tenant_filter(q)
        if since_dt is not None:
            q = q.filter(AuditLog.created_at >= since_dt)
        return int(q.scalar() or 0)

    def _active_user_count():
        q = db.session.query(func.count(User.id)).filter(User.is_active.is_(True))
        if tenant_id is not None:
            q = q.filter(User.tenant_id == tenant_id)
        return int(q.scalar() or 0)

    active_cutoff = now_utc - datetime.timedelta(minutes=30)

    login_q = db.session.query(
        AuditLog.user_id.label("user_id"),
        func.max(AuditLog.created_at).label("last_login"),
    ).filter(login_predicate, AuditLog.user_id.isnot(None))
    login_q = _tenant_filter(login_q).group_by(AuditLog.user_id)
    login_sq = login_q.subquery()

    logout_q = db.session.query(
        AuditLog.user_id.label("user_id"),
        func.max(AuditLog.created_at).label("logout_at"),
    ).filter(logout_predicate, AuditLog.user_id.isnot(None))
    logout_q = _tenant_filter(logout_q).group_by(AuditLog.user_id)
    logout_sq = logout_q.subquery()

    online_now = db.session.query(func.count()).select_from(login_sq).outerjoin(
        logout_sq,
        and_(
            logout_sq.c.user_id == login_sq.c.user_id,
            logout_sq.c.logout_at > login_sq.c.last_login,
        ),
    ).filter(
        login_sq.c.last_login >= active_cutoff,
        logout_sq.c.user_id.is_(None),
    ).scalar() or 0

    return {
        "online_now": int(online_now),
        "active_now": _active_user_count(),
        "last_24h": _unique_login_count(now_utc - datetime.timedelta(hours=24)),
        "last_7d": _unique_login_count(now_utc - datetime.timedelta(days=7)),
        "last_30d": _unique_login_count(now_utc - datetime.timedelta(days=30)),
        "all_time": _unique_login_count(),
    }


def get_user_activity_stats(tenant_id=None):
    """
    Kullanıcı bazlı giriş ve aktivite istatistikleri.
    Her kullanıcı için: hesap durumu, çevrimiçi mi, son giriş, 30 günlük giriş/işlem sayısı.
    """
    now_utc = datetime.datetime.now(datetime.timezone.utc)
    # AuditLog.created_at naive UTC saklanıyor → Python karşılaştırması için naive cutoff
    # (aksi halde "can't compare offset-naive and offset-aware datetimes" TypeError'ı)
    active_cutoff = (now_utc - datetime.timedelta(minutes=30)).replace(tzinfo=None)
    since_30d = (now_utc - datetime.timedelta(days=30)).replace(tzinfo=None)

    login_predicate = or_(
        AuditLog.action.in_(("OTURUM AÇMA", "LOGIN")),
        func.upper(AuditLog.action).like("OTURUM A%MA%"),
        func.upper(AuditLog.action).like("%LOGIN%"),
    )
    logout_predicate = or_(
        AuditLog.action.in_(("OTURUM KAPATMA", "LOGOUT")),
        func.upper(AuditLog.action).like("OTURUM KAPATMA%"),
        func.upper(AuditLog.action).like("%LOGOUT%"),
    )

    user_q = User.query.options(selectinload(User.role))  # N+1 önlemi: rol eager-load
    if tenant_id is not None:
        user_q = user_q.filter(User.tenant_id == tenant_id)
    users = user_q.order_by(User.last_name, User.first_name).all()
    if not users:
        return []

    user_ids = [u.id for u in users]

    last_login_rows = (
        db.session.query(AuditLog.user_id, func.max(AuditLog.created_at).label("last_login"))
        .filter(login_predicate, AuditLog.user_id.in_(user_ids))
        .group_by(AuditLog.user_id)
        .all()
    )
    last_login_map = {r.user_id: r.last_login for r in last_login_rows}

    last_logout_rows = (
        db.session.query(AuditLog.user_id, func.max(AuditLog.created_at).label("last_logout"))
        .filter(logout_predicate, AuditLog.user_id.in_(user_ids))
        .group_by(AuditLog.user_id)
        .all()
    )
    last_logout_map = {r.user_id: r.last_logout for r in last_logout_rows}

    login_30d_rows = (
        db.session.query(AuditLog.user_id, func.count().label("cnt"))
        .filter(login_predicate, AuditLog.user_id.in_(user_ids), AuditLog.created_at >= since_30d)
        .group_by(AuditLog.user_id)
        .all()
    )
    login_30d_map = {r.user_id: r.cnt for r in login_30d_rows}

    actions_30d_rows = (
        db.session.query(AuditLog.user_id, func.count().label("cnt"))
        .filter(
            AuditLog.user_id.in_(user_ids),
            AuditLog.created_at >= since_30d,
            AuditLog.resource_type != "GÜVENLİK",
        )
        .group_by(AuditLog.user_id)
        .all()
    )
    actions_30d_map = {r.user_id: r.cnt for r in actions_30d_rows}

    result = []
    for u in users:
        last_login = last_login_map.get(u.id)
        last_logout = last_logout_map.get(u.id)
        is_online = bool(
            last_login
            and last_login >= active_cutoff
            and (last_logout is None or last_logout < last_login)
        )
        name = f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip()
        result.append({
            "id": u.id,
            "name": name or u.email,
            "email": u.email,
            "role": u.role.name if u.role else None,
            "is_active": u.is_active,
            "is_online": is_online,
            "last_login_at": last_login.isoformat() + "Z" if last_login else None,
            "login_count_30d": login_30d_map.get(u.id, 0),
            "actions_30d": actions_30d_map.get(u.id, 0),
        })

    # Sıralama: çevrimiçi → aktif → son giriş tarihi desc
    result.sort(key=lambda x: x["last_login_at"] or "", reverse=True)
    result.sort(key=lambda x: not x["is_active"])
    result.sort(key=lambda x: not x["is_online"])
    return result


@app_bp.route("/admin/bakim-modu", methods=["GET", "POST"])
@login_required
def admin_bakim_modu():
    """Bakım modu (yalnız platform Admin). Ortam değişkeni kilitliyse yalnız okuma."""
    if not _is_admin():
        return jsonify({"success": False, "message": "Bu işlem için yetkiniz yok."}), 403
    from flask_wtf.csrf import validate_csrf
    from wtforms import ValidationError
    from app.services.maintenance_service import maintenance_status_for_admin, set_maintenance_db

    if request.method == "GET":
        return jsonify({"success": True, "data": maintenance_status_for_admin(current_app)})

    try:
        validate_csrf(request.headers.get("X-CSRFToken"))
    except ValidationError:
        return jsonify({"success": False, "message": "CSRF doğrulaması başarısız."}), 400

    st = maintenance_status_for_admin(current_app)
    if not st.get("can_toggle_db"):
        return jsonify({
            "success": False,
            "message": "Bakım modu ortam değişkeni (MAINTENANCE_MODE / MAINTENANCE_OVERRIDE_OFF) ile kilitli; "
            "panelden değiştirilemez. Sunucu dokümantasyonuna bakın.",
        }), 400
    data = request.get_json() or {}
    enabled = bool(data.get("enabled"))
    try:
        set_maintenance_db(enabled)
        try:
            AuditLogger.log_create(
                "Sistem ayarı",
                0,
                {"maintenance_mode": enabled},
                description="Bakım modu (system_settings)",
            )
        except Exception as log_err:
            current_app.logger.warning(f"[admin_bakim_modu] audit: {log_err}")
        return jsonify({"success": True, "data": maintenance_status_for_admin(current_app)})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_bakim_modu] {e}")
        return jsonify({"success": False, "message": "Kaydedilemedi."}), 500


@app_bp.route("/admin/yonetim-paneli")
@login_required
def yonetim_paneli():
    if not _is_admin_or_tenant_admin():
        return render_template("platform/errors/403.html"), 403
    try:
        tenant_list = get_tenant_list()
        default_tenant_id = current_user.tenant_id if not _is_admin() else None
        return render_template(
            "platform/admin/yonetim_paneli.html",
            tenant_list=tenant_list,
            default_tenant_id=default_tenant_id,
        )
    except Exception as e:
        current_app.logger.error(f"[yonetim_paneli] {e}")
        return render_template("platform/errors/500.html"), 500


@app_bp.route("/admin/yonetim-paneli/istatistik")
@login_required
def yonetim_paneli_istatistik():
    if not _is_admin_or_tenant_admin():
        return jsonify({"success": False, "message": "Bu işlem için yetkiniz yok."}), 403
    try:
        req_tenant_id = request.args.get("tenant_id", type=int)
        tenant_id = current_user.tenant_id if not _is_admin() else req_tenant_id
        data = get_login_stats(tenant_id=tenant_id)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        current_app.logger.error(f"[yonetim_paneli_istatistik] {e}")
        return jsonify({"success": False, "message": "İstatistik alınamadı."}), 500


@app_bp.route("/admin/yonetim-paneli/kullanici-detay")
@login_required
def yonetim_paneli_kullanici_detay():
    if not _is_admin_or_tenant_admin():
        return jsonify({"success": False, "message": "Bu işlem için yetkiniz yok."}), 403
    try:
        req_tenant_id = request.args.get("tenant_id", type=int)
        tenant_id = current_user.tenant_id if not _is_admin() else req_tenant_id
        if not tenant_id:
            return jsonify({"success": False, "message": "tenant_id parametresi gerekli."}), 400
        data = get_user_activity_stats(tenant_id=tenant_id)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        current_app.logger.error(f"[yonetim_paneli_kullanici_detay] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Kullanıcı istatistikleri alınamadı."}), 500


@app_bp.route("/admin/yonetim-paneli/aktiviteler")
@login_required
def yonetim_paneli_aktiviteler():
    if not _is_admin_or_tenant_admin():
        return jsonify({"success": False, "message": "Bu işlem için yetkiniz yok."}), 403
    try:
        req_tenant_id = request.args.get("tenant_id", type=int)
        limit = request.args.get("limit", type=int) or 50
        limit = max(1, min(limit, 200))

        tenant_id = current_user.tenant_id if not _is_admin() else req_tenant_id

        q = (
            db.session.query(
                AuditLog.id,
                AuditLog.action,
                AuditLog.resource_type,
                AuditLog.resource_id,
                AuditLog.created_at,
                AuditLog.ip_address,
                User.first_name,
                User.last_name,
                User.email,
            )
            .outerjoin(User, AuditLog.user_id == User.id)
            .filter(AuditLog.resource_type != "GÜVENLİK")
            .order_by(AuditLog.created_at.desc())
        )
        if tenant_id is not None:
            q = q.filter(AuditLog.tenant_id == tenant_id)

        rows = q.limit(limit).all()
        data = []
        for r in rows:
            full_name = f"{(r.first_name or '').strip()} {(r.last_name or '').strip()}".strip()
            data.append(
                {
                    "id": r.id,
                    "action": r.action,
                    "action_label": AKTIVITE_ETIKETLER.get(r.action, r.action),
                    "resource_type": r.resource_type,
                    "resource_icon": RESOURCE_IKONLAR.get(r.resource_type, "📌"),
                    "resource_id": r.resource_id,
                    "created_at": r.created_at.isoformat() + "Z" if r.created_at else None,
                    "ip_address": r.ip_address,
                    "user_name": full_name or (r.email or "Bilinmiyor"),
                    "user_email": r.email,
                }
            )
        return jsonify({"success": True, "data": data})
    except Exception as e:
        current_app.logger.error(f"[yonetim_paneli_aktiviteler] {e}")
        return jsonify({"success": False, "message": "Aktivite kayıtları alınamadı."}), 500




# ── Kullanıcı Yönetimi ────────────────────────────────────────────────────────

@app_bp.route("/admin/users")
@login_required
def admin_users():
    if not _is_manager():
        return render_template("platform/errors/403.html"), 403

    if _is_admin():
        users = User.query.options(selectinload(User.role), selectinload(User.tenant)).order_by(User.tenant_id, User.first_name).all()
    else:
        users = User.query.options(selectinload(User.role), selectinload(User.tenant)).filter_by(tenant_id=current_user.tenant_id).order_by(User.first_name).all()

    roles   = Role.query.filter(Role.name.in_(ASSIGNABLE_ROLES.get(current_user.role.name if current_user.role else "", []))).all()
    tenants = Tenant.query.filter_by(is_active=True).order_by(Tenant.name).all() if _is_admin() else []
    return render_template("platform/admin/users.html", users=users, roles=roles, tenants=tenants)


@app_bp.route("/admin/api/users")
@login_required
def admin_api_users_paginated():
    """Sprint 24: Paginated admin user list — Tomofil gibi 100+ user tenant'larda performans.

    Query params: ?page=1&page_size=50&search=ahmet
    """
    if not _is_manager():
        return jsonify({"success": False, "message": "Yetkisiz"}), 403

    from app.utils.pagination import paginate_query
    q = User.query.options(selectinload(User.role))  # N+1 önlemi: rol eager-load
    if not _is_admin():
        q = q.filter_by(tenant_id=current_user.tenant_id)

    search = (request.args.get("search") or "").strip()
    if search:
        like = f"%{search}%"
        q = q.filter(
            db.or_(
                User.email.ilike(like),
                User.first_name.ilike(like),
                User.last_name.ilike(like),
                User.job_title.ilike(like),
            )
        )

    q = q.order_by(User.tenant_id, User.first_name)

    def _ser(u):
        return {
            "id": u.id,
            "email": u.email,
            "first_name": u.first_name,
            "last_name": u.last_name,
            "is_active": u.is_active,
            "tenant_id": u.tenant_id,
            "role_id": u.role_id,
            "role_name": u.role.name if u.role else None,
            "job_title": u.job_title,
            "department": u.department,
        }

    return jsonify(paginate_query(q, default_page_size=50, max_page_size=200, serializer=_ser))


@app_bp.route("/admin/users/add", methods=["POST"])
@login_required
def admin_users_add():
    if not _is_manager():
        return _403()

    data  = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    if not email:
        return jsonify({"success": False, "message": "E-posta zorunludur."}), 400

    if User.query.filter_by(email=email).first():
        return jsonify({"success": False, "message": "Bu e-posta zaten kayıtlı."}), 400

    # Rol atama yetki kontrolü
    role = Role.query.get(data.get("role_id"))
    if role:
        allowed = ASSIGNABLE_ROLES.get(current_user.role.name if current_user.role else "", [])
        if role.name not in allowed:
            return jsonify({"success": False, "message": "Bu rolü atama yetkiniz yok."}), 403

    # tenant_admin tekil kontrolü
    tid  = int(data.get("tenant_id") or current_user.tenant_id or 0)
    if role and role.name == "tenant_admin":
        existing = User.query.join(Role).filter(
            User.tenant_id == tid, Role.name == "tenant_admin", User.is_active.is_(True)
        ).first()
        if existing:
            return jsonify({"success": False, "message": "Bu kurumda zaten aktif bir Kurum Yöneticisi var."}), 400

    try:
        u = User(
            email=email,
            password_hash=generate_password_hash(data.get("password") or ("Kp_" + __import__("secrets").token_urlsafe(12))),
            first_name=data.get("first_name", "").strip(),
            last_name=data.get("last_name", "").strip(),
            tenant_id=tid or None,
            role_id=data.get("role_id"),
            job_title=data.get("job_title", "").strip() or None,
            department=data.get("department", "").strip() or None,
        )
        db.session.add(u)
        db.session.commit()
        try:
            AuditLogger.log_create(
                "Kullanıcı Yönetimi",
                u.id,
                {"email": u.email, "tenant_id": u.tenant_id, "role_id": u.role_id},
            )
        except Exception as e:
            current_app.logger.error(f"Audit log hatası: {e}")
        return jsonify({"success": True, "message": "Kullanıcı oluşturuldu.", "id": u.id})
    except Exception as e:
        db.session.rollback()
        if is_pk_duplicate(e, "users"):
            try:
                sync_pg_sequence_if_needed("users", "id")
                u = User(
                    email=email,
                    password_hash=generate_password_hash(data.get("password") or ("Kp_" + __import__("secrets").token_urlsafe(12))),
                    first_name=(data.get("first_name") or "").strip(),
                    last_name=(data.get("last_name") or "").strip(),
                    tenant_id=tid or None,
                    role_id=data.get("role_id"),
                    job_title=(data.get("job_title") or "").strip() or None,
                    department=(data.get("department") or "").strip() or None,
                )
                db.session.add(u)
                db.session.commit()
                try:
                    AuditLogger.log_create(
                        "Kullanıcı Yönetimi",
                        u.id,
                        {"email": u.email, "tenant_id": u.tenant_id, "role_id": u.role_id},
                    )
                except Exception as e:
                    current_app.logger.error(f"Audit log hatası: {e}")
                return jsonify({"success": True, "message": "Kullanıcı oluşturuldu.", "id": u.id})
            except Exception as e2:
                db.session.rollback()
                current_app.logger.error(f"[admin_users_add/retry] {e2}")
        current_app.logger.error(f"[admin_users_add] {e}")
        return jsonify({"success": False, "message": "Kayıt sırasında hata oluştu."}), 500


@app_bp.route("/admin/users/edit/<int:user_id>", methods=["POST"])
@login_required
def admin_users_edit(user_id):
    if not _is_manager():
        return _403()

    u = User.query.get_or_404(user_id)
    if not _is_admin() and u.tenant_id != current_user.tenant_id:
        return _403()
    if (
        not _is_admin()
        and u.role
        and u.role.name == "tenant_admin"
    ):
        return jsonify({"success": False, "message": "Kurum Yöneticisi hesabını sadece Admin düzenleyebilir."}), 403

    data = request.get_json() or {}
    try:
        u.first_name  = data.get("first_name", u.first_name or "").strip()
        u.last_name   = data.get("last_name",  u.last_name  or "").strip()
        u.job_title   = data.get("job_title",  u.job_title)
        u.department  = data.get("department", u.department)
        if data.get("role_id"):
            new_role = Role.query.get(int(data["role_id"]))
            if new_role:
                allowed = ASSIGNABLE_ROLES.get(current_user.role.name if current_user.role else "", [])
                if new_role.name not in allowed:
                    return jsonify({"success": False, "message": "Bu rolü atama yetkiniz yok."}), 403
            u.role_id = int(data["role_id"])
        if data.get("password"):
            u.password_hash = generate_password_hash(data["password"])
        db.session.commit()
        try:
            AuditLogger.log_update(
                "Kullanıcı Yönetimi",
                u.id,
                {},
                {
                    "first_name": u.first_name,
                    "last_name": u.last_name,
                    "job_title": u.job_title,
                    "department": u.department,
                    "role_id": u.role_id,
                },
            )
        except Exception as e:
            current_app.logger.error(f"Audit log hatası: {e}")
        return jsonify({"success": True, "message": "Kullanıcı güncellendi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_users_edit] {e}")
        return jsonify({"success": False, "message": "Güncelleme sırasında hata oluştu."}), 500


@app_bp.route("/admin/users/toggle/<int:user_id>", methods=["POST"])
@login_required
def admin_users_toggle(user_id):
    if not _is_manager():
        return _403()

    u = User.query.get_or_404(user_id)
    if not _is_admin() and u.tenant_id != current_user.tenant_id:
        return _403()
    if (
        not _is_admin()
        and u.role
        and u.role.name == "tenant_admin"
    ):
        return jsonify({"success": False, "message": "Kurum Yöneticisi hesabını sadece Admin pasife/aktife alabilir."}), 403
    if u.id == current_user.id:
        return jsonify({"success": False, "message": "Kendi hesabınızı pasife alamazsınız."}), 400

    try:
        u.is_active = not u.is_active
        db.session.commit()
        state = "aktifleştirildi" if u.is_active else "pasife alındı"
        return jsonify({"success": True, "message": f"Kullanıcı {state}.", "is_active": u.is_active})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_users_toggle] {e}")
        return jsonify({"success": False, "message": "İşlem sırasında hata oluştu."}), 500


# ── 2FA Sıfırlama (Admin) ────────────────────────────────────────────────────

@app_bp.route("/admin/users/<int:user_id>/2fa-reset", methods=["POST"])
@login_required
def admin_users_2fa_reset(user_id):
    """Bir kullanıcının 2FA'sını admin olarak sıfırla.

    Senaryo: kullanıcı telefonunu kaybetti, backup kodlarını da bulamıyor.
    Admin onun 2FA'sını sıfırlayarak hesabına erişimi geri kazandırır.
    """
    if not _is_manager():
        return _403()

    u = User.query.get_or_404(user_id)
    # Tenant admin yalnızca kendi tenant'ında
    if not _is_admin() and u.tenant_id != current_user.tenant_id:
        return _403()
    # Tenant admin başka bir tenant_admin'in 2FA'sını sıfırlayamaz
    if (
        not _is_admin()
        and u.role
        and u.role.name == "tenant_admin"
        and u.id != current_user.id
    ):
        return jsonify({"success": False, "message": "Kurum Yöneticisi 2FA'sını sadece Admin sıfırlayabilir."}), 403

    if not u.totp_enabled:
        return jsonify({"success": False, "message": "Kullanıcının 2FA'sı zaten devre dışı."}), 400

    try:
        u.totp_enabled = False
        u.totp_secret = None
        u.totp_backup_codes_json = None
        db.session.commit()

        # Audit log
        try:
            from app.utils.audit_logger import AuditLogger
            AuditLogger.log(
                action="2FA_ADMIN_RESET", resource_type="GÜVENLİK",
                resource_id=u.id,
                description=f"Admin {current_user.email} tarafından {u.email} 2FA sıfırlandı",
            )
        except Exception as _audit_err:
            current_app.logger.error("[audit] 2FA reset audit kaydı başarısız: %s", _audit_err)

        return jsonify({
            "success": True,
            "message": f"{u.email} kullanıcısının 2FA'sı sıfırlandı.",
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_users_2fa_reset] {e}")
        return jsonify({"success": False, "message": "Sıfırlama sırasında hata oluştu."}), 500


# ── Toplu Kullanıcı İçe Aktarma ───────────────────────────────────────────────

@app_bp.route("/admin/users/bulk-import", methods=["POST"])
@login_required
def admin_users_bulk_import():
    if not _is_manager():
        return _403()

    file = request.files.get("file")
    if not file:
        return jsonify({"success": False, "message": "Dosya seçilmedi."}), 400

    if request.content_length and request.content_length > 10 * 1024 * 1024:
        return jsonify({"success": False, "message": "Dosya boyutu 10 MB sınırını aşıyor."}), 400

    filename = (file.filename or "").lower()
    try:
        import io
        standard_role = Role.query.filter_by(name="standard_user").first()
        tid = current_user.tenant_id
        rows = []

        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file.stream.read()), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for ws_row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in ws_row])))
        else:
            # CSV fallback
            import csv
            stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
            rows = list(csv.DictReader(stream))

        # Mevcut e-postaları tek sorguda topla (N+1 önlemi)
        _candidate_emails = [
            (r.get("E-posta") or r.get("email") or "").strip().lower()
            for r in rows
        ]
        _candidate_emails = [e for e in _candidate_emails if e]
        _existing = set(e[0] for e in db.session.query(User.email).filter(
            User.email.in_(_candidate_emails)
        ).all()) if _candidate_emails else set()

        created, skipped = 0, 0
        for row in rows:
            email = (row.get("E-posta") or row.get("email") or "").strip().lower()
            if not email:
                skipped += 1
                continue
            if email in _existing:
                skipped += 1
                continue
            _existing.add(email)  # aynı dosyada tekrarı engelle
            raw_pass = (row.get("Sifre") or row.get("Şifre") or row.get("password") or "").strip()
            password = raw_pass if (raw_pass and len(raw_pass) >= 8) else ("Kp_" + __import__("secrets").token_urlsafe(12))
            u = User(
                email=email,
                password_hash=generate_password_hash(password),
                first_name=(row.get("Ad") or row.get("first_name") or "").strip(),
                last_name=(row.get("Soyad") or row.get("last_name") or "").strip(),
                job_title=(row.get("Unvan") or row.get("job_title") or "").strip() or None,
                phone_number=(row.get("Telefon") or row.get("phone_number") or "").strip() or None,
                tenant_id=tid,
                role_id=standard_role.id if standard_role else None,
            )
            db.session.add(u)
            created += 1

        db.session.commit()
        return jsonify({"success": True, "message": f"{created} kullanıcı oluşturuldu, {skipped} atlandı."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_users_bulk_import] {e}")
        return jsonify({"success": False, "message": "İçe aktarma sırasında hata oluştu."}), 500


@app_bp.route("/admin/users/sample-excel")
@login_required
def admin_users_sample_excel():
    """Toplu kullanıcı içe aktarma için örnek Excel dosyası indir."""
    import io
    import openpyxl
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kullanicilar"
    ws.append(["Ad", "Soyad", "E-posta", "Şifre", "Unvan", "Telefon"])
    ws.append(["Ahmet", "Yılmaz", "ahmet@ornek.com", "Gizli123!", "Mühendis", "5551234567"])
    ws.append(["Ayşe", "Kaya", "ayse@ornek.com", "Gizli123!", "Uzman", "5559876543"])
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="kullanici_sablonu.xlsx",
    )


# ── Kurum Yönetimi ────────────────────────────────────────────────────────────

@app_bp.route("/admin/tenants")
@login_required
def admin_tenants():
    if not _is_manager():
        return render_template("platform/errors/403.html"), 403

    if _is_admin():
        tenants = Tenant.query.order_by(Tenant.tenant_type.desc(), Tenant.name).all()
    else:
        tenants = Tenant.query.filter_by(id=current_user.tenant_id).all()

    try:
        from app.models.saas import SubscriptionPackage
        packages = SubscriptionPackage.query.filter_by(is_active=True).order_by(SubscriptionPackage.name).all()
    except Exception:
        packages = []

    # Bayi/Holding listesi — yeni tenant açarken parent seç için
    parent_candidates = []
    if _is_admin():
        parent_candidates = Tenant.query.filter(
            Tenant.tenant_type.in_(["dealer", "holding"]),
            Tenant.is_active.is_(True),
            Tenant.parent_tenant_id.is_(None),  # iç içe yasak
        ).order_by(Tenant.name).all()

    # M-19: N+1 önlemi — kullanıcı sayılarını tek sorguda topla
    _tenant_ids = [t.id for t in tenants]
    _user_counts = dict(
        db.session.query(User.tenant_id, func.count(User.id))
        .filter(User.tenant_id.in_(_tenant_ids), User.is_active == True)
        .group_by(User.tenant_id)
        .all()
    ) if _tenant_ids else {}
    total_users = sum(_user_counts.values())
    return render_template(
        "platform/admin/tenants.html",
        tenants=tenants, packages=packages,
        total_users=total_users,
        user_counts=_user_counts,
        parent_candidates=parent_candidates,
    )


@app_bp.route("/admin/tenants/add", methods=["POST"])
@login_required
def admin_tenants_add():
    if not _is_admin():
        return _403()

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"success": False, "message": "Kurum adı zorunludur."}), 400

    # Bayi/Holding hiyerarşi: tip + parent doğrulama
    from app.utils.tenant_scope import (
        validate_tenant_type, can_be_parent, check_sub_tenant_limit,
        TENANT_TYPE_NORMAL,
    )
    try:
        tenant_type = validate_tenant_type(data.get("tenant_type"))
    except ValueError as ve:
        return jsonify({"success": False, "message": str(ve)}), 400

    parent_id = data.get("parent_tenant_id")
    parent_tenant = None
    if parent_id:
        try:
            parent_id = int(parent_id)
        except (TypeError, ValueError):
            return jsonify({"success": False, "message": "Parent tenant ID geçersiz."}), 400
        parent_tenant = Tenant.query.get(parent_id)
        ok, err = can_be_parent(parent_tenant)
        if not ok:
            return jsonify({"success": False, "message": err}), 400
        # Alt tenant her zaman 'normal' tipi olur (iç içe hiyerarşi yasak)
        if tenant_type != TENANT_TYPE_NORMAL:
            return jsonify({"success": False, "message": "Alt-tenant tipi 'normal' olmalı (iç içe hiyerarşi yok)."}), 400
        ok, err = check_sub_tenant_limit(parent_tenant)
        if not ok:
            return jsonify({"success": False, "message": err}), 400

    try:
        def _int(val, default=None):
            """Sayısal dönüşüm — hatalı değerde ValueError yerine None/default döner."""
            try:
                return int(val) if val not in (None, "") else default
            except (TypeError, ValueError):
                raise ValueError(f"'{val}' geçerli bir sayı değil.")

        t = Tenant(
            name=name,
            short_name=(data.get("short_name") or "").strip() or None,
            sector=data.get("sector") or None,
            activity_area=data.get("activity_area") or None,
            employee_count=_int(data.get("employee_count")),
            contact_email=data.get("contact_email") or None,
            phone_number=data.get("phone_number") or None,
            website_url=data.get("website_url") or None,
            tax_office=data.get("tax_office") or None,
            tax_number=data.get("tax_number") or None,
            max_user_count=_int(data.get("max_user_count"), default=5),
            package_id=_int(data.get("package_id")),
            tenant_type=tenant_type,
            parent_tenant_id=parent_tenant.id if parent_tenant else None,
            sub_tenant_limit=_int(data.get("sub_tenant_limit")),
        )
        if data.get("license_end_date"):
            from datetime import date
            t.license_end_date = date.fromisoformat(data["license_end_date"])
        db.session.add(t)
        db.session.commit()
        try:
            AuditLogger.log_create(
                "Kurum Yönetimi",
                t.id,
                {"name": t.name, "short_name": t.short_name},
            )
        except Exception as e:
            current_app.logger.error(f"Audit log hatası: {e}")
        return jsonify({"success": True, "message": "Kurum oluşturuldu.", "id": t.id})
    except ValueError as e:
        return jsonify({"success": False, "message": f"Geçersiz sayısal alan: {e}"}), 400
    except Exception as e:
        db.session.rollback()
        if is_pk_duplicate(e, "tenants"):
            try:
                sync_pg_sequence_if_needed("tenants", "id")
                t = Tenant(
                    name=name,
                    short_name=(data.get("short_name") or "").strip() or None,
                    sector=data.get("sector") or None,
                    activity_area=data.get("activity_area") or None,
                    employee_count=int(data["employee_count"]) if data.get("employee_count") else None,
                    contact_email=data.get("contact_email") or None,
                    phone_number=data.get("phone_number") or None,
                    website_url=data.get("website_url") or None,
                    tax_office=data.get("tax_office") or None,
                    tax_number=data.get("tax_number") or None,
                    max_user_count=int(data["max_user_count"]) if data.get("max_user_count") else 5,
                    package_id=int(data["package_id"]) if data.get("package_id") else None,
                )
                if data.get("license_end_date"):
                    from datetime import date
                    t.license_end_date = date.fromisoformat(data["license_end_date"])
                db.session.add(t)
                db.session.commit()
                try:
                    AuditLogger.log_create(
                        "Kurum Yönetimi",
                        t.id,
                        {"name": t.name, "short_name": t.short_name},
                    )
                except Exception as e:
                    current_app.logger.error(f"Audit log hatası: {e}")
                return jsonify({"success": True, "message": "Kurum oluşturuldu.", "id": t.id})
            except Exception as e2:
                db.session.rollback()
                current_app.logger.error(f"[admin_tenants_add/retry] {e2}")
        current_app.logger.error(f"[admin_tenants_add] {e}")
        return jsonify({"success": False, "message": "Kayıt sırasında hata oluştu."}), 500


@app_bp.route("/admin/tenants/edit/<int:tenant_id>", methods=["POST"])
@login_required
def admin_tenants_edit(tenant_id):
    if not _is_manager():
        return _403()
    if not _is_admin() and tenant_id != current_user.tenant_id:
        return _403()

    t    = Tenant.query.get_or_404(tenant_id)
    data = request.get_json() or {}
    try:
        t.name            = (data.get("name") or t.name).strip()
        if "short_name" in data:
            t.short_name = ((data.get("short_name") or "").strip() or None)
        t.sector          = data.get("sector") or None
        t.activity_area   = data.get("activity_area") or None
        t.employee_count  = int(data["employee_count"]) if data.get("employee_count") else None
        t.contact_email   = data.get("contact_email") or None
        t.phone_number    = data.get("phone_number") or None
        t.website_url     = data.get("website_url") or None
        t.tax_office      = data.get("tax_office") or None
        t.tax_number      = data.get("tax_number") or None
        t.max_user_count  = int(data["max_user_count"]) if data.get("max_user_count") else t.max_user_count
        if data.get("license_end_date"):
            from datetime import date
            t.license_end_date = date.fromisoformat(data["license_end_date"])
        if data.get("package_id"):
            t.package_id = int(data["package_id"])

        # ── Bayi/Holding tipi değişimi (sadece Platform Admin) ──
        if "tenant_type" in data and _is_admin():
            from app.utils.tenant_scope import validate_tenant_type, TENANT_TYPE_NORMAL
            new_type = validate_tenant_type(data["tenant_type"])
            # Bayi/Holding → normal'e döndürme: aktif alt-tenant varsa engelle
            if t.tenant_type in ("dealer", "holding") and new_type == TENANT_TYPE_NORMAL:
                active_subs = Tenant.query.filter_by(parent_tenant_id=t.id, is_active=True).count()
                if active_subs > 0:
                    return jsonify({
                        "success": False,
                        "message": f"Aktif alt-tenant'ı olan kurum 'normal'a düşürülemez ({active_subs} alt-tenant). Önce alt-tenantları transfer/pasifleştir.",
                    }), 400
            if new_type != t.tenant_type:
                try:
                    AuditLogger.log(
                        action="TENANT_TYPE_CHANGE", resource_type="Kurum Yönetimi",
                        resource_id=t.id,
                        description=f"Tenant {t.name}: tip {t.tenant_type} → {new_type}",
                    )
                except Exception:
                    pass
                t.tenant_type = new_type

        if "sub_tenant_limit" in data and _is_admin():
            v = data["sub_tenant_limit"]
            t.sub_tenant_limit = int(v) if v else None

        db.session.commit()
        try:
            AuditLogger.log_update(
                "Kurum Yönetimi",
                t.id,
                {},
                {
                    "name": t.name,
                    "short_name": t.short_name,
                    "sector": t.sector,
                    "activity_area": t.activity_area,
                },
            )
        except Exception as e:
            current_app.logger.error(f"Audit log hatası: {e}")
        return jsonify({"success": True, "message": "Kurum güncellendi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_tenants_edit] {e}")
        return jsonify({"success": False, "message": "Güncelleme sırasında hata oluştu."}), 500


@app_bp.route("/admin/tenants/logo/<int:tenant_id>", methods=["POST"])
@login_required
def admin_tenants_logo(tenant_id):
    if not _is_manager():
        return _403()
    if not _is_admin() and tenant_id != current_user.tenant_id:
        return _403()

    t = Tenant.query.get_or_404(tenant_id)
    f = request.files.get("logo")
    if not f or not getattr(f, "filename", None):
        return jsonify({"success": False, "message": "Logo dosyası seçilmedi."}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in _TENANT_LOGO_EXT:
        return jsonify(
            {"success": False, "message": "Geçersiz dosya türü. PNG, JPG, WEBP, GIF veya SVG yükleyin."}
        ), 400
    blob = f.read()
    if len(blob) > _TENANT_LOGO_MAX_BYTES:
        return jsonify({"success": False, "message": "Dosya en fazla 2 MB olabilir."}), 400

    # Sprint 2.5: Magic byte doğrulaması — SVG XSS engeller
    from app.utils.upload_security import validate_uploaded_image, safe_filename
    ok, msg, detected_ext = validate_uploaded_image(
        blob, _TENANT_LOGO_EXT, accept_svg=(".svg" in _TENANT_LOGO_EXT)
    )
    if not ok:
        current_app.logger.warning(
            f"[logo_upload] reject tenant={tenant_id} user={current_user.id}: {msg}"
        )
        return jsonify({"success": False, "message": msg}), 400

    # Sprint 2.6: Path traversal koruma — fname tamamen sayı + sabit ext
    folder = _tenant_logos_dir()
    safe_ext = f".{detected_ext}"
    fname = safe_filename(f"{tenant_id}{safe_ext}", fallback=f"{tenant_id}.png")
    dest = os.path.join(folder, fname)
    # Çift kontrol: dest folder'ın içinde olmalı
    if not os.path.abspath(dest).startswith(os.path.abspath(folder)):
        current_app.logger.error(f"[logo_upload] path traversal blocked: {dest}")
        return jsonify({"success": False, "message": "Geçersiz dosya yolu."}), 400
    try:
        for old in os.listdir(folder):
            if old.startswith(f"{tenant_id}.") and old != fname:
                try:
                    os.remove(os.path.join(folder, old))
                except OSError:
                    pass
        with open(dest, "wb") as out:
            out.write(blob)
        t.logo_path = fname
        t.logo_updated_at = datetime.datetime.now(datetime.timezone.utc)
        db.session.commit()
        return jsonify({"success": True, "message": "Kurum logosu güncellendi.", "logo_path": fname})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_tenants_logo] {e}")
        return jsonify({"success": False, "message": "Logo yüklenirken hata oluştu."}), 500


@app_bp.route("/admin/tenants/toggle/<int:tenant_id>", methods=["POST"])
@login_required
def admin_tenants_toggle(tenant_id):
    if not _is_admin():
        return _403()

    t = Tenant.query.get_or_404(tenant_id)
    try:
        t.is_active = not t.is_active
        db.session.commit()
        state = "aktifleştirildi" if t.is_active else "arşivlendi"
        return jsonify({"success": True, "message": f"Kurum {state}.", "is_active": t.is_active})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_tenants_toggle] {e}")
        return jsonify({"success": False, "message": "İşlem sırasında hata oluştu."}), 500


# ── Paket Yönetimi ────────────────────────────────────────────────────────────

@app_bp.route("/admin/packages")
@login_required
def admin_packages():
    if not _is_admin():
        return render_template("platform/errors/403.html"), 403

    try:
        from app.models.saas import SubscriptionPackage, SystemModule
        packages = SubscriptionPackage.query.order_by(SubscriptionPackage.name).all()
        modules  = SystemModule.query.order_by(SystemModule.name).all()
    except Exception:
        packages, modules = [], []

    return render_template("platform/admin/packages.html", packages=packages, modules=modules)


# ── Kart Keşfi (SaaS hiyerarşi 4. katman) ─────────────────────────────────────

@app_bp.route("/admin/cards/discover", methods=["POST"])
@login_required
def admin_cards_discover():
    """Template'lerdeki data-card-* işaretlerini tarayıp KART katmanını seed et."""
    if not _is_admin():
        return _403()
    try:
        from app.services.card_discovery_service import discover_cards
        r = discover_cards(dry_run=False)
        if not r.get("ok"):
            return jsonify({"success": False, "message": r.get("error", "Keşif başarısız.")}), 500
        return jsonify({
            "success": True,
            "message": f"{r['files']} dosya tarandı · {r['cards']} yeni kart · {r['data_sources']} yeni veri kaynağı.",
            "stats": {"files": r["files"], "cards": r["cards"], "data_sources": r["data_sources"]},
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_cards_discover] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Kart keşfi sırasında hata oluştu."}), 500


@app_bp.route("/admin/hierarchy", methods=["GET"])
@login_required
def admin_hierarchy_page():
    """SaaS 4-katman hiyerarşi yönetim sayfası (gör/keşfet)."""
    if not _is_admin():
        return render_template("platform/errors/403.html"), 403
    return render_template("platform/admin/hierarchy.html")


@app_bp.route("/admin/api/hierarchy", methods=["GET"])
@login_required
def admin_api_hierarchy():
    """SaaS 4-katman ağacı: paket → modül → bileşen → kart → veri kaynağı."""
    if not _is_admin():
        return _403()
    from app.models.saas import (
        SubscriptionPackage, SystemComponent, SystemCard, ModuleComponentSlug,
    )
    cards_by_comp = {}
    for card in SystemCard.query.filter_by(is_active=True).order_by(SystemCard.sira, SystemCard.id).all():
        ds = [
            {"id": d.id, "data_key": d.data_key, "label": d.label,
             "required_component_code": d.required_component_code}
            for d in card.data_sources.filter_by(is_active=True).all()
        ]
        cards_by_comp.setdefault(card.component_id, []).append({
            "id": card.id, "code": card.code, "name": card.name,
            "sira": card.sira, "data_sources": ds,
        })
    comp_by_code = {c.code: c for c in SystemComponent.query.all()}
    mod_components = {}
    for mcs in ModuleComponentSlug.query.all():
        mod_components.setdefault(mcs.module_id, []).append(mcs.component_slug)

    packages = []
    for pkg in SubscriptionPackage.query.order_by(SubscriptionPackage.id).all():
        mods = []
        for m in pkg.modules:
            comps = []
            for ccode in sorted(set(mod_components.get(m.id, []))):
                comp = comp_by_code.get(ccode)
                comps.append({
                    "code": ccode,
                    "name": comp.name if comp else ccode,
                    "cards": cards_by_comp.get(comp.id, []) if comp else [],
                })
            mods.append({"id": m.id, "code": m.code, "name": m.name, "components": comps})
        packages.append({"id": pkg.id, "code": pkg.code, "name": pkg.name, "modules": mods})

    return jsonify({"success": True, "packages": packages})


# ── KART + VERİ katmanı düzenleme (4-katman hiyerarşi) ────────────────────────

@app_bp.route("/admin/api/components/list", methods=["GET"])
@login_required
def admin_api_components_list():
    """Bileşen kodları (dropdown'lar için: kart bileşeni, veri required_component)."""
    if not _is_admin():
        return _403()
    from app.models.saas import SystemComponent
    comps = SystemComponent.query.filter_by(is_active=True).order_by(SystemComponent.name).all()
    return jsonify({"success": True, "components": [
        {"id": c.id, "code": c.code, "name": c.name} for c in comps
    ]})


@app_bp.route("/admin/api/cards/<int:card_id>", methods=["POST"])
@login_required
def admin_api_card_update(card_id):
    """Kart düzenle: ad, sıra, bileşen (component_code)."""
    if not _is_admin():
        return _403()
    from app.models.saas import SystemCard, SystemComponent
    card = SystemCard.query.get_or_404(card_id)
    data = request.get_json(silent=True) or {}
    try:
        if "name" in data:
            name = (data.get("name") or "").strip()
            if not name:
                return jsonify({"success": False, "message": "Kart adı boş olamaz."}), 400
            card.name = name
        if "sira" in data:
            try:
                card.sira = int(data.get("sira") or 0)
            except (TypeError, ValueError):
                pass
        if "component_code" in data:
            ccode = (data.get("component_code") or "").strip()
            if ccode:
                comp = SystemComponent.query.filter_by(code=ccode).first()
                card.component_id = comp.id if comp else None
            else:
                card.component_id = None
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_api_card_update] {e}")
        return jsonify({"success": False, "message": "Kart güncellenemedi."}), 500


@app_bp.route("/admin/api/data-sources/<int:ds_id>", methods=["POST"])
@login_required
def admin_api_data_source_update(ds_id):
    """Veri kaynağı düzenle: required_component_code (hangi pakete tabi), label."""
    if not _is_admin():
        return _403()
    from app.models.saas import CardDataSource
    ds = CardDataSource.query.get_or_404(ds_id)
    data = request.get_json(silent=True) or {}
    try:
        if "required_component_code" in data:
            rc = (data.get("required_component_code") or "").strip()
            ds.required_component_code = rc or None  # boş = kısıtsız
        if "label" in data:
            ds.label = (data.get("label") or "").strip() or None
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_api_data_source_update] {e}")
        return jsonify({"success": False, "message": "Veri kaynağı güncellenemedi."}), 500


@app_bp.route("/admin/api/cards/<int:card_id>/data-sources", methods=["POST"])
@login_required
def admin_api_data_source_add(card_id):
    """Karta yeni veri kaynağı ekle."""
    if not _is_admin():
        return _403()
    from app.models.saas import SystemCard, CardDataSource
    SystemCard.query.get_or_404(card_id)
    data = request.get_json(silent=True) or {}
    dk = (data.get("data_key") or "").strip()
    if not dk:
        return jsonify({"success": False, "message": "Veri anahtarı zorunludur."}), 400
    if CardDataSource.query.filter_by(card_id=card_id, data_key=dk).first():
        return jsonify({"success": False, "message": "Bu veri anahtarı zaten var."}), 400
    try:
        ds = CardDataSource(
            card_id=card_id, data_key=dk,
            required_component_code=(data.get("required_component_code") or "").strip() or None,
            label=(data.get("label") or "").strip() or None, is_active=True,
        )
        db.session.add(ds)
        db.session.commit()
        return jsonify({"success": True, "id": ds.id})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_api_data_source_add] {e}")
        return jsonify({"success": False, "message": "Veri kaynağı eklenemedi."}), 500


@app_bp.route("/admin/api/data-sources/<int:ds_id>/delete", methods=["POST"])
@login_required
def admin_api_data_source_delete(ds_id):
    """Veri kaynağını sil (soft)."""
    if not _is_admin():
        return _403()
    from app.models.saas import CardDataSource
    ds = CardDataSource.query.get_or_404(ds_id)
    try:
        ds.is_active = False
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_api_data_source_delete] {e}")
        return jsonify({"success": False, "message": "Silinemedi."}), 500


# ── Bileşen Senkronizasyonu ───────────────────────────────────────────────────

@app_bp.route("/admin/components/sync", methods=["POST"])
@login_required
def admin_components_sync():
    if not _is_admin():
        return _403()

    try:
        from app.models.saas import RouteRegistry
        from flask import current_app as app

        # Mevcut endpoint'leri tek sorguda topla (N+1 önlemi: 500+ route × 1 query)
        _existing_endpoints = {r[0] for r in db.session.query(RouteRegistry.endpoint).all()}
        synced = 0
        for rule in app.url_map.iter_rules():
            slug = rule.endpoint
            if slug not in _existing_endpoints:
                db.session.add(RouteRegistry(
                    endpoint=slug,
                    url_pattern=str(rule),
                    methods=",".join(sorted(rule.methods or [])),
                ))
                _existing_endpoints.add(slug)
                synced += 1
        db.session.commit()
        return jsonify({"success": True, "message": f"{synced} yeni bileşen kaydedildi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_components_sync] {e}")
        return jsonify({"success": False, "message": "Senkronizasyon sırasında hata oluştu."}), 500


@app_bp.route("/admin/components/update", methods=["POST"])
@login_required
def admin_components_update():
    if not _is_admin():
        return _403()

    data = request.get_json() or {}
    try:
        from app.models.saas import RouteRegistry
        # Tüm endpoint'leri tek sorguda topla (N+1 önlemi)
        _endpoints = [item.get("endpoint") for item in data.get("items", []) if item.get("endpoint")]
        _recs = {r.endpoint: r for r in RouteRegistry.query.filter(
            RouteRegistry.endpoint.in_(_endpoints)
        ).all()} if _endpoints else {}
        for item in data.get("items", []):
            rec = _recs.get(item.get("endpoint"))
            if rec and item.get("component_slug"):
                rec.component_slug = item["component_slug"]
        db.session.commit()
        return jsonify({"success": True, "message": "Bileşen slug'ları güncellendi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_components_update] {e}")
        return jsonify({"success": False, "message": "Güncelleme sırasında hata oluştu."}), 500


# ── Bildirim Merkezi Yönetimi ─────────────────────────────────────────────────

@app_bp.route("/admin/notifications")
@login_required
def admin_notifications():
    if not _is_manager():
        return render_template("platform/errors/403.html"), 403

    from app.models.core import Notification
    from sqlalchemy.orm import joinedload
    if _is_admin():
        notifications = (Notification.query
                         .options(joinedload(Notification.user))
                         .order_by(Notification.created_at.desc()).limit(100).all())
        tenants = Tenant.query.filter_by(is_active=True).order_by(Tenant.name).all()
    else:
        notifications = (Notification.query
                         .options(joinedload(Notification.user))
                         .filter_by(tenant_id=current_user.tenant_id)
                         .order_by(Notification.created_at.desc()).limit(100).all())
        tenants = []

    return render_template("platform/admin/notifications.html",
                           notifications=notifications, tenants=tenants)


@app_bp.route("/admin/notifications/delete/<int:notif_id>", methods=["POST"])
@login_required
def admin_notifications_delete(notif_id):
    if not _is_manager():
        return _403()

    from app.models.core import Notification
    n = Notification.query.get_or_404(notif_id)
    if not _is_admin() and n.tenant_id != current_user.tenant_id:
        return _403()

    try:
        # Soft delete: is_read=True ile işaretle
        n.is_read = True
        db.session.commit()
        return jsonify({"success": True, "message": "Bildirim silindi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_notifications_delete] {e}")
        return jsonify({"success": False, "message": "İşlem sırasında hata oluştu."}), 500


@app_bp.route("/admin/notifications/broadcast", methods=["POST"])
@login_required
def admin_notifications_broadcast():
    if not _is_manager():
        return _403()

    data    = request.get_json() or {}
    title   = (data.get("title") or "").strip()
    message = (data.get("message") or "").strip()
    notif_type = data.get("type", "system_broadcast")

    if not title or not message:
        return jsonify({"success": False, "message": "Başlık ve mesaj zorunludur."}), 400

    try:
        from app.models.core import Notification
        if _is_admin():
            target_users = User.query.filter_by(is_active=True).all()
            if data.get("tenant_id"):
                target_users = User.query.filter_by(
                    is_active=True, tenant_id=int(data["tenant_id"])
                ).all()
        else:
            target_users = User.query.filter_by(
                is_active=True, tenant_id=current_user.tenant_id
            ).all()

        count = 0
        for u in target_users:
            n = Notification(
                user_id=u.id,
                tenant_id=u.tenant_id,
                notification_type=notif_type,
                title=title,
                message=message,
                link=data.get("link") or None,
            )
            db.session.add(n)
            count += 1

        db.session.commit()
        return jsonify({"success": True, "message": f"{count} kullanıcıya bildirim gönderildi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_notifications_broadcast] {e}")
        return jsonify({"success": False, "message": "Bildirim gönderilemedi."}), 500


@app_bp.route("/admin/notifications/stats")
@login_required
def admin_notifications_stats():
    if not _is_manager():
        return _403()

    from app.models.core import Notification
    from sqlalchemy import func
    try:
        q = Notification.query
        if not _is_admin():
            q = q.filter_by(tenant_id=current_user.tenant_id)

        total   = q.count()
        unread  = q.filter_by(is_read=False).count()
        by_type = db.session.query(
            Notification.notification_type, func.count(Notification.id)
        )
        if not _is_admin():
            by_type = by_type.filter(Notification.tenant_id == current_user.tenant_id)
        by_type = by_type.group_by(Notification.notification_type).all()

        return jsonify({
            "success": True,
            "total": total,
            "unread": unread,
            "by_type": {t: c for t, c in by_type},
        })
    except Exception as e:
        current_app.logger.error(f"[admin_notifications_stats] {e}")
        return jsonify({"success": False, "message": "İstatistik alınamadı."}), 500


# ── SaaS Paket / Modül CRUD ───────────────────────────────────────────────────

@app_bp.route("/admin/packages/add", methods=["POST"])
@login_required
def admin_packages_add():
    if not _is_admin():
        return _403()

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    code = (data.get("code") or "").strip().lower().replace(" ", "_")
    if not name or not code:
        return jsonify({"success": False, "message": "Ad ve kod zorunludur."}), 400

    try:
        from app.models.saas import SubscriptionPackage
        if SubscriptionPackage.query.filter_by(code=code).first():
            return jsonify({"success": False, "message": "Bu kod zaten kullanılıyor."}), 400
        pkg = SubscriptionPackage(name=name, code=code, description=data.get("description") or None)
        db.session.add(pkg)
        db.session.commit()
        return jsonify({"success": True, "message": "Paket oluşturuldu.", "id": pkg.id})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_packages_add] {e}")
        return jsonify({"success": False, "message": "Kayıt sırasında hata oluştu."}), 500


@app_bp.route("/admin/packages/edit/<int:pkg_id>", methods=["POST"])
@login_required
def admin_packages_edit(pkg_id):
    if not _is_admin():
        return _403()

    from app.models.saas import SubscriptionPackage
    pkg  = SubscriptionPackage.query.get_or_404(pkg_id)
    data = request.get_json() or {}
    try:
        pkg.name        = (data.get("name") or pkg.name).strip()
        pkg.description = data.get("description") or pkg.description
        db.session.commit()
        return jsonify({"success": True, "message": "Paket güncellendi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_packages_edit] {e}")
        return jsonify({"success": False, "message": "Güncelleme sırasında hata oluştu."}), 500


@app_bp.route("/admin/packages/toggle/<int:pkg_id>", methods=["POST"])
@login_required
def admin_packages_toggle(pkg_id):
    if not _is_admin():
        return _403()

    from app.models.saas import SubscriptionPackage
    pkg = SubscriptionPackage.query.get_or_404(pkg_id)
    try:
        pkg.is_active = not pkg.is_active
        db.session.commit()
        state = "aktifleştirildi" if pkg.is_active else "pasife alındı"
        return jsonify({"success": True, "message": f"Paket {state}.", "is_active": pkg.is_active})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_packages_toggle] {e}")
        return jsonify({"success": False, "message": "İşlem sırasında hata oluştu."}), 500


# ── Paket ↔ Modül atama (paket içeriği yönetimi) ──────────────────────────────

@app_bp.route("/admin/packages/<int:pkg_id>/modules", methods=["GET"])
@login_required
def admin_package_modules(pkg_id):
    """Bir paketin modül atamaları + her modülün bileşenleri.

    UI: pakete tıkla → tüm modüller (içerdikleri işaretli) + bileşen önizleme.
    """
    if not _is_admin():
        return _403()
    from app.models.saas import SubscriptionPackage, SystemModule
    pkg = SubscriptionPackage.query.get_or_404(pkg_id)
    pkg_module_ids = {m.id for m in pkg.modules}
    modules = SystemModule.query.filter_by(is_active=True).order_by(SystemModule.name).all()

    def _bilesenler(m):
        try:
            return sorted(c.component_slug for c in m.component_slugs)
        except Exception:
            return []

    return jsonify({
        "success": True,
        "package": {"id": pkg.id, "name": pkg.name, "code": pkg.code},
        "modules": [
            {
                "id": m.id, "name": m.name, "code": m.code,
                "in_package": m.id in pkg_module_ids,
                "components": _bilesenler(m),
            }
            for m in modules
        ],
    })


@app_bp.route("/admin/packages/<int:pkg_id>/modules/<int:module_id>/toggle", methods=["POST"])
@login_required
def admin_package_module_toggle(pkg_id, module_id):
    """Paket ↔ modül bağını ekle/çıkar (M2M)."""
    if not _is_admin():
        return _403()
    from app.models.saas import SubscriptionPackage, SystemModule
    pkg = SubscriptionPackage.query.get_or_404(pkg_id)
    mod = SystemModule.query.get_or_404(module_id)
    try:
        current = {m.id for m in pkg.modules}
        if module_id in current:
            pkg.modules.remove(mod)
            in_package = False
        else:
            pkg.modules.append(mod)
            in_package = True
        db.session.commit()
        return jsonify({"success": True, "in_package": in_package,
                        "message": f"{mod.name} {'eklendi' if in_package else 'çıkarıldı'}."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_package_module_toggle] {e}")
        return jsonify({"success": False, "message": "İşlem sırasında hata oluştu."}), 500


@app_bp.route("/admin/modules/add", methods=["POST"])
@login_required
def admin_modules_add():
    if not _is_admin():
        return _403()

    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    code = (data.get("code") or "").strip().lower().replace(" ", "_")
    if not name or not code:
        return jsonify({"success": False, "message": "Ad ve kod zorunludur."}), 400

    try:
        from app.models.saas import SystemModule
        if SystemModule.query.filter_by(code=code).first():
            return jsonify({"success": False, "message": "Bu kod zaten kullanılıyor."}), 400
        mod = SystemModule(name=name, code=code, description=data.get("description") or None)
        db.session.add(mod)
        db.session.commit()
        return jsonify({"success": True, "message": "Modül oluşturuldu.", "id": mod.id})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_modules_add] {e}")
        return jsonify({"success": False, "message": "Kayıt sırasında hata oluştu."}), 500


@app_bp.route("/admin/k-radar/weights", methods=["GET"])
@login_required
def admin_k_radar_weights_get():
    """Mevcut K-Radar ağırlıklarını döner."""
    if not _is_admin():
        return _403()
    import json
    from app.models.system_setting import SystemSetting
    tid = current_user.tenant_id
    key = f"k_radar_weights_{tid}"
    row = SystemSetting.query.filter_by(key=key).first()
    defaults = {"ks": 2.0, "kp": 3.0, "kpr": 3.0, "bireysel": 2.0}
    if row:
        try:
            weights = json.loads(row.value)
        except Exception:
            weights = defaults
    else:
        weights = defaults
    return jsonify({"success": True, "weights": weights})


@app_bp.route("/admin/k-radar/weights", methods=["POST"])
@login_required
def admin_k_radar_weights_save():
    """K-Radar ağırlıklarını kaydeder ve cache'i temizler."""
    if not _is_admin():
        return _403()
    import json
    from app.models.system_setting import SystemSetting
    from app.extensions import cache
    payload = request.get_json(silent=True) or {}
    try:
        weights = {
            "ks":       max(0.1, float(payload.get("ks",       2.0))),
            "kp":       max(0.1, float(payload.get("kp",       3.0))),
            "kpr":      max(0.1, float(payload.get("kpr",      3.0))),
            "bireysel": max(0.1, float(payload.get("bireysel", 2.0))),
        }
    except (TypeError, ValueError) as e:
        return jsonify({"success": False, "message": f"Geçersiz değer: {e}"}), 400

    tid = current_user.tenant_id
    key = f"k_radar_weights_{tid}"
    row = SystemSetting.query.filter_by(key=key).first()
    if row:
        row.value = json.dumps(weights)
    else:
        db.session.add(SystemSetting(key=key, value=json.dumps(weights)))
    try:
        db.session.commit()
        # K-Radar cache'i temizle (yeni ağırlıklar hemen geçerli olsun)
        cache.delete_memoized(None)
        cache.clear()
        return jsonify({"success": True, "weights": weights, "message": "Ağırlıklar kaydedildi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_k_radar_weights] {e}")
        return jsonify({"success": False, "message": "Kayıt sırasında hata oluştu."}), 500


@app_bp.route("/admin/modules/toggle/<int:mod_id>", methods=["POST"])
@login_required
def admin_modules_toggle(mod_id):
    if not _is_admin():
        return _403()

    from app.models.saas import SystemModule
    mod = SystemModule.query.get_or_404(mod_id)
    try:
        mod.is_active = not mod.is_active
        db.session.commit()
        state = "aktifleştirildi" if mod.is_active else "pasife alındı"
        return jsonify({"success": True, "message": f"Modül {state}.", "is_active": mod.is_active})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[admin_modules_toggle] {e}")
        return jsonify({"success": False, "message": "İşlem sırasında hata oluştu."}), 500



# Sprint C — Bayi/Holding alt-tenant yönetim sayfası
# TASARIM NOTU: Bu late-import (modül seviyesinin sonunda) kasıtlıdır.
# routes_sub_tenants ve routes_holding, app_bp'yi kayıt ettikten sonra
# route'larını ekler. Circular import riski: bu dosya onları import eder,
# onlar da helpers'tan import yapar — döngüsel zincir yok.
from micro.modules.admin import routes_sub_tenants  # noqa: F401, E402

from micro.modules.admin import routes_holding  # noqa: F401, E402  Sprint D

from micro.modules.admin import routes_admin_tools  # noqa: F401, E402  Admin Araçları (Hata Kontrolü)
