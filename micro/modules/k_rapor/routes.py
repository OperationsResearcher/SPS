"""K-Rapor — Kurumsal Raporlama Merkezi."""

import datetime as _dt

from flask import render_template, jsonify, request, current_app
from flask_login import login_required, current_user
from sqlalchemy import func

from platform_core import app_bp
from extensions import db


# ── Sprint 50: Custom Dashboard Builder — Widget Registry ────────────────────

@app_bp.route("/k-rapor/api/dashboard/widgets")
@login_required
def k_rapor_dashboard_widgets():
    """Kullanılabilir widget'ları listele (kategori filtreli)."""
    from app.services.dashboard_widgets import list_widgets, categories
    category = request.args.get("category")
    return jsonify({
        "success": True,
        "widgets": list_widgets(category),
        "categories": categories(),
    })


# ── Sprint 46: KPI Forecasting ────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/forecast/<int:kpi_id>")
@login_required
def k_rapor_api_forecast(kpi_id):
    """KPI trend forecasting — linear regression + güven aralığı."""
    from app.services.forecast_service import forecast_kpi
    from app.models.process import ProcessKpi, Process

    kpi = (
        ProcessKpi.query.join(Process)
        .filter(ProcessKpi.id == kpi_id, Process.tenant_id == current_user.tenant_id)
        .first()
    )
    if not kpi:
        return jsonify({"success": False, "message": "KPI bulunamadı"}), 404

    periods = request.args.get("periods", 3, type=int)
    periods = max(1, min(periods, 12))
    try:
        result = forecast_kpi(kpi_id, periods_ahead=periods)
        result["kpi_name"] = kpi.name
        result["kpi_code"] = kpi.code
        return jsonify(result)
    except Exception as e:
        current_app.logger.error(f"[forecast] {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500


# ── Ana Sayfa ─────────────────────────────────────────────────────────────────

@app_bp.route("/k-rapor")
@login_required
def k_rapor():
    """K-Rapor ana sayfası."""
    from app.models.process import Process, ProcessKpi, KpiData
    from app.models.core import Strategy
    from app.services.plan_year_service import get_active_plan_year_for_user

    tid = current_user.tenant_id
    year = _dt.date.today().year
    active_py = get_active_plan_year_for_user(current_user)
    if active_py:
        year = active_py.year

    surec_sayisi = Process.query.filter_by(tenant_id=tid, is_active=True).count()
    pg_sayisi    = (
        ProcessKpi.query.join(Process)
        .filter(Process.tenant_id == tid, Process.is_active == True, ProcessKpi.is_active == True)
        .count()
    )
    strateji_sayisi = Strategy.query.filter_by(tenant_id=tid, is_active=True).count()
    veri_sayisi = (
        KpiData.query.join(ProcessKpi).join(Process)
        .filter(Process.tenant_id == tid, KpiData.is_active == True, KpiData.year == year)
        .count()
    )

    return render_template(
        "platform/k_rapor/index.html",
        year=year,
        active_plan_year=active_py,
        surec_sayisi=surec_sayisi,
        pg_sayisi=pg_sayisi,
        strateji_sayisi=strateji_sayisi,
        veri_sayisi=veri_sayisi,
    )


# ── Rapor 1: Kurumsal Performans ──────────────────────────────────────────────

@app_bp.route("/k-rapor/api/kurumsal")
@login_required
def k_rapor_api_kurumsal():
    """Vizyon skoru + strateji bazlı başarı + en iyi/kötü süreçler."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.services.score_engine_service import compute_vision_score
        from app.services.plan_year_service import get_plan_year
        plan_year = get_plan_year(tid, year)
        result = compute_vision_score(tid, year=year, persist_pg_scores=False, plan_year=plan_year)

        # En iyi / en kötü 5 süreç
        process_scores = result.get("process_scores") or {}
        from app.models.process import Process
        if process_scores:
            processes = Process.query.filter(
                Process.id.in_(list(process_scores.keys())),
                Process.is_active == True,
            ).all()
            proc_map = {p.id: p for p in processes}
            sorted_scores = sorted(
                [(pid, sc) for pid, sc in process_scores.items()],
                key=lambda x: x[1], reverse=True
            )
            def _proc_row(pid, sc):
                p = proc_map.get(pid)
                return {
                    "id": pid,
                    "code": p.code if p else "",
                    "name": p.name if p else f"#{pid}",
                    "score": round(sc, 1),
                }
            result["top5"]    = [_proc_row(pid, sc) for pid, sc in sorted_scores[:5]]
            result["bottom5"] = [_proc_row(pid, sc) for pid, sc in sorted_scores[-5:] if sc < 100]

        # Süreç sağlık dağılımı
        if process_scores:
            psc = list(process_scores.values())
            result["pg_saglik"] = {
                "yesil":   sum(1 for s in psc if s >= 80),
                "sari":    sum(1 for s in psc if 50 <= s < 80),
                "kirmizi": sum(1 for s in psc if s < 50),
                "toplam":  len(psc),
            }

        # Strateji adlarını zenginleştir (ID yerine kod+başlık)
        raw_ss = result.get("strategy_scores") or {}
        if raw_ss:
            from app.models.core import Strategy
            strat_objs = Strategy.query.filter(Strategy.id.in_(list(raw_ss.keys()))).all()
            strat_name_map = {s.id: {"code": s.code or "", "title": s.title} for s in strat_objs}
            result["strategy_scores_detail"] = [
                {"id": sid, "code": strat_name_map.get(sid, {}).get("code", ""),
                 "title": strat_name_map.get(sid, {}).get("title", f"#{sid}"),
                 "score": round(sc, 1)}
                for sid, sc in sorted(raw_ss.items(), key=lambda x: -x[1])
            ]

        return jsonify({"success": True, "data": result, "year": year})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_kurumsal] {e}")
        return jsonify({"success": False, "message": "Kurumsal veri alınamadı."}), 500


# ── Rapor 2: Süreç & PG Isı Haritası ─────────────────────────────────────────

@app_bp.route("/k-rapor/api/surec-pg")
@login_required
def k_rapor_api_surec_pg():
    """Süreçler × dönemler ısı haritası verisi."""
    tid    = current_user.tenant_id
    year   = request.args.get("year", _dt.date.today().year, type=int)
    period = request.args.get("period", "ceyrek")  # aylik / ceyrek / yillik
    try:
        from app.models.process import Process, ProcessKpi, KpiData
        processes = (
            Process.query
            .filter_by(tenant_id=tid, is_active=True, parent_id=None)
            .order_by(Process.code)
            .all()
        )
        kpis = (
            ProcessKpi.query.join(Process)
            .filter(Process.tenant_id == tid, Process.is_active == True, ProcessKpi.is_active == True)
            .all()
        )
        kpi_map = {k.id: k for k in kpis}

        if period == "aylik":
            labels = [f"{m}. Ay" for m in range(1, 13)]
            period_keys = list(range(1, 13))
        elif period == "yillik":
            labels = [str(year)]
            period_keys = [1]
        else:  # ceyrek
            labels = ["Ç1", "Ç2", "Ç3", "Ç4"]
            period_keys = [1, 2, 3, 4]

        kpi_ids = [k.id for k in kpis]
        # period_type kolonu DB'de olmayabilir — güvenli sorgu
        try:
            data_rows = (
                KpiData.query
                .filter(
                    KpiData.process_kpi_id.in_(kpi_ids),
                    KpiData.year == year,
                    KpiData.period_type == period,
                    KpiData.is_active == True,
                )
                .all()
            )
        except Exception:
            db.session.rollback()
            # period_type kolonu yoksa tüm yıl verisini çek
            data_rows = (
                KpiData.query
                .filter(
                    KpiData.process_kpi_id.in_(kpi_ids),
                    KpiData.year == year,
                    KpiData.is_active == True,
                )
                .all()
            )
        data_index = {}
        for d in data_rows:
            key = (d.process_kpi_id, getattr(d, 'period_no', None) or 1)
            data_index[key] = d

        rows = []
        for proc in processes:
            proc_kpis = [k for k in kpis if k.process_id == proc.id]
            if not proc_kpis:
                continue
            cells = []
            for pno in period_keys:
                vals = []
                for k in proc_kpis:
                    entry = data_index.get((k.id, pno))
                    if entry and entry.actual_value is not None:
                        try:
                            vals.append(float(entry.actual_value))
                        except (ValueError, TypeError):
                            pass
                cells.append({
                    "has_data": len(vals) > 0,
                    "avg_val":  round(sum(vals) / len(vals), 1) if vals else None,
                    "count":    len(vals),
                })
            rows.append({
                "id":    proc.id,
                "code":  proc.code or "",
                "name":  proc.name,
                "kpi_count": len(proc_kpis),
                "kpi_ids":   [k.id for k in proc_kpis],
                "cells": cells,
            })

        return jsonify({"success": True, "data": rows, "labels": labels, "year": year, "period": period})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_surec_pg] {e}")
        return jsonify({"success": False, "message": "Süreç/PG verisi alınamadı."}), 500


# ── Rapor 3: PG Trend ─────────────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/trend/<int:kpi_id>")
@login_required
def k_rapor_api_trend(kpi_id):
    """KPI zaman serisi — analytics_service'e delege."""
    from app.models.process import ProcessKpi, Process
    kpi = (
        ProcessKpi.query.join(Process)
        .filter(ProcessKpi.id == kpi_id, Process.tenant_id == current_user.tenant_id,
                Process.is_active == True, ProcessKpi.is_active == True)
        .first_or_404()
    )
    try:
        from app.services.analytics_service import AnalyticsService
        import datetime as _dt2
        today = _dt2.date.today()
        start_date = request.args.get("start_date")
        end_date   = request.args.get("end_date")
        # Tarih parse — yoksa son 12 ay
        try:
            sd = _dt2.datetime.strptime(start_date, "%Y-%m-%d") if start_date else _dt2.datetime(today.year, 1, 1)
            ed = _dt2.datetime.strptime(end_date,   "%Y-%m-%d") if end_date   else _dt2.datetime(today.year, 12, 31)
        except (ValueError, TypeError):
            sd = _dt2.datetime(today.year, 1, 1)
            ed = _dt2.datetime(today.year, 12, 31)
        result = AnalyticsService.get_performance_trend(
            kpi.process_id,
            kpi_id,
            sd,
            ed,
            frequency=request.args.get("frequency", "monthly"),
        )
        return jsonify({"success": True, "data": result, "kpi_name": kpi.name, "kpi_code": kpi.code})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_trend] {e}")
        return jsonify({"success": False, "message": "Trend verisi alınamadı."}), 500


# ── Rapor 4: Stratejik Uyum ───────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/uyum")
@login_required
def k_rapor_api_uyum():
    """Strateji → Alt Strateji → Süreç katkı ağacı."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.core import Strategy, SubStrategy
        from app.models.process import Process, ProcessKpi
        from app.services.score_engine_service import compute_vision_score
        from app.services.plan_year_service import get_plan_year

        plan_year = get_plan_year(tid, year)
        vision = compute_vision_score(tid, year=year, persist_pg_scores=False, plan_year=plan_year)
        process_scores  = vision.get("process_scores") or {}
        strategy_scores = vision.get("strategy_scores") or {}
        ss_scores       = vision.get("sub_strategy_scores") or {}

        strategies = Strategy.query.filter_by(tenant_id=tid, is_active=True).order_by(Strategy.code).all()
        tree = []
        for s in strategies:
            s_score = strategy_scores.get(s.id, 0.0)
            ss_list = []
            for ss in (s.sub_strategies or []):
                if not ss.is_active:
                    continue
                linked = [p for p in (ss.processes or []) if p.is_active and p.tenant_id == tid]
                proc_list = []
                for p in linked:
                    proc_list.append({
                        "id":    p.id,
                        "code":  p.code or "",
                        "name":  p.name,
                        "score": round(process_scores.get(p.id, 0.0), 1),
                    })
                ss_list.append({
                    "id":    ss.id,
                    "code":  ss.code or "",
                    "title": ss.title,
                    "score": round(ss_scores.get(ss.id, 0.0), 1),
                    "processes": proc_list,
                })
            tree.append({
                "id":    s.id,
                "code":  s.code or "",
                "title": s.title,
                "score": round(s_score, 1),
                "sub_strategies": ss_list,
            })

        return jsonify({"success": True, "data": tree, "vision_score": vision.get("vision_score", 0), "year": year})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_uyum] {e}")
        return jsonify({"success": False, "message": "Stratejik uyum verisi alınamadı."}), 500


# ── Rapor 5: Faaliyet ─────────────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/faaliyet")
@login_required
def k_rapor_api_faaliyet():
    """Faaliyet tamamlanma oranı, geciken ve tamamlanan."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.process import Process, ProcessActivity
        today = _dt.date.today()
        activities = (
            ProcessActivity.query.join(Process)
            .filter(
                Process.tenant_id == tid,
                Process.is_active == True,
                ProcessActivity.is_active == True,
            )
            .all()
        )

        toplam      = len(activities)
        tamamlanan  = sum(1 for a in activities if a.status == "Tamamlandı")
        geciken     = sum(
            1 for a in activities
            if a.status not in ("Tamamlandı", "İptal")
            and a.end_date and a.end_date < today
        )
        devam       = sum(1 for a in activities if a.status == "Devam Ediyor")
        planlandi   = sum(1 for a in activities if a.status == "Planlandı")

        # Aylık tamamlanma (ActivityTrack)
        from app.models.process import ActivityTrack
        tracks = (
            ActivityTrack.query.join(ProcessActivity).join(Process)
            .filter(
                Process.tenant_id == tid,
                ActivityTrack.year == year,
                ActivityTrack.completed == True,
            )
            .all()
        )
        aylik = [0] * 12
        for t in tracks:
            if 1 <= t.month <= 12:
                aylik[t.month - 1] += 1

        # Geciken faaliyet listesi
        geciken_liste = []
        for a in activities:
            if a.status not in ("Tamamlandı", "İptal") and a.end_date and a.end_date < today:
                proc = a.process
                geciken_liste.append({
                    "id":       a.id,
                    "name":     a.name,
                    "surec":    proc.name if proc else "",
                    "end_date": str(a.end_date),
                    "status":   a.status,
                    "progress": a.progress or 0,
                })
        geciken_liste.sort(key=lambda x: x["end_date"])

        # Proje portföy verisi
        try:
            from app.models.project import PlanProject
            from app.models.plan_year import PlanYear
            py = PlanYear.query.filter_by(tenant_id=tid, year=year).first()
            projeler = []
            proje_ozet = {"toplam": 0, "tamamlanan": 0, "devam": 0, "planlandi": 0, "geciken": 0}
            if py:
                projects = PlanProject.query.filter_by(tenant_id=tid, plan_year_id=py.id, is_active=True).all()
                proje_ozet["toplam"] = len(projects)
                proje_ozet["tamamlanan"] = sum(1 for p in projects if p.status == "Tamamlandı")
                proje_ozet["devam"] = sum(1 for p in projects if p.status == "Devam Ediyor")
                proje_ozet["planlandi"] = sum(1 for p in projects if p.status == "Planlandı")
                proje_ozet["geciken"] = sum(
                    1 for p in projects
                    if p.status not in ("Tamamlandı", "İptal")
                    and p.end_date and p.end_date < today
                )
                for p in sorted(projects, key=lambda x: x.progress or 0, reverse=True)[:15]:
                    projeler.append({
                        "id": p.id, "name": p.name, "status": p.status,
                        "progress": p.progress or 0,
                        "start_date": str(p.start_date) if p.start_date else None,
                        "end_date": str(p.end_date) if p.end_date else None,
                    })
        except (ImportError, Exception) as _proj_err:
            current_app.logger.warning(f"[k_rapor_api_faaliyet] proje verisi alınamadı: {_proj_err}")
            db.session.rollback()
            projeler = []
            proje_ozet = {"toplam": 0, "tamamlanan": 0, "devam": 0, "planlandi": 0, "geciken": 0}

        return jsonify({
            "success": True,
            "data": {
                "toplam":     toplam,
                "tamamlanan": tamamlanan,
                "geciken":    geciken,
                "devam":      devam,
                "planlandi":  planlandi,
                "tamamlanma_orani": round(tamamlanan / toplam * 100, 1) if toplam else 0,
                "aylik_tamamlanan": aylik,
                "geciken_liste":    geciken_liste[:20],
                "proje_ozet":       proje_ozet,
                "projeler":         projeler,
            },
            "year": year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_faaliyet] {e}")
        return jsonify({"success": False, "message": "Faaliyet verisi alınamadı."}), 500


# ── Rapor 6: Bireysel Performans ──────────────────────────────────────────────

@app_bp.route("/k-rapor/api/bireysel")
@login_required
def k_rapor_api_bireysel():
    """Kullanıcı bazlı bireysel PG — özet + detay + sağlık."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.process import IndividualPerformanceIndicator as IPI, IndividualKpiData
        from app.models.core import User
        from app.models.plan_year import PlanYear

        users    = User.query.filter_by(tenant_id=tid, is_active=True).all()
        user_map = {u.id: u for u in users}
        uid_list = [u.id for u in users]

        # Plan year filtresi — istenen yılda IPI yoksa fallback
        py = PlanYear.query.filter_by(tenant_id=tid, year=year).first()
        if py:
            pgs = IPI.query.filter(
                IPI.user_id.in_(uid_list), IPI.is_active == True, IPI.plan_year_id == py.id
            ).all()
            if not pgs:
                pgs = IPI.query.filter(IPI.user_id.in_(uid_list), IPI.is_active == True).all()
        else:
            pgs = IPI.query.filter(IPI.user_id.in_(uid_list), IPI.is_active == True).all()

        pg_ids = [pg.id for pg in pgs]
        pg_by_user: dict = {}
        for pg in pgs:
            pg_by_user.setdefault(pg.user_id, []).append(pg)

        latest_by_pg: dict = {}
        all_by_pg: dict = {}
        data_rows = (
            IndividualKpiData.query
            .filter(IndividualKpiData.individual_pg_id.in_(pg_ids), IndividualKpiData.year == year)
            .all()
        ) if pg_ids else []
        for d in data_rows:
            all_by_pg.setdefault(d.individual_pg_id, []).append(d)
            prev = latest_by_pg.get(d.individual_pg_id)
            if prev is None or (d.data_date and (prev.data_date is None or d.data_date > prev.data_date)):
                latest_by_pg[d.individual_pg_id] = d

        # user_id → toplam giriş sayısı + son giriş tarihi
        # (girişi yapan user_id = IndividualKpiData.user_id)
        from collections import defaultdict
        giris_cnt: dict  = defaultdict(int)
        son_giris: dict  = {}
        for d in data_rows:
            if d.user_id:
                giris_cnt[d.user_id] += 1
                prev_date = son_giris.get(d.user_id)
                if d.data_date and (prev_date is None or d.data_date > prev_date):
                    son_giris[d.user_id] = d.data_date

        # Kullanıcı özet satırları
        rows = []
        for uid, user_pgs in pg_by_user.items():
            u = user_map.get(uid)
            if not u:
                continue
            scores = [float(latest_by_pg[pg.id].status_percentage)
                      for pg in user_pgs
                      if pg.id in latest_by_pg and latest_by_pg[pg.id].status_percentage is not None]
            # Kaç IPI'ya herhangi bir veri girilmiş
            veri_girilmis = sum(1 for pg in user_pgs if pg.id in all_by_pg)
            ort = round(sum(scores) / len(scores), 1) if scores else None
            rows.append({
                "user_id":        uid,
                "ad":             f"{u.first_name or ''} {u.last_name or ''}".strip() or u.email,
                "email":          u.email,
                "pg_sayisi":      len(user_pgs),
                "veri_girilmis":  veri_girilmis,       # kaç IPI'ya veri girilmiş
                "toplam_giris":   giris_cnt.get(uid, 0), # bu kullanıcının girdiği toplam satır
                "son_giris":      str(son_giris[uid]) if uid in son_giris else None,
                "ort_basari":     ort,
            })
        rows.sort(key=lambda x: x["ort_basari"] or -1, reverse=True)

        # IPI detay listesi
        ipi_detail = []
        for pg in pgs[:200]:
            u   = user_map.get(pg.user_id)
            d   = latest_by_pg.get(pg.id)
            pct = d.status_percentage if d else None
            ipi_detail.append({
                "pg_id":  pg.id,
                "ad":     f"{u.first_name or ''} {u.last_name or ''}".strip() if u else "—",
                "code":   pg.code or "",
                "name":   pg.name,
                "hedef":  pg.target_value,
                "gercek": d.actual_value if d else None,
                "pct":    round(float(pct), 1) if pct is not None else None,
                "status": d.status if d else None,
                "source": pg.source or "—",
            })

        all_pcts = [r["pct"] for r in ipi_detail if r["pct"] is not None]
        saglik = {
            "yesil":   sum(1 for p in all_pcts if p >= 80),
            "sari":    sum(1 for p in all_pcts if 50 <= p < 80),
            "kirmizi": sum(1 for p in all_pcts if p < 50),
            "veri_yok": len(pgs) - len(all_pcts),
        }

        return jsonify({
            "success":    True,
            "data":       rows,
            "ipi_detail": ipi_detail,
            "saglik":     saglik,
            "toplam_pg":  len(pgs),
            "year":       year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_bireysel] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Bireysel veri alınamadı."}), 500


# ── Rapor 7: Veri Giriş Durumu ────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/veri-durumu")
@login_required
def k_rapor_api_veri_durumu():
    """Aktif plan yılında cari döneme ait PG veri giriş durumu."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.process import Process, ProcessKpi, KpiData
        from app.models.core import User

        kpis = (
            ProcessKpi.query.join(Process)
            .filter(
                Process.tenant_id == tid,
                Process.is_active == True,
                ProcessKpi.is_active == True,
            )
            .all()
        )
        kpi_ids = [k.id for k in kpis]
        kpi_map = {k.id: k for k in kpis}

        latest_data = (
            KpiData.query
            .filter(
                KpiData.process_kpi_id.in_(kpi_ids),
                KpiData.year == year,
                KpiData.is_active == True,
            )
            .order_by(KpiData.data_date.desc())
            .all()
        )
        entered_ids = {d.process_kpi_id for d in latest_data}
        latest_by_kpi = {}
        for d in latest_data:
            if d.process_kpi_id not in latest_by_kpi:
                latest_by_kpi[d.process_kpi_id] = d

        user_ids = {d.user_id for d in latest_data if d.user_id}
        users = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()} if user_ids else {}

        # Süreç bazlı grupla
        proc_ids = list({k.process_id for k in kpis})
        procs = {p.id: p for p in Process.query.filter(Process.id.in_(proc_ids)).all()}

        girilen, girilmeyen = [], []
        for k in kpis:
            proc = procs.get(k.process_id)
            d = latest_by_kpi.get(k.id)
            row = {
                "kpi_id":      k.id,
                "kpi_code":    k.code or "",
                "kpi_name":    k.name,
                "surec_code":  proc.code if proc else "",
                "surec_name":  proc.name if proc else "",
            }
            if d:
                u = users.get(d.user_id)
                row.update({
                    "son_tarih":  str(d.data_date) if d.data_date else None,
                    "son_deger":  d.actual_value,
                    "giren":      f"{u.first_name or ''} {u.last_name or ''}".strip() if u else "—",
                })
                girilen.append(row)
            else:
                girilmeyen.append(row)

        toplam = len(kpis)
        return jsonify({
            "success": True,
            "data": {
                "toplam":          toplam,
                "girilen_sayisi":  len(girilen),
                "girilmeyen_sayisi": len(girilmeyen),
                "tamamlanma_orani": round(len(girilen) / toplam * 100, 1) if toplam else 0,
                "girilen":         girilen[:50],
                "girilmeyen":      girilmeyen[:50],
            },
            "year": year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_veri_durumu] {e}")
        return jsonify({"success": False, "message": "Veri durumu alınamadı."}), 500


# ── Rapor 8: Risk ─────────────────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/risk")
@login_required
def k_rapor_api_risk():
    """Risk matrisi + darboğaz + süreç olgunluk."""
    tid = current_user.tenant_id
    try:
        try:
            from app.models.k_radar_domain import RiskHeatmapItem, BottleneckLog, ProcessMaturity
        except ImportError:
            return jsonify({"success": True, "data": {"risk_listesi": [], "darbogaz": [], "olgunluk": []}})
        from app.models.process import Process

        risks = RiskHeatmapItem.query.filter_by(tenant_id=tid).order_by(
            RiskHeatmapItem.rpn.desc()
        ).limit(50).all()

        risk_list = []
        for r in risks:
            risk_list.append({
                "id":          r.id,
                "title":       r.title,
                "probability": r.probability,
                "impact":      r.impact,
                "rpn":         r.rpn,
                "status":      r.status or "—",
                "source_type": r.source_type or "—",
            })

        bottlenecks = BottleneckLog.query.filter_by(tenant_id=tid).order_by(
            BottleneckLog.triggered_at.desc()
        ).limit(20).all()
        bn_list = []
        proc_ids = {b.process_id for b in bottlenecks if b.process_id}
        proc_map = {p.id: p for p in Process.query.filter(Process.id.in_(proc_ids)).all()} if proc_ids else {}
        for b in bottlenecks:
            p = proc_map.get(b.process_id)
            bn_list.append({
                "id":           b.id,
                "surec":        p.name if p else "—",
                "severity":     b.severity or "—",
                "note":         b.note or "",
                "triggered_at": str(b.triggered_at)[:16] if b.triggered_at else None,
                "resolved_at":  str(b.resolved_at)[:16] if b.resolved_at else None,
                "cozuldu":      b.resolved_at is not None,
            })

        maturity = ProcessMaturity.query.filter_by(tenant_id=tid).all()
        mat_list = []
        mat_proc_ids = {m.process_id for m in maturity if m.process_id}
        mat_proc_map = {p.id: p for p in Process.query.filter(Process.id.in_(mat_proc_ids)).all()} if mat_proc_ids else {}
        for m in maturity:
            p = mat_proc_map.get(m.process_id)
            mat_list.append({
                "surec":          p.name if p else "—",
                "maturity_level": m.maturity_level,
                "dimension":      m.dimension or "—",
            })

        return jsonify({
            "success": True,
            "data": {
                "risk_listesi":  risk_list,
                "darbogaz":      bn_list,
                "olgunluk":      mat_list,
            },
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_risk] {e}")
        return jsonify({"success": False, "message": "Risk verisi alınamadı."}), 500


# ── Rapor 9: Denetim ──────────────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/denetim")
@login_required
def k_rapor_api_denetim():
    """Audit log özeti — kim ne yaptı."""
    tid = current_user.tenant_id
    gun = request.args.get("gun", 30, type=int)
    gun = min(gun, 180)
    since = _dt.datetime.utcnow() - _dt.timedelta(days=gun)
    try:
        from app.models.audit import AuditLog

        logs = (
            AuditLog.query
            .filter(
                AuditLog.tenant_id == tid,
                AuditLog.created_at >= since,
            )
            .order_by(AuditLog.created_at.desc())
            .limit(200)
            .all()
        )

        kayit_listesi = []
        for log in logs:
            kayit_listesi.append({
                "id":            log.id,
                "kullanici":     log.username or "—",
                "action":        log.action or "—",
                "resource_type": log.resource_type or "—",
                "resource_id":   log.resource_id,
                "description":   (log.description or "")[:120],
                "tarih":         str(log.created_at)[:16] if log.created_at else None,
            })

        # İşlem tipi dağılımı
        from collections import Counter
        action_counts = Counter(log.action for log in logs)
        # Aktif kullanıcılar
        user_counts = Counter(log.username for log in logs if log.username)

        return jsonify({
            "success": True,
            "data": {
                "kayitlar":       kayit_listesi,
                "action_dagilim": dict(action_counts.most_common(10)),
                "aktif_kullanicilar": [
                    {"kullanici": k, "islem_sayisi": v}
                    for k, v in user_counts.most_common(10)
                ],
                "toplam_islem": len(logs),
            },
            "gun": gun,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_denetim] {e}")
        return jsonify({"success": False, "message": "Denetim verisi alınamadı."}), 500


# ── Rapor 10: Uyarı Merkezi ───────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/uyari")
@login_required
def k_rapor_api_uyari():
    """Uyarı merkezi — kritik PG'ler, geciken faaliyetler, yüksek riskler."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.process import Process, ProcessKpi, KpiData, ProcessActivity
        try:
            from app.models.k_radar_domain import RiskHeatmapItem
            _has_risk = True
        except ImportError:
            _has_risk = False
        today = _dt.date.today()

        # ── Kritik PG'ler (gerçekleşen %50'nin altında) ────────────────────────
        kpis = (
            ProcessKpi.query.join(Process)
            .filter(Process.tenant_id == tid, Process.is_active == True, ProcessKpi.is_active == True)
            .all()
        )
        kpi_map = {k.id: k for k in kpis}
        proc_ids_map = {k.id: k.process_id for k in kpis}

        kpi_ids = [k.id for k in kpis]
        data_rows = (
            KpiData.query
            .filter(KpiData.process_kpi_id.in_(kpi_ids), KpiData.year == year, KpiData.is_active == True)
            .order_by(KpiData.data_date.desc())
            .all()
        ) if kpi_ids else []

        latest_kpi: dict = {}
        for d in data_rows:
            if d.process_kpi_id not in latest_kpi:
                latest_kpi[d.process_kpi_id] = d

        proc_map = {p.id: p for p in Process.query.filter(
            Process.id.in_({k.process_id for k in kpis}), Process.is_active == True
        ).all()} if kpis else {}

        kritik_pg = []
        for kpi_id, d in latest_kpi.items():
            kpi = kpi_map.get(kpi_id)
            if not kpi or not kpi.target_value or not d.actual_value:
                continue
            try:
                target = float(kpi.target_value)
                actual = float(d.actual_value)
                if target <= 0:
                    continue
                if kpi.direction == "lower_is_better":
                    pct = round(min(100.0, target / actual * 100), 1) if actual > 0 else 0.0
                else:
                    pct = round(min(100.0, actual / target * 100), 1)
                if pct < 50:
                    proc = proc_map.get(kpi.process_id)
                    kritik_pg.append({
                        "kpi_id": kpi_id,
                        "code":   kpi.code or "",
                        "name":   kpi.name,
                        "surec":  proc.name if proc else "—",
                        "hedef":  kpi.target_value,
                        "gercek": d.actual_value,
                        "pct":    pct,
                    })
            except (ValueError, TypeError):
                pass
        kritik_pg.sort(key=lambda x: x["pct"])

        # ── Geciken faaliyetler ─────────────────────────────────────────────────
        geciken_faal = []
        acts = (
            ProcessActivity.query.join(Process)
            .filter(
                Process.tenant_id == tid, Process.is_active == True,
                ProcessActivity.is_active == True,
                ProcessActivity.status.notin_(["Tamamlandı", "İptal"]),
            ).all()
        )
        for a in acts:
            if a.end_date and a.end_date < today:
                geciken_faal.append({
                    "id":          a.id,
                    "name":        a.name,
                    "surec":       a.process.name if a.process else "—",
                    "end_date":    str(a.end_date),
                    "gecikme_gun": (today - a.end_date).days,
                    "status":      a.status,
                    "progress":    a.progress or 0,
                })
        geciken_faal.sort(key=lambda x: -x["gecikme_gun"])

        # ── Yüksek risk (RPN > 10) ─────────────────────────────────────────────
        if _has_risk:
            yuksek_risk = [{
                "id":          r.id,
                "title":       r.title,
                "probability": r.probability,
                "impact":      r.impact,
                "rpn":         r.rpn,
                "status":      r.status or "—",
            } for r in (
                RiskHeatmapItem.query
                .filter(RiskHeatmapItem.tenant_id == tid, RiskHeatmapItem.is_active == True,
                        RiskHeatmapItem.rpn > 10)
                .order_by(RiskHeatmapItem.rpn.desc()).limit(15).all()
            )]
        else:
            yuksek_risk = []

        return jsonify({
            "success": True,
            "data": {
                "kritik_pg":           kritik_pg[:20],
                "geciken_faaliyetler": geciken_faal[:20],
                "yuksek_risk":         yuksek_risk,
                "ozet": {
                    "kritik_pg_sayisi":       len(kritik_pg),
                    "geciken_faaliyet_sayisi": len(geciken_faal),
                    "yuksek_risk_sayisi":      len(yuksek_risk),
                },
            },
            "year": year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_uyari] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Uyarı verisi alınamadı."}), 500


# ── Rapor 11: K-Vektör Ağırlık Dağılımı ─────────────────────────────────────

@app_bp.route("/k-rapor/api/k-vektor")
@login_required
def k_rapor_api_k_vektor():
    """K-Vektör kota dağılımı ve strateji skorları (compute_k_vektor_bundle)."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        try:
            from app.services.k_vektor_engine import compute_k_vektor_bundle
        except ImportError:
            return jsonify({"success": False, "message": "K-Vektör servisi henüz aktif değil."}), 200
        from app.services.plan_year_service import get_plan_year
        from app.models.core import Strategy, SubStrategy
        from sqlalchemy import or_, and_

        plan_year = get_plan_year(tid, year)
        bundle = compute_k_vektor_bundle(
            tenant_id=tid,
            year=year,
            as_of=_dt.date.today(),
            persist_pg_scores=False,
            plan_year=plan_year,
        )

        quotas_main    = bundle.get("k_vektor_quotas_main", {})   # str keys, 1000-scale
        strat_scores   = bundle.get("strategy_scores", {})        # int keys
        sub_scores     = bundle.get("sub_strategy_scores", {})    # int keys

        strat_ids = [int(k) for k in quotas_main.keys()]
        strategies = {s.id: s for s in Strategy.query.filter(Strategy.id.in_(strat_ids)).all()} if strat_ids else {}

        strat_data = []
        for sid_str, quota in sorted(quotas_main.items(), key=lambda x: -float(x[1])):
            sid = int(sid_str)
            s   = strategies.get(sid)
            pct = round(float(quota) / 1000 * 100, 1)
            strat_data.append({
                "id":          sid,
                "code":        s.code  if s else "",
                "title":       s.title if s else f"#{sid}",
                "quota_1000":  round(float(quota), 1),
                "pct":         pct,
                "score":       round(strat_scores.get(sid, 0.0), 1),
            })

        # Alt stratejiler
        if plan_year:
            sub_list = (
                SubStrategy.query.join(Strategy)
                .filter(
                    Strategy.is_active == True,
                    SubStrategy.is_active == True,
                    or_(
                        Strategy.plan_year_id == plan_year.id,
                        and_(Strategy.plan_year_id.is_(None), Strategy.tenant_id == tid),
                    ),
                ).all()
            )
        else:
            sub_list = (
                SubStrategy.query.join(Strategy)
                .filter(Strategy.tenant_id == tid, Strategy.is_active == True, SubStrategy.is_active == True)
                .all()
            )

        sub_data = sorted([{
            "id":          ss.id,
            "parent_id":   ss.strategy_id,
            "parent_code": (strategies.get(ss.strategy_id) or type("", (), {"code": ""})()).code,
            "code":        ss.code or "",
            "title":       ss.title,
            "score":       round(sub_scores.get(ss.id, 0.0), 1),
        } for ss in sub_list], key=lambda x: -x["score"])

        return jsonify({"success": True, "data": {"strateji": strat_data, "alt_strateji": sub_data}})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_k_vektor] {e}", exc_info=True)
        return jsonify({"success": False, "message": "K-Vektör verisi alınamadı."}), 500


# ── Rapor 11: EVM (Kazanılmış Değer) ─────────────────────────────────────────

@app_bp.route("/k-rapor/api/evm")
@login_required
def k_rapor_api_evm():
    """EVM snapshot — CPI / SPI zaman serisi."""
    tid = current_user.tenant_id
    try:
        try:
            from app.models.k_radar_domain import EvmSnapshot
        except ImportError:
            return jsonify({"success": True, "data": []})

        snaps = (
            EvmSnapshot.query
            .filter_by(tenant_id=tid, is_active=True)
            .order_by(EvmSnapshot.snapshot_date.desc())
            .limit(60)
            .all()
        )
        # Proje adlarını çek
        proj_ids = {s.project_id for s in snaps if s.project_id}
        proj_names = {}
        if proj_ids:
            try:
                from app.models.project import Project as ProjModel
                for p in ProjModel.query.filter(ProjModel.id.in_(proj_ids)).all():
                    proj_names[p.id] = p.name
            except Exception:
                pass
        data = [{
            "id":            s.id,
            "project_id":    s.project_id,
            "project_name":  proj_names.get(s.project_id, f"Proje #{s.project_id}"),
            "snapshot_date": str(s.snapshot_date),
            "pv":  s.pv,
            "ev":  s.ev,
            "ac":  s.ac,
            "spi": round(s.spi, 2) if s.spi is not None else None,
            "cpi": round(s.cpi, 2) if s.cpi is not None else None,
        } for s in snaps]

        return jsonify({"success": True, "data": data})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_evm] {e}")
        return jsonify({"success": False, "message": "EVM verisi alınamadı."}), 500


# ── Rapor 12: Stratejik Analiz Özeti ─────────────────────────────────────────

@app_bp.route("/k-rapor/api/stratejik-analiz")
@login_required
def k_rapor_api_stratejik_analiz():
    """SWOT / TOWS / PESTEL / Porter özet — yıl yoksa en son analizi döner."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        import json as _json
        from app.models.swot import SwotAnalysis, TowsAnalysis, PestelAnalysis, PorterFiveForcesAnalysis
        from app.models.plan_year import PlanYear

        # Önce istenen yılın plan_year'ını bul; yoksa tenant'ın tüm plan_year'larını ara
        py = PlanYear.query.filter_by(tenant_id=tid, year=year).first()
        all_py_ids = [r.id for r in PlanYear.query.filter_by(tenant_id=tid).order_by(PlanYear.year.desc()).all()]

        def _cnt(js):
            try: return len(_json.loads(js or "[]"))
            except Exception: return 0

        def _score(js):
            try: return _json.loads(js or "{}").get("score")
            except Exception: return None

        def _year_label(plan_year_id):
            p = PlanYear.query.get(plan_year_id)
            return p.year if p else None

        def _latest(model, py_id_pref):
            """Önce istenen plan_year, yoksa tenant'ın herhangi bir plan_year'ı."""
            if py_id_pref:
                r = model.query.filter_by(tenant_id=tid, plan_year_id=py_id_pref).first()
                if r:
                    return r
            if all_py_ids:
                return (
                    model.query
                    .filter(model.tenant_id == tid, model.plan_year_id.in_(all_py_ids))
                    .order_by(model.updated_at.desc())
                    .first()
                )
            return None

        py_id = py.id if py else None
        swot   = _latest(SwotAnalysis, py_id)
        tows   = _latest(TowsAnalysis, py_id)
        pestel = _latest(PestelAnalysis, py_id)
        porter = _latest(PorterFiveForcesAnalysis, py_id)

        return jsonify({
            "success": True,
            "data": {
                "swot": {
                    "mevcut":        swot is not None,
                    "year_label":    _year_label(swot.plan_year_id) if swot else None,
                    "strengths":     _cnt(swot.strengths)     if swot else 0,
                    "weaknesses":    _cnt(swot.weaknesses)    if swot else 0,
                    "opportunities": _cnt(swot.opportunities) if swot else 0,
                    "threats":       _cnt(swot.threats)       if swot else 0,
                    "guncelleme":    str(swot.updated_at)[:10] if swot and swot.updated_at else None,
                },
                "tows": {
                    "mevcut":     tows is not None,
                    "year_label": _year_label(tows.plan_year_id) if tows else None,
                    "so": _cnt(tows.so_strategies) if tows else 0,
                    "st": _cnt(tows.st_strategies) if tows else 0,
                    "wo": _cnt(tows.wo_strategies) if tows else 0,
                    "wt": _cnt(tows.wt_strategies) if tows else 0,
                    "guncelleme": str(tows.updated_at)[:10] if tows and tows.updated_at else None,
                },
                "pestel": {
                    "mevcut":        pestel is not None,
                    "year_label":    _year_label(pestel.plan_year_id) if pestel else None,
                    "political":     _cnt(pestel.political)     if pestel else 0,
                    "economic":      _cnt(pestel.economic)      if pestel else 0,
                    "social":        _cnt(pestel.social)        if pestel else 0,
                    "technological": _cnt(pestel.technological) if pestel else 0,
                    "environmental": _cnt(pestel.environmental) if pestel else 0,
                    "legal":         _cnt(pestel.legal)         if pestel else 0,
                    "guncelleme":    str(pestel.updated_at)[:10] if pestel and pestel.updated_at else None,
                },
                "porter": {
                    "mevcut":      porter is not None,
                    "year_label":  _year_label(porter.plan_year_id) if porter else None,
                    "rivalry":     _score(porter.rivalry_intensity)  if porter else None,
                    "supplier":    _score(porter.supplier_power)     if porter else None,
                    "buyer":       _score(porter.buyer_power)        if porter else None,
                    "new_entrant": _score(porter.new_entrant_threat) if porter else None,
                    "substitute":  _score(porter.substitute_threat)  if porter else None,
                    "guncelleme":  str(porter.updated_at)[:10] if porter and porter.updated_at else None,
                },
            },
            "year": year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_stratejik_analiz] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Stratejik analiz verisi alınamadı."}), 500


# ── Rapor 13: Paydaş Haritası ─────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/paydas")
@login_required
def k_rapor_api_paydas():
    """Paydaş haritası + anket özeti."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.k_radar_domain import StakeholderMap, StakeholderSurvey
        from app.models.plan_year import PlanYear
        from collections import defaultdict

        py    = PlanYear.query.filter_by(tenant_id=tid, year=year).first()
        py_id = py.id if py else None

        stakeholders = StakeholderMap.query.filter_by(tenant_id=tid, is_active=True).all()
        if py_id:
            stakeholders = [s for s in stakeholders if s.plan_year_id in (py_id, None)]

        surveys = StakeholderSurvey.query.filter_by(tenant_id=tid, is_active=True).all()

        stk_list = [{
            "id": s.id, "name": s.name, "role": s.role or "—",
            "influence": s.influence, "interest": s.interest,
            "strategy": s.strategy or "—",
        } for s in stakeholders]

        type_scores = defaultdict(list)
        for s in surveys:
            if s.score is not None:
                type_scores[s.stakeholder_type].append(s.score)
        type_avg = [
            {"tip": t, "ort_skor": round(sum(v) / len(v), 1), "sayi": len(v)}
            for t, v in type_scores.items()
        ]

        return jsonify({
            "success": True,
            "data": {
                "paydas_listesi":   stk_list,
                "anket_ozeti":      type_avg,
                "toplam_paydas":    len(stk_list),
                "toplam_anket":     len(surveys),
            },
            "year": year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_paydas] {e}")
        return jsonify({"success": False, "message": "Paydaş verisi alınamadı."}), 500


# ── Rapor 14: Rekabet & A3 ────────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/rekabet")
@login_required
def k_rapor_api_rekabet():
    """Rekabetçi analiz + A3 raporları özeti."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.k_radar_domain import CompetitorAnalysis, A3Report
        from app.models.plan_year import PlanYear
        from collections import defaultdict

        py    = PlanYear.query.filter_by(tenant_id=tid, year=year).first()
        py_id = py.id if py else None

        competitors = CompetitorAnalysis.query.filter_by(tenant_id=tid, is_active=True).all()
        if py_id:
            competitors = [c for c in competitors if c.plan_year_id in (py_id, None)]

        a3_reports = (
            A3Report.query.filter_by(tenant_id=tid, is_active=True)
            .order_by(A3Report.created_at.desc()).limit(20).all()
        )

        comp_dim = defaultdict(list)
        for c in competitors:
            comp_dim[c.competitor_name].append({
                "dimension": c.dimension or "—",
                "our_score": c.our_score,
                "their_score": c.their_score,
            })
        comp_list = [{"competitor": k, "dimensions": v} for k, v in comp_dim.items()]

        a3_list = [{
            "id": r.id,
            "problem":        (r.problem or "")[:120],
            "countermeasures": (r.countermeasures or "")[:80],
            "source_type":    r.source_type or "—",
            "tarih":          str(r.created_at)[:10] if r.created_at else None,
        } for r in a3_reports]

        return jsonify({
            "success": True,
            "data": {
                "rekabet_listesi": comp_list,
                "a3_listesi":      a3_list,
                "toplam_rakip":    len(comp_list),
                "toplam_a3":       len(a3_list),
            },
            "year": year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_rekabet] {e}")
        return jsonify({"success": False, "message": "Rekabet/A3 verisi alınamadı."}), 500


# ── Export: Excel ─────────────────────────────────────────────────────────────

@app_bp.route("/k-rapor/api/export-excel")
@login_required
def k_rapor_api_export_excel():
    """Seçili tab verisini Excel olarak indir."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    tab  = request.args.get("tab", "kurumsal")
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file

        wb = Workbook()
        ws = wb.active

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="6366F1")
        center_align = Alignment(horizontal="center")

        def _hdr(sheet, cols):
            sheet.append(cols)
            for cell in sheet[1]:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align

        if tab == "veri-durumu":
            ws.title = "Veri Durumu"
            from app.models.process import Process, ProcessKpi, KpiData
            from app.models.core import User
            kpis = (ProcessKpi.query.join(Process)
                    .filter(Process.tenant_id == tid, Process.is_active == True,
                            ProcessKpi.is_active == True).all())
            kpi_ids = [k.id for k in kpis]
            latest_data = (KpiData.query
                           .filter(KpiData.process_kpi_id.in_(kpi_ids),
                                   KpiData.year == year, KpiData.is_active == True)
                           .order_by(KpiData.data_date.desc()).all())
            latest_by_kpi = {}
            for d in latest_data:
                if d.process_kpi_id not in latest_by_kpi:
                    latest_by_kpi[d.process_kpi_id] = d
            user_ids = {d.user_id for d in latest_data if d.user_id}
            users = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()} if user_ids else {}
            proc_ids = list({k.process_id for k in kpis})
            procs = {p.id: p for p in Process.query.filter(Process.id.in_(proc_ids)).all()}
            _hdr(ws, ["PG Kodu", "PG Adı", "Süreç", "Son Tarih", "Son Değer", "Giren", "Durum"])
            for k in kpis:
                proc = procs.get(k.process_id)
                d    = latest_by_kpi.get(k.id)
                u    = users.get(d.user_id) if d and d.user_id else None
                ws.append([
                    k.code or "", k.name,
                    proc.name if proc else "",
                    str(d.data_date) if d and d.data_date else "",
                    d.actual_value if d else "",
                    f"{u.first_name or ''} {u.last_name or ''}".strip() if u else "",
                    "Girildi" if d else "Eksik",
                ])

        elif tab == "bireysel":
            ws.title = "Bireysel Performans"
            from app.models.process import IndividualPerformanceIndicator as IPI, IndividualKpiData
            from app.models.core import User
            from app.models.plan_year import PlanYear
            users    = User.query.filter_by(tenant_id=tid, is_active=True).all()
            user_map = {u.id: u for u in users}
            uid_list = [u.id for u in users]
            py  = PlanYear.query.filter_by(tenant_id=tid, year=year).first()
            pgs = IPI.query.filter(
                IPI.user_id.in_(uid_list), IPI.is_active == True,
                *([IPI.plan_year_id == py.id] if py else [])
            ).all()
            pg_ids    = [pg.id for pg in pgs]
            data_rows = IndividualKpiData.query.filter(
                IndividualKpiData.individual_pg_id.in_(pg_ids),
                IndividualKpiData.year == year
            ).all() if pg_ids else []
            latest_by_pg = {}
            for d in data_rows:
                prev = latest_by_pg.get(d.individual_pg_id)
                if prev is None or (d.data_date and (prev.data_date is None or d.data_date > prev.data_date)):
                    latest_by_pg[d.individual_pg_id] = d
            _hdr(ws, ["Kullanıcı", "E-posta", "PG Kodu", "PG Adı", "Hedef", "Gerçekleşen", "Başarı %", "Kaynak"])
            for pg in pgs:
                u   = user_map.get(pg.user_id)
                d   = latest_by_pg.get(pg.id)
                pct = None
                if d and d.status_percentage is not None:
                    try:
                        pct = round(float(d.status_percentage), 1)
                    except Exception:
                        pass
                ws.append([
                    f"{u.first_name or ''} {u.last_name or ''}".strip() if u else "",
                    u.email if u else "",
                    pg.code or "", pg.name,
                    pg.target_value,
                    d.actual_value if d else "",
                    pct,
                    pg.source or "",
                ])

        elif tab == "faaliyet":
            ws.title = "Faaliyetler"
            from app.models.process import Process, ProcessActivity
            today      = _dt.date.today()
            activities = (ProcessActivity.query.join(Process)
                          .filter(Process.tenant_id == tid, Process.is_active == True,
                                  ProcessActivity.is_active == True).all())
            _hdr(ws, ["Faaliyet", "Süreç", "Durum", "Başlangıç", "Bitiş", "İlerleme %", "Gecikme (gün)"])
            for a in activities:
                gecikme = (
                    (today - a.end_date).days
                    if (a.end_date and a.end_date < today and a.status not in ("Tamamlandı", "İptal"))
                    else 0
                )
                ws.append([
                    a.name,
                    a.process.name if a.process else "",
                    a.status,
                    str(a.start_date) if a.start_date else "",
                    str(a.end_date)   if a.end_date   else "",
                    a.progress or 0,
                    gecikme,
                ])

        else:  # kurumsal — süreç skorları
            ws.title = "Kurumsal Performans"
            from app.services.score_engine_service import compute_vision_score
            from app.services.plan_year_service import get_plan_year
            from app.models.process import Process
            plan_year      = get_plan_year(tid, year)
            result         = compute_vision_score(tid, year=year, persist_pg_scores=False, plan_year=plan_year)
            process_scores = result.get("process_scores") or {}
            proc_map       = {}
            if process_scores:
                for p in Process.query.filter(
                    Process.id.in_(list(process_scores.keys())), Process.is_active == True
                ).all():
                    proc_map[p.id] = p
            _hdr(ws, ["Süreç Kodu", "Süreç Adı", "Skor", "Durum"])
            for pid, sc in sorted(process_scores.items(), key=lambda x: -x[1]):
                p     = proc_map.get(pid)
                durum = "Hedefte" if sc >= 80 else ("Riskli" if sc >= 50 else "Kritik")
                ws.append([p.code if p else "", p.name if p else f"#{pid}", round(sc, 1), durum])

        # Sütun genişliklerini otomatik ayarla
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"k-rapor-{tab}-{year}.xlsx"
        return send_file(
            buf, as_attachment=True, download_name=fname,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_export_excel] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Excel oluşturulamadı."}), 500


# ══════════════════════════════════════════════════════════════════════════════
# KPI ANOMALI TESPİT + SLACK BİLDİRİM (Sprint 14-15)
# ══════════════════════════════════════════════════════════════════════════════

@app_bp.route("/k-rapor/anomalies")
@login_required
def k_rapor_anomalies_page():
    """KPI anomali UI sayfası."""
    return render_template("platform/k_rapor/anomalies.html")


@app_bp.route("/k-rapor/api/anomalies")
@login_required
def k_rapor_api_anomalies():
    """Aktif tenant için KPI anomali listesi (Z-score tabanlı)."""
    from app.services.kpi_anomaly_service import detect_anomalies_for_tenant

    tid = current_user.tenant_id
    threshold = request.args.get("threshold", 2.0, type=float)
    limit = request.args.get("limit", 50, type=int)

    try:
        anomalies = detect_anomalies_for_tenant(tid, threshold=threshold, limit=limit)
        return jsonify({
            "success": True,
            "threshold": threshold,
            "count": len(anomalies),
            "data": [a.to_dict() for a in anomalies],
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_anomalies] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Anomali tespiti başarısız."}), 500


@app_bp.route("/k-rapor/api/webhook/test", methods=["POST"])
@login_required
def k_rapor_webhook_test():
    """Sprint 45: Test mesajı gönder (Slack/Teams/Discord).

    Body: {"provider": "slack|teams|discord", "webhook_url": "...", "message": "Test"}
    """
    payload = request.get_json(silent=True) or {}
    provider = (payload.get("provider") or "slack").lower()
    url = (payload.get("webhook_url") or "").strip()
    msg = payload.get("message") or f"🚀 Kokpitim test mesajı — {current_user.email}"

    if not url:
        return jsonify({"success": False, "message": "webhook_url gerekli"}), 400

    from app.services.slack_notification import dispatch_webhook
    result = dispatch_webhook(provider, msg, url)
    return jsonify(result)


@app_bp.route("/k-rapor/api/anomalies/notify-webhook", methods=["POST"])
@login_required
def k_rapor_anomalies_notify_webhook():
    """Sprint 45: Anomalileri generic webhook ile gönder.

    Body: {"provider": "slack|teams|discord", "webhook_url": "...",
           "severity_min": "medium", "limit": 5}
    """
    from app.services.kpi_anomaly_service import detect_anomalies_for_tenant
    from app.services.slack_notification import dispatch_webhook

    tid = current_user.tenant_id
    payload = request.get_json(silent=True) or {}
    provider = (payload.get("provider") or "slack").lower()
    webhook = (payload.get("webhook_url") or "").strip()
    if not webhook:
        return jsonify({"success": False, "message": "webhook_url gerekli"}), 400

    sev_min = payload.get("severity_min", "medium").lower()
    limit = int(payload.get("limit", 5))
    rank = {"low": 1, "medium": 2, "high": 3}
    min_r = rank.get(sev_min, 2)

    try:
        anomalies = detect_anomalies_for_tenant(tid, threshold=2.0, limit=50)
        filtered = [a for a in anomalies if rank.get(a.severity, 0) >= min_r][:limit]
        if not filtered:
            return jsonify({"success": True, "message": "Bildirilecek anomali yok.", "sent": 0})

        lines = [f"⚠️ *{len(filtered)} KPI anomalisi tespit edildi* (severity ≥ {sev_min})"]
        for a in filtered:
            arrow = "↓" if a.z_score < 0 else "↑"
            lines.append(
                f"• `{a.process_code}` {a.kpi_name}: {a.latest_value} {arrow} "
                f"(ort: {a.mean}, z={a.z_score})"
            )
        msg = "\n".join(lines)

        result = dispatch_webhook(provider, msg, webhook)
        return jsonify({**result, "anomaly_count": len(filtered)})
    except Exception as e:
        current_app.logger.error(f"[notify_webhook] {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500


@app_bp.route("/k-rapor/api/anomalies/notify-slack", methods=["POST"])
@login_required
def k_rapor_api_anomalies_notify_slack():
    """Tespit edilen anomalileri Slack'e gönder.

    Body: {"webhook_url": "...", "severity_min": "medium", "limit": 5}
    """
    from app.services.kpi_anomaly_service import detect_anomalies_for_tenant
    from app.services.slack_notification import send_slack_message, format_anomaly_blocks

    tid = current_user.tenant_id
    payload = request.get_json(silent=True) or {}
    webhook = (payload.get("webhook_url") or "").strip() or None
    severity_min = payload.get("severity_min", "medium").lower()
    limit = int(payload.get("limit", 5))

    severity_rank = {"low": 1, "medium": 2, "high": 3}
    min_rank = severity_rank.get(severity_min, 2)

    try:
        anomalies = detect_anomalies_for_tenant(tid, threshold=2.0, limit=50)
        filtered = [a for a in anomalies if severity_rank.get(a.severity, 0) >= min_rank][:limit]

        if not filtered:
            return jsonify({"success": True, "message": "Bildirilecek anomali yok.", "sent": 0})

        # Header mesajı + her anomali için block
        from app.models.core import Tenant
        tenant = Tenant.query.get(tid)
        header_msg = (
            f":bar_chart: *Kokpitim Anomali Raporu* — {tenant.name if tenant else 'Kurum'}\n"
            f"{len(filtered)} anomali tespit edildi (severity ≥ {severity_min})"
        )
        result = send_slack_message(header_msg, tenant_id=tid, webhook_url=webhook)
        if not result["success"]:
            return jsonify({"success": False, "message": f"Slack: {result['message']}"}), 400

        # Her anomali için ayrı block message
        sent = 1
        for a in filtered:
            blocks = format_anomaly_blocks(a.to_dict())
            r = send_slack_message(f"Anomali: {a.kpi_name}", tenant_id=tid,
                                    webhook_url=webhook, blocks=blocks)
            if r["success"]:
                sent += 1

        return jsonify({"success": True, "sent": sent, "anomaly_count": len(filtered)})
    except Exception as e:
        current_app.logger.error(f"[anomalies_notify_slack] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Slack gönderimi başarısız."}), 500


# ══════════════════════════════════════════════════════════════════════════════
# EMAIL DIGEST (Sprint 18)
# ══════════════════════════════════════════════════════════════════════════════

@app_bp.route("/k-rapor/api/digest/preview")
@login_required
def k_rapor_digest_preview():
    """Digest mail içeriğini HTML olarak preview et."""
    from app.services.email_digest_service import render_digest_html
    try:
        html = render_digest_html(current_user.tenant_id)
        if not html:
            return jsonify({"success": False, "message": "Tenant bulunamadı"}), 404
        return html  # render HTML
    except Exception as e:
        current_app.logger.error(f"[digest_preview] {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500


@app_bp.route("/k-rapor/api/digest/send", methods=["POST"])
@login_required
def k_rapor_digest_send():
    """Digest mail'i tenant yöneticilerine gönder.

    Sadece tenant_admin + executive_manager + Admin role'leri tetikleyebilir.
    Payload (opsiyonel): {"recipients": ["email1", "email2"]}
    """
    _MANAGE = ("tenant_admin", "executive_manager", "Admin")
    if not current_user.role or current_user.role.name not in _MANAGE:
        return jsonify({"success": False, "message": "Yetkisiz"}), 403

    from app.services.email_digest_service import send_digest
    payload = request.get_json(silent=True) or {}
    recipients = payload.get("recipients") or None

    try:
        result = send_digest(current_user.tenant_id, recipients=recipients)
        return jsonify(result)
    except Exception as e:
        current_app.logger.error(f"[digest_send] {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# PDF EXPORT (Sprint 11.3 — app/utils/pdf_export.py altyapısı)
# ══════════════════════════════════════════════════════════════════════════════

@app_bp.route("/k-rapor/api/export-pdf")
@login_required
def k_rapor_api_export_pdf():
    """K-Rapor kurumsal özet PDF (executive summary).

    Yapı: Tenant header + Genel özet tablosu + Süreç skorları + KPI durumu
    """
    from app.utils.pdf_export import make_pdf, kvp_table
    from app.models.process import Process, ProcessKpi
    from app.models.core import Strategy, Tenant
    from flask import send_file
    import io

    tid = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)

    try:
        tenant = Tenant.query.get(tid)
        if not tenant:
            return jsonify({"success": False, "message": "Kurum bulunamadı."}), 404

        # Özet sayılar
        strategy_count = Strategy.query.filter_by(tenant_id=tid, is_active=True).count()
        process_count = Process.query.filter_by(tenant_id=tid, is_active=True).count()
        kpi_count = (
            ProcessKpi.query.join(Process)
            .filter(Process.tenant_id == tid, ProcessKpi.is_active == True)
            .count()
        )

        sections = [
            {
                "heading": "Genel Özet",
                "body": (
                    f"Kurum vizyonu, stratejik plan dönemi ve operasyonel performansın "
                    f"konsolide görünümü. Rapor tarihi: {_dt.date.today().strftime('%d %B %Y')}."
                ),
                "table": kvp_table([
                    ("Plan yılı", year),
                    ("Stratejik hedef sayısı", strategy_count),
                    ("Aktif süreç sayısı", process_count),
                    ("Toplam performans göstergesi", kpi_count),
                    ("Aktif kullanıcı sayısı", tenant.employee_count or "—"),
                ]),
            },
        ]
        if tenant.vision:
            sections.insert(0, {"heading": "Vizyon", "body": tenant.vision})
        if tenant.purpose:
            sections.insert(1, {"heading": "Amaç (Misyon)", "body": tenant.purpose})

        pdf_bytes = make_pdf(
            title="K-Rapor Kurumsal Özet",
            sections=sections,
            tenant_name=tenant.name,
            footer=f"Bu rapor Kokpitim platformu tarafından otomatik üretilmiştir · {_dt.date.today().isoformat()}",
        )

        return send_file(
            io.BytesIO(pdf_bytes),
            as_attachment=True,
            download_name=f"k-rapor-ozet-{year}.pdf",
            mimetype="application/pdf",
        )
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_export_pdf] {e}", exc_info=True)
        return jsonify({"success": False, "message": "PDF oluşturulamadı."}), 500


# ══════════════════════════════════════════════════════════════════════════════
# YENİ RAPORLAR (1-8)
# ══════════════════════════════════════════════════════════════════════════════

# ── Yeni Rapor 1: PG Performans Dağılımı (Histogram) ─────────────────────────

@app_bp.route("/k-rapor/api/pg-dagilim")
@login_required
def k_rapor_api_pg_dagilim():
    """Tüm PG'lerin başarı yüzdesi dağılımı — histogram + özet."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.process import Process, ProcessKpi, KpiData

        kpis = (
            ProcessKpi.query.join(Process)
            .filter(Process.tenant_id == tid, Process.is_active == True, ProcessKpi.is_active == True)
            .all()
        )
        kpi_ids = [k.id for k in kpis]
        if not kpi_ids:
            return jsonify({"success": True, "data": {"buckets": [], "ozet": {}, "scatter": []}})

        latest_data = (
            KpiData.query
            .filter(KpiData.process_kpi_id.in_(kpi_ids), KpiData.year == year, KpiData.is_active == True)
            .order_by(KpiData.data_date.desc())
            .all()
        )
        latest_by_kpi = {}
        for d in latest_data:
            if d.process_kpi_id not in latest_by_kpi:
                latest_by_kpi[d.process_kpi_id] = d

        kpi_map  = {k.id: k for k in kpis}
        proc_ids = {k.process_id for k in kpis}
        procs    = {p.id: p for p in Process.query.filter(Process.id.in_(proc_ids)).all()}

        # Başarı yüzdelerini hesapla
        scores = []
        scatter = []
        for kpi_id, d in latest_by_kpi.items():
            kpi = kpi_map.get(kpi_id)
            if not kpi or not kpi.target_value or not d.actual_value:
                continue
            try:
                target = float(kpi.target_value)
                actual = float(d.actual_value)
                if target <= 0:
                    continue
                if getattr(kpi, 'direction', None) == 'lower_is_better':
                    pct = round(min(100.0, target / actual * 100), 1) if actual > 0 else 0.0
                else:
                    pct = round(min(100.0, actual / target * 100), 1)
                scores.append(pct)
                proc = procs.get(kpi.process_id)
                scatter.append({
                    "kpi_id":   kpi_id,
                    "kpi_name": kpi.name,
                    "surec":    proc.name if proc else "—",
                    "pct":      pct,
                    "actual":   actual,
                    "target":   target,
                })
            except (ValueError, TypeError):
                pass

        # Histogram bucket'ları: 0-10, 10-20, ..., 90-100
        buckets = [{"label": f"%{i*10}–{(i+1)*10}", "min": i*10, "max": (i+1)*10, "count": 0} for i in range(10)]
        for s in scores:
            idx = min(int(s // 10), 9)
            buckets[idx]["count"] += 1

        toplam = len(scores)
        ozet = {
            "toplam":    toplam,
            "ort":       round(sum(scores) / toplam, 1) if toplam else 0,
            "medyan":    round(sorted(scores)[toplam // 2], 1) if toplam else 0,
            "yesil":     sum(1 for s in scores if s >= 80),
            "sari":      sum(1 for s in scores if 50 <= s < 80),
            "kirmizi":   sum(1 for s in scores if s < 50),
            "veri_yok":  len(kpi_ids) - toplam,
        }
        scatter.sort(key=lambda x: x["pct"])

        return jsonify({"success": True, "data": {"buckets": buckets, "ozet": ozet, "scatter": scatter[:100]}})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_pg_dagilim] {e}", exc_info=True)
        return jsonify({"success": False, "message": "PG dağılım verisi alınamadı."}), 500


# ── Yeni Rapor 2: Süreç Faaliyet Matrisi ─────────────────────────────────────

@app_bp.route("/k-rapor/api/faaliyet-matris")
@login_required
def k_rapor_api_faaliyet_matris():
    """Her süreç için faaliyet durumu dağılımı — yatay bar chart."""
    tid  = current_user.tenant_id
    try:
        from app.models.process import Process, ProcessActivity
        today = _dt.date.today()

        processes = Process.query.filter_by(tenant_id=tid, is_active=True).order_by(Process.code).all()
        proc_ids  = [p.id for p in processes]
        if not proc_ids:
            return jsonify({"success": True, "data": []})

        activities = (
            ProcessActivity.query
            .filter(ProcessActivity.process_id.in_(proc_ids), ProcessActivity.is_active == True)
            .all()
        )

        # Süreç bazlı grupla
        from collections import defaultdict
        proc_acts = defaultdict(list)
        for a in activities:
            proc_acts[a.process_id].append(a)

        rows = []
        for p in processes:
            acts = proc_acts.get(p.id, [])
            if not acts:
                continue
            tamamlanan = sum(1 for a in acts if a.status == "Tamamlandı")
            devam      = sum(1 for a in acts if a.status == "Devam Ediyor")
            planlandi  = sum(1 for a in acts if a.status == "Planlandı")
            geciken    = sum(1 for a in acts
                            if a.status not in ("Tamamlandı", "İptal")
                            and a.end_date and a.end_date < today)
            iptal      = sum(1 for a in acts if a.status == "İptal")
            toplam     = len(acts)
            rows.append({
                "id":          p.id,
                "code":        p.code or "",
                "name":        p.name,
                "toplam":      toplam,
                "tamamlanan":  tamamlanan,
                "devam":       devam,
                "planlandi":   planlandi,
                "geciken":     geciken,
                "iptal":       iptal,
                "tamamlanma_pct": round(tamamlanan / toplam * 100, 1) if toplam else 0,
            })

        rows.sort(key=lambda x: -x["geciken"])
        return jsonify({"success": True, "data": rows})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_faaliyet_matris] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Faaliyet matris verisi alınamadı."}), 500


# ── Yeni Rapor 3: Veri Giriş Aktivite Takvimi ────────────────────────────────

@app_bp.route("/k-rapor/api/aktivite-takvim")
@login_required
def k_rapor_api_aktivite_takvim():
    """Son 365 günde günlük veri giriş sayısı — GitHub heatmap tarzı."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.process import Process, ProcessKpi, KpiData
        from sqlalchemy import func

        # Süreç PG verileri
        kpis = (
            ProcessKpi.query.join(Process)
            .filter(Process.tenant_id == tid, Process.is_active == True, ProcessKpi.is_active == True)
            .all()
        )
        kpi_ids = [k.id for k in kpis]

        # Günlük sayım
        daily_counts = {}
        if kpi_ids:
            rows = (
                db.session.query(KpiData.data_date, func.count(KpiData.id))
                .filter(
                    KpiData.process_kpi_id.in_(kpi_ids),
                    KpiData.year == year,
                    KpiData.is_active == True,
                )
                .group_by(KpiData.data_date)
                .all()
            )
            for date, cnt in rows:
                if date:
                    daily_counts[str(date)] = cnt

        # Bireysel PG verileri de ekle
        try:
            from app.models.process import IndividualPerformanceIndicator as IPI, IndividualKpiData
            from app.models.core import User
            uid_list = [u.id for u in User.query.filter_by(tenant_id=tid, is_active=True).all()]
            if uid_list:
                ipi_ids = [pg.id for pg in IPI.query.filter(IPI.user_id.in_(uid_list), IPI.is_active == True).all()]
                if ipi_ids:
                    ipi_rows = (
                        db.session.query(IndividualKpiData.data_date, func.count(IndividualKpiData.id))
                        .filter(IndividualKpiData.individual_pg_id.in_(ipi_ids), IndividualKpiData.year == year)
                        .group_by(IndividualKpiData.data_date)
                        .all()
                    )
                    for date, cnt in ipi_rows:
                        if date:
                            key = str(date)
                            daily_counts[key] = daily_counts.get(key, 0) + cnt
        except Exception:
            pass

        # Max değer (renk skalası için)
        max_val = max(daily_counts.values()) if daily_counts else 1

        return jsonify({
            "success": True,
            "data": {
                "daily": daily_counts,
                "max_val": max_val,
                "toplam_gun": len(daily_counts),
                "toplam_giris": sum(daily_counts.values()),
            },
            "year": year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_aktivite_takvim] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Aktivite takvim verisi alınamadı."}), 500


# ── Yeni Rapor 4: Kurum Karşılaştırma ────────────────────────────────────────

@app_bp.route("/k-rapor/api/kurum-karsilastirma")
@login_required
def k_rapor_api_kurum_karsilastirma():
    """Kurumları ortalama PG başarısına göre karşılaştır (yalnız Admin)."""
    from app.models.core import User
    role_name = current_user.role.name if current_user.role else ""
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.process import Process, ProcessKpi, KpiData
        from app.models.core import Tenant

        # Admin tüm kurumları görür, diğerleri sadece kendi kurumunu
        if role_name == "Admin":
            tenants = Tenant.query.filter_by(is_active=True).all()
        else:
            tenants = Tenant.query.filter_by(id=current_user.tenant_id, is_active=True).all()

        results = []
        for tenant in tenants:
            kpis = (
                ProcessKpi.query.join(Process)
                .filter(Process.tenant_id == tenant.id, Process.is_active == True, ProcessKpi.is_active == True)
                .all()
            )
            kpi_ids = [k.id for k in kpis]
            if not kpi_ids:
                results.append({
                    "tenant_id": tenant.id,
                    "name": tenant.short_name or tenant.name,
                    "pg_sayisi": 0,
                    "veri_girilen": 0,
                    "ort_basari": None,
                    "yesil": 0, "sari": 0, "kirmizi": 0,
                })
                continue

            latest_data = (
                KpiData.query
                .filter(KpiData.process_kpi_id.in_(kpi_ids), KpiData.year == year, KpiData.is_active == True)
                .order_by(KpiData.data_date.desc())
                .all()
            )
            latest_by_kpi = {}
            for d in latest_data:
                if d.process_kpi_id not in latest_by_kpi:
                    latest_by_kpi[d.process_kpi_id] = d

            kpi_map = {k.id: k for k in kpis}
            scores = []
            for kpi_id, d in latest_by_kpi.items():
                kpi = kpi_map.get(kpi_id)
                if not kpi or not kpi.target_value or not d.actual_value:
                    continue
                try:
                    target = float(kpi.target_value)
                    actual = float(d.actual_value)
                    if target <= 0:
                        continue
                    if getattr(kpi, 'direction', None) == 'lower_is_better':
                        pct = round(min(100.0, target / actual * 100), 1) if actual > 0 else 0.0
                    else:
                        pct = round(min(100.0, actual / target * 100), 1)
                    scores.append(pct)
                except (ValueError, TypeError):
                    pass

            results.append({
                "tenant_id":   tenant.id,
                "name":        tenant.short_name or tenant.name,
                "pg_sayisi":   len(kpi_ids),
                "veri_girilen": len(latest_by_kpi),
                "ort_basari":  round(sum(scores) / len(scores), 1) if scores else None,
                "yesil":       sum(1 for s in scores if s >= 80),
                "sari":        sum(1 for s in scores if 50 <= s < 80),
                "kirmizi":     sum(1 for s in scores if s < 50),
            })

        results.sort(key=lambda x: x["ort_basari"] or -1, reverse=True)
        return jsonify({"success": True, "data": results, "year": year, "is_admin": role_name == "Admin"})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_kurum_karsilastirma] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Kurum karşılaştırma verisi alınamadı."}), 500


# ── Yeni Rapor 5: Strateji Kapsama Analizi ───────────────────────────────────

@app_bp.route("/k-rapor/api/strateji-kapsama")
@login_required
def k_rapor_api_strateji_kapsama():
    """Hangi stratejilerin altında süreç yok, hangi süreçler stratejisiz."""
    tid  = current_user.tenant_id
    year = request.args.get("year", _dt.date.today().year, type=int)
    try:
        from app.models.core import Strategy, SubStrategy
        from app.models.process import Process

        strategies   = Strategy.query.filter_by(tenant_id=tid, is_active=True).order_by(Strategy.code).all()
        sub_strats   = SubStrategy.query.join(Strategy).filter(
            Strategy.tenant_id == tid, Strategy.is_active == True, SubStrategy.is_active == True
        ).all()
        all_processes = Process.query.filter_by(tenant_id=tid, is_active=True).all()

        # Alt strateji → bağlı süreç sayısı
        ss_proc_count = {}
        for ss in sub_strats:
            linked = [p for p in (ss.processes or []) if p.is_active and p.tenant_id == tid]
            ss_proc_count[ss.id] = len(linked)

        # Strateji bazlı özet
        strat_data = []
        for s in strategies:
            subs = [ss for ss in sub_strats if ss.strategy_id == s.id]
            total_procs = sum(ss_proc_count.get(ss.id, 0) for ss in subs)
            bos_subs    = [ss for ss in subs if ss_proc_count.get(ss.id, 0) == 0]
            strat_data.append({
                "id":          s.id,
                "code":        s.code or "",
                "title":       s.title,
                "alt_strateji_sayisi": len(subs),
                "bagli_surec_sayisi":  total_procs,
                "bos_alt_strateji":    len(bos_subs),
                "bos_alt_strat_list":  [{"code": ss.code or "", "title": ss.title} for ss in bos_subs],
                "durum": "tam" if total_procs > 0 and len(bos_subs) == 0
                         else ("kismi" if total_procs > 0 else "bos"),
            })

        # Hiçbir alt stratejiye bağlı olmayan süreçler
        linked_proc_ids = set()
        for ss in sub_strats:
            for p in (ss.processes or []):
                if p.is_active and p.tenant_id == tid:
                    linked_proc_ids.add(p.id)

        stratejisiz = [
            {"id": p.id, "code": p.code or "", "name": p.name}
            for p in all_processes if p.id not in linked_proc_ids
        ]

        ozet = {
            "toplam_strateji":    len(strategies),
            "tam_kapsamli":       sum(1 for s in strat_data if s["durum"] == "tam"),
            "kismi_kapsamli":     sum(1 for s in strat_data if s["durum"] == "kismi"),
            "bos_strateji":       sum(1 for s in strat_data if s["durum"] == "bos"),
            "stratejisiz_surec":  len(stratejisiz),
            "toplam_surec":       len(all_processes),
        }

        return jsonify({
            "success": True,
            "data": {"stratejiler": strat_data, "stratejisiz_surecler": stratejisiz, "ozet": ozet},
            "year": year,
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_strateji_kapsama] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Strateji kapsama verisi alınamadı."}), 500


# ── Yeni Rapor 6: Faaliyet Sorumlu Analizi ───────────────────────────────────

@app_bp.route("/k-rapor/api/sorumlu-analiz")
@login_required
def k_rapor_api_sorumlu_analiz():
    """Kişi başına faaliyet yükü, tamamlanma ve gecikme oranları."""
    tid  = current_user.tenant_id
    try:
        from app.models.process import Process, ProcessActivity
        from app.models.core import User
        today = _dt.date.today()

        users    = User.query.filter_by(tenant_id=tid, is_active=True).all()
        user_map = {u.id: u for u in users}

        # Atanan faaliyetler
        from app.models.process import ProcessActivityAssignee
        assignees = (
            ProcessActivityAssignee.query
            .join(ProcessActivity)
            .join(Process)
            .filter(Process.tenant_id == tid, Process.is_active == True, ProcessActivity.is_active == True)
            .all()
        )

        from collections import defaultdict
        user_acts = defaultdict(list)
        act_ids   = {a.activity_id for a in assignees}
        acts_map  = {}
        if act_ids:
            for act in ProcessActivity.query.filter(ProcessActivity.id.in_(act_ids)).all():
                acts_map[act.id] = act

        for asgn in assignees:
            act = acts_map.get(asgn.activity_id)
            if act:
                user_acts[asgn.user_id].append(act)

        rows = []
        for uid, acts in user_acts.items():
            u = user_map.get(uid)
            if not u:
                continue
            toplam     = len(acts)
            tamamlanan = sum(1 for a in acts if a.status == "Tamamlandı")
            geciken    = sum(1 for a in acts
                            if a.status not in ("Tamamlandı", "İptal")
                            and a.end_date and a.end_date < today)
            devam      = sum(1 for a in acts if a.status == "Devam Ediyor")
            rows.append({
                "user_id":        uid,
                "ad":             f"{u.first_name or ''} {u.last_name or ''}".strip() or u.email,
                "email":          u.email,
                "toplam":         toplam,
                "tamamlanan":     tamamlanan,
                "geciken":        geciken,
                "devam":          devam,
                "tamamlanma_pct": round(tamamlanan / toplam * 100, 1) if toplam else 0,
                "gecikme_pct":    round(geciken / toplam * 100, 1) if toplam else 0,
            })

        rows.sort(key=lambda x: -x["toplam"])
        return jsonify({"success": True, "data": rows})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_sorumlu_analiz] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Sorumlu analiz verisi alınamadı."}), 500


# ── Yeni Rapor 7: SWOT/TOWS Trend ────────────────────────────────────────────

@app_bp.route("/k-rapor/api/swot-trend")
@login_required
def k_rapor_api_swot_trend():
    """Yıllar içinde SWOT madde sayıları ve TOWS strateji sayıları."""
    tid = current_user.tenant_id
    try:
        import json as _json
        from app.models.swot import SwotAnalysis, TowsAnalysis
        from app.models.plan_year import PlanYear

        plan_years = PlanYear.query.filter_by(tenant_id=tid).order_by(PlanYear.year).all()
        py_map     = {py.id: py.year for py in plan_years}

        swots = SwotAnalysis.query.filter_by(tenant_id=tid).all()
        tows  = TowsAnalysis.query.filter_by(tenant_id=tid).all()

        def _cnt(js):
            try:
                return len(_json.loads(js or "[]"))
            except Exception:
                return 0

        swot_by_year = {}
        for s in swots:
            yr = py_map.get(s.plan_year_id) if s.plan_year_id else None
            if not yr:
                continue
            swot_by_year[yr] = {
                "strengths":     _cnt(s.strengths),
                "weaknesses":    _cnt(s.weaknesses),
                "opportunities": _cnt(s.opportunities),
                "threats":       _cnt(s.threats),
            }

        tows_by_year = {}
        for t in tows:
            yr = py_map.get(t.plan_year_id) if t.plan_year_id else None
            if not yr:
                continue
            tows_by_year[yr] = {
                "so": _cnt(t.so_strategies),
                "st": _cnt(t.st_strategies),
                "wo": _cnt(t.wo_strategies),
                "wt": _cnt(t.wt_strategies),
            }

        years = sorted(set(list(swot_by_year.keys()) + list(tows_by_year.keys())))
        trend = []
        for yr in years:
            sw = swot_by_year.get(yr, {})
            tw = tows_by_year.get(yr, {})
            trend.append({
                "year":          yr,
                "strengths":     sw.get("strengths", 0),
                "weaknesses":    sw.get("weaknesses", 0),
                "opportunities": sw.get("opportunities", 0),
                "threats":       sw.get("threats", 0),
                "so": tw.get("so", 0), "st": tw.get("st", 0),
                "wo": tw.get("wo", 0), "wt": tw.get("wt", 0),
                "swot_toplam": sum(sw.values()) if sw else 0,
                "tows_toplam": sum(tw.values()) if tw else 0,
            })

        return jsonify({"success": True, "data": trend})
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_swot_trend] {e}", exc_info=True)
        return jsonify({"success": False, "message": "SWOT trend verisi alınamadı."}), 500


# ── Yeni Rapor 8: Bildirim Analizi ───────────────────────────────────────────

@app_bp.route("/k-rapor/api/bildirim-analiz")
@login_required
def k_rapor_api_bildirim_analiz():
    """Bildirim türü dağılımı, okunma oranı, günlük trend."""
    tid = current_user.tenant_id
    try:
        from app.models.core import Notification
        from sqlalchemy import func
        from collections import Counter

        notifs = Notification.query.filter_by(tenant_id=tid).all()
        if not notifs:
            return jsonify({"success": True, "data": {
                "toplam": 0, "okunma_orani": 0,
                "tur_dagilim": [], "gunluk": [], "son_7_gun": 0,
            }})

        toplam   = len(notifs)
        okunan   = sum(1 for n in notifs if getattr(n, 'is_read', False))
        okunmayan = toplam - okunan

        # Tür dağılımı
        tur_counter = Counter(getattr(n, 'notification_type', None) or getattr(n, 'type', 'genel') for n in notifs)
        tur_dagilim = [{"tur": k or "genel", "sayi": v} for k, v in tur_counter.most_common(10)]

        # Son 30 günlük günlük trend
        today = _dt.date.today()
        gunluk_map = {}
        son_7 = 0
        for n in notifs:
            created = getattr(n, 'created_at', None)
            if not created:
                continue
            d = created.date() if hasattr(created, 'date') else None
            if d:
                key = str(d)
                gunluk_map[key] = gunluk_map.get(key, 0) + 1
                if (today - d).days <= 7:
                    son_7 += 1

        # Son 30 gün
        gunluk = []
        for i in range(29, -1, -1):
            d = today - _dt.timedelta(days=i)
            gunluk.append({"tarih": str(d), "sayi": gunluk_map.get(str(d), 0)})

        return jsonify({
            "success": True,
            "data": {
                "toplam":        toplam,
                "okunan":        okunan,
                "okunmayan":     okunmayan,
                "okunma_orani":  round(okunan / toplam * 100, 1) if toplam else 0,
                "tur_dagilim":   tur_dagilim,
                "gunluk":        gunluk,
                "son_7_gun":     son_7,
            },
        })
    except Exception as e:
        current_app.logger.error(f"[k_rapor_api_bildirim_analiz] {e}", exc_info=True)
        return jsonify({"success": False, "message": "Bildirim analiz verisi alınamadı."}), 500
