"""Süreç modülü — KPI veri (PGV) API."""

from __future__ import annotations
from flask_babel import gettext as _

from datetime import datetime, timezone, date, timedelta
from io import BytesIO

from flask import render_template, jsonify, request, current_app, redirect, abort, send_file, url_for
from flask_login import login_required, current_user
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload, selectinload

from platform_core import app_bp
from app.models import db
from sqlalchemy import or_, func as _sqla_func
from app.models.process import (
    Process,
    ProcessSubStrategyLink,
    ProcessKpi,
    ProcessActivity,
    ProcessActivityAssignee,
    ProcessActivityReminder,
    ActivityTrack,
    KpiData,
    KpiDataAudit,
    FavoriteKpi,
)
from app.models.core import User, Strategy, Tenant
from app.services.plan_year_service import (
    get_plan_year, get_kpi_configs_bulk, get_active_plan_year_for_user, list_plan_years,
    upsert_kpi_year_config,
)
from app.services.date_sovereign import (
    resolve_plan_year_for_date,
    entity_exists_in_year,
    build_existence_error,
    build_cross_year_notice,
    get_view_year,
    plan_year_writable,
    build_sealed_error,
)
from app.services.score_engine_service import compute_process_scores_internal
from app.utils.audit_logger import AuditLogger
from app.utils.db_sequence import is_pk_duplicate, sync_kpi_data_related_sequences, sync_pg_sequence_if_needed
from app.utils.process_utils import (
    validate_process_parent_id,
    last_day_of_period,
    data_date_to_period_keys,
    validate_same_tenant_sub_strategies,
)
from utils.karne_hesaplamalar import (
    hesapla_basari_puani as _hesapla_basari_puani,
    parse_basari_puani_araliklari as _parse_bpa,
)

from app_platform.modules.surec.permissions import (
    accessible_processes_filter,
    can_crud_process_entity,
    user_can_access_process,
    user_can_crud_pg_and_activity,
    user_can_edit_kpi_data_row,
    user_can_edit_process_record,
    user_can_enter_pgv,
)
from micro.modules.surec.helpers import (
    _apply_sub_strategy_links,
    _is_kpi_data_audit_pk_duplicate,
    _is_notification_pk_duplicate,
    _is_process_activity_pk_duplicate,
    _latest_delete_audit_by_kpi_data_ids,
    _latest_update_audit_by_kpi_data_ids,
    _parent_options_with_depth,
    _process_for_user,
    muhur_engeli,
    _user_can_add_activity,
    _user_can_manage_activity,
    _user_display_name,
    _users_pick_json,
)

@app_bp.route("/k-plan/process/api/kpi-data/add", methods=["POST"])
@login_required
def surec_api_kpi_data_add():
    data = request.get_json() or {}
    try:
        kpi_id = int(data.get("kpi_id") or 0)
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": _("Geçersiz kpi_id.")}), 400
    if not kpi_id:
        return jsonify({"success": False, "message": "kpi_id gerekli."}), 400
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id,
        Process.is_active.is_(True),
    ).first_or_404()
    proc = _process_for_user(kpi.process_id)
    if not proc or not user_can_enter_pgv(current_user, proc):
        return jsonify({"success": False, "message": "PG verisi girme yetkiniz yok."}), 403

    # ─── Tarih egemen plan year kontrolü (Faz 1) ──────────────────────────
    # Önce data_date_val'i belirle, sonra ondan plan_year resolve et.
    try:
        year_val = int(data.get("year", datetime.now().year))
        pt = (data.get("period_type") or "yillik").lower().strip()
        pn = int(data.get("period_no") or 1)
        pm = data.get("period_month")
        period_month = int(pm) if pm is not None and str(pm).strip() else None

        data_date_val = None
        if data.get("data_date"):
            data_date_val = datetime.strptime(data["data_date"], "%Y-%m-%d").date()
        else:
            last_day = last_day_of_period(year_val, pt, pn, period_month)
            if last_day is not None:
                data_date_val = last_day
        if data_date_val is None:
            data_date_val = date.today()
        # year_val tarihten türesin — kullanıcı manuel year göndermişse bile data_date öncelikli
        year_val = data_date_val.year

        # Plan year toggle açıksa kontrolleri yap (yoksa eski davranış)
        tenant_obj = db.session.get(Tenant, current_user.tenant_id)
        if tenant_obj:  # K5: yıl bazlılık koşulsuz
            target_py = resolve_plan_year_for_date(current_user.tenant_id, data_date_val)
            # MÜHÜR (K8): kapalı yıla veri girilemez. Mutlak kural, istisna yok.
            # target_py None ise buraya girilmez — o durumu aşağıdaki
            # entity_exists_in_year zaten 409 + açıklayıcı mesajla yakalıyor.
            if target_py is not None and not plan_year_writable(target_py):
                return jsonify(build_sealed_error(target_py)), 423
            if not entity_exists_in_year(kpi, target_py):
                return jsonify(build_existence_error(
                    entity=kpi,
                    entity_label=f"{kpi.code or ''} {kpi.name}".strip(),
                    data_date=data_date_val,
                    target_plan_year=target_py,
                    entity_kind="PG",
                )), 409
            if not entity_exists_in_year(proc, target_py):
                return jsonify(build_existence_error(
                    entity=proc,
                    entity_label=f"{proc.code or ''} {proc.name}".strip(),
                    data_date=data_date_val,
                    target_plan_year=target_py,
                    entity_kind="süreç",
                )), 409
            cross_year_notice = build_cross_year_notice(
                view_year=get_view_year(current_user),
                target_year=data_date_val.year,
            )
        else:
            cross_year_notice = None

        entry = None
        for attempt in (1, 2):
            # Sprint 19.1: SAVEPOINT ile atomic KpiData + KpiDataAudit
            # Audit fail ederse KpiData da SAVEPOINT rollback ile geri alınır.
            try:
                with db.session.begin_nested():  # SAVEPOINT
                    entry = KpiData(
                        process_kpi_id=kpi.id,
                        year=year_val,
                        data_date=data_date_val,
                        period_type=pt,
                        period_no=pn,
                        period_month=period_month,
                        target_value=data.get("target_value"),
                        actual_value=data.get("actual_value", ""),
                        description=data.get("description"),
                        user_id=current_user.id,
                    )
                    db.session.add(entry)
                    db.session.flush()
                    db.session.add(KpiDataAudit(
                        kpi_data_id=entry.id,
                        action_type="CREATE",
                        new_value=entry.actual_value,
                        action_detail="Micro platform veri girişi",
                        user_id=current_user.id,
                    ))
                    # begin_nested() context exit'inde SAVEPOINT RELEASE; hata olursa ROLLBACK TO
                db.session.commit()
                break
            except Exception as e:
                db.session.rollback()
                if attempt == 1 and (
                    is_pk_duplicate(e, "kpi_data") or _is_kpi_data_audit_pk_duplicate(e)
                ):
                    sync_kpi_data_related_sequences()
                    db.session.commit()
                    continue
                raise

        try:
            from app.services.score_engine_service import recalc_on_pg_data_change
            recalc_on_pg_data_change(current_user.tenant_id, year_val)
        except Exception as svc_err:
            current_app.logger.warning(f"[surec_api_kpi_data_add] score_engine: {svc_err}")

        try:
            from app.services.process_deviation_service import check_pg_performance_deviation
            check_pg_performance_deviation(entry.id)
        except Exception as svc_err:
            current_app.logger.warning(f"[surec_api_kpi_data_add] deviation_service: {svc_err}")

        try:
            AuditLogger.log_create(
                "PG Veri Girişi",
                entry.id,
                {"kpi_id": kpi.id, "year": entry.year, "period_type": entry.period_type, "period_no": entry.period_no},
            )
        except Exception as e:
            current_app.logger.error(f"Audit log hatası: {e}")
        resp = {"success": True, "message": "Veri kaydedildi."}
        if cross_year_notice:
            resp["notice"] = cross_year_notice
        return jsonify(resp)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[surec_api_kpi_data_add] {e}")
        return jsonify({"success": False, "message": _("İşlem tamamlanamadı.")}), 400


# Sprint 44 — Bulk Excel import
@app_bp.route("/k-plan/process/api/kpi-data/bulk-import", methods=["POST"])
@login_required
def surec_api_kpi_data_bulk_import():
    """Excel'den KPI veri toplu import (dry-run + commit modu)."""
    from app.services.bulk_import_service import import_kpi_data_from_excel
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"success": False, "message": _("Dosya seçilmedi")}), 400
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"success": False, "message": _("Yalnızca .xlsx / .xls dosyası kabul edilir.")}), 400
    # Content-Length header ile erken boyut kontrolü (tam okumadan önce)
    content_length = request.content_length
    if content_length and content_length > 5 * 1024 * 1024:
        return jsonify({"success": False, "message": "Dosya en fazla 5 MB"}), 400
    dry_run = (request.form.get("dry_run", "true").lower() in ("true", "1", "on"))
    blob = f.read()
    if len(blob) > 5 * 1024 * 1024:
        return jsonify({"success": False, "message": "Dosya en fazla 5 MB"}), 400
    try:
        result = import_kpi_data_from_excel(
            blob, tenant_id=current_user.tenant_id,
            user_id=current_user.id, dry_run=dry_run,
        )
        return jsonify(result)
    except Exception as e:
        current_app.logger.error(f"[bulk_import] {e}", exc_info=True)
        return jsonify({"success": False, "message": _("Sunucu hatası oluştu.")}), 500


@app_bp.route("/k-plan/process/api/kpi-data/template.xlsx")
@login_required
def surec_api_kpi_data_template():
    """Boş Excel şablonu indir."""
    from app.services.bulk_import_service import make_kpi_template_excel
    from flask import send_file
    import io
    blob = make_kpi_template_excel(current_user.tenant_id)
    return send_file(
        io.BytesIO(blob), as_attachment=True,
        download_name="kpi-data-template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app_bp.route("/k-plan/process/api/kpi-data/list/<int:kpi_id>", methods=["GET"])
@login_required
def surec_api_kpi_data_list(kpi_id):
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id,
        Process.is_active.is_(True),
    ).first_or_404()
    proc = _process_for_user(kpi.process_id)
    if not proc or not user_can_access_process(current_user, proc):
        abort(403)
    year = request.args.get("year", datetime.now().year, type=int)
    q = (
        KpiData.query.filter_by(process_kpi_id=kpi.id, year=year, is_active=True)
        .order_by(KpiData.data_date)
    )
    if not user_can_crud_pg_and_activity(current_user, proc):
        q = q.filter_by(user_id=current_user.id)
    entries = q.all()
    return jsonify({
        "success": True,
        "data": [
            {
                "id": e.id,
                "year": e.year,
                "data_date": str(e.data_date),
                "period_type": e.period_type,
                "period_no": e.period_no,
                "actual_value": e.actual_value,
                "target_value": e.target_value,
                "description": e.description,
                "user_id": e.user_id,
            }
            for e in entries
        ],
    })


@app_bp.route("/k-plan/process/api/kpi-data/history/<int:kpi_id>", methods=["GET"])
@login_required
def surec_api_kpi_data_history(kpi_id):
    """
    PG’ye ait tüm yıllar + silinmiş kayıtlar (salt okuma listesi).
    Üye: herkesin geçmişini görür; düzenle/sil yalnız yetkili satırlarda.
    """
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id,
        Process.is_active.is_(True),
    ).first_or_404()
    proc = _process_for_user(kpi.process_id)
    if not proc or not user_can_access_process(current_user, proc):
        abort(403)

    entries = (
        KpiData.query.filter_by(process_kpi_id=kpi.id, is_active=True)
        .options(joinedload(KpiData.user))
        .order_by(KpiData.data_date.desc(), KpiData.id.desc())
        .all()
    )
    del_audits = _latest_delete_audit_by_kpi_data_ids([e.id for e in entries])
    upd_audits = _latest_update_audit_by_kpi_data_ids([e.id for e in entries])
    user_ids = {e.user_id for e in entries}
    for e in entries:
        if e.deleted_by_id:
            user_ids.add(e.deleted_by_id)
        au = del_audits.get(e.id)
        if au and au.user_id:
            user_ids.add(au.user_id)
        uau = upd_audits.get(e.id)
        if uau and uau.user_id:
            user_ids.add(uau.user_id)
    users_map = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()} if user_ids else {}

    out = []
    for e in entries:
        del_audit = del_audits.get(e.id)
        upd_audit = upd_audits.get(e.id)
        del_at = e.deleted_at
        del_by_id = e.deleted_by_id
        if not e.is_active and del_at is None and del_audit:
            del_at = del_audit.created_at
            del_by_id = del_audit.user_id
        del_user = users_map.get(del_by_id) if del_by_id else None
        upd_user = users_map.get(upd_audit.user_id) if upd_audit and upd_audit.user_id else None
        can_edit = bool(e.is_active and user_can_edit_kpi_data_row(current_user, e, proc))
        can_delete = can_edit
        out.append({
            "id": e.id,
            "year": e.year,
            "data_date": str(e.data_date),
            "period_type": e.period_type,
            "period_no": e.period_no,
            "period_month": e.period_month,
            "actual_value": e.actual_value,
            "target_value": e.target_value,
            "description": e.description or "",
            "user_id": e.user_id,
            "entered_by_name": _user_display_name(users_map.get(e.user_id) or e.user),
            "recorded_at": e.created_at.isoformat() if getattr(e, "created_at", None) else None,
            "last_updated_at": upd_audit.created_at.isoformat() if upd_audit else None,
            "last_updated_by_name": _user_display_name(upd_user) if upd_user else None,
            "is_active": e.is_active,
            "deleted_at": del_at.isoformat() if del_at else None,
            "deleted_by_id": del_by_id,
            "deleted_by_name": _user_display_name(del_user) if del_user else None,
            "can_edit": can_edit,
            "can_delete": can_delete,
        })

    return jsonify({
        "success": True,
        "current_user_id": current_user.id,
        "data": out,
    })


@app_bp.route("/k-plan/process/api/kpi-data/update/<int:data_id>", methods=["POST", "PUT"])
@login_required
def surec_api_kpi_data_update(data_id):
    entry = KpiData.query.join(ProcessKpi).join(Process).filter(
        KpiData.id == data_id,
        Process.tenant_id == current_user.tenant_id,
        Process.is_active.is_(True),
    ).first_or_404()
    if not entry.is_active:
        return jsonify({"success": False, "message": _("Silinmiş veri düzenlenemez.")}), 400
    proc = _process_for_user(entry.process_kpi.process_id)
    if not proc or not user_can_edit_kpi_data_row(current_user, entry, proc):
        return jsonify({"success": False, "message": _("Bu veriyi güncelleme yetkiniz yok.")}), 403
    # MÜHÜR (K8): mühürlü yılın verisi düzenlenemez
    engel = muhur_engeli(entry.process_kpi)
    if engel:
        return jsonify(engel[0]), engel[1]
    data = request.get_json() or {}
    old_actual = entry.actual_value or ""
    old_desc = entry.description or ""
    old_target = entry.target_value or ""

    if "actual_value" in data:
        entry.actual_value = str(data["actual_value"]) if data.get("actual_value") is not None else ""
    if "description" in data:
        entry.description = data["description"]
    if "target_value" in data:
        entry.target_value = data["target_value"]

    new_actual = entry.actual_value or ""
    new_desc = entry.description or ""
    new_target = entry.target_value or ""

    changed_labels = []
    if "actual_value" in data and str(old_actual) != str(new_actual):
        changed_labels.append("gerçekleşen")
    if "description" in data and str(old_desc) != str(new_desc):
        changed_labels.append("açıklama")
    if "target_value" in data and str(old_target) != str(new_target):
        changed_labels.append("hedef")

    try:
        for attempt in (1, 2):
            try:
                if changed_labels:
                    _hedef_degisti = "hedef" in changed_labels
                    db.session.add(KpiDataAudit(
                        kpi_data_id=entry.id,
                        action_type="UPDATE",
                        old_value=old_actual,
                        new_value=new_actual,
                        # TASK-261: hedef değişikliğinin NE'den NE'ye olduğu
                        # eskiden kayboluyordu — yalnız action_detail'e "hedef"
                        # etiketi düşüyordu. Hedef Radarı (TASK-262) bu iki
                        # değere dayanıyor. Hedef değişmediyse NULL bırakılır
                        # (her satıra kopyalamak "değişti" sanısı yaratırdı).
                        old_target=old_target if _hedef_degisti else None,
                        new_target=new_target if _hedef_degisti else None,
                        action_detail="Micro platform PGV güncelleme: " + ", ".join(changed_labels),
                        user_id=current_user.id,
                    ))
                db.session.commit()
                break
            except Exception as e:
                db.session.rollback()
                if attempt == 1 and (
                    is_pk_duplicate(e, "kpi_data") or _is_kpi_data_audit_pk_duplicate(e)
                ):
                    sync_kpi_data_related_sequences()
                    db.session.commit()
                    continue
                raise
        try:
            from app.services.score_engine_service import recalc_on_pg_data_change
            recalc_on_pg_data_change(current_user.tenant_id, entry.year)
        except Exception as svc_err:
            current_app.logger.warning(f"[surec_api_kpi_data_update] score_engine: {svc_err}")
        try:
            from app.services.process_deviation_service import check_pg_performance_deviation
            check_pg_performance_deviation(data_id)
        except Exception as svc_err:
            current_app.logger.warning(f"[surec_api_kpi_data_update] deviation: {svc_err}")
        try:
            AuditLogger.log_update(
                "PG Veri Girişi",
                entry.id,
                {},
                {"actual_value": entry.actual_value, "target_value": entry.target_value, "description": entry.description},
            )
        except Exception as e:
            current_app.logger.error(f"Audit log hatası: {e}")
        return jsonify({"success": True, "message": _("Veri güncellendi.")})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[surec_api_kpi_data_update] {e}")
        return jsonify({"success": False, "message": _("İşlem tamamlanamadı.")}), 400


@app_bp.route("/k-plan/process/api/kpi-data/delete/<int:data_id>", methods=["POST", "DELETE"])
@login_required
def surec_api_kpi_data_delete(data_id):
    entry = KpiData.query.join(ProcessKpi).join(Process).filter(
        KpiData.id == data_id,
        Process.tenant_id == current_user.tenant_id,
        Process.is_active.is_(True),
    ).first_or_404()
    if not entry.is_active:
        return jsonify({"success": False, "message": _("Veri zaten silinmiş.")}), 400
    proc = _process_for_user(entry.process_kpi.process_id)
    if not proc or not user_can_edit_kpi_data_row(current_user, entry, proc):
        return jsonify({"success": False, "message": "Bu veriyi silme yetkiniz yok."}), 403
    # MÜHÜR (K8): mühürlü yılın verisi silinemez
    engel = muhur_engeli(entry.process_kpi)
    if engel:
        return jsonify(engel[0]), engel[1]
    try:
        now = datetime.now(timezone.utc)
        for attempt in (1, 2):
            try:
                db.session.add(KpiDataAudit(
                    kpi_data_id=entry.id,
                    action_type="DELETE",
                    old_value=entry.actual_value,
                    new_value=None,
                    action_detail="Micro platform PGV silme (soft)",
                    user_id=current_user.id,
                ))
                entry.is_active = False
                entry.deleted_at = now
                entry.deleted_by_id = current_user.id
                db.session.commit()
                break
            except Exception as e:
                db.session.rollback()
                if attempt == 1 and (
                    is_pk_duplicate(e, "kpi_data") or _is_kpi_data_audit_pk_duplicate(e)
                ):
                    sync_kpi_data_related_sequences()
                    db.session.commit()
                    continue
                raise
        try:
            from app.services.score_engine_service import recalc_on_pg_data_change
            recalc_on_pg_data_change(current_user.tenant_id, entry.year)
        except Exception as svc_err:
            current_app.logger.warning(f"[surec_api_kpi_data_delete] score_engine: {svc_err}")
        return jsonify({"success": True, "message": "Veri silindi."})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"[surec_api_kpi_data_delete] {e}")
        return jsonify({"success": False, "message": _("İşlem tamamlanamadı.")}), 400


@app_bp.route("/k-plan/process/api/kpi-data/detail", methods=["GET"])
@login_required
def surec_api_kpi_data_detail():
    """Kök karne «veri detay» modalı ile uyumlu: periyot bazlı kayıtlar + audit."""
    kpi_id = request.args.get("kpi_id", type=int)
    period_key = request.args.get("period_key", "")
    year = request.args.get("year", datetime.now().year, type=int)
    if kpi_id is None:
        return jsonify({"success": False, "message": "kpi_id gerekli."}), 400
    if not period_key:
        return jsonify({"success": False, "message": "period_key gerekli."}), 400

    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id,
        Process.is_active.is_(True),
    ).first_or_404()
    proc = _process_for_user(kpi.process_id)
    if not proc or not user_can_access_process(current_user, proc):
        abort(403)

    entries_raw = (
        KpiData.query.filter_by(process_kpi_id=kpi.id, year=year)
        .filter(KpiData.is_active.is_(True))
        .order_by(KpiData.data_date.desc())
        .all()
    )
    rows = []
    entry_ids = []
    for e in entries_raw:
        keys = data_date_to_period_keys(e.data_date, year)
        if period_key not in keys:
            continue
        entry_ids.append(e.id)
        u = e.user
        user_name = _user_display_name(u) if u else "Bilinmiyor"
        rows.append({
            "id": e.id,
            "data_date": str(e.data_date),
            "created_at": e.created_at.isoformat() if e.created_at else None,
            "actual_value": e.actual_value,
            "target_value": e.target_value,
            "description": e.description,
            "user": user_name,
        })

    audits = []
    if entry_ids:
        audit_rows = (
            KpiDataAudit.query.filter(KpiDataAudit.kpi_data_id.in_(entry_ids))
            .options(joinedload(KpiDataAudit.user))
            .order_by(KpiDataAudit.created_at.desc())
            .all()
        )
        for a in audit_rows:
            u = a.user
            user_name = _user_display_name(u) if u else "Bilinmiyor"
            action_label = {"CREATE": "Ekleme", "UPDATE": "Değiştirme", "DELETE": "Silme"}.get(
                a.action_type, a.action_type
            )
            audits.append({
                "id": a.id,
                "kpi_data_id": a.kpi_data_id,
                "action_type": a.action_type,
                "action_label": action_label,
                "old_value": a.old_value,
                "new_value": a.new_value,
                "action_detail": a.action_detail,
                "user": user_name,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            })

    return jsonify({
        "success": True,
        "kpi_name": kpi.name,
        "period_key": period_key,
        "year": year,
        "entries": rows,
        "audits": audits,
    })


@app_bp.route("/k-plan/process/api/kpi/<int:kpi_id>/score-detail", methods=["GET"])
@login_required
def surec_api_kpi_score_detail(kpi_id: int):
    """KPI başarı puanı hesaplama detayını döner (şeffaflık için)."""
    from utils.karne_hesaplamalar import (
        hesapla_basari_puani as _hesapla,
        parse_basari_puani_araliklari as _parse_bpa,
    )
    kpi = ProcessKpi.query.filter_by(id=kpi_id, is_active=True).first()
    if not kpi:
        return jsonify({"success": False, "message": _("KPI bulunamadı.")}), 404
    proc = kpi.process
    if not proc or proc.tenant_id != current_user.tenant_id:
        return jsonify({"success": False, "message": "Yetkisiz."}), 403

    year = request.args.get("year", date.today().year, type=int)
    entries = (
        KpiData.query
        .filter_by(process_kpi_id=kpi_id, year=year, is_active=True)
        .order_by(KpiData.data_date.desc())
        .limit(12)
        .all()
    )

    actual_values = []
    for e in entries:
        try:
            actual_values.append(float(e.actual_value))
        except (TypeError, ValueError):
            pass

    method = (kpi.data_collection_method or "Ortalama").lower()
    if method in ("toplama", "toplam") and actual_values:
        rollup = sum(actual_values)
    elif method == "ortalama" and actual_values:
        rollup = round(sum(actual_values) / len(actual_values), 4)
    elif actual_values:
        rollup = actual_values[0]
    else:
        rollup = None

    bpa_raw = kpi.basari_puani_araliklari
    score = None
    bpa_parsed = None
    if rollup is not None and bpa_raw:
        try:
            bpa_parsed = _parse_bpa(bpa_raw) if isinstance(bpa_raw, str) else bpa_raw
            score = _hesapla(rollup, bpa_parsed, kpi.direction or "Increasing")
        except Exception as e:
            current_app.logger.warning(f"[surec_api_kpi_score_detail] suppressed: {e}")

    return jsonify({
        "success": True,
        "kpi": {
            "id": kpi.id,
            "name": kpi.name,
            "code": kpi.code,
            "target_value": kpi.target_value,
            "unit": kpi.unit,
            "direction": kpi.direction or "Increasing",
            "data_collection_method": kpi.data_collection_method or "Ortalama",
            "weight": kpi.weight,
        },
        "year": year,
        "entry_count": len(entries),
        "actual_values": actual_values[:6],
        "rollup_value": rollup,
        "rollup_method": method,
        "basari_puani_araliklari": bpa_parsed,
        "score": score,
        "explanation": (
            f"Son {len(actual_values)} veri {method} yöntemiyle birleştirildi → "
            f"{'%0.2f' % rollup if rollup is not None else '?'} {kpi.unit or ''}. "
            f"Başarı puanı aralıkları kullanılarak puan hesaplandı: {score if score is not None else 'Hesaplanamadı'}."
        ) if actual_values else "Bu KPI için henüz veri girilmemiş.",
    })


@app_bp.route("/k-plan/process/api/kpi-data/proje-gorevleri", methods=["GET"])
@login_required
def surec_api_kpi_data_proje_gorevleri():
    """Kök API ile uyumlu; proje modülü entegrasyonu yoksa boş liste."""
    _ = request.args.get("kpi_id")
    return jsonify({"success": True, "gorevler": []})


# ──────────────────────────────────────────────────
# Toplu KPI Veri Girişi — Excel şablon indir / yükle
# ──────────────────────────────────────────────────

@app_bp.route("/k-plan/process/api/kpi-data/bulk-template", methods=["GET"])
@login_required
def surec_api_kpi_bulk_template():
    """Tenant'ın aktif KPI listesini Excel şablon olarak indirir."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"success": False, "message": _("openpyxl kurulu değil.")}), 500

    tid = current_user.tenant_id
    kpis = (
        ProcessKpi.query
        .join(Process, Process.id == ProcessKpi.process_id)
        .filter(Process.tenant_id == tid, Process.is_active.is_(True), ProcessKpi.is_active.is_(True))
        .order_by(Process.code, ProcessKpi.code)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Veri Girişi"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    info_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    headers = ["kpi_id", "Süreç Kodu", "KPI Kodu", "KPI Adı", "Birim", "Hedef",
               "Gerçekleşen Değer (*)", "Veri Tarihi (*) (YYYY-AA-GG)", "Dönem Tipi", "Açıklama"]
    ws.append(headers)
    for col_idx, _hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    today_str = date.today().isoformat()
    for kpi in kpis:
        proc = kpi.process
        ws.append([
            kpi.id,
            getattr(proc, "code", "") or "",
            kpi.code or "",
            kpi.name or "",
            kpi.unit or "",
            kpi.target_value or "",
            "",
            today_str,
            kpi.period or "Aylık",
            "",
        ])
        row = ws.max_row
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = info_fill

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 25

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"kpi_veri_sablon_{date.today().isoformat()}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )




# ──────────────────────────────────────────────────
# API — Faaliyet CRUD ve Aylık Takip
# ──────────────────────────────────────────────────
