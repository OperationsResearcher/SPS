# -*- coding: utf-8 -*-
"""
SR6 Danışmanlık Hizmetleri Yönetimi Süreç Karnesi Excel Dosyasından Süreç ve PG Import Scripti
Düzeltilmiş versiyon - Excel'deki iki satırlı (Fiili/Hedef) yapıyı destekler
"""

import sys
import os
from datetime import datetime, date
import json

# Windows konsol encoding sorununu çöz
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

try:
    import openpyxl
except ImportError:
    print("openpyxl kütüphanesi bulunamadı. Lütfen yükleyin: pip install openpyxl")
    sys.exit(1)

# Flask app context için
from __init__ import create_app
from extensions import db
from models import (
    Kurum, Surec, SurecPerformansGostergesi, SurecFaaliyet,
    AnaStrateji, AltStrateji, User
)

EXCEL_FILE = 'SR6 Danışmanlık Hizmetleri Yönetimi Süreç Karnesi.xlsx'


def analyze_excel_structure():
    """Excel dosyasının yapısını analiz et"""
    print(f"📊 Excel dosyası analiz ediliyor: {EXCEL_FILE}")
    
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Hata: Dosya bulunamadı: {EXCEL_FILE}")
        return None
    
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    print(f"✅ Excel dosyası açıldı. Sayfalar: {wb.sheetnames}")
    
    # İlk sayfayı al
    ws = wb.active
    print(f"\n📄 Aktif sayfa: {ws.title}")
    print(f"   Toplam satır: {ws.max_row}, Toplam sütun: {ws.max_column}\n")
    
    return wb, ws


def find_header_row(ws):
    """Tablo başlık satırını bul"""
    print("\n🔍 Başlık satırı aranıyor...")
    
    header_keywords = ['Ana Strateji', 'Alt Strateji', 'Gösterge', 'Göst. Türü', 
                       'Hedef Belirl. Yön.', 'Göst. Ağırlığı', 'Birim', 'Ölçüm Per.']
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50), 1):
        row_values = [str(cell.value) if cell.value else "" for cell in row[:20]]
        row_text = " ".join(row_values)
        
        match_count = sum(1 for keyword in header_keywords if keyword in row_text)
        if match_count >= 5:
            print(f"✅ Başlık satırı bulundu: Satır {row_idx}")
            return row_idx, row
    
    print("❌ Başlık satırı bulunamadı!")
    return None, None


def parse_excel_process(wb, ws, header_row_idx, kurum_id):
    """Excel'den süreç ve PG verilerini parse et"""
    print(f"\n📥 Süreç verileri parse ediliyor (Kurum ID: {kurum_id})...")
    
    # Başlık satırını oku
    header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
    headers = {}
    for col_idx, cell in enumerate(header_row, 1):
        if cell.value:
            header_name = str(cell.value).strip()
            headers[header_name] = col_idx
    
    print(f"\n📋 Başlıklar: {list(headers.keys())}")
    
    # Süreç bilgilerini bul
    surec_ad = "Danışmanlık Hizmetleri Yönetimi"
    surec_kodu = "SR6"
    dokuman_no = None
    rev_no = None
    
    # İlk satırlarda süreç bilgilerini ara
    for row_idx in range(1, header_row_idx):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        row_values = [cell.value for cell in row[:10]]
        
        for idx, value in enumerate(row_values):
            if value:
                val_str = str(value).strip()
                if 'SR6' in val_str and 'DANIŞMANLIK' in val_str.upper():
                    surec_kodu = 'SR6'
                    surec_ad = "Danışmanlık Hizmetleri Yönetimi"
    
    print(f"\n📌 Süreç Bilgileri:")
    print(f"   Kod: {surec_kodu}")
    print(f"   Ad: {surec_ad}")
    print(f"   Döküman No: {dokuman_no or 'Belirtilmemiş'}")
    print(f"   Rev. No: {rev_no or 'Belirtilmemiş'}")
    
    # PG verilerini oku - Excel'de iki satırlı yapı (Fiili/Hedef)
    pg_list = []
    current_pg = None
    
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        
        # "Gösterge" sütununu kontrol et
        gosterge_col = headers.get('Gösterge', None)
        if not gosterge_col:
            continue
        
        gosterge_cell = row[gosterge_col - 1]
        gosterge_value = str(gosterge_cell.value).strip() if gosterge_cell.value else ""
        
        # Fiili/Hedef kontrolü
        fiili_hedef_col = headers.get('Fiili/ Hedef', None)
        fiili_hedef_value = None
        if fiili_hedef_col:
            fiili_hedef_cell = row[fiili_hedef_col - 1]
            fiili_hedef_value = str(fiili_hedef_cell.value).strip() if fiili_hedef_cell.value else ""
        
        # Eğer "Gösterge" sütununda yeni bir PG adı varsa
        if gosterge_value and gosterge_value not in ['Fiili', 'Hedef', '']:
            # Önceki PG'yi kaydet (varsa)
            if current_pg and current_pg.get('ad'):
                pg_list.append(current_pg)
                print(f"   ✅ PG bulundu: {current_pg.get('kodu')} - {current_pg.get('ad')}")
            
            # Yeni PG oluştur
            current_pg = {}
            current_pg['ad'] = gosterge_value
            current_pg['kodu'] = f"PG-{len(pg_list) + 1:02d}"
            
            # Diğer alanları oku
            if 'Ana Strateji' in headers and row[headers['Ana Strateji'] - 1].value:
                current_pg['ana_strateji_kodu'] = str(row[headers['Ana Strateji'] - 1].value).strip()
            if 'Alt Strateji' in headers and row[headers['Alt Strateji'] - 1].value:
                current_pg['alt_strateji_kodu'] = str(row[headers['Alt Strateji'] - 1].value).strip()
            if 'Göst. Türü' in headers and row[headers['Göst. Türü'] - 1].value:
                current_pg['gosterge_turu'] = str(row[headers['Göst. Türü'] - 1].value).strip()
            if 'Hedef Belirl. Yön.' in headers and row[headers['Hedef Belirl. Yön.'] - 1].value:
                current_pg['target_method'] = str(row[headers['Hedef Belirl. Yön.'] - 1].value).strip()
            if 'Göst. Ağırlığı (%)' in headers and row[headers['Göst. Ağırlığı (%)'] - 1].value:
                agirlik = row[headers['Göst. Ağırlığı (%)'] - 1].value
                try:
                    current_pg['agirlik'] = float(agirlik)
                except:
                    try:
                        current_pg['agirlik'] = float(agirlik) / 100.0
                    except:
                        current_pg['agirlik'] = 0
            if 'Birim' in headers and row[headers['Birim'] - 1].value:
                current_pg['olcum_birimi'] = str(row[headers['Birim'] - 1].value).strip()
            if 'Ölçüm Per.' in headers and row[headers['Ölçüm Per.'] - 1].value:
                periyot_str = str(row[headers['Ölçüm Per.'] - 1].value).strip()
                periyot_lower = periyot_str.lower()
                if 'ay' in periyot_lower:
                    if '3' in periyot_str:
                        current_pg['periyot'] = 'ceyrek'
                    elif '6' in periyot_str:
                        current_pg['periyot'] = 'ceyrek'
                    else:
                        current_pg['periyot'] = 'aylik'
                elif 'yıl' in periyot_lower or 'year' in periyot_lower:
                    current_pg['periyot'] = 'yillik'
                elif 'hafta' in periyot_lower:
                    current_pg['periyot'] = 'haftalik'
                elif 'gün' in periyot_lower:
                    current_pg['periyot'] = 'gunluk'
                else:
                    current_pg['periyot'] = 'ceyrek'
            if 'Önceki Yıl Ort.' in headers and row[headers['Önceki Yıl Ort.'] - 1].value:
                onceki_yil = row[headers['Önceki Yıl Ort.'] - 1].value
                try:
                    current_pg['onceki_yil_ortalamasi'] = float(onceki_yil)
                except:
                    pass
        
        # Fiili/Hedef satırlarında hedef değerini al
        elif fiili_hedef_value == 'Hedef' and current_pg:
            # Hedef değerini bul - "1.P" sütunundan ilk dolu olanı al
            for col_name in ['1.P', '2.P', '3.P', '4.P']:
                if col_name in headers:
                    hedef_cell = row[headers[col_name] - 1]
                    if hedef_cell.value:
                        if 'hedef_deger' not in current_pg or not current_pg.get('hedef_deger'):
                            hedef_val = str(hedef_cell.value).strip()
                            current_pg['hedef_deger'] = hedef_val
                            break
    
    # Son PG'yi kaydet
    if current_pg and current_pg.get('ad'):
        pg_list.append(current_pg)
        print(f"   ✅ PG bulundu: {current_pg.get('kodu')} - {current_pg.get('ad')}")
    
    print(f"\n📊 Toplam {len(pg_list)} PG bulundu")
    
    return {
        'surec_kodu': surec_kodu,
        'surec_ad': surec_ad,
        'dokuman_no': dokuman_no,
        'rev_no': rev_no,
        'pg_list': pg_list
    }


def import_to_database(process_data, kurum_id):
    """Parse edilen verileri veritabanına aktar"""
    print(f"\n💾 Veritabanına aktarılıyor (Kurum ID: {kurum_id})...")
    
    app = create_app()
    with app.app_context():
        try:
            kurum = Kurum.query.get(kurum_id)
            if not kurum:
                print(f"❌ Hata: Kurum bulunamadı (ID: {kurum_id})")
                return False
            
            print(f"✅ Kurum bulundu: {kurum.kisa_ad}")
            
            # Mevcut süreci kontrol et
            existing_process = Surec.query.filter_by(
                kurum_id=kurum_id,
                code=process_data['surec_kodu']
            ).first()
            
            if existing_process:
                print(f"⚠️  Süreç zaten mevcut (ID: {existing_process.id}). Güncellenecek...")
                surec = existing_process
                surec.ad = process_data['surec_ad']
                surec.dokuman_no = process_data.get('dokuman_no')
                surec.rev_no = process_data.get('rev_no')
            else:
                surec = Surec(
                    kurum_id=kurum_id,
                    code=process_data['surec_kodu'],
                    ad=process_data['surec_ad'],
                    dokuman_no=process_data.get('dokuman_no'),
                    rev_no=process_data.get('rev_no'),
                    durum='Aktif'
                )
                db.session.add(surec)
                db.session.flush()
                print(f"✅ Yeni süreç oluşturuldu (ID: {surec.id})")
            
            db.session.commit()
            
            # PG'leri ekle/güncelle
            pg_count = 0
            for pg_data in process_data['pg_list']:
                alt_strateji_id = None
                if pg_data.get('alt_strateji_kodu'):
                    # AltStrateji, AnaStrateji üzerinden kuruma bağlı
                    alt_strateji = AltStrateji.query.join(AnaStrateji).filter(
                        AnaStrateji.kurum_id == kurum_id,
                        AltStrateji.code == pg_data['alt_strateji_kodu']
                    ).first()
                    if alt_strateji:
                        alt_strateji_id = alt_strateji.id
                
                existing_pg = SurecPerformansGostergesi.query.filter_by(
                    surec_id=surec.id,
                    kodu=pg_data.get('kodu')
                ).first()
                
                if existing_pg:
                    existing_pg.ad = pg_data['ad']
                    existing_pg.kodu = pg_data.get('kodu')
                    existing_pg.alt_strateji_id = alt_strateji_id
                    existing_pg.gosterge_turu = pg_data.get('gosterge_turu')
                    existing_pg.target_method = pg_data.get('target_method')
                    existing_pg.agirlik = pg_data.get('agirlik', 0)
                    existing_pg.olcum_birimi = pg_data.get('olcum_birimi')
                    existing_pg.periyot = pg_data.get('periyot', 'ceyrek')
                    existing_pg.onceki_yil_ortalamasi = pg_data.get('onceki_yil_ortalamasi')
                    existing_pg.hedef_deger = pg_data.get('hedef_deger')
                    print(f"   🔄 PG güncellendi: {pg_data.get('kodu')} - {pg_data.get('ad')}")
                else:
                    new_pg = SurecPerformansGostergesi(
                        surec_id=surec.id,
                        ad=pg_data['ad'],
                        kodu=pg_data.get('kodu'),
                        alt_strateji_id=alt_strateji_id,
                        gosterge_turu=pg_data.get('gosterge_turu'),
                        target_method=pg_data.get('target_method'),
                        agirlik=pg_data.get('agirlik', 0),
                        olcum_birimi=pg_data.get('olcum_birimi'),
                        periyot=pg_data.get('periyot', 'ceyrek'),
                        onceki_yil_ortalamasi=pg_data.get('onceki_yil_ortalamasi'),
                        hedef_deger=pg_data.get('hedef_deger'),
                        veri_toplama_yontemi='Ortalama'
                    )
                    db.session.add(new_pg)
                    print(f"   ✅ Yeni PG eklendi: {pg_data.get('kodu')} - {pg_data.get('ad')}")
                
                pg_count += 1
                db.session.commit()
            
            print(f"\n✅ İşlem tamamlandı! {pg_count} PG işlendi.")
            return True
            
        except Exception as e:
            db.session.rollback()
            print(f"\n❌ Hata: {e}")
            import traceback
            traceback.print_exc()
            return False


def main(kurum_id=None):
    """Ana fonksiyon"""
    print("=" * 80)
    print("SR6 DANIŞMANLIK HİZMETLERİ YÖNETİMİ SÜREÇ KARNESİ IMPORT")
    print("=" * 80)
    
    result = analyze_excel_structure()
    if not result:
        return
    
    wb, ws = result
    
    header_row_idx, header_row = find_header_row(ws)
    if not header_row_idx:
        print("\n❌ Başlık satırı bulunamadı. İşlem sonlandırılıyor.")
        return
    
    if kurum_id is None:
        kurum_id = 87
    print(f"\n📌 Kurum ID: {kurum_id} (KMF Demo Kurum)")
    
    process_data = parse_excel_process(wb, ws, header_row_idx, kurum_id)
    
    if not process_data or not process_data.get('pg_list'):
        print("\n❌ Süreç verileri parse edilemedi veya PG bulunamadı.")
        return
    
    print("\n" + "=" * 80)
    print("ÖZET:")
    print(f"  Süreç: {process_data['surec_kodu']} - {process_data['surec_ad']}")
    print(f"  PG Sayısı: {len(process_data['pg_list'])}")
    print("=" * 80)
    
    print("\n💾 Veritabanına aktarılıyor...")
    success = import_to_database(process_data, kurum_id)
    
    if success:
        print("\n" + "=" * 80)
        print("✅ İŞLEM BAŞARIYLA TAMAMLANDI!")
        print("=" * 80)
    else:
        print("\n" + "=" * 80)
        print("❌ İŞLEM BAŞARISIZ!")
        print("=" * 80)


if __name__ == '__main__':
    import sys
    kurum_id = None
    if len(sys.argv) > 1:
        try:
            kurum_id = int(sys.argv[1])
        except:
            print(f"⚠️  Geçersiz kurum ID: {sys.argv[1]}. Varsayılan (87) kullanılıyor.")
    
    main(kurum_id=kurum_id)

