"""
Dönemsel Karşılaştırma Raporu — KPI performansını yıl bazında Excel'e aktarır.
"""
from __future__ import annotations
from datetime import date
from io import BytesIO

from app.models import db
from app.models.process import Process, ProcessKpi, KpiData
from app.models.core import Tenant


def generate_period_report(tenant_id: int, year: int, compare_year: int | None = None) -> BytesIO:
    """
    Belirtilen yıl için KPI performans raporunu Excel olarak üretir.
    compare_year verilirse önceki yıl sütunu da eklenir.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl kurulu değil.")

    wb = Workbook()
    ws = wb.active
    ws.title = f"KPI Raporu {year}"

    # Renkler
    h_fill  = PatternFill("solid", fgColor="4F46E5")
    s_fill  = PatternFill("solid", fgColor="EDE9FE")
    ok_fill = PatternFill("solid", fgColor="D1FAE5")
    nok_fill= PatternFill("solid", fgColor="FEE2E2")
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    s_font  = Font(bold=True, color="4C1D95", size=10)
    thin    = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    tenant = Tenant.query.get(tenant_id)
    tenant_name = tenant.name if tenant else f"Tenant #{tenant_id}"

    # Başlık
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{tenant_name} — KPI Performans Raporu {year}"
    ws["A1"].font = Font(bold=True, size=14, color="1E293B")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Rapor tarihi: {date.today().isoformat()}"
    ws["A2"].font = Font(size=10, color="64748B")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Sütun başlıkları
    headers = [
        "Süreç Kodu", "Süreç Adı", "KPI Kodu", "KPI Adı",
        "Birim", "Hedef", f"Ortalama {year}", "Başarı %",
    ]
    if compare_year:
        headers += [f"Ortalama {compare_year}", "Değişim %"]
    headers += ["Durum"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = h_fill; c.font = h_font
        c.alignment = Alignment(horizontal="center")
        c.border = thin
    ws.row_dimensions[4].height = 22

    # Veri
    processes = (
        Process.query
        .filter_by(tenant_id=tenant_id, is_active=True)
        .order_by(Process.code, Process.name)
        .all()
    )

    row_idx = 5
    for proc in processes:
        kpis = ProcessKpi.query.filter_by(process_id=proc.id, is_active=True).all()
        if not kpis:
            continue

        # Süreç başlık satırı
        ws.merge_cells(f"A{row_idx}:J{row_idx}")
        c = ws.cell(row=row_idx, column=1, value=f"{proc.code or ''} — {proc.name}")
        c.fill = s_fill; c.font = s_font
        c.alignment = Alignment(horizontal="left")
        row_idx += 1

        for kpi in kpis:
            def _avg(kpi_id: int, yr: int) -> float | None:
                rows = KpiData.query.filter_by(process_kpi_id=kpi_id, year=yr, is_active=True).all()
                vals = []
                for r in rows:
                    try: vals.append(float(r.actual_value))
                    except (TypeError, ValueError): pass
                return round(sum(vals) / len(vals), 2) if vals else None

            avg_curr = _avg(kpi.id, year)
            try:
                target = float(kpi.target_value or 0)
                ratio  = round((avg_curr / target) * 100, 1) if avg_curr and target else None
            except (TypeError, ValueError):
                ratio = None

            direction = (kpi.direction or "Increasing").lower()
            if ratio is None:
                status = "Veri Yok"
            elif direction != "decreasing":
                status = "Hedefte" if ratio >= 80 else "Hedef Altı"
            else:
                status = "Hedefte" if ratio <= 120 else "Hedef Üstü"

            row_data = [
                proc.code or "", proc.name,
                kpi.code or "", kpi.name,
                kpi.unit or "", kpi.target_value or "",
                avg_curr, ratio,
            ]
            if compare_year:
                avg_prev = _avg(kpi.id, compare_year)
                change = None
                if avg_curr is not None and avg_prev and avg_prev != 0:
                    change = round(((avg_curr - avg_prev) / avg_prev) * 100, 1)
                row_data += [avg_prev, change]
            row_data.append(status)

            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border = thin
                c.alignment = Alignment(horizontal="center" if col > 4 else "left")
                if col == len(row_data):  # Durum sütunu
                    c.fill = ok_fill if status == "Hedefte" else (nok_fill if status == "Hedef Altı" else PatternFill())

            row_idx += 1

    # Sütun genişlikleri
    widths = [10, 28, 10, 30, 8, 10, 14, 12, 14, 12, 12]
    for i, w in enumerate(widths[:ws.max_column], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
