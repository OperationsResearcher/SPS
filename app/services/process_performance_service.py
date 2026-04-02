# -*- coding: utf-8 -*-
"""
Process Performance Service
Süreç karnesi performans verilerinin hesaplandığı "Business Logic" katmanı.
"""
from typing import Dict, Any, Tuple
import time
from datetime import datetime

from flask import current_app
from models import db, Surec, SurecPerformansGostergesi, AltStrateji, surec_liderleri, surec_uyeleri
from services.performance_service import generatePeriyotVerileri, calculateHedefDeger
from utils.karne_hesaplamalar import (
    hesapla_basari_puani, hesapla_agirlikli_basari_puani, parse_basari_puani_araliklari
)
from app.utils.errors import AuthorizationError, ResourceNotFoundError


class ProcessPerformanceService:
    @staticmethod
    def get_process_performance(user, surec_id: int, yil: int, periyot: str, ay: int = None, debug_mode: bool = False) -> Tuple[Dict[str, Any], int]:
        """
        Sürecin performans göstergelerini ve çeyreklik verilerini hesaplar.
        Tanrı nesnelerden (God Objects) iş mantığını ayırmak (Decoupling) için oluşturulmuştur.
        """
        baslangic_zamani = time.time()
        
        # 1. Kaynak ve Yetki Kontrolleri
        surec = Surec.query.options(db.joinedload(Surec.kurum)).get(surec_id)
        if not surec:
            raise ResourceNotFoundError("Süreç bulunamadı.")
            
        if user.sistem_rol == 'admin':
            pass
        elif user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            if surec.kurum_id != user.kurum_id:
                raise AuthorizationError("Farklı kurumun sürecine erişim reddedildi.")
        else:
            lider_mi = db.session.query(surec_liderleri).filter(
                surec_liderleri.c.surec_id == surec_id,
                surec_liderleri.c.user_id == user.id
            ).first() is not None
            uye_mi = db.session.query(surec_uyeleri).filter(
                surec_uyeleri.c.surec_id == surec_id,
                surec_uyeleri.c.user_id == user.id
            ).first() is not None
            
            if not (lider_mi or uye_mi):
                raise AuthorizationError("Bu süreçte yetkiniz yok.")
                
        # 2. Veri Çekimi
        surec_pgler = SurecPerformansGostergesi.query.options(
            db.joinedload(SurecPerformansGostergesi.alt_strateji)
        ).filter_by(surec_id=surec_id).all()
        
        gostergeler = []
        hata_listesi = []
        
        # 3. Hesaplama (Business Logic)
        for pg in surec_pgler:
            try:
                hesaplama_yontemi = getattr(pg, 'veri_toplama_yontemi', 'Ortalama')
                
                veriler = generatePeriyotVerileri(
                    periyot=periyot,
                    pg_id=pg.id,
                    yil=yil,
                    pg_hedef_deger=pg.hedef_deger,
                    pg_periyot=pg.periyot,
                    hesaplama_yontemi=hesaplama_yontemi,
                    ay=ay
                )
                
                gosterim_hedef_deger = calculateHedefDeger(
                    pg.hedef_deger, pg.periyot, periyot, hesaplama_yontemi
                )
                
                ana_strateji_kodu = None
                alt_strateji_kodu = None
                if pg.alt_strateji:
                    alt_strateji_kodu = getattr(pg.alt_strateji, 'code', None)
                    if pg.alt_strateji.ana_strateji:
                        ana_strateji_kodu = getattr(pg.alt_strateji.ana_strateji, 'code', None)
                
                onceki_yil_ortalamasi = getattr(pg, 'onceki_yil_ortalamasi', None)
                basari_puani_araliklari = parse_basari_puani_araliklari(getattr(pg, 'basari_puani_araliklari', None))
                
                yil_sonu_gerceklesen = None
                if periyot == 'ceyrek' and veriler:
                    ceyrek_degerler = []
                    for veri in veriler:
                        if veri.get('gerceklesen_deger') is not None:
                            try:
                                ceyrek_degerler.append(float(veri['gerceklesen_deger']))
                            except (ValueError, TypeError):
                                pass
                    if ceyrek_degerler:
                        if hesaplama_yontemi in ['Toplama', 'SUM']:
                            yil_sonu_gerceklesen = sum(ceyrek_degerler)
                        else:  # Ortalama
                            yil_sonu_gerceklesen = sum(ceyrek_degerler) / len(ceyrek_degerler)
                
                basari_puani = None
                agirlikli_basari_puani = None
                
                if yil_sonu_gerceklesen is not None and basari_puani_araliklari:
                    direction = getattr(pg, 'direction', 'Increasing')
                    basari_puani = hesapla_basari_puani(yil_sonu_gerceklesen, basari_puani_araliklari, direction)
                    try:
                        pg.basari_puani = basari_puani
                        agirlik = getattr(pg, 'agirlik', 0)
                        agirlikli_basari_puani = hesapla_agirlikli_basari_puani(basari_puani, agirlik)
                        pg.agirlikli_basari_puani = agirlikli_basari_puani
                        db.session.commit()
                    except Exception as e:
                        db.session.rollback()
                        current_app.logger.warning(f"Başarı puanı kaydedilirken hata: {e}")
                
                gostergeler.append({
                    'id': pg.id,
                    'surec_pg_id': pg.id,
                    'ad': pg.ad,
                    'kodu': getattr(pg, 'kodu', f'PG-{pg.id}'),
                    'hedef_deger': pg.hedef_deger,
                    'gosterim_hedef_deger': round(gosterim_hedef_deger, 2) if gosterim_hedef_deger is not None else None,
                    'olcum_birimi': pg.olcum_birimi,
                    'periyot': pg.periyot,
                    'agirlik': getattr(pg, 'agirlik', 0),
                    'onemli': getattr(pg, 'onemli', False),
                    'veri_toplama_yontemi': hesaplama_yontemi,
                    'veriler': veriler,
                    'ana_strateji_kodu': ana_strateji_kodu,
                    'alt_strateji_kodu': alt_strateji_kodu,
                    'gosterge_turu': getattr(pg, 'gosterge_turu', None),
                    'target_method': getattr(pg, 'target_method', None),
                    'onceki_yil_ortalamasi': round(onceki_yil_ortalamasi, 2) if onceki_yil_ortalamasi is not None else None,
                    'basari_puani': basari_puani,
                    'agirlikli_basari_puani': round(agirlikli_basari_puani, 2) if agirlikli_basari_puani is not None else None,
                    'basari_puani_araliklari': basari_puani_araliklari,
                    'yil_sonu_gerceklesen': round(yil_sonu_gerceklesen, 2) if yil_sonu_gerceklesen is not None else None
                })
            except Exception as pg_error:
                import traceback
                error_trace = traceback.format_exc()
                hata_listesi.append({
                    'pg_id': pg.id,
                    'pg_ad': pg.ad,
                    'hata': str(pg_error),
                    'traceback': error_trace if debug_mode else None
                })
                continue
                
        gecen_sure = time.time() - baslangic_zamani
        current_app.logger.info(f"Süreç Karnesi Servisi tamamlandı, geçen süre: {gecen_sure:.2f} saniye")
        
        response_data = {
            'success': True,
            'gostergeler': gostergeler
        }
        
        if debug_mode and hata_listesi:
            response_data['hatalar'] = hata_listesi
            response_data['toplam_pg'] = len(surec_pgler)
            response_data['basarili_pg'] = len(gostergeler)
            response_data['hatali_pg'] = len(hata_listesi)
            
        return response_data, 200

    @staticmethod
    def get_process_activities(user, surec_id: int, yil: int) -> Tuple[Dict[str, Any], int]:
        from models import db, SurecFaaliyet, BireyselFaaliyet, FaaliyetTakip
        
        surec_faaliyetler = SurecFaaliyet.query.filter_by(surec_id=surec_id).all()
        surec_faaliyet_ids = [f.id for f in surec_faaliyetler]
        
        bireysel_map = {}
        if surec_faaliyet_ids:
            bireyseller = BireyselFaaliyet.query.filter(
                BireyselFaaliyet.user_id == user.id,
                BireyselFaaliyet.kaynak == 'Süreç',
                BireyselFaaliyet.kaynak_surec_id == surec_id,
                BireyselFaaliyet.kaynak_surec_faaliyet_id.in_(surec_faaliyet_ids)
            ).all()
            bireysel_map = {b.kaynak_surec_faaliyet_id: b for b in bireyseller if b.kaynak_surec_faaliyet_id}

        takip_map = {}
        bireysel_ids = [b.id for b in bireysel_map.values()]
        if bireysel_ids:
            takip_rows = FaaliyetTakip.query.filter(
                FaaliyetTakip.user_id == user.id,
                FaaliyetTakip.yil == yil,
                FaaliyetTakip.bireysel_faaliyet_id.in_(bireysel_ids)
            ).all()
            for row in takip_rows:
                takip_map.setdefault(row.bireysel_faaliyet_id, []).append({
                    'ay': row.ay,
                    'gerceklesti': bool(row.gerceklesti)
                })

        faaliyetler = []
        for faaliyet in surec_faaliyetler:
            bireysel = bireysel_map.get(faaliyet.id)
            bireysel_id = bireysel.id if bireysel else 0
            faaliyetler.append({
                'id': bireysel_id,
                'surec_faaliyet_id': faaliyet.id,
                'ad': faaliyet.ad,
                'oneri': faaliyet.aciklama,
                'takip': takip_map.get(bireysel_id, [])
            })
        
        return {'success': True, 'faaliyetler': faaliyetler}, 200

    @staticmethod
    def create_bireysel_faaliyet_from_surec(user, surec_id: int, surec_faaliyet_id: int) -> Tuple[Dict[str, Any], int]:
        from models import db, SurecFaaliyet, BireyselFaaliyet
        from app.utils.errors import ResourceNotFoundError
        
        surec_faaliyet = SurecFaaliyet.query.filter_by(id=surec_faaliyet_id, surec_id=surec_id).first()
        if not surec_faaliyet:
            raise ResourceNotFoundError("Faaliyet bulunamadı.")

        mevcut = BireyselFaaliyet.query.filter_by(
            user_id=user.id,
            kaynak='Süreç',
            kaynak_surec_id=surec_id,
            kaynak_surec_faaliyet_id=surec_faaliyet_id
        ).first()
        if mevcut:
            return {'success': True, 'bireysel_faaliyet_id': mevcut.id}, 200

        yeni = BireyselFaaliyet(
            user_id=user.id,
            ad=surec_faaliyet.ad,
            aciklama=surec_faaliyet.aciklama,
            baslangic_tarihi=surec_faaliyet.baslangic_tarihi,
            bitis_tarihi=surec_faaliyet.bitis_tarihi,
            durum=surec_faaliyet.durum,
            ilerleme=surec_faaliyet.ilerleme,
            kaynak='Süreç',
            kaynak_surec_id=surec_id,
            kaynak_surec_faaliyet_id=surec_faaliyet_id,
            kurum_id=surec_faaliyet.surec.kurum_id if surec_faaliyet.surec else user.kurum_id
        )
        db.session.add(yeni)
        try:
            db.session.commit()
            return {'success': True, 'bireysel_faaliyet_id': yeni.id}, 200
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'Bireysel Faaliyet oluşturma hatası: {str(e)}')
            raise Exception(f'Veritabanı hatası: {str(e)}')

    @staticmethod
    def save_activity_tracking(user, faaliyet_id: int, data: Dict[str, Any]) -> Tuple[Dict[str, Any], int]:
        from models import db, BireyselFaaliyet, FaaliyetTakip
        from datetime import datetime, date
        from app.utils.errors import ResourceNotFoundError, AuthorizationError

        yil = int(data.get('yil') or datetime.now().year)
        ay = int(data.get('ay') or 0)
        if ay < 1 or ay > 12:
            raise ValueError("Ay 1-12 aralığında olmalı")
        
        gerceklesti = bool(data.get('gerceklesti'))

        faaliyet = BireyselFaaliyet.query.get(faaliyet_id)
        if not faaliyet:
            raise ResourceNotFoundError("Faaliyet bulunamadı")
        
        if faaliyet.user_id != user.id:
            raise AuthorizationError("Bu faaliyet üzerinde yetkiniz yok")

        takip = FaaliyetTakip.query.filter_by(
            bireysel_faaliyet_id=faaliyet_id,
            user_id=user.id,
            yil=yil,
            ay=ay
        ).first()

        if not takip:
            takip = FaaliyetTakip(
                bireysel_faaliyet_id=faaliyet_id,
                user_id=user.id,
                yil=yil,
                ay=ay,
                gerceklesti=gerceklesti
            )
            db.session.add(takip)
        else:
            takip.gerceklesti = gerceklesti

        takip.tamamlanma_tarihi = date.today() if gerceklesti else None

        db.session.commit()
        return {'success': True}, 200

    @staticmethod
    def save_performance_data(user, surec_id: int, data: Dict[str, Any]) -> Tuple[Dict[str, Any], int]:
        from models import (
            db, Surec, User, BireyselPerformansGostergesi,
            PerformansGostergeVeri, SurecPerformansGostergesi
        )
        from services.performance_service import (
            get_last_friday_of_month, get_last_friday_of_quarter, 
            get_last_friday_of_year, get_last_friday_of_week, get_last_friday_of_day,
            hesapla_durum
        )
        from datetime import datetime
        from app.utils.errors import ResourceNotFoundError, AuthorizationError

        yil = int(data.get('yil') or datetime.now().year)
        pg_verileri = data.get('pg_verileri') or []

        for veri in pg_verileri:
            pg_id = int(veri.get('pg_id', 0))
            surec_pg_id = int(veri.get('surec_pg_id', 0))

            veri_sahibi_id = veri.get('user_id') or veri.get('kullanici_id')
            veri_sahibi_id = int(veri_sahibi_id) if veri_sahibi_id else int(user.id)

            if user.sistem_rol != 'admin':
                veri_sahibi = User.query.get(veri_sahibi_id)
                if not veri_sahibi:
                    raise ResourceNotFoundError("Kullanıcı bulunamadı")
                if veri_sahibi.kurum_id != user.kurum_id:
                    raise AuthorizationError("Bu kullanıcı üzerinde işlem yetkiniz yok")

            surec_pg = None
            if surec_pg_id > 0:
                surec_pg = SurecPerformansGostergesi.query.get(surec_pg_id)
                if not surec_pg or surec_pg.surec_id != surec_id:
                    raise ValueError("Geçersiz süreç performans göstergesi")

            periyot_tipi = veri.get('periyot_tipi')
            periyot_no = veri.get('periyot_no')
            ceyrek = int(veri.get('ceyrek', 0))
            ay = int(veri.get('ay', 0))
            field = veri.get('field')
            value = veri.get('value')
            aciklama = veri.get('aciklama', '')

            if pg_id == 0 and surec_pg_id > 0:
                if not surec_pg:
                    continue

                bireysel_pg = BireyselPerformansGostergesi(
                    user_id=veri_sahibi_id,
                    ad=surec_pg.ad,
                    aciklama=surec_pg.aciklama,
                    hedef_deger=surec_pg.hedef_deger,
                    olcum_birimi=surec_pg.olcum_birimi,
                    periyot=surec_pg.periyot,
                    kaynak='Süreç',
                    kaynak_surec_id=surec_id,
                    kaynak_surec_pg_id=surec_pg_id,
                    agirlik=surec_pg.agirlik if hasattr(surec_pg, 'agirlik') else 0,
                    onemli=surec_pg.onemli if hasattr(surec_pg, 'onemli') else False,
                    kodu=surec_pg.kodu if hasattr(surec_pg, 'kodu') else None
                )
                db.session.add(bireysel_pg)
                db.session.flush()
                pg_id = bireysel_pg.id

            if pg_id == 0:
                continue

            if periyot_tipi:
                if periyot_tipi == 'ceyrek':
                    ceyrek = int(periyot_no) if periyot_no else None
                    ay = None
                elif periyot_tipi == 'aylik':
                    ay = int(periyot_no) if periyot_no else None
                    ceyrek = None
                elif periyot_tipi == 'yillik':
                    ceyrek = None
                    ay = None
                else:
                    ay = int(veri.get('ay', 0)) if veri.get('ay') else None
            elif ceyrek > 0:
                periyot_tipi = 'ceyrek'
                periyot_no = ceyrek
                ay = None
            elif ay > 0:
                periyot_tipi = 'aylik'
                periyot_no = ay
                ceyrek = None
            else:
                periyot_tipi = 'yillik'
                periyot_no = None
                ceyrek = None
                ay = None

            if periyot_tipi == 'ceyrek' and periyot_no:
                veri_tarihi = get_last_friday_of_quarter(int(periyot_no), yil)
            elif periyot_tipi == 'aylik' and periyot_no:
                veri_tarihi = get_last_friday_of_month(int(periyot_no), yil)
            elif periyot_tipi == 'yillik':
                veri_tarihi = get_last_friday_of_year(yil)
            elif periyot_tipi == 'haftalik' and periyot_no and ay:
                veri_tarihi = get_last_friday_of_week(int(ay), int(periyot_no), yil)
            elif periyot_tipi == 'gunluk' and periyot_no and ay:
                veri_tarihi = get_last_friday_of_day(int(ay), int(periyot_no), yil)
            else:
                veri_tarihi = datetime.now().date()

            bireysel_pg = BireyselPerformansGostergesi.query.get(pg_id)
            if not bireysel_pg:
                raise ResourceNotFoundError("Performans göstergesi bulunamadı")
            if bireysel_pg.user_id != veri_sahibi_id:
                raise AuthorizationError("Bu gösterge üzerinde yetkiniz yok")

            if bireysel_pg.kaynak == 'Süreç':
                if bireysel_pg.kaynak_surec_id != surec_id:
                    raise AuthorizationError("Bu süreçte yetkiniz yok")
            elif bireysel_pg.kaynak_surec_pg_id:
                kaynak_surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
                if not kaynak_surec_pg or kaynak_surec_pg.surec_id != surec_id:
                    raise AuthorizationError("Bu süreçte yetkiniz yok")

            pg_veri = PerformansGostergeVeri.query.filter_by(
                bireysel_pg_id=pg_id,
                user_id=veri_sahibi_id,
                yil=yil,
                veri_tarihi=veri_tarihi
            ).first()

            if not pg_veri:
                pg_veri = PerformansGostergeVeri(
                    bireysel_pg_id=pg_id,
                    user_id=veri_sahibi_id,
                    yil=yil,
                    veri_tarihi=veri_tarihi,
                    giris_periyot_tipi=periyot_tipi,
                    giris_periyot_no=periyot_no,
                    giris_periyot_ay=ay,
                    ceyrek=ceyrek if ceyrek else None,
                    ay=ay if ay else None,
                    created_by=user.id
                )
                db.session.add(pg_veri)
            else:
                pg_veri.giris_periyot_tipi = periyot_tipi
                pg_veri.giris_periyot_no = periyot_no
                pg_veri.giris_periyot_ay = ay
                pg_veri.ceyrek = ceyrek if ceyrek else None
                pg_veri.ay = ay if ay else None
                pg_veri.updated_by = user.id

            if field == 'hedef':
                pg_veri.hedef_deger = value
            elif field == 'gerceklesen':
                pg_veri.gerceklesen_deger = value
                pg_veri.veri_tarihi = veri_tarihi
                if aciklama:
                    pg_veri.aciklama = aciklama
                pg_veri.updated_by = user.id

                if pg_veri.hedef_deger and pg_veri.gerceklesen_deger:
                    durum, durum_yuzdesi = hesapla_durum(
                        float(pg_veri.hedef_deger) if pg_veri.hedef_deger else None,
                        float(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger else None
                    )
                    pg_veri.durum = durum
                    pg_veri.durum_yuzdesi = durum_yuzdesi

        db.session.commit()

        try:
            kurum_id = getattr(user, 'kurum_id', None)
            if kurum_id:
                from services.score_engine_service import recalc_on_pg_data_change
                recalc_on_pg_data_change(kurum_id)
        except Exception as ex:
            current_app.logger.debug(f"Skor motoru güncellemesi atlandı: {ex}")

        return {'success': True, 'message': 'Veriler başarıyla kaydedildi'}, 200




    @staticmethod
    def get_pg_veri_detay_list(user, surec_pg_id, ceyrek, yil):
        """Süreç karnesinde bir PG'nin kullanıcı bazlı veri detaylarını getir"""
        try:
        
            if not surec_pg_id:
                return dict({'success': False, 'message': 'Eksik parametreler'}), 400

            # Bu endpoint kullanıcı bazlı veri döndürdüğü için yönetim rolleri ile sınırla
            surec_pg = SurecPerformansGostergesi.query.get(surec_pg_id)
            if not surec_pg:
                return dict({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
            surec = Surec.query.get(surec_pg.surec_id)
            if not surec:
                return dict({'success': False, 'message': 'Süreç bulunamadı'}), 404

            if user.sistem_rol == 'admin':
                pass
            elif user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
                if surec.kurum_id != user.kurum_id:
                    return dict({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403
            else:
                return dict({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
            # Bu süreç PG'sine bağlı tüm bireysel PG'leri bul
            bireysel_pgler = BireyselPerformansGostergesi.query.filter_by(
                kaynak_surec_pg_id=surec_pg_id,
                kaynak='Süreç'
            ).all()
        
            if not bireysel_pgler:
                return dict({'success': True, 'veriler': []})
        
            # Her bir bireysel PG için bu periyottaki veri girişlerini topla
            tum_veriler = []
            for bireysel_pg in bireysel_pgler:
                veri_girisleri = []
            
                # Periyot tipine göre sorgu yap
                if ceyrek == 'yillik':
                    # Yıllık veri
                    veri_girisleri = PerformansGostergeVeri.query.filter_by(
                        bireysel_pg_id=bireysel_pg.id,
                        yil=yil,
                        ceyrek=None,
                        ay=None
                    ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
                elif ceyrek and ceyrek.isdigit() and int(ceyrek) >= 1 and int(ceyrek) <= 4:
                    # Çeyreklik veri
                    ceyrek_no = int(ceyrek)
                    # Çeyreklik kaydı ve bu çeyreğin aylarındaki veriler
                    ceyrek_aylari = get_ceyrek_aylari(ceyrek_no)
                    veri_girisleri = PerformansGostergeVeri.query.filter(
                        PerformansGostergeVeri.bireysel_pg_id == bireysel_pg.id,
                        PerformansGostergeVeri.yil == yil
                    ).filter(
                        or_(
                            and_(PerformansGostergeVeri.ceyrek == ceyrek_no, PerformansGostergeVeri.ay.is_(None)),  # Çeyreklik kayıt
                            PerformansGostergeVeri.ay.in_(ceyrek_aylari)  # Bu çeyreğin aylarındaki veriler
                        )
                    ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
                elif ceyrek and ceyrek.isdigit() and int(ceyrek) >= 1 and int(ceyrek) <= 12:
                    # Aylık veri
                    ay_no = int(ceyrek)
                    # Aylık kayıt ve bu ayın tüm verileri (haftalık/günlük dahil)
                    veri_girisleri = PerformansGostergeVeri.query.filter_by(
                        bireysel_pg_id=bireysel_pg.id,
                        yil=yil,
                        ay=ay_no
                    ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
            
                    # Eğer bu ay için veri yoksa, çeyreklik veriyi kontrol et
                    if not veri_girisleri:
                        ilgili_ceyrek = get_ay_ceyreği(ay_no)
                        if ilgili_ceyrek:
                            ceyrek_verileri = PerformansGostergeVeri.query.filter_by(
                                bireysel_pg_id=bireysel_pg.id,
                                yil=yil,
                                ceyrek=ilgili_ceyrek,
                                ay=None
                            ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
                            if ceyrek_verileri:
                                veri_girisleri = ceyrek_verileri
                
                    # Eğer hala veri yoksa, yıllık veriyi kontrol et
                    if not veri_girisleri:
                        yillik_verileri = PerformansGostergeVeri.query.filter_by(
                            bireysel_pg_id=bireysel_pg.id,
                            yil=yil,
                            ceyrek=None,
                            ay=None
                        ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
                        if yillik_verileri:
                            veri_girisleri = yillik_verileri
                elif ceyrek and ceyrek.isdigit() and int(ceyrek) >= 1 and int(ceyrek) <= 5:
                    # Haftalık veri - hafta numarası
                    hafta_no = int(ceyrek)
                    # Mevcut ayın haftalarını al
                    mevcut_ay = datetime.now().month
                    haftalar = get_ay_haftalari(mevcut_ay, yil)
                    if 1 <= hafta_no <= len(haftalar):
                        hafta_bilgi = haftalar[hafta_no - 1]
                        hafta_baslangic = hafta_bilgi['baslangic_tarih']
                        hafta_bitis = hafta_bilgi['bitis_tarih']
                        veri_girisleri = PerformansGostergeVeri.query.filter(
                            PerformansGostergeVeri.bireysel_pg_id == bireysel_pg.id,
                            PerformansGostergeVeri.yil == yil,
                            PerformansGostergeVeri.veri_tarihi >= hafta_baslangic,
                            PerformansGostergeVeri.veri_tarihi <= hafta_bitis
                        ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
                elif ceyrek and ceyrek.isdigit() and int(ceyrek) >= 1 and int(ceyrek) <= 31:
                    # Günlük veri - gün numarası
                    gun_no = int(ceyrek)
                    # Mevcut ayın günlerini al
                    mevcut_ay = datetime.now().month
                    gunler = get_ay_gunleri(mevcut_ay, yil)
                    if 1 <= gun_no <= len(gunler):
                        gun_tarih = gunler[gun_no - 1]['tarih']
                        veri_girisleri = PerformansGostergeVeri.query.filter_by(
                            bireysel_pg_id=bireysel_pg.id,
                            yil=yil,
                            veri_tarihi=gun_tarih
                        ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
                else:
                    # Varsayılan: çeyrek olarak kabul et
                    if ceyrek and ceyrek.isdigit():
                        ceyrek_no = int(ceyrek)
                        veri_girisleri = PerformansGostergeVeri.query.filter_by(
                            bireysel_pg_id=bireysel_pg.id,
                            yil=yil,
                            ceyrek=ceyrek_no
                        ).order_by(PerformansGostergeVeri.veri_tarihi.desc()).all()
            
                current_app.logger.info(f"Bireysel PG {bireysel_pg.id} için {len(veri_girisleri)} veri bulundu (Periyot: {ceyrek})")
            
                for veri in veri_girisleri:
                    kullanici = User.query.get(veri.user_id)
                    if kullanici:
                        if kullanici.first_name and kullanici.last_name:
                            kullanici_adi = f"{kullanici.first_name} {kullanici.last_name}"
                        else:
                            kullanici_adi = kullanici.username
                    else:
                        kullanici_adi = 'Bilinmiyor'
                
                    current_app.logger.info(f"Veri: user_id={veri.user_id}, kullanici={kullanici_adi}, tarih={veri.veri_tarihi}, ay={veri.ay}, ceyrek={veri.ceyrek}")
                
                    tum_veriler.append({
                        'id': veri.id,
                        'kullanici_id': veri.user_id,
                        'kullanici_adi': kullanici_adi,
                        'hedef_deger': veri.hedef_deger,
                        'gerceklesen_deger': veri.gerceklesen_deger,
                        'durum': veri.durum,
                        'durum_yuzdesi': veri.durum_yuzdesi,
                        'veri_tarihi': veri.veri_tarihi.isoformat() if veri.veri_tarihi else None,
                        'aciklama': veri.aciklama if veri.aciklama else None,
                        'created_at': veri.created_at.isoformat() if veri.created_at else None
                    })
        
            # Tarihe göre sırala (en yeni en üstte)
            tum_veriler.sort(key=lambda x: x['veri_tarihi'] or '', reverse=True)
        
            current_app.logger.info(f"Toplam {len(tum_veriler)} veri döndürülüyor")
        
            return dict({
                'success': True,
                'veriler': tum_veriler
            })
        
        except Exception as e:
            current_app.logger.error(f'PG veri detay getirme hatası: {str(e)}', exc_info=True)
            return dict({'success': False, 'message': str(e)}), 500


################################################################################

    @staticmethod
    def export_surec_karnesi_excel_service(user, surec_id, yil):
        """Süreç karnesi verilerini Excel olarak dışa aktarır."""
    
        if not surec_id or not yil:
            return dict({"success": False, "message": "Süreç ID ve Yıl gereklidir."}), 400
    
        surec = Surec.query.get(surec_id)
        if not surec:
            return dict({"success": False, "message": "Süreç bulunamadı."}), 404
    
        # Check user permission
        if user.sistem_rol == 'admin':
            # Admin tüm süreçleri görüntüleyebilir
            pass
        elif user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            # Kurum yöneticileri kendi kurumlarındaki süreçleri görüntüleyebilir
            if surec.kurum_id != user.kurum_id:
                return dict({"success": False, "message": "Bu süreci görüntüleme yetkiniz yok."}), 403
        else:
            # Normal kullanıcı için lider/üye kontrolü
            lider_mi = db.session.query(surec_liderleri).filter(
                surec_liderleri.c.surec_id == surec_id,
                surec_liderleri.c.user_id == user.id
            ).first() is not None
        
            uye_mi = db.session.query(surec_uyeleri).filter(
                surec_uyeleri.c.surec_id == surec_id,
                surec_uyeleri.c.user_id == user.id
            ).first() is not None
        
            if not (lider_mi or uye_mi):
                return dict({"success": False, "message": "Bu süreci görüntüleme yetkiniz yok."}), 403
    
        try:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except Exception as import_err:
                current_app.logger.error(f'openpyxl import hatası: {import_err}')
                return dict({
                    "success": False,
                    "message": "Excel aktarımı için openpyxl gerekli. Lütfen bağımlılığı yükleyin."
                }), 500

            # Create workbook and sheets
            wb = Workbook()
            ws_pg = wb.active
            ws_pg.title = "Performans Göstergeleri"
            ws_faaliyet = wb.create_sheet("Faaliyetler")
        
            # --- Sheet 1: Performans Göstergeleri ---
        
            # Headers
            pg_headers = [
                "Kodu", "Performans Adı", "Hedef", "Periyot", "Ağırlık (%)",
                "I. Çeyrek Hedef", "I. Çeyrek Gerç.", "I. Çeyrek Durum",
                "II. Çeyrek Hedef", "II. Çeyrek Gerç.", "II. Çeyrek Durum",
                "III. Çeyrek Hedef", "III. Çeyrek Gerç.", "III. Çeyrek Durum",
                "IV. Çeyrek Hedef", "IV. Çeyrek Gerç.", "IV. Çeyrek Durum"
            ]
            ws_pg.append(pg_headers)
        
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for cell in ws_pg[1]:
                cell.font = header_font
                cell.fill = header_fill
        
            # Fetch PG data
            pg_query = SurecPerformansGostergesi.query.filter_by(surec_id=surec_id).order_by(SurecPerformansGostergesi.id)
            performans_gostergeleri = pg_query.all()
        
            for pg in performans_gostergeleri:
                row_data = [
                    pg.kodu if hasattr(pg, 'kodu') and pg.kodu else f"PG-{pg.id}",
                    pg.ad,
                    pg.hedef_deger,
                    pg.periyot,
                    pg.agirlik if hasattr(pg, 'agirlik') else 0,
                ]
                # Fetch data for each quarter
                for ceyrek in range(1, 5):
                    # This is a simplified data fetching. A more robust implementation would query PerformansGostergeVeri
                    # and aggregate the results based on the logic in the frontend.
                    # For now, we'll leave these blank as a placeholder.
                    row_data.extend(["", "", ""]) # Hedef, Gerç., Durum
                ws_pg.append(row_data)
        
            # --- Sheet 2: Faaliyetler ---
            faaliyet_headers = ["No", "Faaliyet Adı", "Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
            ws_faaliyet.append(faaliyet_headers)
        
            # Style headers
            for cell in ws_faaliyet[1]:
                cell.font = header_font
                cell.fill = header_fill
        
            # Fetch Faaliyet data
            faaliyet_query = SurecFaaliyet.query.filter_by(surec_id=surec_id).order_by(SurecFaaliyet.id)
            faaliyetler = faaliyet_query.all()
        
            for i, faaliyet in enumerate(faaliyetler):
                row_data = [i + 1, faaliyet.ad]
                # Fetch monthly completion data
                # Note: FaaliyetTakip modeli ile ilişki kontrol edilmeli
                for ay in range(1, 13):
                    row_data.append("")  # Placeholder - gerçek veri için FaaliyetTakip sorgusu gerekli
                ws_faaliyet.append(row_data)
        
            # Save to a memory buffer
            output = BytesIO()
            wb.save(output)
            output.seek(0)
        
            filename = f"surec_karnesi_{surec.ad.replace(' ', '_')}_{yil}.xlsx"
        
            return tuple([output, filename])
        except Exception as e:
            current_app.logger.error(f'Excel export hatası: {e}')
            return dict({"success": False, "message": str(e)}), 500


################################################################################

    @staticmethod
    def get_pg_veri_detay(user, veri_id):
        """Tek bir PG verisinin detaylarını ve audit log'unu getir"""
        try:
            # Veriyi bul
            pg_veri = PerformansGostergeVeri.query.get(veri_id)
            if not pg_veri:
                return dict({'success': False, 'message': 'Veri bulunamadı'}), 404
        
            # Yetki kontrolü - verinin ait olduğu bireysel PG'yi kontrol et
            bireysel_pg = BireyselPerformansGostergesi.query.get(pg_veri.bireysel_pg_id)
            if not bireysel_pg:
                return dict({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
        
            # Süreç bazlı PG ise süreç yetkisi kontrol et
            surec_id = None
            if bireysel_pg.kaynak == 'Süreç' and bireysel_pg.kaynak_surec_id:
                surec_id = bireysel_pg.kaynak_surec_id
            elif bireysel_pg.kaynak_surec_pg_id:
                kaynak_surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
                surec_id = kaynak_surec_pg.surec_id if kaynak_surec_pg else None

            is_surec_verisi = surec_id is not None
            if is_surec_verisi:
                # Kurum yöneticisi kontrolü
                kurum_yoneticisi_mi = user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']
                if kurum_yoneticisi_mi:
                    surec = Surec.query.get(surec_id)
                    if not surec:
                        return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                    if user.sistem_rol != 'admin' and surec.kurum_id != user.kurum_id:
                        return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                else:
                    # Normal kullanıcı için lider/üye kontrolü
                    lider_mi = db.session.query(surec_liderleri).filter(
                        surec_liderleri.c.surec_id == surec_id,
                        surec_liderleri.c.user_id == user.id
                    ).first() is not None

                    uye_mi = db.session.query(surec_uyeleri).filter(
                        surec_uyeleri.c.surec_id == surec_id,
                        surec_uyeleri.c.user_id == user.id
                    ).first() is not None

                    if not (lider_mi or uye_mi):
                        return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
        
            # Audit log'ları getir
            audit_logs = PerformansGostergeVeriAudit.query.filter_by(
                pg_veri_id=veri_id
            ).order_by(PerformansGostergeVeriAudit.islem_tarihi.desc()).all()
        
            # Kullanıcı bilgilerini al
            veri_sahibi = User.query.get(pg_veri.user_id)
            veri_sahibi_adi = ''
            if veri_sahibi:
                if veri_sahibi.first_name and veri_sahibi.last_name:
                    veri_sahibi_adi = f"{veri_sahibi.first_name} {veri_sahibi.last_name}"
                else:
                    veri_sahibi_adi = veri_sahibi.username
        
            # Audit log'ları formatla
            audit_list = []
            for audit in audit_logs:
                islem_yapan = User.query.get(audit.islem_yapan_user_id)
                islem_yapan_adi = ''
                if islem_yapan:
                    if islem_yapan.first_name and islem_yapan.last_name:
                        islem_yapan_adi = f"{islem_yapan.first_name} {islem_yapan.last_name}"
                    else:
                        islem_yapan_adi = islem_yapan.username
            
                audit_list.append({
                    'id': audit.id,
                    'islem_tipi': audit.islem_tipi,
                    'eski_deger': audit.eski_deger,
                    'yeni_deger': audit.yeni_deger,
                    'degisiklik_aciklama': audit.degisiklik_aciklama,
                    'islem_yapan': islem_yapan_adi,
                    'islem_tarihi': audit.islem_tarihi.strftime('%d.%m.%Y %H:%M:%S') if audit.islem_tarihi else None
                })
        
            # Son güncelleyen bilgisini audit log'dan al
            son_guncelleyen = None
            son_guncelleme_tarihi = None
            if audit_list and len(audit_list) > 0:
                son_islem = audit_list[0]  # En son işlem
                if son_islem['islem_tipi'] == 'GUNCELLE':
                    son_guncelleyen = son_islem['islem_yapan']
                    son_guncelleme_tarihi = son_islem['islem_tarihi']
        
            # Veri bilgilerini formatla (frontend'in beklediği alan adlarıyla)
            veri_data = {
                'id': pg_veri.id,
                'bireysel_pg_id': pg_veri.bireysel_pg_id,
                'pg_adi': bireysel_pg.ad,
                'yil': pg_veri.yil,
                'veri_tarihi': pg_veri.veri_tarihi.strftime('%d.%m.%Y') if pg_veri.veri_tarihi else None,
                'giris_periyot_tipi': pg_veri.giris_periyot_tipi,
                'giris_periyot_no': pg_veri.giris_periyot_no,
                'giris_periyot_ay': pg_veri.giris_periyot_ay,
                'hedef_deger': pg_veri.hedef_deger,
                'gerceklesen_deger': str(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger is not None else '-',
                'durum': pg_veri.durum,
                'durum_yuzdesi': pg_veri.durum_yuzdesi,
                'aciklama': pg_veri.aciklama,
                'olusturan': veri_sahibi_adi,  # Frontend'in beklediği alan adı
                'olusturma_tarihi': pg_veri.created_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.created_at else None,  # Frontend'in beklediği alan adı
                'guncelleyen': son_guncelleyen,  # Son güncelleyen audit log'dan
                'guncelleme_tarihi': son_guncelleme_tarihi,  # Son güncelleme tarihi audit log'dan
                # Geriye dönük uyumluluk için eski alan adları da ekle
                'veri_sahibi': veri_sahibi_adi,
                'created_at': pg_veri.created_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.created_at else None,
                'updated_at': pg_veri.updated_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.updated_at else None
            }
        
            # Yetki bilgileri
            if is_surec_verisi:
                # Süreç karnesi yazma işlemleri sadece yönetim rolleri
                duzenleyebilir = user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']
                silebilir = duzenleyebilir
            else:
                # Süreç dışı (bireysel) KPI'larda mevcut davranışı koru
                duzenleyebilir = (pg_veri.user_id == user.id) or (user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim'])
                silebilir = duzenleyebilir

            yetki = {
                'duzenleyebilir': duzenleyebilir,
                'silebilir': silebilir
            }
        
            # Hesaplama yöntemi bilgisini ekle (bireysel PG'den)
            hesaplama_yontemi = None
            if bireysel_pg.kaynak_surec_pg_id:
                surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
                if surec_pg:
                    hesaplama_yontemi = surec_pg.veri_toplama_yontemi
        
            veri_data['hesaplama_yontemi'] = hesaplama_yontemi
            veri_data['surec_pg_id'] = bireysel_pg.kaynak_surec_pg_id if bireysel_pg.kaynak == 'Süreç' else None
        
            return dict({
                'success': True,
                'veri': veri_data,
                'audit_log': audit_list,
                'yetki': yetki,
                'bireysel_pg_id': bireysel_pg.id,
                'surec_pg_id': bireysel_pg.kaynak_surec_pg_id if bireysel_pg.kaynak == 'Süreç' else None
            })
        except Exception as e:
            current_app.logger.error(f'PG veri detay hatası: {e}')
            import traceback
            current_app.logger.error(traceback.format_exc())
            return dict({'success': False, 'message': str(e)}), 500


################################################################################

    @staticmethod
    def get_pg_veri_detay_toplu(user, veri_idleri):
        """Birden fazla PG verisinin detaylarını toplu olarak getir"""
        try:
            if False:
                return dict({'success': False, 'message': 'Content-Type application/json olmalı'}), 400

        
            current_app.logger.info(f'Toplu veri detay isteği alındı: {data}')
        
            if not data:
                return dict({'success': False, 'message': 'Request body boş'}), 400
        
        
            current_app.logger.info(f'Veri ID listesi: {veri_idleri}, Tip: {type(veri_idleri)}')
        
            if not veri_idleri:
                return dict({'success': False, 'message': 'Veri ID listesi boş'}), 400
        
            if not isinstance(veri_idleri, list):
                return dict({'success': False, 'message': f'Veri ID listesi geçersiz tip: {type(veri_idleri)}'}), 400
        
            # Liste elemanlarını integer'a çevir
            try:
                veri_idleri = [int(vid) for vid in veri_idleri]
            except (ValueError, TypeError) as e:
                current_app.logger.error(f'Veri ID dönüştürme hatası: {e}')
                return dict({'success': False, 'message': f'Veri ID\'leri integer\'a dönüştürülemedi: {e}'}), 400
        
            # Verileri getir
            pg_veriler = PerformansGostergeVeri.query.filter(PerformansGostergeVeri.id.in_(veri_idleri)).all()

            if not pg_veriler:
                return dict({'success': False, 'message': 'Veriler bulunamadı'}), 404

            pg_veri_by_id = {v.id: v for v in pg_veriler}
            missing_ids = [vid for vid in veri_idleri if vid not in pg_veri_by_id]
            if missing_ids:
                return dict({'success': False, 'message': 'Bazı veriler bulunamadı'}), 404
        
            # Her veri için yetki kontrolü ve detayları topla (kısmi sızıntıyı engellemek için strict)
            veriler_listesi = []
            for veri_id in veri_idleri:
                pg_veri = pg_veri_by_id[veri_id]
                # Yetki kontrolü
                bireysel_pg = BireyselPerformansGostergesi.query.get(pg_veri.bireysel_pg_id)
                if not bireysel_pg:
                    return dict({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
            
                # Süreç bazlı PG ise süreç yetkisi kontrol et
                surec_id = None
                if bireysel_pg.kaynak == 'Süreç' and bireysel_pg.kaynak_surec_id:
                    surec_id = bireysel_pg.kaynak_surec_id
                elif bireysel_pg.kaynak_surec_pg_id:
                    kaynak_surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
                    surec_id = kaynak_surec_pg.surec_id if kaynak_surec_pg else None

                if surec_id is not None:
                    surec = Surec.query.get(surec_id)
                    if not surec:
                        return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403

                    if user.sistem_rol == 'admin':
                        pass
                    elif user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
                        if surec.kurum_id != user.kurum_id:
                            return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                    else:
                        if surec.kurum_id != user.kurum_id:
                            return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403

                        lider_mi = db.session.query(surec_liderleri).filter(
                            surec_liderleri.c.surec_id == surec_id,
                            surec_liderleri.c.user_id == user.id
                        ).first() is not None

                        uye_mi = db.session.query(surec_uyeleri).filter(
                            surec_uyeleri.c.surec_id == surec_id,
                            surec_uyeleri.c.user_id == user.id
                        ).first() is not None

                        if not (lider_mi or uye_mi):
                            return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                else:
                    # Süreç dışı (bireysel) KPI verisi
                    if user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
                        if user.sistem_rol != 'admin':
                            veri_sahibi = User.query.get(pg_veri.user_id)
                            if veri_sahibi and veri_sahibi.kurum_id != user.kurum_id:
                                return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                    else:
                        if pg_veri.user_id != user.id:
                            return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
            
                # Kullanıcı bilgilerini al
                veri_sahibi = User.query.get(pg_veri.user_id)
                veri_sahibi_adi = ''
                if veri_sahibi:
                    if veri_sahibi.first_name and veri_sahibi.last_name:
                        veri_sahibi_adi = f"{veri_sahibi.first_name} {veri_sahibi.last_name}"
                    else:
                        veri_sahibi_adi = veri_sahibi.username
            
                # Audit log'ları getir (tüm geçmişi göster)
                audit_logs = PerformansGostergeVeriAudit.query.filter_by(
                    pg_veri_id=pg_veri.id
                ).order_by(PerformansGostergeVeriAudit.islem_tarihi.desc()).all()
            
                audit_list = []
                for audit in audit_logs:
                    islem_yapan = User.query.get(audit.islem_yapan_user_id)
                    islem_yapan_adi = ''
                    if islem_yapan:
                        if islem_yapan.first_name and islem_yapan.last_name:
                            islem_yapan_adi = f"{islem_yapan.first_name} {islem_yapan.last_name}"
                        else:
                            islem_yapan_adi = islem_yapan.username
                
                    audit_list.append({
                        'id': audit.id,
                        'islem_tipi': audit.islem_tipi,
                        'eski_deger': audit.eski_deger,
                        'yeni_deger': audit.yeni_deger,
                        'degisiklik_aciklama': audit.degisiklik_aciklama,
                        'islem_yapan': islem_yapan_adi,
                        'islem_tarihi': audit.islem_tarihi.strftime('%d.%m.%Y %H:%M:%S') if audit.islem_tarihi else None
                    })
            
                # Frontend'in beklediği yapıya uygun formatla
                veriler_listesi.append({
                    'veri': {
                        'id': pg_veri.id,
                        'bireysel_pg_id': pg_veri.bireysel_pg_id,
                        'pg_adi': bireysel_pg.ad,
                        'yil': pg_veri.yil,
                        'veri_tarihi': pg_veri.veri_tarihi.strftime('%d.%m.%Y') if pg_veri.veri_tarihi else None,
                        'giris_periyot_tipi': pg_veri.giris_periyot_tipi,
                        'giris_periyot_no': pg_veri.giris_periyot_no,
                        'giris_periyot_ay': pg_veri.giris_periyot_ay,
                        'hedef_deger': pg_veri.hedef_deger,
                        'gerceklesen_deger': str(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger is not None else '-',
                        'durum': pg_veri.durum,
                        'durum_yuzdesi': pg_veri.durum_yuzdesi,
                        'aciklama': pg_veri.aciklama,
                        'olusturan': veri_sahibi_adi,
                        'olusturma_tarihi': pg_veri.created_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.created_at else None,
                        'guncelleyen': None,  # Son güncelleyen audit log'dan alınacak
                        'guncelleme_tarihi': pg_veri.updated_at.strftime('%d.%m.%Y %H:%M:%S') if pg_veri.updated_at else None
                    },
                    'audit_log': audit_list,
                    'yetki': {
                        'can_edit': True,
                        'can_delete': True
                    }
                })
            
                # Son güncelleyen bilgisini audit log'dan al
                if audit_list and len(audit_list) > 0:
                    son_islem = audit_list[0]  # En son işlem
                    if son_islem['islem_tipi'] == 'GUNCELLE':
                        veriler_listesi[-1]['veri']['guncelleyen'] = son_islem['islem_yapan']
                        veriler_listesi[-1]['veri']['guncelleme_tarihi'] = son_islem['islem_tarihi']
        
            # Yetki bilgileri
            # (Bu endpoint süreç karnesi veri detayında kullanılıyor; yazma işlemleri yönetim rolleri ile sınırlandırılır)
            yetki = {
                'duzenleyebilir': user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim'],
                'silebilir': user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']
            }
        
            return dict({
                'success': True,
                'veriler': veriler_listesi,
                'yetki': yetki
            })
        except Exception as e:
            current_app.logger.error(f'Toplu PG veri detay hatası: {e}')
            import traceback
            current_app.logger.error(traceback.format_exc())
            return dict({'success': False, 'message': str(e)}), 500


################################################################################

    @staticmethod
    def update_pg_veri(user, veri_id, gerceklesen_deger, aciklama):
        """PG verisini güncelle - Süreç lideri tüm kullanıcıların verilerini güncelleyebilir"""
        try:
            # Veriyi bul
            pg_veri = PerformansGostergeVeri.query.get(veri_id)
            if not pg_veri:
                return dict({'success': False, 'message': 'Veri bulunamadı'}), 404
        
            # Bireysel PG'yi bul
            bireysel_pg = BireyselPerformansGostergesi.query.get(pg_veri.bireysel_pg_id)
            if not bireysel_pg:
                return dict({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
        
            # Yetki kontrolü
            # Süreç karnesi verileri için yazma işlemleri sadece yönetim rolleri
            surec_id = None
            if bireysel_pg.kaynak == 'Süreç' and bireysel_pg.kaynak_surec_id:
                surec_id = bireysel_pg.kaynak_surec_id
            elif bireysel_pg.kaynak_surec_pg_id:
                kaynak_surec_pg = SurecPerformansGostergesi.query.get(bireysel_pg.kaynak_surec_pg_id)
                surec_id = kaynak_surec_pg.surec_id if kaynak_surec_pg else None

            is_surec_verisi = surec_id is not None
            kurum_yoneticisi_mi = user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']

            if is_surec_verisi:
                if not kurum_yoneticisi_mi:
                    return dict({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
                surec = Surec.query.get(surec_id)
                if not surec:
                    return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                if user.sistem_rol != 'admin' and surec.kurum_id != user.kurum_id:
                    return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
            else:
                # Süreç dışı (bireysel) KPI davranışını koru
                if kurum_yoneticisi_mi:
                    # Yönetim rolleri kendi kurumundaki verileri güncelleyebilir
                    veri_sahibi = User.query.get(pg_veri.user_id)
                    if veri_sahibi and user.sistem_rol != 'admin' and veri_sahibi.kurum_id != user.kurum_id:
                        return dict({'success': False, 'message': 'Bu veriye erişim yetkiniz yok'}), 403
                else:
                    if pg_veri.user_id != user.id:
                        return dict({'success': False, 'message': 'Bu veriyi güncelleme yetkiniz yok'}), 403
        
            # Güncelleme verilerini al
        
        
        
        
            # Eski değerleri kaydet (audit log için)
            eski_gerceklesen = pg_veri.gerceklesen_deger
            eski_aciklama = pg_veri.aciklama
        
            # Değerleri güncelle
            if gerceklesen_deger is not None:
                try:
                    # String'den float'a çevir
                    pg_veri.gerceklesen_deger = float(gerceklesen_deger) if gerceklesen_deger else None
                except (ValueError, TypeError):
                    return dict({'success': False, 'message': 'Geçersiz gerçekleşen değer formatı'}), 400
        
            if aciklama is not None:
                pg_veri.aciklama = aciklama
        
            # Durumu hesapla
            if pg_veri.hedef_deger and pg_veri.gerceklesen_deger is not None:
                durum, durum_yuzdesi = hesapla_durum(
                    float(pg_veri.hedef_deger) if pg_veri.hedef_deger else None,
                    float(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger else None
                )
                pg_veri.durum = durum
                pg_veri.durum_yuzdesi = durum_yuzdesi
        
            pg_veri.updated_by = user.id
            pg_veri.updated_at = datetime.utcnow()
        
            # Audit log oluştur
            degisiklik_aciklama = []
            if eski_gerceklesen != pg_veri.gerceklesen_deger:
                degisiklik_aciklama.append(f'Gerçekleşen değer: {eski_gerceklesen} → {pg_veri.gerceklesen_deger}')
            if eski_aciklama != pg_veri.aciklama:
                degisiklik_aciklama.append(f'Açıklama güncellendi')
        
            if degisiklik_aciklama:
                audit = PerformansGostergeVeriAudit(
                    pg_veri_id=pg_veri.id,
                    islem_tipi='GUNCELLE',
                    eski_deger=str(eski_gerceklesen) if eski_gerceklesen is not None else None,
                    yeni_deger=str(pg_veri.gerceklesen_deger) if pg_veri.gerceklesen_deger is not None else None,
                    degisiklik_aciklama='; '.join(degisiklik_aciklama),
                    islem_yapan_user_id=user.id,
                    islem_tarihi=datetime.utcnow()
                )
                db.session.add(audit)
        
            db.session.commit()
        
            # Skor Motoru: PG verisi güncellenince Vizyon puanını tüm hiyerarşide yeniden hesapla
            try:
                kurum_id = getattr(user, 'kurum_id', None)
                if kurum_id:
                    from services.score_engine_service import recalc_on_pg_data_change
                    recalc_on_pg_data_change(kurum_id)
            except Exception as ex:
                current_app.logger.debug("Skor motoru güncellemesi atlandı: %s", ex)
        
            return dict({
                'success': True,
                'message': 'Veri başarıyla güncellendi'
            })
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'PG veri güncelleme hatası: %s', e)
            import traceback
            current_app.logger.error(traceback.format_exc())
            return dict({'success': False, 'message': str(e)}), 500


################################################################################

