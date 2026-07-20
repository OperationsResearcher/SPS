"""Bulk Excel import servisi (Sprint 44).

KPI, Process, Project için Excel veya CSV bulk import.
Validation + dry-run preview + commit.

Kullanım:
    from app.services.bulk_import_service import import_kpi_data_from_excel
    result = import_kpi_data_from_excel(file_path, tenant_id=27, dry_run=True)
    # {valid_rows: 95, error_rows: 5, errors: [...], dry_run: True}
"""
from __future__ import annotations

import io
from collections import Counter
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook

from extensions import db
from app.constants.periods import normalize_period_type
from app.models.process import Process, ProcessKpi, KpiData


# ─── KPI Data bulk import ────────────────────────────────────────────────────

KPI_DATA_COLUMNS = [
    "kpi_code", "year", "period_type", "period_no",
    "actual_value", "target_value", "description"
]


def import_kpi_data_from_excel(
    file_bytes: bytes,
    tenant_id: int,
    user_id: int,
    dry_run: bool = True,
    skip_header: bool = True,
) -> dict:
    """Excel'den KPI veri girişi.

    Excel formatı:
    | kpi_code | year | period_type | period_no | actual_value | target_value | description |
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return {"success": False, "message": f"Excel açılamadı: {e}", "valid_rows": 0, "errors": []}

    # KPI code → ProcessKpi.id map (tenant scope)
    # H-29: duplicate kodları Counter ile tespit et, map'e sadece tekil kodları koy
    _kpis = (
        db.session.query(ProcessKpi)
        .join(Process)
        .filter(Process.tenant_id == tenant_id, ProcessKpi.is_active.is_(True))
        .all()
    )
    _code_counts = Counter(k.code for k in _kpis if k.code)
    _duplicate_codes = {code for code, cnt in _code_counts.items() if cnt > 1}
    kpi_map = {k.code: k.id for k in _kpis if k.code and k.code not in _duplicate_codes}

    # MÜHÜR (K8, yıl bazlı Faz 2): kurumun mühürlü yılları — tek sorguda,
    # döngü içinde sorgulamak N+1 olurdu.
    from app.models.plan_year import PlanYear
    from app.services.date_sovereign import KAPALI_DURUMLAR
    _muhurlu_yillar = {
        py.year for py in PlanYear.query
        .filter(PlanYear.tenant_id == tenant_id,
                PlanYear.status.in_(KAPALI_DURUMLAR))
        .all()
    }

    valid: list[dict] = []
    errors: list[dict] = []
    row_iter = ws.iter_rows(values_only=True)
    if skip_header:
        next(row_iter, None)

    for row_idx, row in enumerate(row_iter, start=2 if skip_header else 1):
        if not any(c is not None and str(c).strip() for c in row):
            continue  # boş satır

        try:
            cells = list(row) + [None] * (7 - len(row))

            # M-17: Excel injection koruması — = + - @ ile başlayan string değerlere ' prefix ekle
            cells = [
                ("'" + c if isinstance(c, str) and c and c[0] in ('=', '+', '-', '@') else c)
                for c in cells
            ]

            kpi_code = str(cells[0] or "").strip()
            # kpi_code injection-prefix'i kaldır (kodun kendisi değişmemeli)
            if kpi_code.startswith("'"):
                kpi_code = kpi_code[1:]
            year_val = cells[1]
            period_type = str(cells[2] or "Aylık").strip()
            period_no = cells[3]
            actual = cells[4]
            target = cells[5]
            desc = str(cells[6] or "").strip() or None

            # Validation
            if not kpi_code:
                raise ValueError("kpi_code boş")
            if kpi_code in _duplicate_codes:
                raise ValueError(f"Veritabanında duplicate KPI kodu: '{kpi_code}' — önce DB'yi düzeltin")
            if kpi_code not in kpi_map:
                raise ValueError(f"Tanımsız KPI kodu: '{kpi_code}'")
            if year_val is None:
                raise ValueError("year boş")
            try:
                year_int = int(year_val)
            except (ValueError, TypeError):
                raise ValueError(f"Geçersiz year: {year_val}")
            if period_type not in ("Aylık", "Çeyreklik", "Yıllık", "Haftalık"):
                raise ValueError(f"Geçersiz period_type: {period_type}")
            try:
                period_no_int = int(period_no) if period_no else 1
            except (ValueError, TypeError):
                raise ValueError(f"Geçersiz period_no: {period_no}")
            if actual in (None, ""):
                raise ValueError("actual_value boş")
            # MÜHÜR (K8, yıl bazlı Faz 2): mühürlü yıla toplu içe aktarım da
            # yapılamaz. Satır bazında reddedilir ki kullanıcı hangi yılın
            # kapalı olduğunu görsün — dosyanın tamamı sessizce düşmesin.
            if year_int in _muhurlu_yillar:
                raise ValueError(
                    f"{year_int} plan dönemi mühürlü — bu yıla veri aktarılamaz"
                )

            # Veri tarihi: yıl + period_no → tahmini date
            month = period_no_int if period_type == "Aylık" else (period_no_int * 3 if period_type == "Çeyreklik" else 12)
            month = max(1, min(12, month))
            data_date = date(year_int, month, 28)

            valid.append({
                "process_kpi_id": kpi_map[kpi_code],
                "year": year_int,
                "data_date": data_date,
                # DB'ye kanonik biçim yazılır (TASK-253). Excel'de kullanıcı
                # Türkçe yazar ("Aylık") — doğrulama (satır ~105) ve ay hesabı
                # (~115) o biçimde kalır; sapma yalnız DB sınırında düzelir.
                # NOT: bu servis `bulk_insert_mappings` kullanıyor → ORM `set`
                # event'leri ÇALIŞMAZ, bu yüzden normalize burada elle yapılır.
                "period_type": normalize_period_type(period_type),
                "period_no": period_no_int,
                "period_month": month,
                "target_value": float(str(target).replace(",", ".")) if target not in (None, "") else None,
                "actual_value": float(str(actual).replace(",", ".")),
                "description": desc,
                "user_id": user_id,
                "is_active": True,
                "row_idx": row_idx,
                "kpi_code": kpi_code,
            })
        except Exception as e:
            errors.append({"row": row_idx, "error": "Satır işlenemedi.", "data": list(row)[:7]})

    result = {
        "success": True,
        "dry_run": dry_run,
        "valid_rows": len(valid),
        "error_rows": len(errors),
        "errors": errors[:50],  # ilk 50
    }

    if dry_run or not valid:
        return result

    # Commit
    rows_to_insert = [{k: v for k, v in r.items() if k not in ("row_idx", "kpi_code")} for r in valid]
    try:
        db.session.bulk_insert_mappings(KpiData, rows_to_insert)
        db.session.commit()
        result["inserted"] = len(rows_to_insert)
    except Exception as e:
        db.session.rollback()
        result["success"] = False
        result["message"] = f"Insert hatası: {e}"
        result["inserted"] = 0

    return result


def make_kpi_template_excel(tenant_id: int) -> bytes:
    """Boş şablon Excel üret (header + örnek satır)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Verileri"
    ws.append([
        "kpi_code (zorunlu)",
        "year (zorunlu, integer)",
        "period_type (Aylık/Çeyreklik/Yıllık)",
        "period_no (1-12/1-4/1)",
        "actual_value (zorunlu)",
        "target_value",
        "description"
    ])
    ws.append(["P2M-01", 2026, "Aylık", 1, 87.5, 88, "Ocak ölçümü"])
    ws.append(["P2M-02", 2026, "Aylık", 1, 96.2, 98, ""])
    # Tablo formatla
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
