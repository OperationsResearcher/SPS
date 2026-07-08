# -*- coding: utf-8 -*-
"""Kurulum verisi import sihirbazı servisi (TASK-235).

Tasarım + mutabakat: docs/paketler/KURULUM-IMPORT-SIHIRBAZI.md
- Tek .xlsx, 3 sayfa: Surecler / PG_Tanimlari / Strateji (opsiyonel)
- dry_run: hiçbir yazma yapmadan satır-satır plan (add/update/error)
- apply: TEK transaction (TASK-228 dersi: kısmi başarı yok, hata → tam rollback)
- Upsert: Süreç=(tenant_id, code), PG=(process_id, code), Strateji=title — mevcutsa GÜNCELLE
- Silme asla yapılmaz.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from extensions import db
from app.models.core import Strategy, SubStrategy
from app.models.process import Process, ProcessKpi, ProcessSubStrategyLink

SHEET_PROCESSES = "Surecler"
SHEET_KPIS = "PG_Tanimlari"
SHEET_STRATEGY = "Strateji"

VALID_PERIODS = ("Aylık", "Çeyreklik", "Yıllık")
VALID_COLLECTION = ("Toplama", "Ortalama", "Son Değer")
VALID_GOSTERGE = ("İyileştirme", "Koruma", "Bilgi Amaçlı")


def _clean(val):
    """Hücre değerini normalize et; Excel injection prefix'lerini etkisizleştir."""
    if val is None:
        return ""
    s = str(val).strip()
    if s and s[0] in ("=", "+", "@"):
        s = "'" + s
    return s


def _num(val, field, row, errors, lo=None, hi=None):
    """Sayısal alan parse; hatada errors'a yazar, None döner."""
    if val in (None, ""):
        return None
    try:
        f = float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        errors.append({"row": row, "sheet": None, "error": f"{field} sayısal değil: {val!r}"})
        return None
    if lo is not None and (f < lo or f > hi):
        errors.append({"row": row, "sheet": None, "error": f"{field} {lo}-{hi} aralığında olmalı: {f}"})
        return None
    return f


def parse_workbook(file_bytes: bytes, tenant_id: int) -> dict:
    """Çalışma kitabını parse edip yazma YAPMADAN plan üretir.

    Dönen plan:
    {
      "success": bool, "message": str|None,
      "processes": [{"row", "action": add|update, "code", "name", "parent_code", "weight", ...}],
      "kpis":      [{"row", "action", "process_code", "code", "name", ...}],
      "strategies":[{"row", "action", "title", "sub_title", "process_codes", "weight"}],
      "errors":    [{"sheet", "row", "error"}],
      "summary":   {"add": n, "update": n, "error": n}
    }
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "message": f"Excel açılamadı: {e}", "errors": []}

    plan = {"success": True, "message": None, "processes": [], "kpis": [],
            "strategies": [], "errors": []}
    errors = plan["errors"]

    # Mevcut kayıtlar (upsert kararı için)
    existing_proc = {
        p.code: p.id
        for p in Process.query.filter_by(tenant_id=tenant_id, is_active=True).all()
        if p.code
    }
    existing_kpis = {}  # (process_code, kpi_code) -> kpi_id
    for k, pcode in (
        db.session.query(ProcessKpi, Process.code)
        .join(Process, ProcessKpi.process_id == Process.id)
        .filter(Process.tenant_id == tenant_id, ProcessKpi.is_active.is_(True))
        .all()
    ):
        if k.code and pcode:
            existing_kpis[(pcode, k.code)] = k.id

    # ── Sayfa 1: Surecler ──────────────────────────────────────────────────
    file_proc_codes = set()
    if SHEET_PROCESSES in wb.sheetnames:
        ws = wb[SHEET_PROCESSES]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)  # başlık
        for idx, row in enumerate(rows, start=2):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            cells = list(row) + [None] * (6 - len(row))
            code = _clean(cells[0])
            name = _clean(cells[1])
            parent_code = _clean(cells[2]) or None
            if not code:
                errors.append({"sheet": SHEET_PROCESSES, "row": idx, "error": "Süreç Kodu boş"})
                continue
            if not name:
                errors.append({"sheet": SHEET_PROCESSES, "row": idx, "error": "Süreç Adı boş"})
                continue
            if len(name) > 200:
                errors.append({"sheet": SHEET_PROCESSES, "row": idx, "error": "Süreç Adı 200 karakteri aşıyor"})
                continue
            if code in file_proc_codes:
                errors.append({"sheet": SHEET_PROCESSES, "row": idx, "error": f"Dosyada tekrarlanan Süreç Kodu: {code}"})
                continue
            _err_ref = []
            weight = _num(cells[3], "Ağırlık", idx, _err_ref, 0, 100)
            for e in _err_ref:
                e["sheet"] = SHEET_PROCESSES
                errors.append(e)
            if _err_ref:
                continue
            file_proc_codes.add(code)
            plan["processes"].append({
                "row": idx,
                "action": "update" if code in existing_proc else "add",
                "code": code, "name": name, "parent_code": parent_code,
                "weight": weight,
                "document_no": _clean(cells[4]) or None,
                "revision_no": _clean(cells[5]) or None,
            })

        # Üst süreç referans + döngü kontrolü (dosya içi zincir)
        parent_of = {p["code"]: p["parent_code"] for p in plan["processes"]}
        for p in plan["processes"]:
            pc = p["parent_code"]
            if pc and pc not in file_proc_codes and pc not in existing_proc:
                errors.append({"sheet": SHEET_PROCESSES, "row": p["row"],
                               "error": f"Üst Süreç Kodu bulunamadı: {pc}"})
                p["action"] = "error"
            # döngü: kendi zincirinde kendine dönüyor mu
            seen, cur = set(), pc
            while cur and cur in parent_of:
                if cur in seen or cur == p["code"]:
                    errors.append({"sheet": SHEET_PROCESSES, "row": p["row"],
                                   "error": f"Üst süreç döngüsü: {p['code']} → {pc}"})
                    p["action"] = "error"
                    break
                seen.add(cur)
                cur = parent_of.get(cur)

    # ── Sayfa 2: PG_Tanimlari ─────────────────────────────────────────────
    file_kpi_keys = set()
    if SHEET_KPIS in wb.sheetnames:
        ws = wb[SHEET_KPIS]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        for idx, row in enumerate(rows, start=2):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            cells = list(row) + [None] * (9 - len(row))
            pcode = _clean(cells[0])
            kcode = _clean(cells[1])
            name = _clean(cells[2])
            period = _clean(cells[5])
            if not pcode or not kcode or not name:
                errors.append({"sheet": SHEET_KPIS, "row": idx,
                               "error": "Süreç Kodu, PG Kodu ve PG Adı zorunlu"})
                continue
            if pcode not in file_proc_codes and pcode not in existing_proc:
                errors.append({"sheet": SHEET_KPIS, "row": idx,
                               "error": f"Süreç Kodu bulunamadı: {pcode}"})
                continue
            if not period or period not in VALID_PERIODS:
                errors.append({"sheet": SHEET_KPIS, "row": idx,
                               "error": f"Periyot geçersiz: {period!r} — {'/'.join(VALID_PERIODS)} olmalı"})
                continue
            collection = _clean(cells[6]) or "Ortalama"
            if collection not in VALID_COLLECTION:
                errors.append({"sheet": SHEET_KPIS, "row": idx,
                               "error": f"Toplama Yöntemi geçersiz: {collection!r} — {'/'.join(VALID_COLLECTION)}"})
                continue
            gosterge = _clean(cells[7]) or None
            if gosterge and gosterge not in VALID_GOSTERGE:
                errors.append({"sheet": SHEET_KPIS, "row": idx,
                               "error": f"Gösterge Türü geçersiz: {gosterge!r} — {'/'.join(VALID_GOSTERGE)}"})
                continue
            if (pcode, kcode) in file_kpi_keys:
                errors.append({"sheet": SHEET_KPIS, "row": idx,
                               "error": f"Dosyada tekrarlanan PG: {pcode}/{kcode}"})
                continue
            _err_ref = []
            weight = _num(cells[8], "Ağırlık", idx, _err_ref, 0, 100)
            for e in _err_ref:
                e["sheet"] = SHEET_KPIS
                errors.append(e)
            if _err_ref:
                continue
            file_kpi_keys.add((pcode, kcode))
            plan["kpis"].append({
                "row": idx,
                "action": "update" if (pcode, kcode) in existing_kpis else "add",
                "process_code": pcode, "code": kcode, "name": name,
                "target_value": _clean(cells[3]) or None,
                "unit": _clean(cells[4]) or None,
                "period": period,
                "data_collection_method": collection,
                "gosterge_turu": gosterge,
                "weight": weight,
            })

    # ── Sayfa 3: Strateji (opsiyonel) ─────────────────────────────────────
    if SHEET_STRATEGY in wb.sheetnames:
        ws = wb[SHEET_STRATEGY]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        for idx, row in enumerate(rows, start=2):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            cells = list(row) + [None] * (3 - len(row))
            title = _clean(cells[0])
            sub_title = _clean(cells[1]) or None
            proc_codes_raw = _clean(cells[2])
            if not title:
                errors.append({"sheet": SHEET_STRATEGY, "row": idx, "error": "Ana Strateji Adı boş"})
                continue
            proc_codes = [c.strip() for c in proc_codes_raw.split(",") if c.strip()] if proc_codes_raw else []
            bad = [c for c in proc_codes if c not in file_proc_codes and c not in existing_proc]
            if bad:
                errors.append({"sheet": SHEET_STRATEGY, "row": idx,
                               "error": f"Bağlı Süreç Kodları bulunamadı: {', '.join(bad)}"})
                continue
            if proc_codes and not sub_title:
                errors.append({"sheet": SHEET_STRATEGY, "row": idx,
                               "error": "Süreç bağlantısı için Alt Strateji Adı zorunlu (bağ alt stratejiye kurulur)"})
                continue
            plan["strategies"].append({
                "row": idx, "action": "upsert",
                "title": title, "sub_title": sub_title,
                "process_codes": proc_codes,
            })

    ok_items = ([p for p in plan["processes"] if p["action"] != "error"]
                + [k for k in plan["kpis"]] + plan["strategies"])
    plan["summary"] = {
        "add": sum(1 for i in ok_items if i.get("action") == "add"),
        "update": sum(1 for i in ok_items if i.get("action") == "update"),
        "strategy_rows": len(plan["strategies"]),
        "error": len(errors),
    }
    return plan


def apply_workbook(file_bytes: bytes, tenant_id: int, user_id: int,
                   skip_errors: bool = False, plan_year_id: int | None = None) -> dict:
    """Planı TEK transaction'da uygular.

    skip_errors=False iken herhangi bir hata varsa hiçbir şey yazılmaz.
    Strateji sayfası plan_year_id gerektirir; yoksa sayfa atlanır (uyarıyla).
    """
    plan = parse_workbook(file_bytes, tenant_id)
    if not plan.get("success"):
        return plan
    if plan["errors"] and not skip_errors:
        return {**plan, "applied": False,
                "message": "Hatalı satırlar var; düzeltin veya 'hatalıları atla' seçin."}

    warnings = []
    if plan["strategies"] and not plan_year_id:
        warnings.append("Aktif plan yılı yok — Strateji sayfası atlandı.")
        plan["strategies"] = []

    try:
        # 1) Süreçler — önce hepsi (parent bağı ikinci geçişte)
        code_to_proc = {
            p.code: p
            for p in Process.query.filter_by(tenant_id=tenant_id, is_active=True).all()
            if p.code
        }
        applied_proc = {"add": 0, "update": 0}
        valid_procs = [p for p in plan["processes"] if p["action"] in ("add", "update")]
        for item in valid_procs:
            proc = code_to_proc.get(item["code"])
            if proc is None:
                proc = Process(tenant_id=tenant_id, code=item["code"], is_active=True)
                db.session.add(proc)
                applied_proc["add"] += 1
            else:
                applied_proc["update"] += 1
            proc.name = item["name"]
            for f in ("weight", "document_no", "revision_no"):
                if item.get(f) is not None:
                    setattr(proc, f, item[f])
            code_to_proc[item["code"]] = proc
        db.session.flush()
        for item in valid_procs:  # parent bağı (id'ler flush ile oluştu)
            if item["parent_code"]:
                parent = code_to_proc.get(item["parent_code"])
                if parent is not None:
                    code_to_proc[item["code"]].parent_id = parent.id

        # 2) PG tanımları
        applied_kpi = {"add": 0, "update": 0}
        kpi_lookup = {}
        for k, pcode in (
            db.session.query(ProcessKpi, Process.code)
            .join(Process, ProcessKpi.process_id == Process.id)
            .filter(Process.tenant_id == tenant_id, ProcessKpi.is_active.is_(True))
            .all()
        ):
            if k.code and pcode:
                kpi_lookup[(pcode, k.code)] = k
        for item in plan["kpis"]:
            proc = code_to_proc.get(item["process_code"])
            if proc is None:
                continue  # parse aşamasında yakalanmış olmalı
            kpi = kpi_lookup.get((item["process_code"], item["code"]))
            if kpi is None:
                kpi = ProcessKpi(process_id=proc.id, code=item["code"], is_active=True)
                db.session.add(kpi)
                applied_kpi["add"] += 1
            else:
                applied_kpi["update"] += 1
            kpi.name = item["name"]
            for f in ("target_value", "unit", "period", "data_collection_method",
                      "gosterge_turu", "weight"):
                if item.get(f) is not None:
                    setattr(kpi, f, item[f])

        # 3) Strateji (opsiyonel) — upsert ad ile, aktif plan yılında
        applied_str = {"strategy": 0, "sub": 0, "link": 0}
        if plan["strategies"] and plan_year_id:
            db.session.flush()
            str_by_title = {
                s.title: s
                for s in Strategy.query.filter_by(
                    tenant_id=tenant_id, plan_year_id=plan_year_id, is_active=True).all()
            }
            for item in plan["strategies"]:
                strategy = str_by_title.get(item["title"])
                if strategy is None:
                    strategy = Strategy(tenant_id=tenant_id, plan_year_id=plan_year_id,
                                        title=item["title"], is_active=True)
                    db.session.add(strategy)
                    db.session.flush()
                    str_by_title[item["title"]] = strategy
                    applied_str["strategy"] += 1
                if item["sub_title"]:
                    sub = SubStrategy.query.filter_by(
                        strategy_id=strategy.id, title=item["sub_title"], is_active=True).first()
                    if sub is None:
                        sub = SubStrategy(strategy_id=strategy.id, title=item["sub_title"],
                                          plan_year_id=plan_year_id, is_active=True)
                        db.session.add(sub)
                        db.session.flush()
                        applied_str["sub"] += 1
                    for pc in item["process_codes"]:
                        proc = code_to_proc.get(pc)
                        if proc is None:
                            continue
                        link = ProcessSubStrategyLink.query.filter_by(
                            process_id=proc.id, sub_strategy_id=sub.id).first()
                        if link is None:
                            db.session.add(ProcessSubStrategyLink(
                                process_id=proc.id, sub_strategy_id=sub.id))
                            applied_str["link"] += 1

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        import logging
        logging.getLogger(__name__).error("[setup-import] apply failed", exc_info=True)
        return {"success": False, "applied": False,
                "message": f"Uygulama hatası — hiçbir değişiklik yazılmadı: {e}"}

    return {
        "success": True, "applied": True, "warnings": warnings,
        "errors": plan["errors"] if skip_errors else [],
        "result": {"processes": applied_proc, "kpis": applied_kpi, "strategy": applied_str},
    }


def make_setup_template_excel() -> bytes:
    """3 sayfalı kurulum şablonu (başlıklar + örnek satırlar + periyot dropdown)."""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    ws = wb.active
    ws.title = SHEET_PROCESSES
    ws.append(["Süreç Kodu (*)", "Süreç Adı (*)", "Üst Süreç Kodu",
               "Ağırlık (0-100)", "Doküman No", "Revizyon No"])
    ws.append(["SR1", "Satış Süreci", "", 30, "DOK-001", "R0"])
    ws.append(["SR1.1", "Teklif Hazırlama", "SR1", 50, "", ""])

    ws2 = wb.create_sheet(SHEET_KPIS)
    ws2.append(["Süreç Kodu (*)", "PG Kodu (*)", "PG Adı (*)", "Hedef", "Birim",
                "Periyot (*)", "Toplama Yöntemi", "Gösterge Türü", "Ağırlık (0-100)"])
    ws2.append(["SR1", "PG-01", "Müşteri memnuniyeti", "90", "%", "Aylık", "Ortalama", "İyileştirme", 40])
    ws2.append(["SR1.1", "PG-02", "Teklif dönüş süresi", "3", "gün", "Aylık", "Ortalama", "Koruma", 60])
    dv = DataValidation(type="list", formula1=f'"{",".join(VALID_PERIODS)}"', allow_blank=False)
    ws2.add_data_validation(dv)
    dv.add("F2:F500")
    dv2 = DataValidation(type="list", formula1=f'"{",".join(VALID_COLLECTION)}"', allow_blank=True)
    ws2.add_data_validation(dv2)
    dv2.add("G2:G500")

    ws3 = wb.create_sheet(SHEET_STRATEGY)
    ws3.append(["Ana Strateji Adı (*)", "Alt Strateji Adı", "Bağlı Süreç Kodları (virgülle)"])
    ws3.append(["Müşteri odaklı büyüme", "Satış kanallarını genişlet", "SR1, SR1.1"])

    for sheet in (ws, ws2, ws3):
        for col in sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 3, 42)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
