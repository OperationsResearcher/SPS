from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file, current_app
from flask_login import login_required, current_user
from sqlalchemy import func
from sqlalchemy.orm import joinedload, selectinload
from app.models import db
from app.utils.process_utils import (
    process_descendant_ids,
    validate_process_parent_id,
    last_day_of_period,
    data_date_to_period_keys,
)
from app.models.core import User, SubStrategy, Strategy
from app.models.process import (
    Process,
    ProcessSubStrategyLink,
    ProcessKpi,
    ProcessActivity,
    ProcessActivityAssignee,
    ProcessActivityReminder,
    KpiData,
    KpiDataAudit,
    ActivityTrack,
    IndividualPerformanceIndicator,
    IndividualActivity,
    IndividualKpiData,
    IndividualActivityTrack,
    FavoriteKpi,
)
from app.services.cache_service import CacheService
from app.utils.db_sequence import is_pk_duplicate, sync_kpi_data_related_sequences
from datetime import datetime, timezone, date, timedelta
from io import BytesIO


process_bp = Blueprint('process_bp', __name__, url_prefix='/process')


_PRIVILEGED_ACTIVITY_ASSIGN_ROLES = {'tenant_admin', 'executive_manager', 'Admin'}


def _parse_local_datetime(value):
    if not value:
        return None
    val = str(value).strip()
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S'):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


def _activity_datetime_fallback(start_date, end_date):
    start_at = datetime.combine(start_date, datetime.min.time()) if start_date else None
    end_at = datetime.combine(end_date, datetime.max.time().replace(microsecond=0)) if end_date else None
    return start_at, end_at


def _activity_allowed_assignee_ids(process):
    allowed = set()
    for u in (process.leaders + process.members + process.owners):
        if u and u.is_active:
            allowed.add(int(u.id))
    return allowed


def _is_process_leader(process):
    return any(int(u.id) == int(current_user.id) for u in (process.leaders or []))


def _can_manage_activity(process, activity):
    if _can_crud_pg_and_activity(process):
        return True
    return any(int(l.user_id) == int(current_user.id) for l in (activity.assignment_links or []))


def _role_name() -> str:
    return current_user.role.name if current_user and current_user.role else ""


def _is_privileged_process_role() -> bool:
    return _role_name() in _PRIVILEGED_ACTIVITY_ASSIGN_ROLES


def _user_assigned_to_process(process) -> bool:
    if not process:
        return False
    uid = int(current_user.id)
    return (
        any(int(u.id) == uid for u in (process.leaders or []))
        or any(int(u.id) == uid for u in (process.members or []))
        or any(int(u.id) == uid for u in (process.owners or []))
    )


def _can_access_process(process) -> bool:
    return _is_privileged_process_role() or _user_assigned_to_process(process)


def _can_edit_process_record(process) -> bool:
    if _is_privileged_process_role():
        return True
    uid = int(current_user.id)
    return (
        any(int(u.id) == uid for u in (process.leaders or []))
        or any(int(u.id) == uid for u in (process.owners or []))
    )


def _can_crud_pg_and_activity(process) -> bool:
    return _can_edit_process_record(process)


def _can_enter_pgv(process) -> bool:
    return _can_access_process(process)


def _can_edit_kpi_data_row(entry, process) -> bool:
    if _is_privileged_process_role() or _can_edit_process_record(process):
        return True
    return bool(entry and int(entry.user_id) == int(current_user.id))


def _can_assign_multiple(process):
    role_name = current_user.role.name if current_user.role else ''
    return role_name in _PRIVILEGED_ACTIVITY_ASSIGN_ROLES or _is_process_leader(process)


def _normalize_assignee_ids(raw_ids, process):
    # JSON list veya tek değer gelebilir
    if raw_ids is None:
        return []
    if not isinstance(raw_ids, list):
        raw_ids = [raw_ids]
    ids = []
    for rid in raw_ids:
        try:
            ids.append(int(rid))
        except (TypeError, ValueError):
            continue
    # sıralı + uniq koru
    uniq = []
    seen = set()
    for uid in ids:
        if uid not in seen:
            seen.add(uid)
            uniq.append(uid)
    if not uniq:
        return []
    if _can_assign_multiple(process):
        allowed = _activity_allowed_assignee_ids(process)
        return [uid for uid in uniq if uid in allowed]
    # Süreç üyesi/lider olmayan kullanıcı sadece kendine atayabilir
    return [int(current_user.id)]


def _normalize_reminder_offsets(offsets_raw):
    if not offsets_raw:
        return []
    if not isinstance(offsets_raw, list):
        offsets_raw = [offsets_raw]
    out = []
    seen = set()
    for val in offsets_raw:
        try:
            m = int(val)
        except (TypeError, ValueError):
            continue
        if m < 0 or m > 60 * 24 * 60:
            continue
        if m not in seen:
            seen.add(m)
            out.append(m)
    out.sort(reverse=True)
    return out


def _replace_activity_assignees(activity, assignee_ids):
    ProcessActivityAssignee.query.filter_by(activity_id=activity.id).delete(synchronize_session=False)
    for idx, uid in enumerate(assignee_ids, start=1):
        db.session.add(ProcessActivityAssignee(
            activity_id=activity.id,
            user_id=uid,
            order_no=idx,
            assigned_by=current_user.id,
            assigned_at=datetime.now(timezone.utc),
        ))


def _replace_activity_reminders(activity, offsets, notify_email):
    ProcessActivityReminder.query.filter_by(activity_id=activity.id).delete(synchronize_session=False)
    if not activity.start_at:
        return
    for off in offsets:
        remind_at = activity.start_at - timedelta(minutes=off)
        db.session.add(ProcessActivityReminder(
            activity_id=activity.id,
            minutes_before=off,
            remind_at=remind_at,
            channel_email=bool(notify_email),
        ))


@process_bp.route('/')
@login_required
def index():
    """Süreç Yönetim Paneli — birebir eski surec_panel mantığı, sayfalama ile."""
    
    # Query with eager loading (cache disabled to avoid DetachedInstanceError)
    all_processes = Process.query.options(
        joinedload(Process.leaders),
        joinedload(Process.members),
        joinedload(Process.owners),
        selectinload(Process.process_sub_strategy_links).joinedload(ProcessSubStrategyLink.sub_strategy),
        selectinload(Process.kpis)  # Eager load KPIs to prevent lazy load errors
    ).filter_by(
        tenant_id=current_user.tenant_id,
        is_active=True
    ).order_by(Process.code.asc()).all()
    if not _is_privileged_process_role():
        all_processes = [p for p in all_processes if _user_assigned_to_process(p)]

    # Hiyerarşi: root (parent_id=None) ve child'lar
    all_ids = {p.id for p in all_processes}
    roots_all = [p for p in all_processes if p.parent_id is None or p.parent_id not in all_ids]

    # Sayfalama (kök süreçler üzerinden)
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 15, type=int), 50)
    per_page = max(5, per_page)
    total_roots = len(roots_all)
    pagination = type('Pagination', (), {
        'page': page,
        'per_page': per_page,
        'total': total_roots,
        'pages': (total_roots + per_page - 1) // per_page if per_page else 1,
        'has_prev': page > 1,
        'has_next': page * per_page < total_roots,
        'prev_num': page - 1,
        'next_num': page + 1,
    })()
    start = (page - 1) * per_page
    roots = roots_all[start:start + per_page]

    children_map = {}  # {parent_id: [child, ...]}
    for p in all_processes:
        if p.parent_id and p.parent_id in all_ids:
            children_map.setdefault(p.parent_id, []).append(p)

    # Orphan listesi (parent'ı bu listede bulunmayan çocuklar)
    orphans = [p for p in all_processes if p.parent_id is not None and p.parent_id not in all_ids]

    # Kullanıcılar (aynı tenant'taki) — lider/üye atama için
    users = User.query.filter_by(tenant_id=current_user.tenant_id, is_active=True).all()

    # Üst süreç seçenekleri (with depth) — yeni süreç eklerken parent picker için
    def _collect_with_depth(node_list, depth=0):
        result = []
        for p in node_list:
            result.append((p, depth))
            kids = children_map.get(p.id, [])
            result.extend(_collect_with_depth(kids, depth + 1))
        return result

    parent_options_with_depth = _collect_with_depth(roots_all)

    strategies = Strategy.query.filter_by(
        tenant_id=current_user.tenant_id,
        is_active=True
    ).order_by(Strategy.code).all()
    strategies_json = [
        {
            'id': s.id,
            'code': s.code or '',
            'title': s.title or '',
            'sub_strategies': [{'id': ss.id, 'code': ss.code or '', 'title': ss.title or ''} for ss in (s.sub_strategies or []) if ss.is_active]
        }
        for s in strategies
    ]

    return render_template(
        'process/panel.html',
        processes=all_processes,
        processes_roots=roots,
        processes_roots_all=roots_all,
        processes_children=children_map,
        processes_orphans=orphans,
        users=users,
        strategies=strategies,
        strategies_json=strategies_json,
        parent_options_with_depth=parent_options_with_depth,
        pagination=pagination,
    )


@process_bp.route('/<int:process_id>/karne')
@login_required
def karne(process_id):
    """Süreç Karnesi — birebir eski surec_karnesi mantığı."""
    process = Process.query.filter_by(
        id=process_id,
        tenant_id=current_user.tenant_id,
        is_active=True
    ).first_or_404()
    if not _can_access_process(process):
        abort(403)

    # Tüm süreçler — karne içindeki süreç seçici için
    all_processes = Process.query.filter_by(
        tenant_id=current_user.tenant_id,
        is_active=True
    ).order_by(Process.code.asc()).all()
    if not _is_privileged_process_role():
        all_processes = [p for p in all_processes if _user_assigned_to_process(p)]

    # Performans Göstergeleri
    kpis = ProcessKpi.query.filter_by(process_id=process.id, is_active=True).all()

    favorite_kpi_ids = {f.process_kpi_id for f in FavoriteKpi.query.filter_by(user_id=current_user.id).all()}

    # Faaliyetler
    activities = ProcessActivity.query.filter_by(process_id=process.id, is_active=True).all()

    # Alt stratejiler (PG'ye bağlamak için)
    sub_strategies = (
        SubStrategy.query
        .join(SubStrategy.strategy)
        .filter(Strategy.tenant_id == current_user.tenant_id)
        .all()
    )

    current_year = datetime.now().year
    process_users = []
    seen_ids = set()
    for u in (process.leaders + process.members + process.owners):
        if not u or not u.is_active or u.id in seen_ids:
            continue
        seen_ids.add(u.id)
        process_users.append(u)

    return render_template(
        'process/karne.html',
        process=process,
        all_processes=all_processes,
        kpis=kpis,
        activities=activities,
        process_users=process_users,
        sub_strategies=sub_strategies,
        current_year=current_year,
        favorite_kpi_ids=favorite_kpi_ids
    )


# ──────────────────────────────────────────────────
# API — Process CRUD
# ──────────────────────────────────────────────────

@process_bp.route('/api/add', methods=['POST'])
@login_required
def add_process():
    data = request.get_json()
    if not _is_privileged_process_role():
        return jsonify({'success': False, 'message': 'Süreç oluşturma yetkiniz yok.'}), 403
    try:
        parent_id_raw = data.get('parent_id') or None
        parent_id = validate_process_parent_id(None, parent_id_raw, current_user.tenant_id)
        new_process = Process(
            tenant_id=current_user.tenant_id,
            parent_id=parent_id,
            code=data.get('code'),
            name=data.get('name'),
            english_name=data.get('english_name'),
            description=data.get('description'),
            document_no=data.get('document_no'),
            revision_no=data.get('revision_no'),
            status=data.get('status', 'Aktif'),
            progress=int(data.get('progress', 0) or 0),
            start_boundary=data.get('start_boundary'),
            end_boundary=data.get('end_boundary'),
        )
        for field, attr in [
            ('revision_date', 'revision_date'),
            ('first_publish_date', 'first_publish_date'),
            ('start_date', 'start_date'),
            ('end_date', 'end_date'),
        ]:
            if data.get(field):
                setattr(new_process, attr, datetime.strptime(data[field], '%Y-%m-%d').date())

        db.session.add(new_process)
        db.session.flush()  # get id for M2M

        # Leaders / Members
        for uid in (data.get('leader_ids') or []):
            u = User.query.filter_by(id=int(uid), tenant_id=current_user.tenant_id, is_active=True).first()
            if u:
                new_process.leaders.append(u)
        for uid in (data.get('member_ids') or []):
            u = User.query.filter_by(id=int(uid), tenant_id=current_user.tenant_id, is_active=True).first()
            if u:
                new_process.members.append(u)

        db.session.commit()
        return jsonify({'success': True, 'message': 'Süreç başarıyla oluşturuldu.', 'id': new_process.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/get/<int:process_id>', methods=['GET'])
@login_required
def get_process(process_id):
    """Süreç bilgilerini düzenle modalı için döner."""
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    if not _can_edit_process_record(p):
        return jsonify({'success': False, 'message': 'Bu süreci düzenleme yetkiniz yok.'}), 403
    sub_strategy_links = [
        {'sub_strategy_id': link.sub_strategy_id, 'contribution_pct': link.contribution_pct}
        for link in p.process_sub_strategy_links
    ]
    return jsonify({
        'success': True,
        'process': {
            'id': p.id,
            'code': p.code,
            'name': p.name,
            'english_name': p.english_name,
            'description': p.description,
            'document_no': p.document_no,
            'revision_no': p.revision_no,
            'revision_date': str(p.revision_date) if p.revision_date else '',
            'first_publish_date': str(p.first_publish_date) if p.first_publish_date else '',
            'start_date': str(p.start_date) if p.start_date else '',
            'end_date': str(p.end_date) if p.end_date else '',
            'status': p.status,
            'progress': p.progress,
            'parent_id': p.parent_id,
            'start_boundary': p.start_boundary,
            'end_boundary': p.end_boundary,
            'leader_ids': [u.id for u in p.leaders],
            'member_ids': [u.id for u in p.members],
            'sub_strategy_links': sub_strategy_links,
        }
    })


@process_bp.route('/api/update/<int:process_id>', methods=['POST'])
@login_required
def update_process(process_id):
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    if not _can_edit_process_record(p):
        return jsonify({'success': False, 'message': 'Bu süreci güncelleme yetkiniz yok.'}), 403
    data = request.get_json()
    try:
        p.code = data.get('code', p.code)
        p.name = data.get('name', p.name)
        p.english_name = data.get('english_name', p.english_name)
        p.description = data.get('description', p.description)
        p.document_no = data.get('document_no', p.document_no)
        p.revision_no = data.get('revision_no', p.revision_no)
        p.status = data.get('status', p.status)
        p.progress = int(data.get('progress', p.progress) or 0)
        p.start_boundary = data.get('start_boundary', p.start_boundary)
        p.end_boundary = data.get('end_boundary', p.end_boundary)

        parent_id_raw = data.get('parent_id')
        p.parent_id = validate_process_parent_id(p.id, parent_id_raw, current_user.tenant_id)

        for field, attr in [
            ('revision_date', 'revision_date'),
            ('first_publish_date', 'first_publish_date'),
            ('start_date', 'start_date'),
            ('end_date', 'end_date'),
        ]:
            if data.get(field):
                setattr(p, attr, datetime.strptime(data[field], '%Y-%m-%d').date())

        # Reset and reassign M2M
        if 'leader_ids' in data:
            p.leaders = [
                u for u in (
                    User.query.filter_by(id=int(i), tenant_id=current_user.tenant_id, is_active=True).first()
                    for i in data['leader_ids']
                ) if u
            ]
        if 'member_ids' in data:
            p.members = [
                u for u in (
                    User.query.filter_by(id=int(i), tenant_id=current_user.tenant_id, is_active=True).first()
                    for i in data['member_ids']
                ) if u
            ]

        # Alt Strateji bağlantıları (katkı yüzdesi ile)
        if 'sub_strategy_links' in data:
            for link in list(p.process_sub_strategy_links):
                db.session.delete(link)
            from app.utils.process_utils import validate_same_tenant_sub_strategies
            for item in data['sub_strategy_links']:
                ss_id = item.get('sub_strategy_id')
                if not ss_id:
                    continue
                sub_strategies = validate_same_tenant_sub_strategies(current_user.tenant_id, [int(ss_id)])
                if sub_strategies:
                    pct = item.get('contribution_pct')
                    pct_float = float(pct) if pct is not None and str(pct).strip() != '' else None
                    link = ProcessSubStrategyLink(
                        process_id=p.id,
                        sub_strategy_id=int(ss_id),
                        contribution_pct=pct_float
                    )
                    db.session.add(link)

        db.session.commit()
        return jsonify({'success': True, 'message': 'Süreç güncellendi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/delete/<int:process_id>', methods=['POST'])
@login_required
def delete_process(process_id):
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    if not _is_privileged_process_role():
        return jsonify({'success': False, 'message': 'Süreç silme yetkiniz yok.'}), 403
    try:
        p.is_active = False
        p.deleted_at = datetime.now(timezone.utc)
        p.deleted_by = current_user.id
        db.session.commit()
        return jsonify({'success': True, 'message': 'Süreç silindi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


# ──────────────────────────────────────────────────
# API — KPI (Performans Göstergesi) CRUD
# ──────────────────────────────────────────────────

@process_bp.route('/api/kpi/add', methods=['POST'])
@login_required
def add_kpi():
    data = request.get_json()
    process_id = data.get('process_id') or data.get('surec_id')
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    if not _can_crud_pg_and_activity(p):
        return jsonify({'success': False, 'message': 'PG ekleme yetkiniz yok.'}), 403
    try:
        new_kpi = ProcessKpi(
            process_id=p.id,
            name=data.get('name'),
            code=data.get('code'),
            description=data.get('description'),
            target_value=data.get('target_value') or data.get('hedef_deger'),
            unit=data.get('unit') or data.get('olcum_birimi'),
            period=data.get('period') or data.get('periyot', 'Aylık'),
            direction=data.get('direction', 'Increasing'),
            weight=float(data.get('weight') or data.get('agirlik') or 0),
            data_collection_method=data.get('calculation_method') or data.get('data_collection_method', 'Ortalama'),
            is_important=bool(data.get('onemli', False)),
            # Eski proje uyumluluğu
            gosterge_turu=data.get('gosterge_turu') or None,
            target_method=data.get('target_method') or None,
            basari_puani_araliklari=data.get('basari_puani_araliklari') or None,
            onceki_yil_ortalamasi=float(data['onceki_yil_ortalamasi']) if data.get('onceki_yil_ortalamasi') else None,
        )
        if data.get('sub_strategy_id'):
            new_kpi.sub_strategy_id = int(data['sub_strategy_id'])
        if data.get('start_date'):
            new_kpi.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        if data.get('end_date'):
            new_kpi.end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()

        db.session.add(new_kpi)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Performans göstergesi eklendi.', 'id': new_kpi.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/kpi/get/<int:kpi_id>', methods=['GET'])
@login_required
def get_kpi(kpi_id):
    """Tek bir KPI'nin detaylı bilgisini döner (edit modal için)."""
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_access_process(kpi.process):
        return jsonify({'success': False, 'message': 'Bu PG için erişim yetkiniz yok.'}), 403
    return jsonify({
        'success': True,
        'kpi': {
            'id': kpi.id,
            'name': kpi.name,
            'code': kpi.code,
            'description': kpi.description,
            'target_value': kpi.target_value,
            'unit': kpi.unit,
            'period': kpi.period,
            'direction': kpi.direction,
            'weight': kpi.weight,
            'data_collection_method': kpi.data_collection_method,
            'gosterge_turu': kpi.gosterge_turu,
            'target_method': kpi.target_method,
            'basari_puani_araliklari': kpi.basari_puani_araliklari,
            'onceki_yil_ortalamasi': kpi.onceki_yil_ortalamasi,
            'sub_strategy_id': kpi.sub_strategy_id,
        }
    })


@process_bp.route('/api/kpi/update/<int:kpi_id>', methods=['POST'])
@login_required
def update_kpi(kpi_id):
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_crud_pg_and_activity(kpi.process):
        return jsonify({'success': False, 'message': 'PG güncelleme yetkiniz yok.'}), 403
    data = request.get_json()
    try:
        kpi.name = data.get('name', kpi.name)
        kpi.code = data.get('code', kpi.code)
        kpi.description = data.get('description', kpi.description)
        kpi.target_value = data.get('target_value', kpi.target_value)
        kpi.unit = data.get('unit', kpi.unit)
        kpi.period = data.get('period', kpi.period)
        kpi.direction = data.get('direction', kpi.direction)
        kpi.weight = float(data.get('weight') or kpi.weight or 0)
        kpi.data_collection_method = data.get('calculation_method') or data.get('data_collection_method', kpi.data_collection_method)
        # Eski proje uyumluluğu
        kpi.gosterge_turu = data.get('gosterge_turu', kpi.gosterge_turu)
        kpi.target_method = data.get('target_method', kpi.target_method)
        kpi.basari_puani_araliklari = data.get('basari_puani_araliklari', kpi.basari_puani_araliklari)
        if data.get('onceki_yil_ortalamasi') is not None:
            kpi.onceki_yil_ortalamasi = float(data['onceki_yil_ortalamasi']) if data['onceki_yil_ortalamasi'] != '' else None
        if data.get('sub_strategy_id'):
            kpi.sub_strategy_id = int(data['sub_strategy_id'])
        elif 'sub_strategy_id' in data and not data['sub_strategy_id']:
            kpi.sub_strategy_id = None
        db.session.commit()
        return jsonify({'success': True, 'message': 'PG güncellendi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/kpi/delete/<int:kpi_id>', methods=['POST'])
@login_required
def delete_kpi(kpi_id):
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_crud_pg_and_activity(kpi.process):
        return jsonify({'success': False, 'message': 'PG silme yetkiniz yok.'}), 403
    try:
        kpi.is_active = False
        db.session.commit()
        return jsonify({'success': True, 'message': 'PG silindi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/kpi/list/<int:process_id>', methods=['GET'])
@login_required
def list_kpis(process_id):
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    if not _can_access_process(p):
        return jsonify({'success': False, 'message': 'Bu süreç için erişim yetkiniz yok.'}), 403
    kpis = ProcessKpi.query.filter_by(process_id=p.id, is_active=True).all()
    result = []
    for k in kpis:
        result.append({
            'id': k.id,
            'name': k.name,
            'code': k.code,
            'description': k.description,
            'target_value': k.target_value,
            'unit': k.unit,
            'period': k.period,
            'direction': k.direction,
            'weight': k.weight,
            'sub_strategy_id': k.sub_strategy_id,
            'sub_strategy_title': k.sub_strategy.title if k.sub_strategy else None,
        })
    return jsonify({'success': True, 'kpis': result})


# ──────────────────────────────────────────────────
# API — Activity CRUD
# ──────────────────────────────────────────────────

@process_bp.route('/api/activity/add', methods=['POST'])
@login_required
def add_activity():
    data = request.get_json() or {}
    process_id = data.get('process_id') or data.get('surec_id')
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    if not _can_access_process(p):
        return jsonify({'success': False, 'message': 'Faaliyet ekleme yetkiniz yok.'}), 403
    try:
        start_at = _parse_local_datetime(data.get('start_at')) or _parse_local_datetime(data.get('start_date'))
        end_at = _parse_local_datetime(data.get('end_at')) or _parse_local_datetime(data.get('end_date'))
        if not start_at and data.get('start_date'):
            start_at = datetime.strptime(data['start_date'], '%Y-%m-%d')
        if not end_at and data.get('end_date'):
            end_at = datetime.strptime(data['end_date'], '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        if not start_at or not end_at:
            return jsonify({'success': False, 'message': 'Başlangıç ve bitiş zamanı zorunludur.'}), 400
        if start_at and end_at and end_at <= start_at:
            return jsonify({'success': False, 'message': 'Bitiş zamanı başlangıçtan sonra olmalıdır.'}), 400

        assignee_ids = _normalize_assignee_ids(
            data.get('assignee_ids') if data.get('assignee_ids') is not None else data.get('assigned_user_ids'),
            p,
        )
        if not assignee_ids:
            assignee_ids = [int(current_user.id)]
        reminder_offsets = _normalize_reminder_offsets(data.get('reminder_offsets'))

        new_act = ProcessActivity(
            process_id=p.id,
            process_kpi_id=data.get('process_kpi_id') or None,
            name=data.get('name'),
            description=data.get('description'),
            status=data.get('status', 'Planlandı'),
            progress=int(data.get('progress', 0) or 0),
            notify_email=bool(data.get('notify_email', False)),
            auto_complete_enabled=bool(data.get('auto_complete_enabled', True)),
            start_at=start_at,
            end_at=end_at,
        )
        if start_at:
            new_act.start_date = start_at.date()
        if end_at:
            new_act.end_date = end_at.date()
        db.session.add(new_act)
        db.session.flush()
        _replace_activity_assignees(new_act, assignee_ids)
        _replace_activity_reminders(new_act, reminder_offsets, new_act.notify_email)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Faaliyet eklendi.', 'id': new_act.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/activity/delete/<int:act_id>', methods=['POST'])
@login_required
def delete_activity(act_id):
    act = ProcessActivity.query.join(Process).filter(
        ProcessActivity.id == act_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_manage_activity(act.process, act):
        return jsonify({'success': False, 'message': 'Faaliyet silme yetkiniz yok.'}), 403
    try:
        act.is_active = False
        act.status = 'İptal'
        act.cancelled_at = datetime.now(timezone.utc)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Faaliyet silindi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/activity/cancel/<int:act_id>', methods=['POST'])
@login_required
def cancel_activity(act_id):
    act = ProcessActivity.query.join(Process).filter(
        ProcessActivity.id == act_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_manage_activity(act.process, act):
        return jsonify({'success': False, 'message': 'Faaliyet iptal yetkiniz yok.'}), 403
    try:
        act.status = 'İptal'
        act.cancelled_at = datetime.now(timezone.utc)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Faaliyet iptal edildi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/activity/postpone/<int:act_id>', methods=['POST'])
@login_required
def postpone_activity(act_id):
    act = ProcessActivity.query.join(Process).filter(
        ProcessActivity.id == act_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_manage_activity(act.process, act):
        return jsonify({'success': False, 'message': 'Faaliyet erteleme yetkiniz yok.'}), 403
    data = request.get_json() or {}
    new_start = _parse_local_datetime(data.get('start_at'))
    new_end = _parse_local_datetime(data.get('end_at'))
    if not new_start or not new_end:
        return jsonify({'success': False, 'message': 'Yeni başlangıç ve bitiş zamanı zorunludur.'}), 400
    if new_end <= new_start:
        return jsonify({'success': False, 'message': 'Bitiş zamanı başlangıçtan sonra olmalıdır.'}), 400
    try:
        act.start_at = new_start
        act.end_at = new_end
        act.start_date = new_start.date()
        act.end_date = new_end.date()
        act.status = 'Ertelendi'
        act.postponed_at = datetime.now(timezone.utc)
        offsets = [int(r.minutes_before) for r in act.reminders]
        _replace_activity_reminders(act, offsets, act.notify_email)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Faaliyet ertelendi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/activity/complete/<int:act_id>', methods=['POST'])
@login_required
def complete_activity(act_id):
    act = ProcessActivity.query.join(Process).filter(
        ProcessActivity.id == act_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_manage_activity(act.process, act):
        return jsonify({'success': False, 'message': 'Faaliyet tamamlama yetkiniz yok.'}), 403
    try:
        act.status = 'Tamamlandı'
        act.progress = 100
        act.completed_at = datetime.now(timezone.utc)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Faaliyet tamamlandı.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


# ──────────────────────────────────────────────────
# API — KPI Data Entry (Karne veri girişi)
# ──────────────────────────────────────────────────

@process_bp.route('/api/kpi-data/add', methods=['POST'])
@login_required
def add_kpi_data():
    data = request.get_json()
    kpi_id = data.get('kpi_id') or data.get('pg_id')
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_enter_pgv(kpi.process):
        return jsonify({'success': False, 'message': 'PG verisi girme yetkiniz yok.'}), 403
    try:
        year_val = int(data.get('year', datetime.now().year))
        pt = (data.get('period_type') or 'yillik').lower().strip()
        pn = int(data.get('period_no') or 1)
        pm = data.get('period_month')
        period_month = int(pm) if pm is not None and str(pm).strip() else None
        data_date_val = None
        if data.get('data_date'):
            data_date_val = datetime.strptime(data['data_date'], '%Y-%m-%d').date()
        last_day = last_day_of_period(year_val, pt, pn, period_month)
        if last_day is not None:
            data_date_val = last_day
        if data_date_val is None:
            data_date_val = date.today()
        entry = None
        for attempt in (1, 2):
            try:
                entry = KpiData(
                    process_kpi_id=kpi.id,
                    year=year_val,
                    data_date=data_date_val,
                    period_type=pt,
                    period_no=pn,
                    period_month=period_month,
                    target_value=data.get('target_value'),
                    actual_value=data.get('actual_value') or data.get('gerceklesen_deger', ''),
                    description=data.get('description') or data.get('aciklama'),
                    user_id=current_user.id,
                )
                db.session.add(entry)
                db.session.flush()
                db.session.add(KpiDataAudit(
                    kpi_data_id=entry.id,
                    action_type='CREATE',
                    new_value=entry.actual_value,
                    action_detail='Veri Giriş Sihirbazı ile eklendi',
                    user_id=current_user.id,
                ))
                db.session.commit()
                break
            except Exception as e:
                db.session.rollback()
                if attempt == 1 and (
                    is_pk_duplicate(e, 'kpi_data')
                    or is_pk_duplicate(e, 'kpi_data_audits')
                ):
                    sync_kpi_data_related_sequences()
                    db.session.commit()
                    continue
                raise
        try:
            from app.services.score_engine_service import recalc_on_pg_data_change
            recalc_on_pg_data_change(current_user.tenant_id, int(data.get('year', datetime.now().year)))
        except Exception as e:
            current_app.logger.warning(f'[kpi_data_add] score_engine hatası: {e}')
        try:
            from app.services.process_deviation_service import check_pg_performance_deviation
            check_pg_performance_deviation(entry.id)
        except Exception as e:
            current_app.logger.warning(f'[kpi_data_add] deviation_service hatası: {e}')
        return jsonify({'success': True, 'message': 'Veri kaydedildi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/kpi-data/list/<int:kpi_id>', methods=['GET'])
@login_required
def list_kpi_data(kpi_id):
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_access_process(kpi.process):
        return jsonify({'success': False, 'message': 'Bu PG verilerine erişim yetkiniz yok.'}), 403
    year = request.args.get('year', datetime.now().year, type=int)
    q = KpiData.query.filter_by(process_kpi_id=kpi.id, year=year, is_active=True).order_by(KpiData.data_date)
    if not _can_crud_pg_and_activity(kpi.process):
        q = q.filter_by(user_id=current_user.id)
    entries = q.all()
    return jsonify({
        'success': True,
        'data': [{
            'id': e.id,
            'year': e.year,
            'data_date': str(e.data_date),
            'period_type': e.period_type,
            'period_no': e.period_no,
            'actual_value': e.actual_value,
            'target_value': e.target_value,
            'description': e.description,
        } for e in entries]
    })


@process_bp.route('/api/karne/<int:process_id>', methods=['GET'])
@login_required
def karne_data(process_id):
    """Karne sayfasının AJAX ile çektiği tüm veriyi döner."""
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id, is_active=True).first_or_404()
    if not _can_access_process(p):
        return jsonify({'success': False, 'message': 'Bu süreç için erişim yetkiniz yok.'}), 403
    year = request.args.get('year', datetime.now().year, type=int)

    kpis = ProcessKpi.query.filter_by(process_id=p.id, is_active=True).all()
    activities = ProcessActivity.query.filter_by(process_id=p.id, is_active=True).all()
    favorite_kpi_ids = {f.process_kpi_id for f in FavoriteKpi.query.filter_by(user_id=current_user.id).all()}

    def _parse_float(v):
        if v is None or v == '':
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def _aggregate_entries(entries_with_keys, method):
        by_key = {}
        for key, val_str, dt in entries_with_keys:
            if key not in by_key:
                by_key[key] = []
            by_key[key].append((val_str, _parse_float(val_str), dt))
        result = {}
        for key, items in by_key.items():
            method_lower = (method or '').lower()
            numeric = [(v, d) for _, v, d in items if v is not None]
            if method_lower in ('toplama', 'toplam') and numeric:
                result[key] = str(sum(v for v, _ in numeric))
            elif method_lower in ('ortalama',) and numeric:
                result[key] = str(round(sum(v for v, _ in numeric) / len(numeric), 2))
            else:
                last = max(items, key=lambda x: x[2] or date(1900, 1, 1))
                result[key] = last[0]
        return result

    kpi_list = []
    for k in kpis:
        entries = KpiData.query.filter_by(process_kpi_id=k.id, year=year, is_active=True).order_by(KpiData.data_date).all()
        entries_with_keys = []
        for e in entries:
            keys = data_date_to_period_keys(e.data_date, year)
            for key in keys:
                entries_with_keys.append((key, e.actual_value, e.data_date))
        entries_by_period = _aggregate_entries(entries_with_keys, k.data_collection_method)

        strategy_title = ''
        if k.sub_strategy and k.sub_strategy.strategy:
            strategy_title = k.sub_strategy.strategy.title
        kpi_list.append({
            'id': k.id,
            'name': k.name,
            'code': k.code,
            'target_value': k.target_value,
            'unit': k.unit,
            'period': k.period,
            'data_collection_method': k.data_collection_method or 'Ortalama',
            'direction': k.direction or 'Increasing',
            'weight': k.weight,
            'sub_strategy_id': k.sub_strategy_id,
            'strategy_title': strategy_title or '-',
            'sub_strategy_title': k.sub_strategy.title if k.sub_strategy else '-',
            'basari_puani_araliklari': k.basari_puani_araliklari,
            'entries': entries_by_period,
        })

    act_list = []
    for a in activities:
        # Faaliyet ay takibi
        tracks = ActivityTrack.query.filter_by(activity_id=a.id, year=year).all()
        tracks_map = {t.month: t.completed for t in tracks}
        start_at, end_at = a.start_at, a.end_at
        if not start_at and a.start_date:
            start_at = datetime.combine(a.start_date, datetime.min.time())
        if not end_at and a.end_date:
            end_at = datetime.combine(a.end_date, datetime.max.time().replace(microsecond=0))
        assignee_links = sorted(a.assignment_links, key=lambda x: x.order_no or 0)
        assignees = []
        for link in assignee_links:
            u = link.user
            if not u:
                continue
            full_name = f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip() or (u.email or '')
            assignees.append({
                'id': int(u.id),
                'full_name': full_name,
                'email': u.email,
                'order_no': link.order_no,
            })

        can_manage = _can_manage_activity(p, a)
        act_list.append({
            'id': a.id,
            'name': a.name,
            'description': a.description,
            'status': a.status,
            'progress': a.progress,
            'start_date': str(a.start_date) if a.start_date else None,
            'end_date': str(a.end_date) if a.end_date else None,
            'start_at': start_at.isoformat(timespec='minutes') if start_at else None,
            'end_at': end_at.isoformat(timespec='minutes') if end_at else None,
            'notify_email': bool(a.notify_email),
            'process_kpi_id': a.process_kpi_id,
            'process_kpi_name': a.process_kpi.name if a.process_kpi else None,
            'assignee_ids': [x['id'] for x in assignees],
            'assignees': assignees,
            'first_assignee_id': a.first_assignee_id,
            'reminder_offsets': [int(r.minutes_before) for r in sorted(a.reminders, key=lambda x: x.minutes_before, reverse=True)],
            'monthly_tracks': tracks_map,  # {1: True, 3: True, ...}
            'can_manage': bool(can_manage),
        })

    return jsonify({
        'success': True,
        'process': {
            'id': p.id,
            'name': p.name,
            'document_no': p.document_no,
            'revision_no': p.revision_no,
        },
        'year': year,
        'kpis': kpi_list,
        'activities': act_list,
        'favorite_kpi_ids': list(favorite_kpi_ids),
    })


# ──────────────────────────────────────────────────
# API — Activity Update & Monthly Tracking
# ──────────────────────────────────────────────────

@process_bp.route('/api/activity/get/<int:act_id>', methods=['GET'])
@login_required
def get_activity(act_id):
    act = ProcessActivity.query.join(Process).filter(
        ProcessActivity.id == act_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_access_process(act.process):
        return jsonify({'success': False, 'message': 'Bu faaliyet için erişim yetkiniz yok.'}), 403
    return jsonify({
        'success': True,
        'activity': {
            'id': act.id,
            'name': act.name,
            'description': act.description,
            'status': act.status,
            'progress': act.progress,
            'start_date': str(act.start_date) if act.start_date else '',
            'end_date': str(act.end_date) if act.end_date else '',
            'start_at': act.start_at.strftime('%Y-%m-%dT%H:%M') if act.start_at else '',
            'end_at': act.end_at.strftime('%Y-%m-%dT%H:%M') if act.end_at else '',
            'notify_email': bool(act.notify_email),
            'assignee_ids': [int(l.user_id) for l in sorted(act.assignment_links, key=lambda x: x.order_no or 0)],
            'reminder_offsets': [int(r.minutes_before) for r in sorted(act.reminders, key=lambda x: x.minutes_before, reverse=True)],
        }
    })


@process_bp.route('/api/activity/update/<int:act_id>', methods=['POST'])
@login_required
def update_activity(act_id):
    act = ProcessActivity.query.join(Process).filter(
        ProcessActivity.id == act_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_manage_activity(act.process, act):
        return jsonify({'success': False, 'message': 'Faaliyet güncelleme yetkiniz yok.'}), 403
    data = request.get_json() or {}
    try:
        act.name = data.get('name', act.name)
        act.description = data.get('description', act.description)
        act.status = data.get('status', act.status)
        act.progress = int(data.get('progress', act.progress) or 0)
        if 'notify_email' in data:
            act.notify_email = bool(data.get('notify_email'))
        if 'process_kpi_id' in data:
            act.process_kpi_id = data.get('process_kpi_id') or None

        parsed_start = _parse_local_datetime(data.get('start_at')) or _parse_local_datetime(data.get('start_date'))
        parsed_end = _parse_local_datetime(data.get('end_at')) or _parse_local_datetime(data.get('end_date'))
        if data.get('start_date') and not parsed_start:
            parsed_start = datetime.strptime(data['start_date'], '%Y-%m-%d')
        if data.get('end_date') and not parsed_end:
            parsed_end = datetime.strptime(data['end_date'], '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        if parsed_start:
            act.start_at = parsed_start
            act.start_date = parsed_start.date()
        if parsed_end:
            act.end_at = parsed_end
            act.end_date = parsed_end.date()
        if act.start_at and act.end_at and act.end_at <= act.start_at:
            return jsonify({'success': False, 'message': 'Bitiş zamanı başlangıçtan sonra olmalıdır.'}), 400

        if data.get('assignee_ids') is not None or data.get('assigned_user_ids') is not None:
            assignee_ids = _normalize_assignee_ids(
                data.get('assignee_ids') if data.get('assignee_ids') is not None else data.get('assigned_user_ids'),
                act.process,
            )
            if not assignee_ids:
                assignee_ids = [int(current_user.id)]
            _replace_activity_assignees(act, assignee_ids)

        if 'reminder_offsets' in data:
            reminder_offsets = _normalize_reminder_offsets(data.get('reminder_offsets'))
            _replace_activity_reminders(act, reminder_offsets, act.notify_email)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Faaliyet güncellendi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/activity/track', methods=['POST'])
@login_required
def toggle_activity_track():
    """Faaliyet ayı checkbox toggle (completed/not completed)."""
    data = request.get_json() or {}
    act_id = data.get('activity_id')
    year = int(data.get('year', datetime.now().year))
    month = int(data.get('month', 1))
    completed = bool(data.get('completed', False))

    act = ProcessActivity.query.join(Process).filter(
        ProcessActivity.id == act_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_manage_activity(act.process, act):
        return jsonify({'success': False, 'message': 'Faaliyet takibi için yetkiniz yok.'}), 403

    try:
        track = ActivityTrack.query.filter_by(
            activity_id=act.id,
            year=year,
            month=month
        ).first()

        if track:
            track.completed = completed
            track.user_id = current_user.id
        else:
            track = ActivityTrack(
                activity_id=act.id,
                year=year,
                month=month,
                completed=completed,
                user_id=current_user.id
            )
            db.session.add(track)

        db.session.commit()
        return jsonify({'success': True, 'completed': track.completed})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


# ──────────────────────────────────────────────────
# API — Kullanıcı Süreçleri, Performans Kartım, Excel Export
# ──────────────────────────────────────────────────

@process_bp.route('/api/kullanici-surecleri', methods=['GET'])
@login_required
def kullanici_surecleri():
    """Kullanıcının üye olduğu veya lider olduğu süreçleri döner (karne süreç seçici için)."""
    processes = (
        Process.query.filter(
            Process.tenant_id == current_user.tenant_id,
            Process.is_active == True
        )
        .filter(
            db.or_(
                Process.leaders.any(User.id == current_user.id),
                Process.members.any(User.id == current_user.id),
                Process.owners.any(User.id == current_user.id)
            )
        )
        .order_by(Process.code.asc())
        .all()
    )
    all_processes = Process.query.filter_by(
        tenant_id=current_user.tenant_id,
        is_active=True
    ).order_by(Process.code.asc()).all()
    items = [{
        'id': p.id,
        'code': p.code or '',
        'name': p.name,
        'document_no': p.document_no or '-',
        'revision_no': p.revision_no or '-',
    } for p in processes]
    all_items = [{
        'id': p.id,
        'code': p.code or '',
        'name': p.name,
    } for p in all_processes]
    return jsonify({'success': True, 'surecler': items, 'tum_surecler': all_items})


@process_bp.route('/api/surec/<int:process_id>/uyeler', methods=['GET'])
@login_required
def surec_uyeleri(process_id):
    """Süreç üyelerini (lider, üye, sahip) döner."""
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    uyeler = set()
    for u in p.leaders + p.members + p.owners:
        if u and u.is_active:
            uyeler.add(u)
    result = [{
        'id': u.id,
        'email': u.email,
        'first_name': u.first_name or '',
        'last_name': u.last_name or '',
        'full_name': f"{(u.first_name or '')} {(u.last_name or '')}".strip() or u.email
    } for u in uyeler]
    return jsonify({'success': True, 'uyeler': result})


@process_bp.route('/api/surec/<int:process_id>/saglik-skoru', methods=['GET'])
@login_required
def surec_saglik_skoru(process_id):
    """Süreç sağlık skoru hesaplar (process_health_service ile gelişmiş)."""
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id, is_active=True).first_or_404()
    year = request.args.get('year', datetime.now().year, type=int)
    from app.services.process_health_service import calculate_process_health_score
    result = calculate_process_health_score(p.id, year)
    if not result:
        return jsonify({'success': True, 'skor': 0, 'durum': 'Veri yok', 'detay': {}})
    return jsonify({
        'success': True,
        'skor': result['skor'],
        'durum': result['durum'],
        'detay': result.get('detay', {})
    })


@process_bp.route('/api/surec/<int:process_id>/muda-analiz', methods=['GET'])
@login_required
def surec_muda_analiz(process_id):
    """Süreç Muda (verimsizlik) analizi — 7 Muda tipi."""
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id, is_active=True).first_or_404()
    from app.services.muda_analyzer import MudaAnalyzerService
    findings = MudaAnalyzerService.analyze_process_inefficiency(p.id, current_user.tenant_id)
    return jsonify({
        'success': True,
        'process_id': p.id,
        'findings': findings,
        'muda_types': MudaAnalyzerService.MUDA_TYPES
    })


@process_bp.route('/api/muda-verimlilik-skoru', methods=['GET'])
@login_required
def muda_verimlilik_skoru():
    """Kiracı genel Muda verimlilik skoru (0-100)."""
    from app.services.muda_analyzer import MudaAnalyzerService
    score = MudaAnalyzerService.get_efficiency_score(current_user.tenant_id)
    return jsonify({'success': True, 'skor': score})


@process_bp.route('/api/surec/<int:process_id>/kpi/<int:kpi_id>/dagit', methods=['POST'])
@login_required
def pg_dagit(process_id, kpi_id):
    """PG'yi seçilen kullanıcılara bireysel hedef olarak dağıtır."""
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    kpi = ProcessKpi.query.filter_by(id=kpi_id, process_id=p.id, is_active=True).first_or_404()
    data = request.get_json() or {}
    user_ids = data.get('user_ids', [])
    if not user_ids:
        return jsonify({'success': False, 'message': 'En az bir kullanıcı seçin.'}), 400
    try:
        users = User.query.filter(
            User.id.in_([int(uid) for uid in user_ids]),
            User.tenant_id == current_user.tenant_id
        ).all()
        created = []
        for u in users:
            existing = IndividualPerformanceIndicator.query.filter_by(
                user_id=u.id,
                source_process_kpi_id=kpi.id,
                source='Süreç'
            ).first()
            if not existing:
                ipg = IndividualPerformanceIndicator(
                    user_id=u.id,
                    name=kpi.name,
                    description=kpi.description,
                    code=kpi.code,
                    target_value=kpi.target_value,
                    unit=kpi.unit,
                    period=kpi.period,
                    weight=kpi.weight,
                    is_important=kpi.is_important,
                    direction=kpi.direction or 'Increasing',
                    basari_puani_araliklari=kpi.basari_puani_araliklari,
                    source='Süreç',
                    source_process_id=p.id,
                    source_process_kpi_id=kpi.id
                )
                db.session.add(ipg)
                created.append(u.id)
        db.session.commit()
        return jsonify({'success': True, 'message': f'{len(created)} kullanıcıya dağıtıldı.', 'created': created})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/surec/<int:process_id>/faaliyet/<int:act_id>/create-bireysel', methods=['POST'])
@login_required
def faaliyet_create_bireysel(process_id, act_id):
    """Süreç faaliyetinden kullanıcı için bireysel faaliyet oluşturur."""
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id).first_or_404()
    act = ProcessActivity.query.filter_by(id=act_id, process_id=p.id, is_active=True).first_or_404()
    data = request.get_json() or {}
    target_user_id = data.get('user_id', current_user.id)
    if target_user_id != current_user.id and not (current_user.role and current_user.role.name in ('tenant_admin', 'Admin', 'executive_manager')):
        return jsonify({'success': False, 'message': 'Yetkisiz.'}), 403
    target_user = User.query.filter_by(id=target_user_id, tenant_id=current_user.tenant_id).first_or_404()
    try:
        existing = IndividualActivity.query.filter_by(
            user_id=target_user.id,
            source_process_activity_id=act.id,
            source='Süreç'
        ).first()
        if existing:
            return jsonify({'success': True, 'message': 'Zaten mevcut.', 'id': existing.id})
        ia = IndividualActivity(
            user_id=target_user.id,
            name=act.name,
            description=act.description,
            start_date=act.start_date,
            end_date=act.end_date,
            status=act.status or 'Planlandı',
            progress=act.progress or 0,
            source='Süreç',
            source_process_id=p.id,
            source_process_activity_id=act.id
        )
        db.session.add(ia)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Bireysel faaliyet oluşturuldu.', 'id': ia.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/bireysel-faaliyet/<int:act_id>/takip', methods=['POST'])
@login_required
def bireysel_faaliyet_takip(act_id):
    """Bireysel faaliyet aylık takibini günceller."""
    ia = IndividualActivity.query.filter_by(id=act_id, user_id=current_user.id).first_or_404()
    data = request.get_json() or {}
    year = int(data.get('year', datetime.now().year))
    month = int(data.get('month', 1))
    completed = bool(data.get('completed', False))
    try:
        track = IndividualActivityTrack.query.filter_by(
            individual_activity_id=ia.id,
            year=year,
            month=month
        ).first()
        if track:
            track.completed = completed
        else:
            track = IndividualActivityTrack(
                individual_activity_id=ia.id,
                user_id=current_user.id,
                year=year,
                month=month,
                completed=completed
            )
            db.session.add(track)
        db.session.commit()
        return jsonify({'success': True, 'completed': track.completed})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/favorite-kpi/toggle', methods=['POST'])
@login_required
def favorite_kpi_toggle():
    """Favori KPI ekle/kaldır."""
    data = request.get_json() or {}
    kpi_id = data.get('process_kpi_id') or data.get('kpi_id')
    if not kpi_id:
        return jsonify({'success': False, 'message': 'process_kpi_id gerekli.'}), 400
    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == int(kpi_id),
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    existing = FavoriteKpi.query.filter_by(
        user_id=current_user.id,
        process_kpi_id=kpi.id,
        is_active=True
    ).first()
    if existing:
        # Kural 4: Hard delete yasağı — soft delete uygulanıyor
        existing.is_active = False
        db.session.commit()
        return jsonify({'success': True, 'favorite': False, 'message': 'Favorilerden kaldırıldı.'})
    # Daha önce is_active=False olan kayıt varsa yeniden aktif et
    archived = FavoriteKpi.query.filter_by(
        user_id=current_user.id,
        process_kpi_id=kpi.id,
        is_active=False
    ).first()
    if archived:
        archived.is_active = True
        db.session.commit()
    else:
        fav = FavoriteKpi(user_id=current_user.id, process_kpi_id=kpi.id)
        db.session.add(fav)
        db.session.commit()
    return jsonify({'success': True, 'favorite': True, 'message': 'Favorilere eklendi.'})


@process_bp.route('/api/favorite-kpis', methods=['GET'])
@login_required
def favorite_kpis_list():
    """Kullanıcının favori KPI id listesi."""
    favs = FavoriteKpi.query.filter_by(user_id=current_user.id).all()
    ids = [f.process_kpi_id for f in favs]
    return jsonify({'success': True, 'favorite_kpi_ids': ids})


@process_bp.route('/api/performans-kartim', methods=['GET'])
@login_required
def performans_kartim_api():
    """Bireysel panel için PG ve faaliyet verilerini döner. Favori KPI'lar önce sıralanır."""
    year = request.args.get('year', datetime.now().year, type=int)
    favorite_ids = {f.process_kpi_id for f in FavoriteKpi.query.filter_by(user_id=current_user.id).all()}
    pg_list = IndividualPerformanceIndicator.query.filter_by(
        user_id=current_user.id
    ).all()
    act_list = IndividualActivity.query.filter_by(user_id=current_user.id).all()
    pg_items = []
    for ipg in pg_list:
        entries = IndividualKpiData.query.filter_by(
            individual_pg_id=ipg.id,
            year=year
        ).all()
        entries_by_period = {}
        for e in entries:
            pt = (e.period_type or 'yillik').lower().strip()
            pn = int(e.period_no) if e.period_no is not None else 1
            key = f"{pt}_{pn}"
            entries_by_period[key] = e.actual_value
        pg_items.append({
            'id': ipg.id,
            'name': ipg.name,
            'code': ipg.code,
            'target_value': ipg.target_value,
            'unit': ipg.unit,
            'period': ipg.period,
            'source': ipg.source,
            'source_process_name': ipg.source_process.name if ipg.source_process else None,
            'entries': entries_by_period,
            'is_favorite': ipg.source_process_kpi_id in favorite_ids if ipg.source_process_kpi_id else False
        })
    act_items = []
    for ia in act_list:
        tracks = IndividualActivityTrack.query.filter_by(
            individual_activity_id=ia.id,
            year=year
        ).all()
        tracks_map = {t.month: t.completed for t in tracks}
        act_items.append({
            'id': ia.id,
            'name': ia.name,
            'description': ia.description,
            'status': ia.status,
            'progress': ia.progress,
            'source': ia.source,
            'source_process_name': ia.source_process.name if ia.source_process else None,
            'monthly_tracks': tracks_map
        })
    pg_items.sort(key=lambda x: (0 if x.get('is_favorite') else 1, x.get('name', '')))
    return jsonify({
        'success': True,
        'year': year,
        'performans_gostergeleri': pg_items,
        'faaliyetler': act_items
    })


@process_bp.route('/api/export/surec-karnesi/excel', methods=['GET'])
@login_required
def export_surec_karnesi_excel():
    """Süreç karnesini Excel olarak dışa aktarır."""
    process_id = request.args.get('process_id', type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    if not process_id:
        return jsonify({'success': False, 'message': 'process_id gerekli.'}), 400
    p = Process.query.filter_by(id=process_id, tenant_id=current_user.tenant_id, is_active=True).first_or_404()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Performans"
        ws.append(["#", "Performans Göstergesi", "Hedef", "Birim", "Ç1", "Ç2", "Ç3", "Ç4", "Yıllık"])
        for row in ws.iter_rows(min_row=1, max_row=1):
            for c in row:
                c.font = Font(bold=True)
        kpis = ProcessKpi.query.filter_by(process_id=p.id, is_active=True).all()
        for i, k in enumerate(kpis, 1):
            entries = KpiData.query.filter_by(process_kpi_id=k.id, year=year, is_active=True).all()
            by_period = {}
            for e in entries:
                pt = (e.period_type or 'yillik').lower()
                pn = int(e.period_no) if e.period_no else 1
                by_period[f"{pt}_{pn}"] = e.actual_value
            ws.append([
                i, k.name or '', k.target_value or '', k.unit or '',
                by_period.get('ceyrek_1', ''), by_period.get('ceyrek_2', ''),
                by_period.get('ceyrek_3', ''), by_period.get('ceyrek_4', ''),
                by_period.get('yillik_1', '')
            ])
        ws2 = wb.create_sheet("Faaliyetler")
        ws2.append(["#", "Faaliyet Adı", "Ocak", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"])
        for row in ws2.iter_rows(min_row=1, max_row=1):
            for c in row:
                c.font = Font(bold=True)
        activities = ProcessActivity.query.filter_by(process_id=p.id, is_active=True).all()
        for i, a in enumerate(activities, 1):
            tracks = ActivityTrack.query.filter_by(activity_id=a.id, year=year).all()
            track_map = {t.month: 'X' if t.completed else '' for t in tracks}
            ws2.append([i, a.name or ''] + [track_map.get(m, '') for m in range(1, 13)])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        safe_name = (p.name or "Surec").replace(" ", "_").replace("/", "-")[:50]
        filename = f"surec_karnesi_{safe_name}_{year}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        current_app.logger.error(f'Excel export hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@process_bp.route('/api/individual-kpi-data/add', methods=['POST'])
@login_required
def add_individual_kpi_data():
    """Bireysel PG veri girişi."""
    data = request.get_json()
    ipg_id = data.get('individual_pg_id') or data.get('bireysel_pg_id')
    ipg = IndividualPerformanceIndicator.query.filter_by(id=ipg_id, user_id=current_user.id).first_or_404()
    try:
        dt = data.get('data_date')
        if not dt:
            dt = date.today()
        elif isinstance(dt, str):
            dt = datetime.strptime(dt, '%Y-%m-%d').date()
        entry = IndividualKpiData(
            individual_pg_id=ipg.id,
            year=int(data.get('year', datetime.now().year)),
            data_date=dt,
            period_type=(data.get('period_type') or 'yillik').lower().strip(),
            period_no=int(data.get('period_no') or 1),
            period_month=data.get('period_month'),
            target_value=data.get('target_value'),
            actual_value=data.get('actual_value') or data.get('gerceklesen_deger', ''),
            description=data.get('description') or data.get('aciklama'),
            user_id=current_user.id
        )
        db.session.add(entry)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Veri kaydedildi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/kpi-data/detail', methods=['GET'])
@login_required
def kpi_data_detail():
    """Belirli bir KPI periyodunun detaylı veri listesini ve audit geçmişini döner (veri detay modal)."""
    kpi_id = request.args.get('kpi_id', type=int)
    period_key = request.args.get('period_key', '')   # e.g. 'ceyrek_1', 'aylik_6', 'haftalik_2_6'
    year = request.args.get('year', datetime.now().year, type=int)

    if kpi_id is None:
        return jsonify({'success': False, 'message': 'kpi_id gerekli.'}), 400
    if not period_key:
        return jsonify({'success': False, 'message': 'period_key gerekli.'}), 400

    kpi = ProcessKpi.query.join(Process).filter(
        ProcessKpi.id == kpi_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()

    entries_raw = KpiData.query.filter_by(process_kpi_id=kpi.id, year=year).order_by(KpiData.data_date.desc()).all()
    rows = []
    entry_ids = []
    for e in entries_raw:
        keys = data_date_to_period_keys(e.data_date, year)
        if period_key not in keys:
            continue
        entry_ids.append(e.id)
        if not e.is_active:
            continue
        user_name = (f"{(e.user.first_name or '')} {(e.user.last_name or '')}".strip() or e.user.email) if e.user else 'Bilinmiyor'
        rows.append({
            'id': e.id,
            'data_date': str(e.data_date),
            'created_at': e.created_at.isoformat() if e.created_at else None,
            'actual_value': e.actual_value,
            'target_value': e.target_value,
            'description': e.description,
            'user': user_name,
        })

    audits = []
    if entry_ids:
        audit_rows = KpiDataAudit.query.filter(
            KpiDataAudit.kpi_data_id.in_(entry_ids)
        ).order_by(KpiDataAudit.created_at.desc()).all()
        for a in audit_rows:
            u = a.user
            user_name = (f"{(u.first_name or '')} {(u.last_name or '')}".strip() or u.email) if u else 'Bilinmiyor'
            action_label = {'CREATE': 'Ekleme', 'UPDATE': 'Değiştirme', 'DELETE': 'Silme'}.get(a.action_type, a.action_type)
            audits.append({
                'id': a.id,
                'kpi_data_id': a.kpi_data_id,
                'action_type': a.action_type,
                'action_label': action_label,
                'old_value': a.old_value,
                'new_value': a.new_value,
                'action_detail': a.action_detail,
                'user': user_name,
                'created_at': a.created_at.isoformat() if a.created_at else None,
            })

    return jsonify({
        'success': True,
        'kpi_name': kpi.name,
        'period_key': period_key,
        'year': year,
        'entries': rows,
        'audits': audits,
    })


@process_bp.route('/api/kpi-data/update/<int:data_id>', methods=['POST', 'PUT'])
@login_required
def kpi_data_update(data_id):
    """KPI verisini güncelle (veri detay modal'dan)."""
    entry = KpiData.query.join(ProcessKpi).join(Process).filter(
        KpiData.id == data_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_edit_kpi_data_row(entry, entry.process_kpi.process):
        return jsonify({'success': False, 'message': 'Bu veriyi güncelleme yetkiniz yok.'}), 403
    if not entry.is_active:
        return jsonify({'success': False, 'message': 'Silinmiş veri düzenlenemez.'}), 400

    data = request.get_json() or {}
    old_actual = entry.actual_value
    new_actual = str(data['actual_value']) if data.get('actual_value') is not None else ''
    if 'actual_value' in data:
        entry.actual_value = new_actual
    if 'description' in data:
        entry.description = data['description']
    if 'target_value' in data:
        entry.target_value = data['target_value']

    try:
        if 'actual_value' in data and old_actual != new_actual:
            db.session.add(KpiDataAudit(
                kpi_data_id=entry.id,
                action_type='UPDATE',
                old_value=old_actual,
                new_value=new_actual,
                action_detail='Değer değiştirildi',
                user_id=current_user.id,
            ))
        db.session.commit()
        try:
            from app.services.score_engine_service import recalc_on_pg_data_change
            recalc_on_pg_data_change(current_user.tenant_id)
        except Exception as e:
            current_app.logger.warning(f'[kpi_data_update] score_engine hatası: {e}')
        try:
            from app.services.process_deviation_service import check_pg_performance_deviation
            check_pg_performance_deviation(data_id)
        except Exception as e:
            current_app.logger.warning(f'[kpi_data_update] deviation_service hatası: {e}')
        return jsonify({'success': True, 'message': 'Veri güncellendi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/kpi-data/delete/<int:data_id>', methods=['POST', 'DELETE'])
@login_required
def kpi_data_delete(data_id):
    """KPI verisini soft delete ile sil (is_active=False). Veri Detayları modalında görünür kalır."""
    entry = KpiData.query.join(ProcessKpi).join(Process).filter(
        KpiData.id == data_id,
        Process.tenant_id == current_user.tenant_id
    ).first_or_404()
    if not _can_edit_kpi_data_row(entry, entry.process_kpi.process):
        return jsonify({'success': False, 'message': 'Bu veriyi silme yetkiniz yok.'}), 403
    if not entry.is_active:
        return jsonify({'success': False, 'message': 'Veri zaten silinmiş.'}), 400

    try:
        db.session.add(KpiDataAudit(
            kpi_data_id=entry.id,
            action_type='DELETE',
            old_value=entry.actual_value,
            new_value=None,
            action_detail='Veri silindi (soft delete)',
            user_id=current_user.id,
        ))
        entry.is_active = False
        db.session.commit()
        try:
            from app.services.score_engine_service import recalc_on_pg_data_change
            recalc_on_pg_data_change(current_user.tenant_id)
        except Exception as e:
            current_app.logger.warning(f'[kpi_data_delete] score_engine hatası: {e}')
        try:
            from app.services.process_deviation_service import check_pg_performance_deviation
            check_pg_performance_deviation(data_id)
        except Exception as e:
            current_app.logger.warning(f'[kpi_data_delete] deviation_service hatası: {e}')
        return jsonify({'success': True, 'message': 'Veri silindi.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400


@process_bp.route('/api/kpi-data/proje-gorevleri', methods=['GET'])
@login_required
def kpi_data_proje_gorevleri():
    """PG verisini besleyen proje görevlerini listeler (proje modülü yoksa boş döner)."""
    # Proje modülü henüz entegre değil; placeholder olarak boş liste döner
    return jsonify({'success': True, 'gorevler': []})

