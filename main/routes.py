# -*- coding: utf-8 -*-
"""
Main Routes
Dashboard ve genel sayfa route'ları
"""
from flask import Blueprint, render_template, redirect, url_for, current_app, jsonify, request, flash, send_file
from flask_login import login_required, current_user
from extensions import csrf
from models import (
    db, Surec, Kurum, User, AnaStrateji, AltStrateji, surec_liderleri, surec_uyeleri,
    DashboardLayout, BireyselPerformansGostergesi, SurecPerformansGostergesi,
    PerformansGostergeVeri, PerformansGostergeVeriAudit, BireyselFaaliyet, SurecFaaliyet, UserActivityLog,
    SurecPerformansGostergesi,
    Deger, EtikKural, KalitePolitikasi, Project, Task, TaskImpact, TaskComment, TaskMention,
    ProjectFile, project_related_processes, project_members, project_observers, ProjectRisk,
    MainStrategy, SubStrategy, Process, StrategyProcessMatrix, Project, SurecPerformansGostergesi,
    # Faz 2 Modelleri
    ObjectiveComment, StrategicPlan, PlanItem, GembaWalk,
    Competency, UserCompetency, StrategicRisk, MudaFinding,
    # Faz 3 Modelleri
    CrisisMode, SafetyCheck, SuccessionPlan, OrgScenario, OrgChange, InfluenceScore, MarketIntel,
    WellbeingScore, SimulationScenario, DeepWorkSession,
    Persona, ProductSimulation, SmartContract, DaoProposal, DaoVote, MetaverseDepartment, LegacyKnowledge,
    # Faz 4 Modelleri
    Competitor, GameScenario, DoomsdayScenario, YearlyChronicle,
    # V67 Modelleri
    Activity
)
from datetime import datetime, timedelta, date
from io import BytesIO, StringIO
import json
import os
import re
from utils.task_status import COMPLETED_STATUSES, normalize_task_status
from decorators import role_required

main_bp = Blueprint('main', __name__)


def _get_user_project_role_for_page(project: Project):
    """Sayfa route'ları için proje rolü belirler.

    Returns:
        'manager' | 'member' | 'observer' | None
    """
    user_id = current_user.id

    if project.manager_id == user_id:
        return 'manager'

    member_exists = db.session.query(project_members).filter(
        project_members.c.project_id == project.id,
        project_members.c.user_id == user_id,
    ).first() is not None
    if member_exists:
        return 'member'

    observer_exists = db.session.query(project_observers).filter(
        project_observers.c.project_id == project.id,
        project_observers.c.user_id == user_id,
    ).first() is not None
    if observer_exists:
        return 'observer'

    if hasattr(current_user, 'sistem_rol') and current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        return 'manager'

    # İlişkili süreç liderleri manager kabul edilir
    try:
        related_ids = [p.id for p in (project.related_processes or [])]
        if related_ids:
            is_leader = db.session.query(surec_liderleri).filter(
                surec_liderleri.c.surec_id.in_(related_ids),
                surec_liderleri.c.user_id == user_id,
            ).first() is not None
            if is_leader:
                return 'manager'
    except Exception:
        pass

    return None


@main_bp.route('/')
def index():
    """Ana sayfa - Login sayfasına yönlendir"""
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('auth.login'))


@main_bp.route('/offline')
def offline():
    """Offline sayfası - PWA için"""
    return render_template('offline.html')


def get_mock_data():
    """V67 DEPRECATED: Eski mock data fonksiyonu - Fallback için korunuyor.
    Artık Activity.query kullanılmalı. Bu fonksiyon sadece migration script'inde kullanılıyor.
    """
    # V67: Artık veritabanından çekiyoruz, bu fonksiyon sadece geriye uyumluluk için
    from models import Activity
    activities = Activity.query.order_by(Activity.date.desc()).all()
    
    # Dictionary formatına çevir (template uyumluluğu için)
    result = []
    for activity in activities:
        result.append({
            'id': activity.id,
            'source': activity.source,
            'project': activity.project.name if activity.project else activity.project_name or 'N/A',
            'subject': activity.subject,
            'status': activity.status,
            'priority': activity.priority,
            'date': activity.date.strftime('%Y-%m-%d') if activity.date else None
        })
    
    # Eğer veritabanında hiç kayıt yoksa, eski mock veriyi döndür (ilk kurulum için)
    if not result:
        return [
            {'id': 101, 'source': 'Redmine', 'project': 'Omega V66', 'subject': 'Login Güvenlik Yaması', 'status': 'Açık', 'priority': 'High', 'date': '2025-12-29'},
            {'id': 102, 'source': 'Jira', 'project': 'Mobil App', 'subject': 'Bildirim Hatası', 'status': 'Beklemede', 'priority': 'Normal', 'date': '2025-12-29'},
            {'id': 103, 'source': 'Dahili', 'project': 'Sunucu', 'subject': 'Disk Temizliği', 'status': 'Tamamlandı', 'priority': 'Low', 'date': '2025-12-28'},
            {'id': 104, 'source': 'Redmine', 'project': 'Omega V66', 'subject': 'DB Migrasyonu', 'status': 'Açık', 'priority': 'High', 'date': '2025-12-30'},
            {'id': 105, 'source': 'CRM', 'project': 'Satış', 'subject': 'Müşteri Listesi', 'status': 'Devam', 'priority': 'Normal', 'date': '2025-12-30'}
        ]
    
    return result


@main_bp.route('/redmine')
@login_required
def redmine():
    """
    Faaliyetler Sayfası (V67 - Güncel)
    Artık Activity tablosundan veri çekiyor.
    """
    from models import Activity
    
    # 1. Veritabanından faaliyetleri çek (en yeni en üstte)
    db_activities = Activity.query.order_by(Activity.date.desc()).all()
    
    # 2. Template için dictionary formatına çevir
    activities = []
    for activity in db_activities:
        activities.append({
            'id': activity.id,
            'source': activity.source,
            'project': activity.project.name if activity.project else activity.project_name or 'N/A',
            'subject': activity.subject,
            'status': activity.status,
            'priority': activity.priority,
            'date': activity.date.strftime('%Y-%m-%d') if activity.date else None
        })
    
    return render_template('redmine.html', activities=activities)


@main_bp.route('/dashboard')
@login_required
def dashboard():
    """Dashboard sayfası - Ana gösterge paneli (V67 - Güncel)"""
    from models import Activity, Project
    from sqlalchemy.orm import joinedload
    from services.cache_service import get_cached_dashboard_stats, set_cached_dashboard_stats
    from services.strategic_impact_service import get_strategic_impact_summary
    
    try:
        # Cache'den dashboard istatistiklerini getirmeye çalış
        cache_key = f"dashboard_stats_{current_user.id}"
        cached_data = get_cached_dashboard_stats(current_user.id)
        
        if cached_data:
            current_app.logger.info(f"Dashboard cache'den yüklendi: {cache_key}")
            if 'strategic_impact' not in cached_data:
                cached_data['strategic_impact'] = get_strategic_impact_summary(current_user.kurum_id)
                set_cached_dashboard_stats(current_user.id, cached_data)
            return render_template('dashboard.html', 
                                 stats=cached_data['stats'], 
                                 recent_activities=cached_data['recent_activities'],
                                 strategic_impact=cached_data.get('strategic_impact', []))
        
        # Cache'de yoksa hesapla
        # 1. Veritabanından faaliyetleri çek (eager loading ile N+1 çözümü)
        db_activities = Activity.query.options(joinedload(Activity.project)).all()
        
        # 2. Temel İstatistikler
        total_tasks = len(db_activities)
        completed_tasks = len([a for a in db_activities if a.status == 'Tamamlandı'])
        performance_score = round((completed_tasks / total_tasks) * 100, 1) if total_tasks else 0

        stats = {
            'total_tasks': total_tasks,
            'critical_tasks': len([a for a in db_activities if a.priority == 'High' and a.status not in ['Tamamlandı', 'Kapalı']]),
            'completed_tasks': completed_tasks,
            # Mevcut istatistikleri koru (geriye uyumluluk için)
            'total_projects': db.session.query(db.func.count(Project.id)).scalar() or 0,
            'pending_tasks': len([a for a in db_activities if a.status not in ['Tamamlandı', 'Kapalı']]),
            'performance_score': performance_score,
            # V67: Grafik Verileri (Chart.js için)
            'source_counts': {},
            'priority_counts': {'High': 0, 'Normal': 0, 'Low': 0},
            'process_performance': {'Genel': performance_score}  # Template'in beklediği key
        }
        
        # 3. Grafik verilerini döngüyle doldur
        for a in db_activities:
            # Kaynak Sayımı
            stats['source_counts'][a.source] = stats['source_counts'].get(a.source, 0) + 1
            # Öncelik Sayımı
            if a.priority in stats['priority_counts']:
                stats['priority_counts'][a.priority] += 1
        
        # 4. Son 5 aktiviteyi tarihe göre sırala (Template için dictionary formatına çevir)
        sorted_activities = sorted(db_activities, key=lambda x: x.date, reverse=True)[:5]
        recent_activities = []
        for activity in sorted_activities:
            recent_activities.append({
                'id': activity.id,
                'source': activity.source,
                'project': activity.project.name if activity.project else activity.project_name or 'N/A',
                'subject': activity.subject,
                'status': activity.status,
                'priority': activity.priority,
                'date': activity.date.strftime('%Y-%m-%d') if activity.date else None
            })
        
        strategic_impact = get_strategic_impact_summary(current_user.kurum_id)

        # Cache'e kaydet (5 dakika)
        dashboard_data = {
            'stats': stats,
            'recent_activities': recent_activities,
            'strategic_impact': strategic_impact
        }
        set_cached_dashboard_stats(current_user.id, dashboard_data)
        current_app.logger.info(f"Dashboard verileri cache'e kaydedildi: {cache_key}")
        
        # 5. Verileri Template'e Gönder
        return render_template('dashboard.html', 
                             stats=stats, 
                             recent_activities=recent_activities,
                             strategic_impact=strategic_impact)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Dashboard sayfası render hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        # Hata durumunda boş verilerle dashboard göster
        return render_template('dashboard.html',
                             stats={
                                 'total_tasks': 0,
                                 'critical_tasks': 0,
                                 'completed_tasks': 0,
                                 'total_projects': 0,
                                 'pending_tasks': 0,
                                 'performance_score': 0,
                                 'source_counts': {},
                                 'priority_counts': {'High': 0, 'Normal': 0, 'Low': 0},
                                 'process_performance': []
                             },
                             recent_activities=[],
                             strategic_impact=[])


@main_bp.route('/surec-karnesi')
@login_required
def surec_karnesi():
    """Süreç Karnesi sayfası - Kullanıcı sadece kendi süreçlerini görür"""
    try:
        return render_template('surec_karnesi.html')
    except Exception as e:
        import traceback
        current_app.logger.error(f'Süreç karnesi sayfası render hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Template render hatası: {str(e)}", 500


@main_bp.route('/debug/surec-data')
@login_required
def debug_surec_data():
    """Debug endpoint - kullanıcının süreçlerini ve PG sayılarını göster"""
    try:
        current_app.logger.info(f"=== DEBUG SÜREÇ DATA ===")
        current_app.logger.info(f"User: {current_user.id}, Role: {current_user.sistem_rol}, Kurum: {current_user.kurum_id}")
        
        # Kullanıcının süreçlerini belirle
        if current_user.sistem_rol == 'admin':
            surecler = Surec.query.filter_by(silindi=False).all()
        elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            surecler = Surec.query.filter_by(kurum_id=current_user.kurum_id, silindi=False).all()
        else:
            # Normal kullanıcı için lider/üye kontrolü
            lider_surec_ids = db.session.query(surec_liderleri.c.surec_id).filter(
                surec_liderleri.c.user_id == current_user.id
            ).all()
            lider_surec_ids = [sid[0] for sid in lider_surec_ids]
            
            uye_surec_ids = db.session.query(surec_uyeleri.c.surec_id).filter(
                surec_uyeleri.c.user_id == current_user.id
            ).all()
            uye_surec_ids = [sid[0] for sid in uye_surec_ids]
            
            tum_surec_ids = list(set(lider_surec_ids + uye_surec_ids))
            
            if tum_surec_ids:
                surecler = Surec.query.filter(Surec.id.in_(tum_surec_ids)).all()
            else:
                surecler = []
        
        # Her süreç için PG ve faaliyet sayısını topla
        surec_listesi = []
        for s in surecler:
            pg_count = SurecPerformansGostergesi.query.filter_by(surec_id=s.id).count()
            faaliyet_count = SurecFaaliyet.query.filter_by(surec_id=s.id).count()
            
            pg_listesi = []
            for pg in SurecPerformansGostergesi.query.filter_by(surec_id=s.id).all():
                pg_listesi.append({
                    'id': pg.id,
                    'ad': pg.ad,
                    'kodu': pg.kodu if hasattr(pg, 'kodu') else f'PG-{pg.id}'
                })
            
            faaliyet_listesi = []
            for f in SurecFaaliyet.query.filter_by(surec_id=s.id).all():
                faaliyet_listesi.append({
                    'id': f.id,
                    'ad': f.ad
                })
            
            surec_listesi.append({
                'id': s.id,
                'ad': s.ad,
                'pg_sayisi': pg_count,
                'faaliyet_sayisi': faaliyet_count,
                'pg_listesi': pg_listesi,
                'faaliyet_listesi': faaliyet_listesi
            })
        
        return jsonify({
            'success': True,
            'user_info': {
                'id': current_user.id,
                'username': current_user.username,
                'sistem_rol': current_user.sistem_rol,
                'kurum_id': current_user.kurum_id
            },
            'toplam_surec': len(surecler),
            'surecler': surec_listesi
        })
    except Exception as e:
        import traceback
        current_app.logger.error(f'Debug surec data hatası: {e}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/api/kullanici/sureclerim')
@login_required
def api_kullanici_surecleri():
    """Kullanıcının üye/lider olduğu süreçleri getir - Kurum yöneticileri tüm süreçleri görür"""
    try:
        current_app.logger.info(f"=== SÜREÇLERİM API ÇAĞRILDI ===")
        current_app.logger.info(f"User: {current_user.id}, Role: {current_user.sistem_rol}, Kurum: {current_user.kurum_id}")
        
        # Kurum yöneticileri kendi kurumlarındaki TÜM süreçleri görür; admin tüm kurumları görür
        if current_user.sistem_rol in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            if current_user.sistem_rol == 'admin':
                current_app.logger.info(f"Admin - tüm süreçler getiriliyor")
                surecler = Surec.query.filter_by(silindi=False).all()
            else:
                current_app.logger.info(f"Kurum yöneticisi - kurum süreçleri getiriliyor")
                surecler = Surec.query.filter_by(kurum_id=current_user.kurum_id, silindi=False).all()
            
            return jsonify({
                'success': True,
                'surecler': [{
                    'id': s.id,
                    'ad': s.ad,
                    'dokuman_no': s.dokuman_no or '-',
                    'rev_no': s.rev_no or '-',
                    'is_lider': False,  # Kurum yöneticisi için özel durum
                    'is_uye': False     # Kurum yöneticisi için özel durum
                } for s in surecler],
                'user_role': current_user.sistem_rol  # Kullanıcı rolünü ekle
            })
        
        # Normal kullanıcılar için lider/üye kontrolü
        # Kullanıcının lider olduğu süreçler
        lider_surec_ids = db.session.query(surec_liderleri.c.surec_id).filter(
            surec_liderleri.c.user_id == current_user.id
        ).all()
        lider_surec_ids = [sid[0] for sid in lider_surec_ids]
        
        # Kullanıcının üye olduğu süreçler
        uye_surec_ids = db.session.query(surec_uyeleri.c.surec_id).filter(
            surec_uyeleri.c.user_id == current_user.id
        ).all()
        uye_surec_ids = [sid[0] for sid in uye_surec_ids]
        
        # Tüm süreç ID'leri
        tum_surec_ids = list(set(lider_surec_ids + uye_surec_ids))
        
        current_app.logger.info(f'Kullanıcı {current_user.id} ({current_user.username}) - Lider: {lider_surec_ids}, Üye: {uye_surec_ids}')
        
        # Süreçleri getir (kurum izolasyonu: admin dışı sadece kendi kurumu)
        if tum_surec_ids:
            surecler = Surec.query.filter(
                Surec.id.in_(tum_surec_ids),
                Surec.kurum_id == current_user.kurum_id,
                Surec.silindi == False
            ).all()
        else:
            surecler = []
        
        current_app.logger.info(f'Bulunan süreç sayısı: {len(surecler)}')
        
        return jsonify({
            'success': True,
            'surecler': [{
                'id': s.id,
                'ad': s.ad,
                'dokuman_no': s.dokuman_no or '-',
                'rev_no': s.rev_no or '-',
                'is_lider': s.id in lider_surec_ids,
                'is_uye': s.id in uye_surec_ids
            } for s in surecler],
            'user_role': current_user.sistem_rol  # Kullanıcı rolünü ekle
        })
    except Exception as e:
        current_app.logger.error(f'Süreçler getirilemedi: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/surec/<int:surec_id>/performans-gostergesi/add', methods=['POST'])
@login_required
@csrf.exempt
def surec_performans_gostergesi_add(surec_id):
    """Süreç için yeni performans göstergesi ekle"""
    try:
        # JSON verisini al
        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Geçersiz JSON verisi'}), 400
        
        current_app.logger.info(f'Süreç {surec_id} için performans göstergesi ekleme isteği: {data}')
        
        # Zorunlu alan kontrolü
        if not data.get('ad') or not data.get('ad').strip():
            return jsonify({'success': False, 'message': 'Performans göstergesi adı zorunludur'}), 400
        
        # Yetki kontrolü
        surec = Surec.query.get(surec_id)
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Yazma yetkisi: sadece yönetim rolleri
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
        # Tarih dönüştürme
        baslangic_tarihi = None
        bitis_tarihi = None
        try:
            if data.get('baslangic_tarihi'):
                baslangic_tarihi = datetime.strptime(data.get('baslangic_tarihi'), '%Y-%m-%d').date()
            if data.get('bitis_tarihi'):
                bitis_tarihi = datetime.strptime(data.get('bitis_tarihi'), '%Y-%m-%d').date()
        except ValueError as e:
            return jsonify({'success': False, 'message': f'Geçersiz tarih formatı: {str(e)}'}), 400
        
        # Alt strateji ID kontrolü
        alt_strateji_id = None
        if data.get('alt_strateji_id'):
            try:
                alt_strateji_id = int(data.get('alt_strateji_id'))
                # Alt stratejinin bu sürece ait olduğunu kontrol et
                alt_strateji = AltStrateji.query.get(alt_strateji_id)
                if alt_strateji and alt_strateji not in surec.alt_stratejiler:
                    return jsonify({'success': False, 'message': 'Seçilen alt strateji bu sürece ait değil'}), 400
            except (ValueError, TypeError):
                alt_strateji_id = None
        
        # Yeni performans göstergesi oluştur
        yeni_pg = SurecPerformansGostergesi(
            surec_id=surec_id,
            ad=data.get('ad', '').strip(),
            aciklama=data.get('aciklama', '').strip() if data.get('aciklama') else None,
            hedef_deger=data.get('hedef_deger', '').strip() if data.get('hedef_deger') else None,
            olcum_birimi=data.get('olcum_birimi', '').strip() if data.get('olcum_birimi') else None,
            periyot=data.get('periyot', 'Aylik'),
            veri_toplama_yontemi=data.get('veri_toplama_yontemi', 'Ortalama'),
            baslangic_tarihi=baslangic_tarihi,
            bitis_tarihi=bitis_tarihi,
            direction=data.get('direction', 'Increasing'),
            basari_puani_araliklari=data.get('basari_puani_araliklari') if data.get('basari_puani_araliklari') else None,
            alt_strateji_id=alt_strateji_id
        )
        
        db.session.add(yeni_pg)
        db.session.commit()
        
        current_app.logger.info(f'Süreç {surec_id} için performans göstergesi eklendi: {yeni_pg.id}')
        
        return jsonify({
            'success': True,
            'message': 'Performans göstergesi başarıyla eklendi',
            'pg_id': yeni_pg.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Performans göstergesi ekleme hatası: {e}')
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500


@main_bp.route('/surec/<int:surec_id>/performans-gostergesi/<int:pg_id>', methods=['GET'])
@login_required
@csrf.exempt
def surec_performans_gostergesi_get(surec_id, pg_id):
    """Süreç için performans göstergesi bilgilerini getir"""
    try:
        # Yetki kontrolü
        surec = Surec.query.get(surec_id)
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Yetki notu: admin tüm kurumlarda yetkili olmalı
        if current_user.sistem_rol == 'admin':
            pass
        elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            pass
        else:
            # Normal kullanıcı için lider/üye kontrolü
            lider_mi = db.session.query(surec_liderleri).filter(
                surec_liderleri.c.surec_id == surec_id,
                surec_liderleri.c.user_id == current_user.id
            ).first() is not None
            
            uye_mi = db.session.query(surec_uyeleri).filter(
                surec_uyeleri.c.surec_id == surec_id,
                surec_uyeleri.c.user_id == current_user.id
            ).first() is not None
            
            if not (lider_mi or uye_mi):
                return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403
        
        # Performans göstergesini getir (alt strateji ile birlikte)
        pg = SurecPerformansGostergesi.query.options(
            db.joinedload(SurecPerformansGostergesi.alt_strateji).joinedload(AltStrateji.ana_strateji)
        ).filter_by(
            id=pg_id,
            surec_id=surec_id
        ).first()
        
        if not pg:
            return jsonify({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
        
        return jsonify({
            'success': True,
            'gosterge': {
                'id': pg.id,
                'ad': pg.ad,
                'aciklama': pg.aciklama,
                'hedef_deger': pg.hedef_deger,
                'olcum_birimi': pg.olcum_birimi,
                'periyot': pg.periyot,
                'baslangic_tarihi': pg.baslangic_tarihi.strftime('%Y-%m-%d') if pg.baslangic_tarihi else None,
                'bitis_tarihi': pg.bitis_tarihi.strftime('%Y-%m-%d') if pg.bitis_tarihi else None,
                'veri_toplama_yontemi': pg.veri_toplama_yontemi,
                'agirlik': pg.agirlik,
                'onemli': pg.onemli,
                'kodu': pg.kodu,
                'gosterge_turu': pg.gosterge_turu,
                'target_method': pg.target_method,
                'direction': pg.direction,
                'basari_puani_araliklari': pg.basari_puani_araliklari,
                'onceki_yil_ortalamasi': pg.onceki_yil_ortalamasi,
                'alt_strateji_id': pg.alt_strateji_id,
                'alt_strateji': {
                    'id': pg.alt_strateji.id,
                    'ad': pg.alt_strateji.ad,
                    'ana_strateji': {
                        'id': pg.alt_strateji.ana_strateji.id,
                        'ad': pg.alt_strateji.ana_strateji.ad
                    } if pg.alt_strateji and pg.alt_strateji.ana_strateji else None
                } if pg.alt_strateji else None
            }
        })
    except Exception as e:
        current_app.logger.error(f'Performans göstergesi getirme hatası: {e}')
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500


@main_bp.route('/surec/<int:surec_id>/performans-gostergesi/<int:pg_id>/update', methods=['POST'])
@login_required
@csrf.exempt
def surec_performans_gostergesi_update(surec_id, pg_id):
    """Süreç için performans göstergesini güncelle"""
    try:
        # JSON verisini al
        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'Geçersiz JSON verisi'}), 400
        
        current_app.logger.info(f'Süreç {surec_id} için performans göstergesi {pg_id} güncelleme isteği: {data}')
        
        # Yetki kontrolü
        surec = Surec.query.get(surec_id)
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Yazma yetkisi: sadece yönetim rolleri
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
        # Performans göstergesini getir
        pg = SurecPerformansGostergesi.query.filter_by(
            id=pg_id,
            surec_id=surec_id
        ).first()
        
        if not pg:
            return jsonify({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
        
        # Zorunlu alan kontrolü
        if data.get('ad') and not data.get('ad').strip():
            return jsonify({'success': False, 'message': 'Performans göstergesi adı boş olamaz'}), 400
        
        # Tarih dönüştürme
        baslangic_tarihi = None
        bitis_tarihi = None
        try:
            if data.get('baslangic_tarihi'):
                baslangic_tarihi = datetime.strptime(data.get('baslangic_tarihi'), '%Y-%m-%d').date()
            if data.get('bitis_tarihi'):
                bitis_tarihi = datetime.strptime(data.get('bitis_tarihi'), '%Y-%m-%d').date()
        except ValueError as e:
            return jsonify({'success': False, 'message': f'Geçersiz tarih formatı: {str(e)}'}), 400
        
        # Performans göstergesini güncelle
        if data.get('ad'):
            pg.ad = data.get('ad', '').strip()
        if 'aciklama' in data:
            pg.aciklama = data.get('aciklama', '').strip() if data.get('aciklama') else None
        if 'hedef_deger' in data:
            pg.hedef_deger = data.get('hedef_deger', '').strip() if data.get('hedef_deger') else None
        if 'olcum_birimi' in data:
            pg.olcum_birimi = data.get('olcum_birimi', '').strip() if data.get('olcum_birimi') else None
        if 'periyot' in data:
            pg.periyot = data.get('periyot', 'Aylik')
        if 'veri_toplama_yontemi' in data:
            pg.veri_toplama_yontemi = data.get('veri_toplama_yontemi', 'Ortalama')
        if 'baslangic_tarihi' in data:
            pg.baslangic_tarihi = baslangic_tarihi
        if 'bitis_tarihi' in data:
            pg.bitis_tarihi = bitis_tarihi
        if 'agirlik' in data:
            try:
                pg.agirlik = int(data.get('agirlik', 0))
            except (ValueError, TypeError):
                pg.agirlik = 0
        if 'onemli' in data:
            pg.onemli = bool(data.get('onemli', False))
        if 'kodu' in data:
            pg.kodu = data.get('kodu', '').strip() if data.get('kodu') else None
        if 'gosterge_turu' in data:
            pg.gosterge_turu = data.get('gosterge_turu', '').strip() if data.get('gosterge_turu') else None
        if 'target_method' in data:
            pg.target_method = data.get('target_method', '').strip() if data.get('target_method') else None
        if 'direction' in data:
            pg.direction = data.get('direction', 'Increasing')
        if 'basari_puani_araliklari' in data:
            pg.basari_puani_araliklari = data.get('basari_puani_araliklari') if data.get('basari_puani_araliklari') else None
        if 'onceki_yil_ortalamasi' in data:
            try:
                pg.onceki_yil_ortalamasi = float(data.get('onceki_yil_ortalamasi')) if data.get('onceki_yil_ortalamasi') else None
            except (ValueError, TypeError):
                pg.onceki_yil_ortalamasi = None
        if 'alt_strateji_id' in data:
            try:
                pg.alt_strateji_id = int(data.get('alt_strateji_id')) if data.get('alt_strateji_id') else None
            except (ValueError, TypeError):
                pg.alt_strateji_id = None
        
        db.session.commit()
        
        current_app.logger.info(f'Süreç {surec_id} için performans göstergesi {pg_id} güncellendi')
        
        return jsonify({
            'success': True,
            'message': 'Performans göstergesi başarıyla güncellendi',
            'pg_id': pg.id
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Performans göstergesi güncelleme hatası: {e}')
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500


@main_bp.route('/surec/<int:surec_id>/performans-gostergesi/<int:pg_id>/delete', methods=['DELETE'])
@login_required
@csrf.exempt
def surec_performans_gostergesi_delete(surec_id, pg_id):
    """Süreç için performans göstergesini sil"""
    try:
        # Yetki kontrolü
        surec = Surec.query.get(surec_id)
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Yazma yetkisi: sadece yönetim rolleri
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
        # Performans göstergesini getir
        pg = SurecPerformansGostergesi.query.filter_by(
            id=pg_id,
            surec_id=surec_id
        ).first()
        
        if not pg:
            return jsonify({'success': False, 'message': 'Performans göstergesi bulunamadı'}), 404
        
        pg_ad = pg.ad
        db.session.delete(pg)
        db.session.commit()
        
        current_app.logger.info(f'Süreç {surec_id} için performans göstergesi {pg_id} silindi')
        
        return jsonify({
            'success': True,
            'message': f'"{pg_ad}" performans göstergesi başarıyla silindi'
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Performans göstergesi silme hatası: {e}')
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500


@main_bp.route('/gorevlerim')
@login_required
def gorevlerim():
    """Görevlerim sayfası - Kullanıcıya atanmış faaliyetler"""
    try:
        # Görevlerim sayfası - şimdilik dashboard'a yönlendir
        return redirect(url_for('main.dashboard'))
    except Exception as e:
        import traceback
        current_app.logger.error(f'Görevlerim sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/surec-paneli')
@login_required
def surec_paneli():
    """Süreç Paneli sayfası - Süreç yönetimi ve listesi"""
    from models import Surec, User, Kurum
    from sqlalchemy.orm import joinedload
    
    try:
        # Pagination parametreleri
        page = request.args.get('page', 1, type=int)
        per_page = 20  # Sayfa başına süreç sayısı
        
        base_query = Surec.query.options(
            joinedload(Surec.kurum),
            joinedload(Surec.liderler),
            joinedload(Surec.uyeler)
        )

        # Kurum izolasyonu: admin dışı sadece kendi kurumunun süreçlerini görebilir
        if current_user.sistem_rol == 'admin':
            query = base_query
            current_app.logger.info(
                f"Admin - tüm süreçler getiriliyor (Kurum ID: {current_user.kurum_id})"
            )
        elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            query = base_query.filter(Surec.kurum_id == current_user.kurum_id)
            current_app.logger.info(
                f"Kurum yöneticisi - kurum süreçleri getiriliyor (Kurum ID: {current_user.kurum_id})"
            )
        else:
            # Normal kullanıcılar: sadece kendi kurumunda lider/üye oldukları süreçler
            query = base_query.filter(
                Surec.kurum_id == current_user.kurum_id,
                (Surec.liderler.any(User.id == current_user.id)) |
                (Surec.uyeler.any(User.id == current_user.id))
            )
            current_app.logger.info("Normal kullanıcı - lider/üye süreçler getiriliyor")
        
        # Pagination uygula
        pagination = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        surecler = pagination.items
        
        # Lider ve üye süreçlerini ayrı ayrı al (template için)
        lider_surecler = [s for s in surecler if current_user in s.liderler]
        uye_surecler = [s for s in surecler if current_user in s.uyeler and current_user not in s.liderler]
        
        # Kullanıcı listesi (süreç oluşturma formu için) - kurum izolasyonu
        if current_user.sistem_rol == 'admin':
            kullanicilar = User.query.options(joinedload(User.kurum)).all()
        else:
            kullanicilar = User.query.options(joinedload(User.kurum)).filter_by(kurum_id=current_user.kurum_id).all()
        
        # Ana stratejiler ve alt stratejiler (süreç oluşturma formu için) - kurum izolasyonu
        ana_stratejiler_query = AnaStrateji.query.options(joinedload(AnaStrateji.alt_stratejiler))
        if current_user.sistem_rol != 'admin':
            ana_stratejiler_query = ana_stratejiler_query.filter_by(kurum_id=current_user.kurum_id)
        ana_stratejiler = ana_stratejiler_query.all()

        # Kurum listesi (sadece admin için dropdown'da gerekli)
        kurumlar = []
        if current_user.sistem_rol == 'admin':
            try:
                kurumlar = Kurum.query.order_by(Kurum.kisa_ad).all()
            except Exception:
                kurumlar = Kurum.query.all()
        
        return render_template('surec_panel.html', 
                             surecler=surecler,
                             lider_surecler=lider_surecler,
                             uye_surecler=uye_surecler,
                             kullanicilar=kullanicilar,
                             ana_stratejiler=ana_stratejiler,
                             kurumlar=kurumlar,
                             pagination=pagination)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Süreç Paneli sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/surec/<int:surec_id>')
@login_required
def get_surec(surec_id):
    """Süreç detaylarını getir - viewSurec fonksiyonu için"""
    try:
        surec = Surec.query.options(
            db.joinedload(Surec.liderler),
            db.joinedload(Surec.uyeler),
            db.joinedload(Surec.alt_stratejiler).joinedload(AltStrateji.ana_strateji),
            db.joinedload(Surec.kurum)
        ).get(surec_id)
        
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Yetki kontrolü (admin tüm kurumlarda yetkili)
        if current_user.sistem_rol == 'admin':
            pass
        elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            pass
        else:
            if current_user not in surec.liderler and current_user not in surec.uyeler:
                return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403
        
        # Performans göstergelerini getir
        performans_gostergeleri = SurecPerformansGostergesi.query.filter_by(surec_id=surec_id).all()
        
        return jsonify({
            'success': True,
            'surec': {
                'id': surec.id,
                'ad': surec.ad,
                'dokuman_no': surec.dokuman_no,
                'rev_no': surec.rev_no,
                'rev_tarihi': surec.rev_tarihi.strftime('%Y-%m-%d') if surec.rev_tarihi else None,
                'ilk_yayin_tarihi': surec.ilk_yayin_tarihi.strftime('%Y-%m-%d') if surec.ilk_yayin_tarihi else None,
                'baslangic_tarihi': surec.baslangic_tarihi.strftime('%Y-%m-%d') if surec.baslangic_tarihi else None,
                'bitis_tarihi': surec.bitis_tarihi.strftime('%Y-%m-%d') if surec.bitis_tarihi else None,
                'durum': surec.durum,
                'ilerleme': surec.ilerleme or 0,
                'baslangic_siniri': surec.baslangic_siniri,
                'bitis_siniri': surec.bitis_siniri,
                'aciklama': surec.aciklama,
                'liderler': [{'id': u.id, 'username': u.username} for u in surec.liderler],
                'uyeler': [{'id': u.id, 'username': u.username} for u in surec.uyeler],
                'alt_stratejiler': [{
                    'id': s.id, 
                    'ad': s.ad,
                    'ana_strateji': {'id': s.ana_strateji.id, 'ad': s.ana_strateji.ad} if s.ana_strateji else None
                } for s in surec.alt_stratejiler],
                'kurum': {'id': surec.kurum.id, 'kisa_ad': surec.kurum.kisa_ad} if surec.kurum else None
            },
            'performans_gostergeleri': [{
                'id': pg.id,
                'ad': pg.ad,
                'hedef_deger': pg.hedef_deger,
                # Template uyumu
                'olcum_birimi': pg.olcum_birimi,
                'periyot': pg.periyot,
                # Geriye uyumluluk
                'birim': pg.olcum_birimi
            } for pg in performans_gostergeleri]
        })
    except Exception as e:
        import traceback
        current_app.logger.error(f'Süreç detay hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/surec/<int:surec_id>/faaliyetler')
@login_required
def get_surec_faaliyetler(surec_id):
    """Süreç faaliyetlerini getir"""
    try:
        surec = Surec.query.get(surec_id)
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Yetki kontrolü (admin tüm kurumlarda yetkili)
        if current_user.sistem_rol == 'admin':
            pass
        elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            pass
        else:
            if current_user not in surec.liderler and current_user not in surec.uyeler:
                return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403
        
        faaliyetler = SurecFaaliyet.query.filter_by(surec_id=surec_id).all()
        
        return jsonify({
            'success': True,
            'faaliyetler': [{
                'id': f.id,
                'ad': f.ad,
                'durum': f.durum,
                'ilerleme': f.ilerleme or 0,
                'baslangic_tarihi': f.baslangic_tarihi.strftime('%Y-%m-%d') if f.baslangic_tarihi else None,
                'bitis_tarihi': f.bitis_tarihi.strftime('%Y-%m-%d') if f.bitis_tarihi else None
            } for f in faaliyetler]
        })
    except Exception as e:
        import traceback
        current_app.logger.error(f'Süreç faaliyetleri hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/get-process/<int:process_id>')
@login_required
def admin_get_process(process_id):
    """Süreç bilgilerini getir - editProcess fonksiyonu için"""
    try:
        surec = Surec.query.options(
            db.joinedload(Surec.liderler),
            db.joinedload(Surec.uyeler),
            db.joinedload(Surec.alt_stratejiler).joinedload(AltStrateji.ana_strateji),
            db.joinedload(Surec.kurum)
        ).get(process_id)
        
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Düzenleme yetkisi: sadece yönetim rolleri
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu süreci düzenleme yetkiniz yok'}), 403
        
        return jsonify({
            'success': True,
            'surec': {
                'id': surec.id,
                'ad': surec.ad,
                'dokuman_no': surec.dokuman_no or '',
                'rev_no': surec.rev_no or '',
                'rev_tarihi': surec.rev_tarihi.strftime('%Y-%m-%d') if surec.rev_tarihi else '',
                'ilk_yayin_tarihi': surec.ilk_yayin_tarihi.strftime('%Y-%m-%d') if surec.ilk_yayin_tarihi else '',
                'baslangic_tarihi': surec.baslangic_tarihi.strftime('%Y-%m-%d') if surec.baslangic_tarihi else '',
                'bitis_tarihi': surec.bitis_tarihi.strftime('%Y-%m-%d') if surec.bitis_tarihi else '',
                'durum': surec.durum,
                'ilerleme': surec.ilerleme or 0,
                'baslangic_siniri': surec.baslangic_siniri or '',
                'bitis_siniri': surec.bitis_siniri or '',
                'aciklama': surec.aciklama or '',
                'liderler': [{'id': u.id, 'username': u.username} for u in surec.liderler],
                'uyeler': [{'id': u.id, 'username': u.username} for u in surec.uyeler],
                'alt_stratejiler': [{
                    'id': s.id,
                    'ad': s.ad,
                    'ana_strateji': {'id': s.ana_strateji.id, 'ad': s.ana_strateji.ad} if s.ana_strateji else None
                } for s in surec.alt_stratejiler]
            }
        })
    except Exception as e:
        import traceback
        current_app.logger.error(f'Süreç bilgisi getirme hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/surec/update/<int:surec_id>', methods=['POST'])
@login_required
@csrf.exempt
def update_surec(surec_id):
    """Süreç güncelle - updateSurec fonksiyonu için"""
    try:
        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400
        
        data = request.get_json()
        surec = Surec.query.get(surec_id)
        
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Düzenleme yetkisi: sadece yönetim rolleri
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu süreci düzenleme yetkiniz yok'}), 403
        
        # Süreç bilgilerini güncelle
        surec.ad = data.get('surec_adi', surec.ad)
        surec.dokuman_no = data.get('dokuman_no', surec.dokuman_no)
        surec.rev_no = data.get('rev_no', surec.rev_no)
        surec.rev_tarihi = datetime.strptime(data['rev_tarihi'], '%Y-%m-%d').date() if data.get('rev_tarihi') else surec.rev_tarihi
        surec.ilk_yayin_tarihi = datetime.strptime(data['ilk_yayin_tarihi'], '%Y-%m-%d').date() if data.get('ilk_yayin_tarihi') else surec.ilk_yayin_tarihi
        surec.baslangic_tarihi = datetime.strptime(data['baslangic_tarihi'], '%Y-%m-%d').date() if data.get('baslangic_tarihi') else surec.baslangic_tarihi
        surec.bitis_tarihi = datetime.strptime(data['bitis_tarihi'], '%Y-%m-%d').date() if data.get('bitis_tarihi') else surec.bitis_tarihi
        surec.durum = data.get('durum', surec.durum)
        surec.ilerleme = data.get('ilerleme', surec.ilerleme)
        surec.baslangic_siniri = data.get('baslangic_siniri', surec.baslangic_siniri)
        surec.bitis_siniri = data.get('bitis_siniri', surec.bitis_siniri)
        surec.aciklama = data.get('aciklama', surec.aciklama)
        
        # Liderleri güncelle (kurum izolasyonu)
        lider_ids = data.get('lider_ids', [])
        if lider_ids:
            normalized_lider_ids = [int(x) for x in lider_ids]
            surec.liderler = User.query.filter(
                User.kurum_id == surec.kurum_id,
                User.id.in_(normalized_lider_ids)
            ).all()
        
        # Üyeleri güncelle (kurum izolasyonu)
        uye_ids = data.get('uye_ids', [])
        if uye_ids is not None:
            normalized_uye_ids = [int(x) for x in uye_ids]
            surec.uyeler = User.query.filter(
                User.kurum_id == surec.kurum_id,
                User.id.in_(normalized_uye_ids)
            ).all()
        
        # Alt stratejileri güncelle (kurum izolasyonu)
        strateji_ids = data.get('strateji_ids', None)
        if strateji_ids is None:
            strateji_ids = data.get('alt_strateji_ids', [])
        if strateji_ids:
            normalized_strateji_ids = [int(x) for x in strateji_ids]
            surec.alt_stratejiler = AltStrateji.query.filter(
                AltStrateji.id.in_(normalized_strateji_ids),
                AltStrateji.ana_strateji.has(kurum_id=surec.kurum_id)
            ).all()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Süreç başarıyla güncellendi'
        })
    except Exception as e:
        db.session.rollback()
        import traceback
        current_app.logger.error(f'Süreç güncelleme hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/update-process/<int:process_id>', methods=['PUT'])
@login_required
@csrf.exempt
def admin_update_process(process_id):
    """Süreç güncelle - updateProcess fonksiyonu için"""
    try:
        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400
        
        data = request.get_json()
        surec = Surec.query.get(process_id)
        
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404
        
        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Düzenleme yetkisi: sadece yönetim rolleri
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu süreci düzenleme yetkiniz yok'}), 403
        
        # Süreç bilgilerini güncelle
        surec.ad = data.get('ad', surec.ad)
        surec.dokuman_no = data.get('dokuman_no', surec.dokuman_no)
        surec.rev_no = data.get('rev_no', surec.rev_no)
        if data.get('rev_tarihi'):
            surec.rev_tarihi = datetime.strptime(data['rev_tarihi'], '%Y-%m-%d').date()
        if data.get('ilk_yayin_tarihi'):
            surec.ilk_yayin_tarihi = datetime.strptime(data['ilk_yayin_tarihi'], '%Y-%m-%d').date()
        if data.get('baslangic_tarihi'):
            surec.baslangic_tarihi = datetime.strptime(data['baslangic_tarihi'], '%Y-%m-%d').date()
        if data.get('bitis_tarihi'):
            surec.bitis_tarihi = datetime.strptime(data['bitis_tarihi'], '%Y-%m-%d').date()
        surec.durum = data.get('durum', surec.durum)
        surec.ilerleme = data.get('ilerleme', surec.ilerleme)
        surec.baslangic_siniri = data.get('baslangic_siniri', surec.baslangic_siniri)
        surec.bitis_siniri = data.get('bitis_siniri', surec.bitis_siniri)
        surec.aciklama = data.get('aciklama', surec.aciklama)
        
        # Liderleri güncelle (kurum izolasyonu)
        lider_ids = data.get('lider_ids', [])
        if lider_ids:
            normalized_lider_ids = [int(x) for x in lider_ids]
            surec.liderler = User.query.filter(
                User.kurum_id == surec.kurum_id,
                User.id.in_(normalized_lider_ids)
            ).all()
        
        # Üyeleri güncelle (kurum izolasyonu)
        uye_ids = data.get('uye_ids', [])
        if uye_ids:
            normalized_uye_ids = [int(x) for x in uye_ids]
            surec.uyeler = User.query.filter(
                User.kurum_id == surec.kurum_id,
                User.id.in_(normalized_uye_ids)
            ).all()
        
        # Alt stratejileri güncelle (kurum izolasyonu)
        strateji_ids = data.get('strateji_ids', [])
        if strateji_ids:
            normalized_strateji_ids = [int(x) for x in strateji_ids]
            surec.alt_stratejiler = AltStrateji.query.filter(
                AltStrateji.id.in_(normalized_strateji_ids),
                AltStrateji.ana_strateji.has(kurum_id=surec.kurum_id)
            ).all()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Süreç başarıyla güncellendi'
        })
    except Exception as e:
        db.session.rollback()
        import traceback
        current_app.logger.error(f'Süreç güncelleme hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return jsonify({'success': False, 'message': str(e)}), 500


def _parse_optional_date(value: object):
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    return datetime.strptime(text, '%Y-%m-%d').date()


@main_bp.route('/surec/<int:surec_id>/faaliyet/<int:faaliyet_id>')
@login_required
def get_surec_faaliyet_detay(surec_id: int, faaliyet_id: int):
    """Süreç faaliyeti detayını getir - surec_karnesi.html editFaaliyet() için"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

        surec = Surec.query.get(surec_id)
        if not surec or surec.silindi:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404

        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        faaliyet = SurecFaaliyet.query.filter_by(id=faaliyet_id, surec_id=surec_id).first()
        if not faaliyet:
            return jsonify({'success': False, 'message': 'Faaliyet bulunamadı'}), 404

        return jsonify({
            'success': True,
            'faaliyet': {
                'id': faaliyet.id,
                'surec_id': faaliyet.surec_id,
                'ad': faaliyet.ad,
                'aciklama': faaliyet.aciklama or '',
                'baslangic_tarihi': faaliyet.baslangic_tarihi.strftime('%Y-%m-%d') if faaliyet.baslangic_tarihi else '',
                'bitis_tarihi': faaliyet.bitis_tarihi.strftime('%Y-%m-%d') if faaliyet.bitis_tarihi else '',
                'durum': faaliyet.durum,
                'ilerleme': faaliyet.ilerleme or 0,
            }
        })
    except Exception as e:
        current_app.logger.error(f'Faaliyet detay hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/surec/<int:surec_id>/faaliyet/add', methods=['POST'])
@login_required
@csrf.exempt
def add_surec_faaliyet(surec_id: int):
    """Süreç faaliyeti ekle - surec_karnesi.html addFaaliyet() için"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400
        data = request.get_json() or {}

        surec = Surec.query.get(surec_id)
        if not surec or surec.silindi:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404

        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        ad = (data.get('ad') or '').strip()
        if not ad:
            return jsonify({'success': False, 'message': 'Faaliyet adı zorunludur'}), 400

        yeni = SurecFaaliyet(
            surec_id=surec_id,
            ad=ad,
            aciklama=(data.get('aciklama') or '').strip() or None,
            baslangic_tarihi=_parse_optional_date(data.get('baslangic_tarihi')),
            bitis_tarihi=_parse_optional_date(data.get('bitis_tarihi')),
            durum=(data.get('durum') or 'Planlandı'),
            ilerleme=int(data.get('ilerleme') or 0),
        )
        db.session.add(yeni)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Faaliyet başarıyla eklendi', 'faaliyet_id': yeni.id})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Faaliyet ekleme hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/surec/<int:surec_id>/faaliyet/<int:faaliyet_id>/update', methods=['POST'])
@login_required
@csrf.exempt
def update_surec_faaliyet(surec_id: int, faaliyet_id: int):
    """Süreç faaliyeti güncelle - surec_karnesi.html updateFaaliyet() için"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400
        data = request.get_json() or {}

        surec = Surec.query.get(surec_id)
        if not surec or surec.silindi:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404

        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        faaliyet = SurecFaaliyet.query.filter_by(id=faaliyet_id, surec_id=surec_id).first()
        if not faaliyet:
            return jsonify({'success': False, 'message': 'Faaliyet bulunamadı'}), 404

        ad = (data.get('ad') or faaliyet.ad).strip()
        if not ad:
            return jsonify({'success': False, 'message': 'Faaliyet adı zorunludur'}), 400

        faaliyet.ad = ad
        faaliyet.aciklama = (data.get('aciklama') or '').strip() or None
        faaliyet.baslangic_tarihi = _parse_optional_date(data.get('baslangic_tarihi'))
        faaliyet.bitis_tarihi = _parse_optional_date(data.get('bitis_tarihi'))

        db.session.commit()

        return jsonify({'success': True, 'message': 'Faaliyet başarıyla güncellendi'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Faaliyet güncelleme hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/surec/<int:surec_id>/faaliyet/<int:faaliyet_id>/delete', methods=['DELETE'])
@login_required
@csrf.exempt
def delete_surec_faaliyet(surec_id: int, faaliyet_id: int):
    """Süreç faaliyeti sil - surec_karnesi.html deleteFaaliyet() için"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

        surec = Surec.query.get(surec_id)
        if not surec or surec.silindi:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404

        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        faaliyet = SurecFaaliyet.query.filter_by(id=faaliyet_id, surec_id=surec_id).first()
        if not faaliyet:
            return jsonify({'success': False, 'message': 'Faaliyet bulunamadı'}), 404

        db.session.delete(faaliyet)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Faaliyet başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Faaliyet silme hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


def _ensure_int_list(value: object) -> list[int]:
    if value is None:
        return []
    if isinstance(value, list):
        raw = value
    else:
        raw = [value]
    out: list[int] = []
    for item in raw:
        if item is None:
            continue
        text = str(item).strip()
        if not text:
            continue
        out.append(int(text))
    return out


def _validate_same_kurum_user_ids(kurum_id: int, user_ids: list[int]) -> list[User]:
    if not user_ids:
        return []
    users = User.query.filter(User.kurum_id == kurum_id, User.id.in_(user_ids)).all()
    if len(users) != len(set(user_ids)):
        raise ValueError('Seçilen kullanıcılar kurumla uyuşmuyor veya bulunamadı')
    return users


def _validate_same_kurum_alt_stratejiler(kurum_id: int, strateji_ids: list[int]) -> list[AltStrateji]:
    if not strateji_ids:
        return []
    stratejiler = AltStrateji.query.filter(
        AltStrateji.id.in_(strateji_ids),
        AltStrateji.ana_strateji.has(kurum_id=kurum_id)
    ).all()
    if len(stratejiler) != len(set(strateji_ids)):
        raise ValueError('Seçilen stratejiler kurumla uyuşmuyor veya bulunamadı')
    return stratejiler


@main_bp.route('/surec/get/<int:surec_id>')
@login_required
def surec_get_for_edit(surec_id):
    """Süreç bilgilerini getir - surec_panel.html editSurec() için"""
    try:
        surec = Surec.query.options(
            db.joinedload(Surec.liderler),
            db.joinedload(Surec.uyeler),
            db.joinedload(Surec.alt_stratejiler).joinedload(AltStrateji.ana_strateji),
            db.joinedload(Surec.kurum)
        ).filter_by(id=surec_id, silindi=False).first()

        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404

        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Düzenleme yetkisi: sadece yönetim rolleri
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu süreci düzenleme yetkiniz yok'}), 403

        return jsonify({
            'success': True,
            'surec': {
                'id': surec.id,
                'ad': surec.ad,
                'dokuman_no': surec.dokuman_no or '',
                'rev_no': surec.rev_no or '',
                'rev_tarihi': surec.rev_tarihi.strftime('%Y-%m-%d') if surec.rev_tarihi else '',
                'ilk_yayin_tarihi': surec.ilk_yayin_tarihi.strftime('%Y-%m-%d') if surec.ilk_yayin_tarihi else '',
                'baslangic_tarihi': surec.baslangic_tarihi.strftime('%Y-%m-%d') if surec.baslangic_tarihi else '',
                'bitis_tarihi': surec.bitis_tarihi.strftime('%Y-%m-%d') if surec.bitis_tarihi else '',
                'durum': surec.durum,
                'ilerleme': surec.ilerleme or 0,
                'baslangic_siniri': surec.baslangic_siniri or '',
                'bitis_siniri': surec.bitis_siniri or '',
                'aciklama': surec.aciklama or '',
                'liderler': [{'id': u.id, 'username': u.username} for u in surec.liderler],
                'uyeler': [{'id': u.id, 'username': u.username} for u in surec.uyeler],
                'alt_stratejiler': [{
                    'id': s.id,
                    'ad': s.ad,
                    'ana_strateji': {'id': s.ana_strateji.id, 'ad': s.ana_strateji.ad} if s.ana_strateji else None
                } for s in surec.alt_stratejiler],
                'kurum': {'id': surec.kurum.id, 'kisa_ad': surec.kurum.kisa_ad} if surec.kurum else None
            }
        })
    except Exception as e:
        current_app.logger.error(f'Süreç getirme hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/surec/add-simple', methods=['POST'])
@login_required
@csrf.exempt
def surec_add_simple():
    """Süreç ekle - surec_panel.html form submit için"""
    try:
        # Yetki: normal kullanıcı süreç oluşturamaz
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Süreç oluşturma yetkiniz yok'}), 403

        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400

        data = request.get_json() or {}

        ad = (data.get('surec_adi') or data.get('ad') or '').strip()
        if not ad:
            return jsonify({'success': False, 'message': 'Süreç adı zorunludur'}), 400

        lider_ids = _ensure_int_list(data.get('lider_ids'))
        if not lider_ids:
            return jsonify({'success': False, 'message': 'En az bir lider seçmelisiniz'}), 400

        # Kurum belirleme: admin opsiyonel alabilir; diğerleri kendi kurumuna sabit
        if current_user.sistem_rol == 'admin':
            kurum_id = int(data.get('kurum_id') or current_user.kurum_id)
        else:
            kurum_id = int(current_user.kurum_id)

        kurum = Kurum.query.get(kurum_id)
        if not kurum:
            return jsonify({'success': False, 'message': 'Kurum bulunamadı'}), 404

        uye_ids = _ensure_int_list(data.get('uye_ids'))
        alt_strateji_ids = _ensure_int_list(data.get('alt_strateji_ids') or data.get('strateji_ids'))

        liderler = _validate_same_kurum_user_ids(kurum_id, lider_ids)
        uyeler = _validate_same_kurum_user_ids(kurum_id, uye_ids)
        alt_stratejiler = _validate_same_kurum_alt_stratejiler(kurum_id, alt_strateji_ids)

        surec = Surec(
            kurum_id=kurum_id,
            ad=ad,
            dokuman_no=(data.get('dokuman_no') or '').strip() or None,
            rev_no=(data.get('rev_no') or '').strip() or None,
            rev_tarihi=_parse_optional_date(data.get('rev_tarihi')),
            ilk_yayin_tarihi=_parse_optional_date(data.get('ilk_yayin_tarihi')),
            baslangic_tarihi=_parse_optional_date(data.get('baslangic_tarihi')),
            bitis_tarihi=_parse_optional_date(data.get('bitis_tarihi')),
            durum=(data.get('durum') or 'Aktif'),
            ilerleme=int(data.get('ilerleme') or 0),
            baslangic_siniri=(data.get('baslangic_siniri') or '').strip() or None,
            bitis_siniri=(data.get('bitis_siniri') or '').strip() or None,
            aciklama=(data.get('aciklama') or '').strip() or None,
        )

        surec.liderler = liderler
        surec.uyeler = uyeler
        surec.alt_stratejiler = alt_stratejiler

        db.session.add(surec)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Süreç başarıyla oluşturuldu', 'surec_id': surec.id})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Süreç oluşturma hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/surec/delete/<int:surec_id>', methods=['DELETE'])
@login_required
@csrf.exempt
def surec_delete(surec_id):
    """Süreç soft-delete"""
    try:
        # Yetki: normal kullanıcı süreç silemez
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Süreç silme yetkiniz yok'}), 403

        surec = Surec.query.filter_by(id=surec_id, silindi=False).first()
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404

        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        # Yetki: admin tam; kurum_yoneticisi/ust_yonetim kurum içinde
        if current_user.sistem_rol == 'admin':
            pass
        elif current_user.sistem_rol in ['kurum_yoneticisi', 'ust_yonetim']:
            pass

        surec.silindi = True
        surec.deleted_at = datetime.utcnow()
        surec.deleted_by = current_user.id
        db.session.commit()

        return jsonify({'success': True, 'message': 'Süreç başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Süreç silme hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/create-process', methods=['POST'])
@login_required
@csrf.exempt
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def admin_create_process():
    """Süreç ekle - admin panel / surec panel modal için"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

        if not request.is_json:
            return jsonify({'success': False, 'message': 'Content-Type application/json olmalı'}), 400

        data = request.get_json() or {}

        ad = (data.get('ad') or data.get('surec_adi') or '').strip()
        if not ad:
            return jsonify({'success': False, 'message': 'Süreç adı zorunludur'}), 400

        lider_ids = _ensure_int_list(data.get('lider_ids'))
        if not lider_ids:
            return jsonify({'success': False, 'message': 'En az bir lider seçmelisiniz'}), 400

        if current_user.sistem_rol == 'admin':
            kurum_id = int(data.get('kurum_id') or current_user.kurum_id)
        else:
            kurum_id = int(current_user.kurum_id)

        kurum = Kurum.query.get(kurum_id)
        if not kurum:
            return jsonify({'success': False, 'message': 'Kurum bulunamadı'}), 404

        uye_ids = _ensure_int_list(data.get('uye_ids'))
        strateji_ids = _ensure_int_list(data.get('strateji_ids'))

        liderler = _validate_same_kurum_user_ids(kurum_id, lider_ids)
        uyeler = _validate_same_kurum_user_ids(kurum_id, uye_ids)
        alt_stratejiler = _validate_same_kurum_alt_stratejiler(kurum_id, strateji_ids)

        surec = Surec(
            kurum_id=kurum_id,
            ad=ad,
            dokuman_no=(data.get('dokuman_no') or '').strip() or None,
            rev_no=(data.get('rev_no') or '').strip() or None,
            rev_tarihi=_parse_optional_date(data.get('rev_tarihi')),
            ilk_yayin_tarihi=_parse_optional_date(data.get('ilk_yayin_tarihi')),
            baslangic_tarihi=_parse_optional_date(data.get('baslangic_tarihi')),
            bitis_tarihi=_parse_optional_date(data.get('bitis_tarihi')),
            durum=(data.get('durum') or 'Aktif'),
            ilerleme=int(data.get('ilerleme') or 0),
            baslangic_siniri=(data.get('baslangic_siniri') or '').strip() or None,
            bitis_siniri=(data.get('bitis_siniri') or '').strip() or None,
            aciklama=(data.get('aciklama') or '').strip() or None,
        )

        surec.liderler = liderler
        surec.uyeler = uyeler
        surec.alt_stratejiler = alt_stratejiler

        db.session.add(surec)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Süreç başarıyla oluşturuldu', 'surec_id': surec.id})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Admin süreç oluşturma hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/add-process', methods=['POST'])
@login_required
@csrf.exempt
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def admin_add_process():
    """Admin panel için uyumlu süreç ekleme endpoint'i."""
    return admin_create_process()


@main_bp.route('/admin/delete-process/<int:process_id>', methods=['DELETE'])
@login_required
@csrf.exempt
def admin_delete_process(process_id):
    """Süreç soft-delete - admin panel için"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

        surec = Surec.query.filter_by(id=process_id, silindi=False).first()
        if not surec:
            return jsonify({'success': False, 'message': 'Süreç bulunamadı'}), 404

        # Kurum izolasyonu: admin dışı sadece kendi kurumu
        if current_user.sistem_rol != 'admin' and surec.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu süreçte yetkiniz yok'}), 403

        surec.silindi = True
        surec.deleted_at = datetime.utcnow()
        surec.deleted_by = current_user.id
        db.session.commit()

        return jsonify({'success': True, 'message': 'Süreç başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Admin süreç silme hatası: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/performans-kartim')
@login_required
def performans_kartim():
    """Performans Kartım sayfası - Bireysel performans göstergeleri"""
    try:
        return render_template('bireysel_panel.html')
    except Exception as e:
        import traceback
        current_app.logger.error(f'Performans Kartım sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/kurum-paneli')
@login_required
def kurum_paneli():
    """Kurum Paneli sayfası"""
    # Kurum yöneticileri için
    if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    try:
        # Kurum paneli sayfasını render et
        # Kurum bilgilerini getir
        from models import AnaStrateji, AltStrateji
        # Eager loading ile ilişkili verileri tek seferde çek
        
        # Admin TÜM kurumları, diğerleri sadece kendi kurumunu görsün
        if current_user.sistem_rol == 'admin':
            # Admin - tüm kurumları görecek
            kurumlar = Kurum.query.all()  # Tüm kurumları getir
            kurum = None  # Admin tüm kurumları görecek, belirli bir kurum için filtreleme yok
            ana_stratejiler = AnaStrateji.query.all()
            surecler = Surec.query.options(db.joinedload(Surec.kurum)).all()
        else:
            # Kurum yöneticisi - sadece kendi kurumunu görecek
            kurumlar = [Kurum.query.get(current_user.kurum_id)]
            kurum = kurumlar[0] if kurumlar[0] else None
            ana_stratejiler = AnaStrateji.query.filter_by(kurum_id=current_user.kurum_id).all()
            surecler = Surec.query.options(db.joinedload(Surec.kurum)).filter_by(kurum_id=current_user.kurum_id).all()
        
        # Deger, EtikKural, KalitePolitikasi verilerini getir
        try:
            if current_user.sistem_rol == 'admin':
                degerler = Deger.query.all()
                etik_kurallari = EtikKural.query.all()
                kalite_politikalari = KalitePolitikasi.query.all()
            else:
                degerler = Deger.query.filter_by(kurum_id=current_user.kurum_id).all()
                etik_kurallari = EtikKural.query.filter_by(kurum_id=current_user.kurum_id).all()
                kalite_politikalari = KalitePolitikasi.query.filter_by(kurum_id=current_user.kurum_id).all()
        except Exception as e:
            current_app.logger.error(f"Kurum bilgileri getirilirken hata: {e}")
            degerler = []
            etik_kurallari = []
            kalite_politikalari = []
        
        if current_user.sistem_rol == 'admin':
            uyeler = User.query.all()
        else:
            uyeler = User.query.filter_by(kurum_id=current_user.kurum_id).all()
        
        return render_template('kurum_panel.html',
                             kurum=kurum,
                             kurumlar=kurumlar if current_user.sistem_rol == 'admin' else None,
                             ana_stratejiler=ana_stratejiler,
                             degerler=degerler,
                             etik_kurallari=etik_kurallari,
                             kalite_politikalari=kalite_politikalari,
                             surecler=surecler,
                             uyeler=uyeler)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Kurum Paneli sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/admin/seed_db')
@login_required
def seed_db():
    """Veritabanına demo veri ekle - Sadece admin için"""
    try:
        # Sadece admin rolü için
        if current_user.sistem_rol != 'admin':
            flash('Bu işlem için yetkiniz yok.', 'error')
            return redirect(url_for('main.dashboard'))
        
        from services.seeder import seed_all
        
        results = seed_all(db)
        
        if results['hata']:
            flash(f'Demo veri oluşturulurken hata oluştu: {results["hata"]}', 'error')
        else:
            flash(
                f'Demo veriler başarıyla oluşturuldu! '
                f'Kurum: {results["kurumlar"]}, Kullanıcı: {results["users"]}, '
                f'Ana Strateji: {results["ana_stratejiler"]}, Alt Strateji: {results["alt_stratejiler"]}, '
                f'Süreç: {results["surecler"]}, Proje: {results["projeler"]}, Görev: {results["gorevler"]}',
                'success'
            )
        
        return redirect(url_for('main.admin_panel'))
    except Exception as e:
        current_app.logger.error(f'Seed DB hatası: {e}')
        flash('Demo veri oluşturulurken bir hata oluştu.', 'error')
        return redirect(url_for('main.admin_panel'))


@main_bp.route('/admin-panel')
@login_required
def admin_panel():
    """Admin Panel sayfası - Sistem yöneticileri ve kurum yöneticileri için"""
    # Admin veya kurum_yoneticisi erişebilir
    if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    from models import Kurum, User, Surec, AnaStrateji
    
    # Sistem admini (kurum_id=1) - TÜM kurumlar, TÜM kullanıcılar, TÜM süreçler
    is_system_admin = current_user.sistem_rol == 'admin' and current_user.kurum_id == 1
    
    if is_system_admin:
        # Sistem yöneticisi: Tüm veriler (sadece aktif kayıtlar)
        kurumlar = Kurum.query.filter_by(silindi=False).all()
        kullanicilar = User.query.filter_by(silindi=False).all()
        surecler = Surec.query.filter_by(silindi=False).all()
    else:
        # Kurum yöneticisi: Sadece kendi kurumunun verileri (sadece aktif kayıtlar)
        kurumlar = Kurum.query.filter_by(id=current_user.kurum_id, silindi=False).all()
        kullanicilar = User.query.filter_by(kurum_id=current_user.kurum_id, silindi=False).all()
        surecler = Surec.query.filter_by(kurum_id=current_user.kurum_id, silindi=False).all()
    
    current_app.logger.info(f'Admin panel loading for {current_user.username} (sistem_admin={is_system_admin}): {len(kullanicilar)} users, {len(kurumlar)} organizations')
    
    return render_template('admin_panel.html',
                         kurumlar=kurumlar,
                         kullanicilar=kullanicilar,
                         surecler=surecler,
                         is_system_admin=is_system_admin,
                         total_users=User.query.count(),
                         total_strategies=AnaStrateji.query.count())


@main_bp.route('/admin/download-user-template')
@login_required
def admin_download_user_template():
    """Kullanıcı içe aktarma şablonunu indir (XLSX)."""
    if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

    # Not: CSV yerine gerçek Excel şablonu indirilsin (UTF-8/Unicode sorunsuz).
    from openpyxl import Workbook
    from openpyxl.styles import Font

    # Kullanıcıların kolayca anlayacağı Türkçe başlıklar.
    # Rol/süreç rolü/kurum alanları yükleme sırasında sistem tarafından anlamlandırılacak.
    headers = ['Kullanıcı Adı', 'E-posta', 'Ad', 'Soyad', 'Şifre']
    sample_row = ['ahmet.yilmaz', 'ahmet.yilmaz@acme.com', 'Ahmet', 'Yılmaz', 'Parola123']

    wb = Workbook()
    ws = wb.active
    ws.title = 'kullanicilar'

    ws.append(headers)
    ws.append(sample_row)

    # Header styling
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = 'A2'

    # Reasonable column widths
    widths = {
        'A': 22,  # Kullanıcı Adı
        'B': 32,  # E-posta
        'C': 16,  # Ad
        'D': 16,  # Soyad
        'E': 16,  # Şifre
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='kullanici_sablonu.xlsx'
    )


@main_bp.route('/admin/upload-users-excel', methods=['POST'])
@login_required
def admin_upload_users_excel():
    """Toplu kullanıcı yükleme (Excel .xlsx)."""
    if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

    file = request.files.get('excel_file') or request.files.get('file') or request.files.get('excel')
    if not file or not getattr(file, 'filename', None):
        return jsonify({'success': False, 'message': 'Dosya bulunamadı', 'error_count': 0, 'success_count': 0, 'errors': []}), 400

    filename = (file.filename or '').lower().strip()
    if not filename.endswith('.xlsx'):
        return jsonify({
            'success': False,
            'message': 'Sadece .xlsx Excel dosyası destekleniyor.',
            'error_count': 1,
            'success_count': 0,
            'errors': ['Sadece .xlsx dosyası yükleyin (Excel Şablonu indir ile).']
        }), 400

    try:
        from openpyxl import load_workbook
        from werkzeug.security import generate_password_hash

        def normalize_header(value: object) -> str:
            text = '' if value is None else str(value)
            text = text.strip().lower()
            tr_map = str.maketrans({
                'ı': 'i',
                'ğ': 'g',
                'ü': 'u',
                'ş': 's',
                'ö': 'o',
                'ç': 'c',
            })
            text = text.translate(tr_map)
            # boşluk/punktuasyon sadeleştir
            for ch in [' ', '-', '_', '.', '/', '\\', ':', ';', ',', '(', ')', '[', ']', '{', '}', "'", '"']:
                text = text.replace(ch, '')
            return text

        # Excel'i oku
        wb = load_workbook(file.stream, data_only=True)
        ws = wb.active

        # Başlık satırı
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            return jsonify({'success': False, 'message': 'Excel başlık satırı bulunamadı.', 'error_count': 1, 'success_count': 0, 'errors': ['Excel boş görünüyor.']}), 400

        header_map = {}
        for idx, raw in enumerate(header_cells):
            key = normalize_header(raw)
            if not key:
                continue
            header_map[key] = idx

        # Beklenen kolonlar (Türkçe şablon + toleranslı eşleşmeler)
        aliases = {
            'username': ['kullaniciadi', 'kullanıcıadı', 'kullanici', 'username', 'user', 'kullaniciadi*'],
            'email': ['eposta', 'e-posta', 'mail', 'email', 'epostadresi'],
            'first_name': ['ad', 'isim', 'firstname', 'name'],
            'last_name': ['soyad', 'soyisim', 'lastname', 'surname'],
            'password': ['sifre', 'şifre', 'parola', 'password'],
        }

        def find_col(field: str):
            for alias in aliases[field]:
                alias_key = normalize_header(alias)
                if alias_key in header_map:
                    return header_map[alias_key]
            return None

        col_username = find_col('username')
        col_email = find_col('email')
        col_first = find_col('first_name')
        col_last = find_col('last_name')
        col_password = find_col('password')

        missing = []
        if col_username is None:
            missing.append('Kullanıcı Adı')
        if col_email is None:
            missing.append('E-posta')
        if col_first is None:
            missing.append('Ad')
        if col_last is None:
            missing.append('Soyad')
        if col_password is None:
            missing.append('Şifre')

        if missing:
            return jsonify({
                'success': False,
                'message': 'Excel başlıkları beklenen formatta değil.',
                'success_count': 0,
                'error_count': 1,
                'errors': [f"Eksik kolon(lar): {', '.join(missing)}"]
            }), 400

        success_count = 0
        error_count = 0
        errors = []

        # Satırları işle
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            def cell(col):
                if col is None:
                    return None
                return row[col] if col < len(row) else None

            username = cell(col_username)
            email = cell(col_email)
            first_name = cell(col_first)
            last_name = cell(col_last)
            password = cell(col_password)

            # boş satırları atla
            if all(v is None or str(v).strip() == '' for v in [username, email, first_name, last_name, password]):
                continue

            username = '' if username is None else str(username).strip()
            email = '' if email is None else str(email).strip().lower()
            first_name = '' if first_name is None else str(first_name).strip()
            last_name = '' if last_name is None else str(last_name).strip()
            password = '' if password is None else str(password).strip()

            row_errors = []
            if not username:
                row_errors.append('Kullanıcı Adı boş')
            if not email or '@' not in email:
                row_errors.append('E-posta geçersiz')
            if not first_name:
                row_errors.append('Ad boş')
            if not last_name:
                row_errors.append('Soyad boş')
            if not password:
                row_errors.append('Şifre boş')
            elif len(password) < 4:
                row_errors.append('Şifre çok kısa')

            if row_errors:
                error_count += 1
                if len(errors) < 50:
                    errors.append(f"Satır {row_idx}: " + '; '.join(row_errors))
                continue

            # Aynı username/email var mı?
            existing = User.query.filter(
                (User.username == username) | (User.email == email)
            ).first()
            if existing:
                error_count += 1
                if len(errors) < 50:
                    errors.append(f"Satır {row_idx}: Kullanıcı adı veya e-posta zaten kayıtlı")
                continue

            # Kurum ataması: en basit yorum - dosyayı yükleyen kullanıcının kurumu
            kurum_id = getattr(current_user, 'kurum_id', None)
            if not kurum_id:
                error_count += 1
                if len(errors) < 50:
                    errors.append(f"Satır {row_idx}: Kurum bilgisi bulunamadı (current_user.kurum_id boş)")
                continue

            user = User(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                password_hash=generate_password_hash(password),
                sistem_rol='kurum_kullanici',
                surec_rol=None,
                kurum_id=kurum_id,
            )
            db.session.add(user)
            success_count += 1

        if success_count > 0:
            db.session.commit()
        else:
            db.session.rollback()

        return jsonify({
            'success': True,
            'message': 'Excel içe aktarma tamamlandı.',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Excel işlenirken hata oluştu: {str(e)}',
            'success_count': 0,
            'error_count': 1,
            'errors': ['Excel okunamadı veya format hatalı. Şablonu yeniden indirip tekrar deneyin.']
        }), 400


@main_bp.route('/admin/upload-profile-photo', methods=['POST'])
@login_required
def admin_upload_profile_photo():
    """Admin tarafından kullanıcı profil fotoğrafı yükleme."""
    if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
    
    try:
        import uuid
        from werkzeug.utils import secure_filename
        
        if 'profile_photo' not in request.files:
            return jsonify({'success': False, 'message': 'Dosya seçilmedi!'}), 400
        
        file = request.files['profile_photo']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Dosya seçilmedi!'}), 400
        
        # Dosya tipini kontrol et
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'success': False, 'message': f'Geçersiz dosya tipi! İzin verilen tipler: {", ".join(allowed_extensions)}'}), 400
        
        # Dosya boyutu kontrolü (16MB)
        file.seek(0, 2)  # Dosya sonuna git
        file_size = file.tell()
        file.seek(0)  # Başa dön
        if file_size > 16 * 1024 * 1024:
            return jsonify({'success': False, 'message': 'Dosya boyutu 16MB\'dan büyük olamaz!'}), 400
        
        # Güvenli dosya adı oluştur
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        
        # Upload klasörünü oluştur
        upload_folder = os.path.join('static', 'uploads', 'profiles')
        os.makedirs(upload_folder, exist_ok=True)
        
        # Dosyayı kaydet
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        
        photo_url = f'/static/uploads/profiles/{unique_filename}'
        
        current_app.logger.info(f'Admin profil fotoğrafı yüklendi: {current_user.username} -> {unique_filename}')
        
        return jsonify({
            'success': True,
            'message': 'Profil fotoğrafı başarıyla yüklendi!',
            'photo_url': photo_url
        })
        
    except Exception as e:
        current_app.logger.error(f'Profil fotoğrafı yükleme hatası: {str(e)}')
        return jsonify({'success': False, 'message': f'Dosya yüklenirken hata oluştu: {str(e)}'}), 500


@main_bp.route('/admin/upload-logo', methods=['POST'])
@login_required
def admin_upload_logo():
    """Admin tarafından kurum logosu yükleme."""
    if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

    try:
        import uuid
        from werkzeug.utils import secure_filename

        if 'logo' not in request.files:
            return jsonify({'success': False, 'message': 'Dosya seçilmedi!'}), 400

        file = request.files['logo']
        if not file or file.filename == '':
            return jsonify({'success': False, 'message': 'Dosya seçilmedi!'}), 400

        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}
        if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions):
            return jsonify({'success': False, 'message': f'Geçersiz dosya tipi! İzin verilen tipler: {", ".join(sorted(allowed_extensions))}'}), 400

        # Dosya boyutu kontrolü (16MB)
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        if file_size > 16 * 1024 * 1024:
            return jsonify({'success': False, 'message': "Dosya boyutu 16MB'dan büyük olamaz!"}), 400

        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"

        upload_folder = os.path.join('static', 'uploads', 'logos')
        os.makedirs(upload_folder, exist_ok=True)

        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)

        logo_url = f'/static/uploads/logos/{unique_filename}'
        current_app.logger.info(f'Admin logo yüklendi: {current_user.username} -> {unique_filename}')

        return jsonify({
            'success': True,
            'message': 'Logo başarıyla yüklendi!',
            'logo_url': logo_url
        })
    except Exception as e:
        current_app.logger.error(f'Logo yükleme hatası: {str(e)}')
        return jsonify({'success': False, 'message': f'Dosya yüklenirken hata oluştu: {str(e)}'}), 500


@main_bp.route('/kurum-yonetim')
@login_required
def kurum_yonetim_page():
    """Kurum Yönetimi sayfası"""
    # Kurum yöneticileri için
    if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        flash('Bu sayfaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    try:
        # Basit bir kurum yönetimi sayfası
        kurum = Kurum.query.get(current_user.kurum_id)
        return render_template('base.html', 
                             page_title='Kurum Yönetimi',
                             content=f'<div class="container mt-4"><h2>Kurum Yönetimi</h2><p>Kurum: {kurum.kisa_ad if kurum else "Bilinmiyor"}</p><p>Kurum yönetimi sayfası yakında eklenecek.</p></div>')
    except Exception as e:
        import traceback
        current_app.logger.error(f'Kurum yönetimi sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500

# Proje yönetimi görünümleri: Kanban, Takvim, Gantt
@main_bp.route('/projeler/<int:project_id>/kanban')
@login_required
def project_kanban(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        tasks = Task.query.filter_by(project_id=project_id).all()
        return render_template('projects/kanban.html', project=project, tasks=tasks)
    except Exception as e:
        current_app.logger.error(f'Kanban sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/takvim')
@login_required
def project_calendar(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        tasks = Task.query.filter_by(project_id=project_id).all()
        return render_template('projects/calendar.html', project=project, tasks=tasks)
    except Exception as e:
        current_app.logger.error(f'Takvim sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/gantt')
@login_required
def project_gantt(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        tasks = Task.query.filter_by(project_id=project_id).all()
        return render_template('projects/gantt.html', project=project, tasks=tasks)
    except Exception as e:
        current_app.logger.error(f'Gantt sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/cpm')
@login_required
def project_cpm(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/cpm.html', project=project)
    except Exception as e:
        current_app.logger.error(f'CPM sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/raid')
@login_required
def project_raid(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/raid.html', project=project)
    except Exception as e:
        current_app.logger.error(f'RAID sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/baseline')
@login_required
def project_baseline(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/baseline.html', project=project)
    except Exception as e:
        current_app.logger.error(f'Baseline sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/kapasite')
@login_required
def project_capacity(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/capacity.html', project=project)
    except Exception as e:
        current_app.logger.error(f'Kapasite sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/integrations')
@login_required
def project_integrations(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/integrations.html', project=project)
    except Exception as e:
        current_app.logger.error(f'Integrations sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/kurallar')
@login_required
def project_rules(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/rules.html', project=project)
    except Exception as e:
        current_app.logger.error(f'Kurallar sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/sla')
@login_required
def project_sla(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/sla.html', project=project)
    except Exception as e:
        current_app.logger.error(f'SLA sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/tekrarlayan')
@login_required
def project_recurring(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/recurring.html', project=project)
    except Exception as e:
        current_app.logger.error(f'Recurring sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/calisma-gunleri')
@login_required
def project_workdays(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/workdays.html', project=project)
    except Exception as e:
        current_app.logger.error(f'Çalışma günleri sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/projeler/<int:project_id>/bagimlilik-matrisi')
@login_required
def project_dependency_matrix(project_id):
    try:
        project = Project.query.get_or_404(project_id)
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.dashboard'))
        return render_template('projects/dependency_matrix.html', project=project)
    except Exception as e:
        current_app.logger.error(f'Bağımlılık matrisi sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/portfoy-ozeti')
@login_required
def portfolio_summary():
    try:
        return render_template('projects/portfolio.html')
    except Exception as e:
        current_app.logger.error(f'Portföy özeti sayfası hatası: {e}')
        return f'Hata: {str(e)}', 500


@main_bp.route('/stratejik-planlama-akisi')
@login_required
def stratejik_planlama_akisi():
    """SP Akış Diyagramı sayfası"""
    try:
        kurum = current_user.kurum
        
        # Kurumun değerlerini getir
        degerler = Deger.query.filter_by(kurum_id=kurum.id).all()
        
        # Kurumun etik kurallarını getir
        etik_kurallari = EtikKural.query.filter_by(kurum_id=kurum.id).all()
        
        # Kurumun kalite politikalarını getir
        kalite_politikalari = KalitePolitikasi.query.filter_by(kurum_id=kurum.id).all()
        
        # Kurumun amaç ve vizyon bilgileri (kurum modelinden)
        amac = kurum.amac
        vizyon = kurum.vizyon
        
        # Kurumun ana stratejilerini getir (alt stratejileriyle birlikte)
        ana_stratejiler = AnaStrateji.query.filter_by(kurum_id=kurum.id).all()
        
        # Her ana strateji için alt stratejileri yükle
        for ana_str in ana_stratejiler:
            ana_str.alt_stratejiler = AltStrateji.query.filter_by(ana_strateji_id=ana_str.id).all()
        
        # Kurumun süreçlerini getir (performans göstergeleriyle birlikte)
        surecler = Surec.query.filter_by(kurum_id=kurum.id, silindi=False).all()
        
        # Her süreç için performans göstergelerini yükle
        for surec in surecler:
            surec.performans_gostergeleri = SurecPerformansGostergesi.query.filter_by(surec_id=surec.id).all()

        # --- Skor Hesapları (0-100 normalize) ---
        def _normalize_score(raw_value: int, max_value: int) -> int:
            if not max_value or max_value <= 0:
                return 0
            try:
                val = int(round((float(raw_value) / float(max_value)) * 100.0))
            except Exception:
                return 0
            return max(0, min(100, val))

        # Strateji-Süreç ilişki skorları (A=9, B=3)
        try:
            sub_strategies = AltStrateji.query.join(AnaStrateji).filter(
                AnaStrateji.kurum_id == kurum.id
            ).all()

            matrix_relations = StrategyProcessMatrix.query.join(AltStrateji).join(AnaStrateji).filter(
                AnaStrateji.kurum_id == kurum.id
            ).all()

            relations_map = {}
            for rel in matrix_relations:
                relations_map[(rel.sub_strategy_id, rel.process_id)] = int(rel.relationship_score or 0)

            # Süreç skorları: tüm alt strateji ilişkilerinin toplamı
            process_max_raw = 9 * len(sub_strategies)
            process_scores_norm = {}
            for proc in surecler:
                raw_total = 0
                for sub in sub_strategies:
                    raw_total += relations_map.get((sub.id, proc.id), 0)
                process_scores_norm[proc.id] = _normalize_score(raw_total, process_max_raw)

            # Alt strateji skorları: tüm süreç ilişkilerinin toplamı
            sub_max_raw = 9 * len(surecler)
            sub_scores_norm = {}
            sub_scores_raw = {}
            for sub in sub_strategies:
                raw_total = 0
                for proc in surecler:
                    raw_total += relations_map.get((sub.id, proc.id), 0)
                sub_scores_raw[sub.id] = raw_total
                sub_scores_norm[sub.id] = _normalize_score(raw_total, sub_max_raw)

            # Ana strateji skorları: alt stratejilerinin toplamı
            main_scores_norm = {}
            for main in ana_stratejiler:
                if not getattr(main, 'alt_stratejiler', None):
                    main_scores_norm[main.id] = 0
                    continue
                raw_total = sum(sub_scores_raw.get(sub.id, 0) for sub in (main.alt_stratejiler or []))
                main_max_raw = 9 * len(surecler) * len(main.alt_stratejiler)
                main_scores_norm[main.id] = _normalize_score(raw_total, main_max_raw)

            # KPI (PG) skorları: agirlikli_basari_puani (maks ~5.0) -> 0-100
            kpi_scores_norm = {}
            for proc in surecler:
                for kpi in (getattr(proc, 'performans_gostergeleri', None) or []):
                    if kpi.agirlikli_basari_puani is None:
                        continue
                    try:
                        kpi_scores_norm[kpi.id] = _normalize_score(float(kpi.agirlikli_basari_puani), 5.0)  # type: ignore[arg-type]
                    except Exception:
                        continue
        except Exception as _score_err:
            current_app.logger.warning(f'SP Akış skor hesaplama atlandı: {_score_err}')
            process_scores_norm = {}
            sub_scores_norm = {}
            main_scores_norm = {}
            kpi_scores_norm = {}
        
        return render_template('stratejik_planlama_akisi.html', 
                             kurum=kurum,
                             degerler=degerler,
                             etik_kurallari=etik_kurallari,
                             kalite_politikalari=kalite_politikalari,
                             amac=amac,
                             vizyon=vizyon,
                             ana_stratejiler=ana_stratejiler,
                             surecler=surecler,
                             process_scores_norm=process_scores_norm,
                             sub_scores_norm=sub_scores_norm,
                             main_scores_norm=main_scores_norm,
                             kpi_scores_norm=kpi_scores_norm)
    except Exception as e:
        import traceback
        current_app.logger.error(f'SP Akış Diyagramı sayfası render hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Template render hatası: {str(e)}", 500


@main_bp.route('/stratejik-planlama-akisi/dinamik')
@login_required
def stratejik_planlama_akisi_dinamik():
    """Dinamik SP Akış (graf) sayfası"""
    try:
        kurum = current_user.kurum
        return render_template('stratejik_planlama_akisi_dinamik.html', kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Dinamik SP Akış sayfası hatası: {e}')
        flash('Dinamik SP Akış sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.stratejik_planlama_akisi'))


@main_bp.route('/api/strategic-planning/graph')
@login_required
def api_strategic_planning_graph():
    """Dinamik SP akışı için graf verisini döndürür (nodes/edges + skorlar)."""
    try:
        kurum = current_user.kurum

        def _normalize_score(raw_value, max_value) -> int:
            try:
                if not max_value or float(max_value) <= 0:
                    return 0
                val = int(round((float(raw_value) / float(max_value)) * 100.0))
                return max(0, min(100, val))
            except Exception:
                return 0

        main_strategies = AnaStrateji.query.filter_by(kurum_id=kurum.id).order_by(AnaStrateji.code, AnaStrateji.ad).all()
        sub_strategies = AltStrateji.query.join(AnaStrateji).filter(
            AnaStrateji.kurum_id == kurum.id
        ).order_by(AltStrateji.code, AltStrateji.ad).all()
        processes = Surec.query.filter_by(kurum_id=kurum.id, silindi=False).order_by(Surec.weight.desc(), Surec.code, Surec.ad).all()
        projects = Project.query.filter_by(kurum_id=kurum.id, is_archived=False).order_by(Project.name).all()
        kpis = SurecPerformansGostergesi.query.join(Surec).filter(
            Surec.kurum_id == kurum.id
        ).order_by(Surec.code, SurecPerformansGostergesi.kodu, SurecPerformansGostergesi.ad).all()

        matrix_relations = StrategyProcessMatrix.query.join(AltStrateji).join(AnaStrateji).filter(
            AnaStrateji.kurum_id == kurum.id
        ).all()

        relations_map = {}
        for rel in matrix_relations:
            relations_map[(rel.sub_strategy_id, rel.process_id)] = int(rel.relationship_score or 0)

        # Totals
        process_max_raw = 9 * len(sub_strategies)
        process_totals_raw = {}
        process_scores_norm = {}
        for proc in processes:
            raw_total = 0
            for sub in sub_strategies:
                raw_total += relations_map.get((sub.id, proc.id), 0)
            process_totals_raw[proc.id] = raw_total
            process_scores_norm[proc.id] = _normalize_score(raw_total, process_max_raw)

        sub_max_raw = 9 * len(processes)
        sub_totals_raw = {}
        sub_scores_norm = {}
        for sub in sub_strategies:
            raw_total = 0
            for proc in processes:
                raw_total += relations_map.get((sub.id, proc.id), 0)
            sub_totals_raw[sub.id] = raw_total
            sub_scores_norm[sub.id] = _normalize_score(raw_total, sub_max_raw)

        main_scores_norm = {}
        # main->subs mapping
        subs_by_main = {}
        for sub in sub_strategies:
            subs_by_main.setdefault(sub.ana_strateji_id, []).append(sub)
        for main in main_strategies:
            subs = subs_by_main.get(main.id, [])
            if not subs:
                main_scores_norm[main.id] = 0
                continue
            raw_total = sum(sub_totals_raw.get(s.id, 0) for s in subs)
            main_max_raw = 9 * len(processes) * len(subs)
            main_scores_norm[main.id] = _normalize_score(raw_total, main_max_raw)

        # KPI score: agirlikli_basari_puani (0-5) => 0-100
        kpi_scores_norm = {}
        for kpi in kpis:
            if kpi.agirlikli_basari_puani is None:
                continue
            kpi_scores_norm[kpi.id] = _normalize_score(kpi.agirlikli_basari_puani, 5.0)

        # Nodes / edges for vis-network
        nodes = []
        edges = []
        kurum_panel_url = url_for('main.kurum_paneli')

        def _label_with_score(prefix: str, name: str, score: int) -> str:
            base = f"{prefix} {name}".strip()
            return f"{base}\n({score}%)"

        # Vision node (top level)
        if kurum.vizyon:
            nodes.append({
                'id': 'vision',
                'group': 'vision',
                'shape': 'box',
                'color': '#8b5cf6',
                'label': f"VİZYON\n{kurum.vizyon[:80]}...",
                'title': f"<b>Vizyon</b><br/>{kurum.vizyon}",
                'url': kurum_panel_url,
                'level': 0,
                'font': {'size': 18, 'bold': True}
            })

        # Main strategy nodes
        for main in main_strategies:
            score = main_scores_norm.get(main.id, 0)
            code = (main.code or '').strip()
            prefix = f"{code}" if code else "ANA"
            nodes.append({
                'id': f"main_{main.id}",
                'group': 'main_strategy',
                'shape': 'box',
                'color': '#f97316',
                'label': _label_with_score(prefix, main.ad, score),
                'title': f"<b>{main.ad}</b><br/>Skor: <b>{score}%</b>",
                'url': kurum_panel_url,
                'level': 1,
                'score': score,
            })

        # Sub strategy nodes
        for sub in sub_strategies:
            score = sub_scores_norm.get(sub.id, 0)
            code = (sub.code or '').strip()
            prefix = f"{code}" if code else "ALT"
            nodes.append({
                'id': f"sub_{sub.id}",
                'group': 'sub_strategy',
                'shape': 'box',
                'color': '#0891b2',
                'label': _label_with_score(prefix, sub.ad, score),
                'title': f"<b>{sub.ad}</b><br/>Skor: <b>{score}%</b>",
                'url': kurum_panel_url,
                'level': 2,
                'score': score,
            })
            # main -> sub edge
            edges.append({
                'from': f"main_{sub.ana_strateji_id}",
                'to': f"sub_{sub.id}",
                'arrows': 'to',
                'color': {'color': '#94a3b8'},
                'width': 1,
                'dashes': True,
            })

        # Vision to main strategies edges
        if kurum.vizyon:
            for main in main_strategies:
                edges.append({
                    'from': 'vision',
                    'to': f"main_{main.id}",
                    'arrows': 'to',
                    'color': {'color': '#c084fc'},
                    'width': 3,
                    'dashes': False,
                })

        # Process nodes
        for proc in processes:
            score = process_scores_norm.get(proc.id, 0)
            code = (proc.code or '').strip()
            prefix = f"{code}" if code else "SR"
            nodes.append({
                'id': f"proc_{proc.id}",
                'group': 'process',
                'shape': 'ellipse',
                'color': '#059669',
                'label': _label_with_score(prefix, proc.ad, score),
                'title': f"<b>{proc.ad}</b><br/>Skor: <b>{score}%</b>",
                'url': kurum_panel_url,
                'level': 3,
                'score': score,
            })

        # Sub -> Process edges (A+B)
        for sub in sub_strategies:
            for proc in processes:
                rel_score = relations_map.get((sub.id, proc.id), 0)
                if rel_score not in (3, 9):
                    continue
                label = 'A' if rel_score == 9 else 'B'
                edges.append({
                    'from': f"sub_{sub.id}",
                    'to': f"proc_{proc.id}",
                    'arrows': 'to',
                    'label': f"{label}",
                    'title': f"İlişki Skoru: {rel_score} ({label})",
                    'color': {'color': '#16a34a' if rel_score == 9 else '#f59e0b'},
                    'width': 3 if rel_score == 9 else 1,
                })

        # KPI nodes + Process -> KPI edges
        for kpi in kpis:
            score = kpi_scores_norm.get(kpi.id)
            score_txt = f" ({score}%)" if score is not None else ""
            label = (kpi.kodu or '').strip()
            if label:
                label = f"{label}: {kpi.ad}{score_txt}"
            else:
                label = f"PG: {kpi.ad}{score_txt}"
            nodes.append({
                'id': f"kpi_{kpi.id}",
                'group': 'kpi',
                'shape': 'box',
                'color': '#7c3aed',
                'label': label,
                'title': f"<b>{kpi.ad}</b><br/>Ağırlıklı Puan: <b>{(kpi.agirlikli_basari_puani if kpi.agirlikli_basari_puani is not None else '-')}</b><br/>Skor: <b>{(str(score) + '%' if score is not None else '-')}</b>",
                'url': kurum_panel_url,
                'level': 4,
                'score': score if score is not None else None,
            })
            edges.append({
                'from': f"proc_{kpi.surec_id}",
                'to': f"kpi_{kpi.id}",
                'arrows': 'to',
                'color': {'color': '#c4b5fd'},
                'width': 1,
            })

        # Project nodes + Project -> Process edges
        for project in projects:
            related_processes = list(project.related_processes or [])
            project_raw = sum(process_totals_raw.get(p.id, 0) for p in related_processes)
            project_max = 9 * len(sub_strategies) * len(related_processes)
            project_score = _normalize_score(project_raw, project_max)
            nodes.append({
                'id': f"proj_{project.id}",
                'group': 'project',
                'shape': 'diamond',
                'color': '#f97316',
                'label': _label_with_score('PRJ', project.name, project_score),
                'title': f"<b>{project.name}</b><br/>Skor: <b>{project_score}%</b>",
                'url': kurum_panel_url,
                'level': 4,
                'score': project_score,
            })

            for proc in related_processes:
                edges.append({
                    'from': f"proj_{project.id}",
                    'to': f"proc_{proc.id}",
                    'arrows': 'to',
                    'color': {'color': '#fb923c'},
                    'width': 2,
                })

            # Derived: Project -> SubStrategy edges (only if >0)
            if related_processes:
                max_rel = 9 * len(related_processes)
                for sub in sub_strategies:
                    raw_rel = 0
                    for proc in related_processes:
                        raw_rel += relations_map.get((sub.id, proc.id), 0)
                    if raw_rel <= 0:
                        continue
                    rel_norm = _normalize_score(raw_rel, max_rel)
                    edges.append({
                        'from': f"proj_{project.id}",
                        'to': f"sub_{sub.id}",
                        'arrows': 'to',
                        'label': f"{rel_norm}%",
                        'title': f"Projeden Alt Stratejiye Türetilmiş Etki: {rel_norm}%",
                        'color': {'color': '#a855f7'},
                        'width': 1,
                        'dashes': True,
                    })

        return jsonify({
            'success': True,
            'nodes': nodes,
            'edges': edges,
            'meta': {
                'kurum_id': kurum.id,
                'main_strategies': len(main_strategies),
                'sub_strategies': len(sub_strategies),
                'processes': len(processes),
                'projects': len(projects),
                'kpis': len(kpis),
            }
        })
    except Exception as e:
        import traceback
        current_app.logger.error(f'SP graph API hatası: {e}')
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/api/ai/insights')
@login_required
def api_ai_insights():
    """AI Insight'ları getir"""
    try:
        from services.ai_service import AIService
        insights = AIService.get_insights_for_user(current_user.id, current_user.kurum_id)
        return jsonify({
            'success': True,
            'insights': insights
        })
    except Exception as e:
        current_app.logger.error(f'AI insights hatası: {e}')
        return jsonify({
            'success': False,
            'message': 'Insight\'lar yüklenemedi',
            'insights': []
        }), 500


@main_bp.route('/api/ai/ask', methods=['POST'])
@login_required
def api_ai_ask():
    """AI'ya soru sor"""
    try:
        data = request.get_json()
        question = data.get('question', '').strip()
        
        if not question:
            return jsonify({
                'success': False,
                'message': 'Lütfen bir soru girin'
            }), 400
        
        from services.ai_service import AIService
        answer = AIService.ask_question(current_user.id, question)
        
        return jsonify({
            'success': True,
            'answer': answer['answer'],
            'suggestions': answer.get('suggestions', [])
        })
    except Exception as e:
        current_app.logger.error(f'AI ask hatası: {e}')
        return jsonify({
            'success': False,
            'message': 'Soru işlenemedi'
        }), 500

        return jsonify({
            'success': False,
            'message': 'Soru işlenemedi'
        }), 500

# ==========================================
# KURUM PANELİ CRUD İŞLEMLERİ
# ==========================================

# --- Ana Strateji ---

@main_bp.route('/kurum/ana-stratejiler/add', methods=['POST'])
@login_required
def add_ana_strateji():
    """Yeni ana strateji ekle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        # kurum_panel.html form alanları: name="ad", name="aciklama"
        ad = request.form.get('ana_strateji_ad') or request.form.get('ad')
        aciklama = request.form.get('ana_strateji_aciklama') or request.form.get('aciklama')
        
        if not ad:
            return jsonify({'success': False, 'message': 'Strateji adı zorunludur'}), 400
            
        yeni_strateji = AnaStrateji(
            kurum_id=current_user.kurum_id,
            ad=ad,
            aciklama=aciklama
        )
        
        db.session.add(yeni_strateji)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Ana strateji başarıyla eklendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/ana-stratejiler/update/<int:id>', methods=['POST'])
@login_required
def update_ana_strateji(id):
    """Ana stratejiyi güncelle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        strateji = AnaStrateji.query.get_or_404(id)
        
        if current_user.sistem_rol != 'admin' and strateji.kurum_id != current_user.kurum_id:
             return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        strateji.ad = request.form.get('ana_strateji_ad') or request.form.get('ad') or strateji.ad
        strateji.aciklama = request.form.get('ana_strateji_aciklama') or request.form.get('aciklama') or strateji.aciklama
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Ana strateji başarıyla güncellendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/ana-stratejiler/delete/<int:id>', methods=['POST'])
@login_required
def delete_ana_strateji(id):
    """Ana stratejiyi sil"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        strateji = AnaStrateji.query.get_or_404(id)
        
        if current_user.sistem_rol != 'admin' and strateji.kurum_id != current_user.kurum_id:
             return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        # Alt stratejileri de sil? Cascade delete DB seviyesinde yoksa manuel silmeli
        # Models'de cascade tanımlı mı kontrol etmek lazım ama şimdilik manuel siliyoruz
        AltStrateji.query.filter_by(ana_strateji_id=strateji.id).delete()
        
        db.session.delete(strateji)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Ana strateji başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# --- Alt Strateji ---

@main_bp.route('/kurum/alt-stratejiler/add', methods=['POST'])
@login_required
def add_alt_strateji():
    """Yeni alt strateji ekle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        # kurum_panel.html form alanları: name="ana_strateji_id", name="ad", name="aciklama"
        ana_strateji_id = request.form.get('alt_strateji_ana_id') or request.form.get('ana_strateji_id')
        ad = request.form.get('alt_strateji_ad') or request.form.get('ad')
        aciklama = request.form.get('alt_strateji_aciklama') or request.form.get('aciklama')
        
        if not ad or not ana_strateji_id:
            return jsonify({'success': False, 'message': 'Strateji adı ve ana strateji zorunludur'}), 400
            
        # Ana strateji kontrolü
        ana_strateji = AnaStrateji.query.get(ana_strateji_id)
        if not ana_strateji:
             return jsonify({'success': False, 'message': 'Ana strateji bulunamadı'}), 404
             
        if current_user.sistem_rol != 'admin' and ana_strateji.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403

        yeni_strateji = AltStrateji(
            ana_strateji_id=ana_strateji_id,
            ad=ad,
            aciklama=aciklama
        )
        
        db.session.add(yeni_strateji)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Alt strateji başarıyla eklendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/alt-stratejiler/update/<int:id>', methods=['POST'])
@login_required
def update_alt_strateji(id):
    """Alt stratejiyi güncelle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        strateji = AltStrateji.query.get_or_404(id)
        
        # Parent check
        if strateji.ana_strateji:
             if current_user.sistem_rol != 'admin' and strateji.ana_strateji.kurum_id != current_user.kurum_id:
                  return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        strateji.ad = request.form.get('alt_strateji_ad') or request.form.get('ad') or strateji.ad
        strateji.aciklama = request.form.get('alt_strateji_aciklama') or request.form.get('aciklama') or strateji.aciklama
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Alt strateji başarıyla güncellendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/alt-stratejiler/delete/<int:id>', methods=['POST'])
@login_required
def delete_alt_strateji(id):
    """Alt stratejiyi sil"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        strateji = AltStrateji.query.get_or_404(id)
        
        if strateji.ana_strateji:
             if current_user.sistem_rol != 'admin' and strateji.ana_strateji.kurum_id != current_user.kurum_id:
                  return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        db.session.delete(strateji)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Alt strateji başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# --- Amaç ve Vizyon ---

@main_bp.route('/kurum/update-amac-vizyon', methods=['POST'])
@login_required
def update_amac_vizyon():
    """Kurum amaç ve vizyonunu güncelle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        kurum = Kurum.query.get(current_user.kurum_id)
        if not kurum:
             return jsonify({'success': False, 'message': 'Kurum bulunamadı'}), 404
             
        kurum.amac = request.form.get('amac', kurum.amac)
        kurum.vizyon = request.form.get('vizyon', kurum.vizyon)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Amaç ve vizyon başarıyla güncellendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# --- Değerler (Values) ---

@main_bp.route('/kurum/degerler/add', methods=['POST'])
@login_required
def add_deger():
    """Yeni kurum değeri ekle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        baslik = request.form.get('baslik')
        aciklama = request.form.get('aciklama')
        
        if not baslik:
            return jsonify({'success': False, 'message': 'Başlık zorunludur'}), 400
            
        yeni_deger = Deger(
            kurum_id=current_user.kurum_id,
            baslik=baslik,
            aciklama=aciklama
        )
        
        db.session.add(yeni_deger)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Değer başarıyla eklendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/degerler/update/<int:id>', methods=['POST'])
@login_required
def update_deger(id):
    """Kurum değerini güncelle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        deger = Deger.query.get_or_404(id)
        
        # Sadece kendi kurumunun değerini güncelleyebilir (admin hariç)
        if current_user.sistem_rol != 'admin' and deger.kurum_id != current_user.kurum_id:
             return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        deger.baslik = request.form.get('baslik', deger.baslik)
        deger.aciklama = request.form.get('aciklama', deger.aciklama)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Değer başarıyla güncellendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/degerler/delete/<int:id>', methods=['POST'])
@login_required
def delete_deger(id):
    """Kurum değerini sil"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        deger = Deger.query.get_or_404(id)
        
        # Sadece kendi kurumunun değerini silebilir (admin hariç)
        if current_user.sistem_rol != 'admin' and deger.kurum_id != current_user.kurum_id:
             return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        db.session.delete(deger)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Değer başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# --- Etik Kurallar (Ethics) ---

@main_bp.route('/kurum/etik-kurallari/add', methods=['POST'])
@login_required
def add_etik_kural():
    """Yeni etik kural ekle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        # kurum_panel.html form alanları: name="baslik", name="aciklama"
        baslik = request.form.get('etik_kural_baslik') or request.form.get('baslik')
        aciklama = request.form.get('etik_kural_aciklama') or request.form.get('aciklama')
        
        if not baslik:
            return jsonify({'success': False, 'message': 'Başlık zorunludur'}), 400
            
        yeni_kural = EtikKural(
            kurum_id=current_user.kurum_id,
            baslik=baslik,
            aciklama=aciklama
        )
        
        db.session.add(yeni_kural)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Etik kural başarıyla eklendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/etik-kurallari/update/<int:id>', methods=['POST'])
@login_required
def update_etik_kural(id):
    """Etik kuralı güncelle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        kural = EtikKural.query.get_or_404(id)
        
        if current_user.sistem_rol != 'admin' and kural.kurum_id != current_user.kurum_id:
             return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        kural.baslik = request.form.get('etik_kural_baslik') or request.form.get('baslik') or kural.baslik
        kural.aciklama = request.form.get('etik_kural_aciklama') or request.form.get('aciklama') or kural.aciklama
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Etik kural başarıyla güncellendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/etik-kurallari/delete/<int:id>', methods=['POST'])
@login_required
def delete_etik_kural(id):
    """Etik kuralı sil"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        kural = EtikKural.query.get_or_404(id)
        
        if current_user.sistem_rol != 'admin' and kural.kurum_id != current_user.kurum_id:
             return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        db.session.delete(kural)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Etik kural başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# --- Kalite Politikaları (Quality Policy) ---

@main_bp.route('/kurum/kalite-politikalari/add', methods=['POST'])
@login_required
def add_kalite_politikasi():
    """Yeni kalite politikası ekle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        # kurum_panel.html form alanları: name="baslik", name="aciklama"
        baslik = request.form.get('kalite_politikasi_baslik') or request.form.get('baslik')
        aciklama = request.form.get('kalite_politikasi_aciklama') or request.form.get('aciklama')
        
        if not baslik:
            return jsonify({'success': False, 'message': 'Başlık zorunludur'}), 400
            
        yeni_politika = KalitePolitikasi(
            kurum_id=current_user.kurum_id,
            baslik=baslik,
            aciklama=aciklama
        )
        
        db.session.add(yeni_politika)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Kalite politikası başarıyla eklendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/kalite-politikalari/update/<int:id>', methods=['POST'])
@login_required
def update_kalite_politikasi(id):
    """Kalite politikasını güncelle"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        politika = KalitePolitikasi.query.get_or_404(id)
        
        if current_user.sistem_rol != 'admin' and politika.kurum_id != current_user.kurum_id:
             return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        politika.baslik = request.form.get('kalite_politikasi_baslik') or request.form.get('baslik') or politika.baslik
        politika.aciklama = request.form.get('kalite_politikasi_aciklama') or request.form.get('aciklama') or politika.aciklama
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Kalite politikası başarıyla güncellendi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@main_bp.route('/kurum/kalite-politikalari/delete/<int:id>', methods=['POST'])
@login_required
def delete_kalite_politikasi(id):
    """Kalite politikasını sil"""
    try:
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
            return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
            
        politika = KalitePolitikasi.query.get_or_404(id)
        
        if current_user.sistem_rol != 'admin' and politika.kurum_id != current_user.kurum_id:
             return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
             
        db.session.delete(politika)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Kalite politikası başarıyla silindi'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/ai-chat')
@login_required
def ai_chat():
    """AI Chat sayfası"""
    return render_template('ai_chat.html')


@main_bp.route('/stratejik-asistan')
@login_required
def stratejik_asistan():
    """Stratejik Asistan sayfası"""
    try:
        kurum = current_user.kurum
        return render_template('stratejik_asistan.html', kurum=kurum)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Stratejik Asistan sayfası render hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Template render hatası: {str(e)}", 500


@main_bp.route('/strategy/matrix')
@login_required
def strategy_matrix():
    """Strateji-Süreç Matrisi görüntüleme sayfası"""
    try:
        # Tüm alt stratejileri ana stratejileriyle birlikte çek
        sub_strategies = db.session.query(SubStrategy).join(MainStrategy).filter(
            MainStrategy.kurum_id == current_user.kurum_id
        ).order_by(MainStrategy.code, SubStrategy.code).all()
        
        # Tüm süreçleri çek (weight veya code sırasına göre)
        processes = Process.query.filter_by(kurum_id=current_user.kurum_id).order_by(
            Process.weight.desc(), Process.code
        ).all()
        
        # Tüm matris ilişkilerini çek
        matrix_relations = StrategyProcessMatrix.query.join(SubStrategy).join(MainStrategy).filter(
            MainStrategy.kurum_id == current_user.kurum_id
        ).all()
        
        # İlişkileri dictionary'ye çevir: relations_map[(sub_strategy_id, process_id)] = score
        relations_map = {}
        for relation in matrix_relations:
            key = (relation.sub_strategy_id, relation.process_id)
            relations_map[key] = relation.relationship_score
        
        # Ana stratejileri gruplamak için
        main_strategies = {}
        for sub_strategy in sub_strategies:
            main_id = sub_strategy.ana_strateji_id
            if main_id not in main_strategies:
                main_strategy = MainStrategy.query.get(main_id)
                main_strategies[main_id] = {
                    'strategy': main_strategy,
                    'subs': []
                }
            main_strategies[main_id]['subs'].append(sub_strategy)
        
        # Her süreç için toplam puan hesapla
        process_totals = {}
        for process in processes:
            total_score = 0
            for sub_strategy in sub_strategies:
                key = (sub_strategy.id, process.id)
                score = relations_map.get(key, 0)
                total_score += score
            process_totals[process.id] = total_score
        
        # En yüksek puana sahip ilk 3 süreci belirle (önem derecesi için)
        sorted_processes = sorted(process_totals.items(), key=lambda x: x[1], reverse=True)
        top_3_process_ids = {pid for pid, score in sorted_processes[:3]} if len(sorted_processes) >= 3 else set()
        
        return render_template('strategy/matrix.html',
                             sub_strategies=sub_strategies,
                             processes=processes,
                             relations_map=relations_map,
                             main_strategies=main_strategies,
                             process_totals=process_totals,
                             top_3_process_ids=top_3_process_ids)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Strateji Matrisi sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash(f'Matris görüntülenirken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/strategy/update_cell', methods=['POST'])
@login_required
@csrf.exempt  # AJAX için CSRF muafiyeti
def update_strategy_cell():
    """Strateji-Süreç matris hücresini güncelle (Toggle: Yok->9->3->Sil)"""
    try:
        data = request.get_json()
        sub_strategy_id = data.get('sub_strategy_id')
        process_id = data.get('process_id')
        
        if not sub_strategy_id or not process_id:
            return jsonify({'success': False, 'error': 'Eksik parametre'}), 400
        
        # Kullanıcının kurumuna ait strateji ve süreç olduğunu kontrol et
        sub_strategy = SubStrategy.query.join(MainStrategy).filter(
            SubStrategy.id == sub_strategy_id,
            MainStrategy.kurum_id == current_user.kurum_id
        ).first()
        
        process = Process.query.filter_by(
            id=process_id,
            kurum_id=current_user.kurum_id
        ).first()
        
        if not sub_strategy or not process:
            return jsonify({'success': False, 'error': 'Strateji veya süreç bulunamadı'}), 404
        
        # Mevcut ilişkiyi kontrol et
        relation = StrategyProcessMatrix.query.filter_by(
            sub_strategy_id=sub_strategy_id,
            process_id=process_id
        ).first()
        
        # Toggle mantığı: Yok -> 9 (A) -> 3 (B) -> Sil
        if not relation:
            # İlişki yoksa -> Puanı 9 (A) yap
            new_relation = StrategyProcessMatrix(
                sub_strategy_id=sub_strategy_id,
                process_id=process_id,
                relationship_score=9
            )
            db.session.add(new_relation)
            db.session.commit()
            return jsonify({
                'success': True,
                'new_score': 9,
                'text': 'A',
                'class': 'bg-success-subtle text-success-emphasis',
                'title': 'Puan: 9 (Güçlü İlişki - A)'
            })
        elif relation.relationship_score == 9:
            # Puan 9 ise -> Puanı 3 (B) yap
            relation.relationship_score = 3
            db.session.commit()
            return jsonify({
                'success': True,
                'new_score': 3,
                'text': 'B',
                'class': 'bg-warning-subtle text-warning-emphasis',
                'title': 'Puan: 3 (Zayıf İlişki - B)'
            })
        elif relation.relationship_score == 3:
            # Puan 3 ise -> İlişkiyi sil
            db.session.delete(relation)
            db.session.commit()
            return jsonify({
                'success': True,
                'new_score': None,
                'text': '-',
                'class': 'text-muted',
                'title': 'İlişki yok'
            })
        else:
            # Beklenmeyen puan değeri, 9'a sıfırla
            relation.relationship_score = 9
            db.session.commit()
            return jsonify({
                'success': True,
                'new_score': 9,
                'text': 'A',
                'class': 'bg-success-subtle text-success-emphasis',
                'title': 'Puan: 9 (Güçlü İlişki - A)'
            })
            
    except Exception as e:
        import traceback
        current_app.logger.error(f'Matris hücresi güncelleme hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/strategy/projects')
@login_required
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def strategy_projects():
    """Stratejik Proje Portföyü - Projeleri stratejik puana göre sırala"""
    try:
        # 1. Süreç puanlarını hesapla (strategy_matrix mantığı)
        sub_strategies = db.session.query(SubStrategy).join(MainStrategy).filter(
            MainStrategy.kurum_id == current_user.kurum_id
        ).all()
        
        processes = Process.query.filter_by(kurum_id=current_user.kurum_id).all()
        
        matrix_relations = StrategyProcessMatrix.query.join(SubStrategy).join(MainStrategy).filter(
            MainStrategy.kurum_id == current_user.kurum_id
        ).all()
        
        # İlişkileri dictionary'ye çevir
        relations_map = {}
        for relation in matrix_relations:
            key = (relation.sub_strategy_id, relation.process_id)
            relations_map[key] = relation.relationship_score
        
        # Her süreç için toplam puan hesapla
        process_totals = {}
        for process in processes:
            total_score = 0
            for sub_strategy in sub_strategies:
                key = (sub_strategy.id, process.id)
                score = relations_map.get(key, 0)
                total_score += score
            process_totals[process.id] = total_score
        
        # 2. Tüm projeleri çek
        projects = Project.query.filter_by(kurum_id=current_user.kurum_id, is_archived=False).all()
        
        # 3. Her proje için stratejik puan hesapla
        projects_with_scores = []
        for project in projects:
            # Projenin bağlı olduğu süreçlerin puanlarını topla
            project_score = 0
            related_process_names = []
            for process in project.related_processes:
                process_score = process_totals.get(process.id, 0)
                project_score += process_score
                related_process_names.append({
                    'name': process.ad,
                    'code': process.code,
                    'score': process_score
                })
            
            projects_with_scores.append({
                'project': project,
                'strategic_score': project_score,
                'related_processes': related_process_names
            })
        
        # 4. Puana göre büyükten küçüğe sırala
        projects_with_scores.sort(key=lambda x: x['strategic_score'], reverse=True)
        
        return render_template('strategy/project_portfolio.html',
                             projects_with_scores=projects_with_scores,
                             process_totals=process_totals)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Stratejik Proje Portföyü hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash(f'Proje portföyü görüntülenirken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/strategy/project/<int:id>')
@login_required
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def strategy_project_detail(id):
    """Stratejik Proje Detay Sayfası - Projenin stratejik uyum analizi"""
    try:
        # Projeyi çek
        project = Project.query.filter_by(id=id, kurum_id=current_user.kurum_id).first_or_404()
        
        # Süreç puanlarını hesapla (strategy_matrix mantığı)
        sub_strategies = db.session.query(SubStrategy).join(MainStrategy).filter(
            MainStrategy.kurum_id == current_user.kurum_id
        ).all()
        
        processes = Process.query.filter_by(kurum_id=current_user.kurum_id).all()
        
        matrix_relations = StrategyProcessMatrix.query.join(SubStrategy).join(MainStrategy).filter(
            MainStrategy.kurum_id == current_user.kurum_id
        ).all()
        
        # İlişkileri dictionary'ye çevir
        relations_map = {}
        for relation in matrix_relations:
            key = (relation.sub_strategy_id, relation.process_id)
            relations_map[key] = relation.relationship_score
        
        # Her süreç için toplam puan hesapla
        process_totals = {}
        for process in processes:
            total_score = 0
            for sub_strategy in sub_strategies:
                key = (sub_strategy.id, process.id)
                score = relations_map.get(key, 0)
                total_score += score
            process_totals[process.id] = total_score
        
        # Projenin bağlı süreçleri ve puanları
        project_processes = []
        total_strategic_score = 0
        strong_relations = 0  # A (9 puan)
        weak_relations = 0    # B (3 puan)
        
        for process in project.related_processes:
            process_score = process_totals.get(process.id, 0)
            total_strategic_score += process_score
            
            # Bu süreç için matris ilişkilerini say
            for sub_strategy in sub_strategies:
                key = (sub_strategy.id, process.id)
                score = relations_map.get(key, 0)
                if score == 9:
                    strong_relations += 1
                elif score == 3:
                    weak_relations += 1
            
            project_processes.append({
                'process': process,
                'score': process_score
            })
        
        # Süreçleri puana göre sırala
        project_processes.sort(key=lambda x: x['score'], reverse=True)
        
        # Mevcut bağlı süreç ID'lerini al (modal için)
        related_process_ids = [p.id for p in project.related_processes]
        
        return render_template('strategy/project_detail.html',
                             project=project,
                             project_processes=project_processes,
                             total_strategic_score=total_strategic_score,
                             strong_relations=strong_relations,
                             weak_relations=weak_relations,
                             processes=processes,
                             process_totals=process_totals,
                             related_process_ids=related_process_ids)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje detay sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash(f'Proje detayı görüntülenirken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.strategy_projects'))


@main_bp.route('/strategy/project/<int:id>/update_processes', methods=['POST'])
@login_required
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def strategy_project_update_processes(id):
    """Proje-Süreç ilişkilerini güncelle"""
    try:
        # Projeyi çek ve yetki kontrolü yap
        project = Project.query.filter_by(id=id, kurum_id=current_user.kurum_id).first_or_404()
        
        # Formdan gelen seçili süreç ID'lerini al
        selected_process_ids = request.form.getlist('process_ids')
        selected_process_ids = [int(pid) for pid in selected_process_ids if pid]
        
        # Süreçlerin kullanıcının kurumuna ait olduğunu kontrol et
        valid_processes = Process.query.filter_by(kurum_id=current_user.kurum_id).filter(
            Process.id.in_(selected_process_ids)
        ).all()
        
        # Mevcut ilişkileri temizle
        project.related_processes.clear()
        
        # Yeni seçilen süreçleri ekle
        for process in valid_processes:
            project.related_processes.append(process)
        
        # Değişiklikleri kaydet
        db.session.commit()
        
        flash('Süreç ilişkileri başarıyla güncellendi.', 'success')
        return redirect(url_for('main.strategy_project_detail', id=project.id))
        
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje süreç güncelleme hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        db.session.rollback()
        flash(f'Süreç ilişkileri güncellenirken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.strategy_project_detail', id=id))


@main_bp.route('/strategy/project/add', methods=['POST'])
@login_required
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def strategy_project_add():
    """Yeni proje oluştur"""
    try:
        # Form verilerini al
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        priority = request.form.get('priority', 'Orta')

        # Bildirim ayarları
        reminder_days_raw = (request.form.get('notification_reminder_days') or '').strip()
        overdue_frequency = (request.form.get('notification_overdue_frequency') or 'daily').strip().lower()
        notify_manager = request.form.get('notification_notify_manager') is not None
        notify_observers = request.form.get('notification_notify_observers') is not None
        channel_email = request.form.get('notification_channel_email') is not None

        reminder_days = [7, 3, 1]
        if reminder_days_raw:
            parts = [p.strip() for p in re.split(r"[;,\s]+", reminder_days_raw) if p.strip()]
            parsed_days = []
            for part in parts:
                if part.isdigit():
                    parsed_days.append(int(part))
            parsed_days = sorted(list(set([d for d in parsed_days if d >= 0])), reverse=True)
            if parsed_days:
                reminder_days = parsed_days

        if overdue_frequency not in ['daily', 'off']:
            overdue_frequency = 'daily'

        notification_settings = {
            'reminder_days': reminder_days,
            'overdue_frequency': overdue_frequency,
            'channels': {'in_app': True, 'email': channel_email},
            'notify_manager': notify_manager,
            'notify_observers': notify_observers,
        }
        
        # Validasyon
        if not name:
            flash('Proje adı zorunludur.', 'error')
            return redirect(url_for('main.strategy_projects'))
        
        # Tarihleri parse et
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Yeni proje oluştur
        new_project = Project(
            name=name,
            description=description if description else None,
            start_date=start_date,
            end_date=end_date,
            priority=priority,
            kurum_id=current_user.kurum_id,
            manager_id=current_user.id  # Varsayılan olarak oturum açan kullanıcıyı yönetici yap
        )

        try:
            new_project.notification_settings = json.dumps(notification_settings, ensure_ascii=False)
        except Exception:
            pass
        
        db.session.add(new_project)
        db.session.commit()
        
        flash('Proje başarıyla oluşturuldu.', 'success')
        return redirect(url_for('main.strategy_projects'))
        
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje oluşturma hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        db.session.rollback()
        flash(f'Proje oluşturulurken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.strategy_projects'))


@main_bp.route('/strategy/project/<int:id>/edit', methods=['POST'])
@login_required
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def strategy_project_edit(id):
    """Proje bilgilerini güncelle"""
    try:
        # Projeyi çek ve yetki kontrolü yap
        project = Project.query.filter_by(id=id, kurum_id=current_user.kurum_id).first_or_404()
        
        # Form verilerini al
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        priority = request.form.get('priority', 'Orta')

        # Bildirim ayarları
        reminder_days_raw = (request.form.get('notification_reminder_days') or '').strip()
        overdue_frequency = (request.form.get('notification_overdue_frequency') or '').strip().lower()
        notify_manager = request.form.get('notification_notify_manager') is not None
        notify_observers = request.form.get('notification_notify_observers') is not None
        channel_email = request.form.get('notification_channel_email') is not None

        reminder_days = None
        if reminder_days_raw:
            parts = [p.strip() for p in re.split(r"[;,\s]+", reminder_days_raw) if p.strip()]
            parsed_days = []
            for part in parts:
                if part.isdigit():
                    parsed_days.append(int(part))
            parsed_days = sorted(list(set([d for d in parsed_days if d >= 0])), reverse=True)
            if parsed_days:
                reminder_days = parsed_days

        if overdue_frequency and overdue_frequency not in ['daily', 'off']:
            overdue_frequency = ''
        
        # Validasyon
        if not name:
            flash('Proje adı zorunludur.', 'error')
            return redirect(url_for('main.strategy_projects'))
        
        # Tarihleri parse et
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Projeyi güncelle
        project.name = name
        project.description = description if description else None
        project.start_date = start_date
        project.end_date = end_date
        project.priority = priority

        # notification_settings merge
        try:
            current_settings = project.get_notification_settings()
            if reminder_days is not None:
                current_settings['reminder_days'] = reminder_days
            if overdue_frequency:
                current_settings['overdue_frequency'] = overdue_frequency
            current_settings['notify_manager'] = notify_manager
            current_settings['notify_observers'] = notify_observers
            current_settings.setdefault('channels', {})
            current_settings['channels']['in_app'] = True
            current_settings['channels']['email'] = channel_email
            project.notification_settings = json.dumps(current_settings, ensure_ascii=False)
        except Exception:
            pass
        
        db.session.commit()
        
        flash('Proje başarıyla güncellendi.', 'success')
        return redirect(url_for('main.strategy_projects'))
        
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje güncelleme hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        db.session.rollback()
        flash(f'Proje güncellenirken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.strategy_projects'))


@main_bp.route('/strategy/project/<int:id>/delete', methods=['POST'])
@login_required
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def strategy_project_delete(id):
    """Projeyi sil"""
    try:
        # Projeyi çek ve yetki kontrolü yap
        project = Project.query.filter_by(id=id, kurum_id=current_user.kurum_id).first_or_404()
        
        project_name = project.name
        
        # İlişkili süreçleri temizle
        project.related_processes.clear()
        
        # Projeyi sil
        db.session.delete(project)
        db.session.commit()
        
        flash(f'"{project_name}" projesi başarıyla silindi.', 'success')
        return redirect(url_for('main.strategy_projects'))
        
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje silme hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        db.session.rollback()
        flash(f'Proje silinirken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.strategy_projects'))


@main_bp.route('/strategy/project/<int:id>/clone', methods=['POST'])
@login_required
@role_required(['admin', 'kurum_yoneticisi', 'ust_yonetim'])
def strategy_project_clone(id):
    """Projeyi klonla (kopyala)"""
    try:
        # Kaynak projeyi çek ve yetki kontrolü yap
        source_project = Project.query.filter_by(id=id, kurum_id=current_user.kurum_id).first_or_404()
        
        # Form verilerini al
        new_name = request.form.get('name', '').strip()
        copy_processes = request.form.get('copy_processes') == '1'
        copy_description = request.form.get('copy_description') == '1'
        copy_priority = request.form.get('copy_priority') == '1'
        
        # Validasyon
        if not new_name:
            flash('Yeni proje adı zorunludur.', 'error')
            return redirect(url_for('main.strategy_projects'))
        
        # Yeni proje oluştur
        new_project = Project(
            name=new_name,
            description=source_project.description if copy_description else None,
            priority=source_project.priority if copy_priority else 'Orta',
            kurum_id=current_user.kurum_id,
            manager_id=current_user.id,
            start_date=date.today(),  # Bugünün tarihi
            end_date=None  # Bitiş tarihi boş bırakılabilir veya hesaplanabilir
        )
        
        # Eğer kaynak projenin bitiş tarihi varsa, süreyi hesapla ve yeni bitiş tarihini ayarla
        if source_project.start_date and source_project.end_date:
            duration = (source_project.end_date - source_project.start_date).days
            new_project.end_date = date.today() + timedelta(days=duration)
        
        db.session.add(new_project)
        db.session.flush()  # ID'yi almak için
        
        # Süreç ilişkilerini kopyala
        if copy_processes:
            for process in source_project.related_processes:
                new_project.related_processes.append(process)
        
        db.session.commit()
        
        flash(f'Proje başarıyla kopyalandı: "{new_name}"', 'success')
        return redirect(url_for('main.strategy_projects'))
        
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje klonlama hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        db.session.rollback()
        flash(f'Proje kopyalanırken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.strategy_projects'))


@main_bp.route('/strategy/kpis')
@login_required
def strategy_kpis():
    """KPI Yönetim Paneli - Performans Göstergeleri Listesi"""
    try:
        # Tüm KPI'ları çek
        kpis = SurecPerformansGostergesi.query.join(Surec).filter(
            Surec.kurum_id == current_user.kurum_id
        ).order_by(Surec.code, SurecPerformansGostergesi.ad).all()
        
        # Süreçleri de çek (modal için)
        processes = Process.query.filter_by(kurum_id=current_user.kurum_id).order_by(Process.code).all()
        
        # Her KPI için başarı oranı hesapla
        kpis_with_scores = []
        total_kpis = 0
        achieved_count = 0
        critical_count = 0
        
        for kpi in kpis:
            total_kpis += 1
            
            # Hedef ve gerçekleşen değerleri parse et
            target_value = None
            actual_value = None
            
            if kpi.hedef_deger:
                try:
                    target_value = float(kpi.hedef_deger)
                except (ValueError, TypeError):
                    pass
            
            # Gerçekleşen değer için PerformansGostergeVeri tablosundan en son değeri al
            # Not: SurecPerformansGostergesi süreç bazlı, BireyselPerformansGostergesi bireysel
            # Şimdilik basit yaklaşım: hedef_deger'ı kullan, gerçek veri girişi için ayrı bir mekanizma gerekebilir
            # İleride SurecPerformansGostergesi için özel bir veri tablosu oluşturulabilir
            
            # Başarı oranı hesapla
            success_rate = None
            if target_value and target_value > 0:
                # Şimdilik gerçekleşen değer yoksa başarı oranı hesaplanamaz
                # İleride PerformansGostergeVeri veya yeni bir tablo üzerinden çekilebilir
                success_rate = None
            elif target_value == 0:
                success_rate = None  # Hedef sıfır ise hesaplama yapma
            
            # Kritik kontrolü (başarı oranı %70'in altındaysa kritik)
            is_critical = False
            if success_rate is not None and success_rate < 70:
                critical_count += 1
                is_critical = True
            elif success_rate is not None and success_rate >= 100:
                achieved_count += 1
            
            kpis_with_scores.append({
                'kpi': kpi,
                'target_value': target_value,
                'actual_value': actual_value,
                'success_rate': success_rate,
                'is_critical': is_critical
            })
        
        # Özet istatistikler
        summary_stats = {
            'total': total_kpis,
            'achieved': achieved_count,
            'critical': critical_count,
            'pending': total_kpis - achieved_count - critical_count if total_kpis > 0 else 0
        }
        
        return render_template('strategy/kpi_dashboard.html',
                             kpis_with_scores=kpis_with_scores,
                             summary_stats=summary_stats,
                             processes=processes)
    except Exception as e:
        import traceback
        current_app.logger.error(f'KPI Dashboard hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash(f'KPI paneli görüntülenirken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/strategy/kpi/add', methods=['POST'])
@login_required
def strategy_kpi_add():
    """Yeni KPI ekle"""
    try:
        # Form verilerini al
        name = request.form.get('name', '').strip()
        process_id = request.form.get('process_id')
        unit = request.form.get('unit', '').strip()
        target_value_str = request.form.get('target_value', '').strip()
        direction = request.form.get('direction', 'Artan')
        description = request.form.get('description', '').strip()
        
        # Validasyon
        if not name:
            flash('Gösterge adı zorunludur.', 'error')
            return redirect(url_for('main.strategy_kpis'))
        
        if not process_id:
            flash('Süreç seçimi zorunludur.', 'error')
            return redirect(url_for('main.strategy_kpis'))
        
        # Sürecin kullanıcının kurumuna ait olduğunu kontrol et
        process = Process.query.filter_by(id=int(process_id), kurum_id=current_user.kurum_id).first()
        if not process:
            flash('Geçersiz süreç seçimi.', 'error')
            return redirect(url_for('main.strategy_kpis'))
        
        # Hedef değeri parse et (String olarak saklanacak)
        target_value_str_final = target_value_str if target_value_str else None
        
        # Direction'ı model formatına çevir
        direction_model = 'Increasing' if direction == 'Artan' else 'Decreasing'
        
        # Yeni KPI oluştur
        new_kpi = SurecPerformansGostergesi(
            ad=name,
            surec_id=int(process_id),
            birim=unit if unit else None,
            olcum_birimi=unit if unit else None,  # Eski alan
            unit=unit if unit else None,  # V3.0 alanı
            hedef_deger=target_value_str_final,
            direction=direction_model,  # V3.0 alanı (Increasing/Decreasing)
            aciklama=description if description else None
        )
        
        db.session.add(new_kpi)
        db.session.commit()
        
        flash('KPI başarıyla oluşturuldu.', 'success')
        return redirect(url_for('main.strategy_kpis'))
        
    except Exception as e:
        import traceback
        current_app.logger.error(f'KPI oluşturma hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        db.session.rollback()
        flash(f'KPI oluşturulurken hata oluştu: {str(e)}', 'error')
        return redirect(url_for('main.strategy_kpis'))


# ============================================================================
# PROJE YÖNETİMİ ROTALARI
# ============================================================================

@main_bp.route('/projeler')
@login_required
def projeler():
    """Proje listesi sayfası - Kart görünümü"""
    from sqlalchemy.orm import joinedload
    
    try:
        # Pagination parametreleri
        page = request.args.get('page', 1, type=int)
        per_page = 20  # Sayfa başına proje sayısı
        
        # Kullanıcının kurumundaki tüm projeleri getir (eager loading ile N+1 çözümü)
        pagination = Project.query.options(
            joinedload(Project.manager),
            joinedload(Project.related_processes)
        ).filter_by(
            kurum_id=current_user.kurum_id
        ).paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        projeler = pagination.items
        
        return render_template('project_list.html', 
                             projeler=projeler,
                             pagination=pagination)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje listesi sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/projeler/yeni')
@login_required
def proje_yeni():
    """Yeni proje oluşturma sayfası"""
    try:
        # Kurumdaki tüm süreçleri getir
        surecler = Surec.query.all()
        
        # Kurumdaki tüm kullanıcıları getir
        kullanicilar = User.query.all()
        
        # Şablon olarak kullanılabilecek projeleri getir (aynı kurumdaki projeler)
        sablon_projeler = Project.query.filter_by(kurum_id=current_user.kurum_id).order_by(Project.created_at.desc()).limit(20).all()
        
        return render_template('project_form.html',
                             project=None,
                             surecler=surecler,
                             kullanicilar=kullanicilar,
                             sablon_projeler=sablon_projeler)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Yeni proje sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/projeler/<int:project_id>/duzenle')
@login_required
def proje_duzenle(project_id):
    """Proje düzenleme sayfası"""
    try:
        project = Project.query.get_or_404(project_id)

        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.projeler'))

        user_project_role = _get_user_project_role_for_page(project)
        if user_project_role not in ['manager', 'member']:
            flash('Bu projeyi düzenleme yetkiniz yok.', 'danger')
            return redirect(url_for('main.proje_detay', project_id=project_id))

        surecler = Surec.query.all()
        kullanicilar = User.query.all()
        sablon_projeler = Project.query.filter_by(
            kurum_id=current_user.kurum_id
        ).order_by(Project.created_at.desc()).limit(20).all()

        return render_template(
            'project_form.html',
            project=project,
            surecler=surecler,
            kullanicilar=kullanicilar,
            sablon_projeler=sablon_projeler
        )
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje düzenleme sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/projeler/<int:project_id>')
@login_required
def proje_detay(project_id):
    """Proje detay sayfası - Kanban ve Liste görünümü"""
    from sqlalchemy.orm import joinedload
    
    try:
        project = Project.query.options(
            joinedload(Project.manager),
            joinedload(Project.members)
        ).get_or_404(project_id)
        
        # Yetki kontrolü - Proje kullanıcının kurumunda olmalı
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.projeler'))

        user_project_role = _get_user_project_role_for_page(project)
        if user_project_role is None:
            flash('Bu projede yetkiniz yok.', 'danger')
            return redirect(url_for('main.projeler'))
        
        # Projenin görevlerini getir (eager loading ile assigned_to ve reporter)
        tasks = Task.query.options(
            joinedload(Task.assignee),
            joinedload(Task.reporter)
        ).filter_by(project_id=project_id).order_by(Task.created_at.desc()).all()
        
        # Durumlara göre görevleri grupla (Kanban için)
        def _normalized_status(task):
            return normalize_task_status(task.status) or task.status

        tasks_by_status = {
            'Yapılacak': [t for t in tasks if _normalized_status(t) == 'Yapılacak'],
            'Devam Ediyor': [t for t in tasks if _normalized_status(t) == 'Devam Ediyor'],
            'Beklemede': [t for t in tasks if _normalized_status(t) == 'Beklemede'],
            'Tamamlandı': [t for t in tasks if _normalized_status(t) == 'Tamamlandı']
        }
        
        # Gecikme analizi için bugünün tarihini al
        from datetime import date
        today = date.today()
        
        # Geciken görevleri işaretle
        geciken_gorevler = [
            t for t in tasks
            if t.due_date and t.due_date < today and _normalized_status(t) != 'Tamamlandı'
        ]
        
        return render_template('project_detail.html', 
                             project=project, 
                             tasks=tasks,
                             tasks_by_status=tasks_by_status,
                             today=today,
                             geciken_gorevler=geciken_gorevler,
                             user_project_role=user_project_role)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje detay sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/projeler/<int:project_id>/gorevler/<int:task_id>')
@login_required
def gorev_detay(project_id, task_id):
    """Görev detay/duzenleme sayfası"""
    try:
        project = Project.query.get_or_404(project_id)
        task = Task.query.get_or_404(task_id)
        
        # Yetki kontrolü
        if project.kurum_id != current_user.kurum_id or task.project_id != project_id:
            flash('Bu göreve erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.proje_detay', project_id=project_id))

        user_project_role = _get_user_project_role_for_page(project)
        if user_project_role is None:
            flash('Bu projede yetkiniz yok.', 'danger')
            return redirect(url_for('main.proje_detay', project_id=project_id))

        can_edit = user_project_role in ['manager', 'member']
        
        # Kullanıcının bireysel PG'lerini getir (form için)
        bireysel_pgler_rows = BireyselPerformansGostergesi.query.filter_by(
            user_id=current_user.id
        ).with_entities(
            BireyselPerformansGostergesi.id,
            BireyselPerformansGostergesi.ad,
            BireyselPerformansGostergesi.kodu,
        ).all()
        bireysel_pgler = [
            {'id': pg_id, 'ad': ad, 'kodu': kodu}
            for pg_id, ad, kodu in bireysel_pgler_rows
        ]
        
        # Kullanıcının bireysel faaliyetlerini getir (form için)
        bireysel_faaliyetler_rows = BireyselFaaliyet.query.filter_by(
            user_id=current_user.id
        ).with_entities(
            BireyselFaaliyet.id,
            BireyselFaaliyet.ad,
        ).all()
        bireysel_faaliyetler = [
            {'id': faaliyet_id, 'ad': ad}
            for faaliyet_id, ad in bireysel_faaliyetler_rows
        ]
        
        # Görevin mevcut impact'lerini getir
        task_impacts = TaskImpact.query.filter_by(task_id=task_id).all()
        
        # Görevin yorumlarını getir
        task_comments = TaskComment.query.filter_by(task_id=task_id).order_by(TaskComment.created_at.desc()).all()
        
        # Proje üyelerini ve kurum kullanıcılarını getir (görev atama için)
        proje_uyeleri = project.members if project.members else []
        kurum_kullanicilar = User.query.filter_by(kurum_id=project.kurum_id).all()
        
        return render_template('task_form.html',
                             project=project,
                             task=task,
                             bireysel_pgler=bireysel_pgler,
                             bireysel_faaliyetler=bireysel_faaliyetler,
                             task_impacts=task_impacts,
                             task_comments=task_comments,
                             kullanicilar=kurum_kullanicilar,
                             proje_uyeleri=proje_uyeleri,
                             can_edit=can_edit,
                             user_project_role=user_project_role)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Görev detay sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/projeler/<int:project_id>/gorevler/yeni')
@login_required
def gorev_yeni(project_id):
    """Yeni görev oluşturma sayfası"""
    try:
        project = Project.query.get_or_404(project_id)
        
        # Yetki kontrolü
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projede görev oluşturma yetkiniz yok.', 'danger')
            return redirect(url_for('main.projeler'))

        user_project_role = _get_user_project_role_for_page(project)
        if user_project_role not in ['manager', 'member']:
            flash('Bu projede görev oluşturma yetkiniz yok.', 'danger')
            return redirect(url_for('main.proje_detay', project_id=project_id))
        
        # Kullanıcının bireysel PG'lerini ve faaliyetlerini getir
        bireysel_pgler_rows = BireyselPerformansGostergesi.query.filter_by(
            user_id=current_user.id
        ).with_entities(
            BireyselPerformansGostergesi.id,
            BireyselPerformansGostergesi.ad,
            BireyselPerformansGostergesi.kodu,
        ).all()
        bireysel_pgler = [
            {'id': pg_id, 'ad': ad, 'kodu': kodu}
            for pg_id, ad, kodu in bireysel_pgler_rows
        ]

        bireysel_faaliyetler_rows = BireyselFaaliyet.query.filter_by(
            user_id=current_user.id
        ).with_entities(
            BireyselFaaliyet.id,
            BireyselFaaliyet.ad,
        ).all()
        bireysel_faaliyetler = [
            {'id': faaliyet_id, 'ad': ad}
            for faaliyet_id, ad in bireysel_faaliyetler_rows
        ]
        
        # Proje üyelerini ve kurum kullanıcılarını getir (görev atama için)
        proje_uyeleri = project.members if project.members else []
        kurum_kullanicilar = User.query.filter_by(kurum_id=project.kurum_id).all()
        
        return render_template('task_form.html',
                             project=project,
                             task=None,  # Yeni görev
                             bireysel_pgler=bireysel_pgler,
                             bireysel_faaliyetler=bireysel_faaliyetler,
                             task_impacts=[],
                             kullanicilar=kurum_kullanicilar,
                             proje_uyeleri=proje_uyeleri,
                             can_edit=True,
                             user_project_role=user_project_role)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Yeni görev sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/dokuman-merkezi')
@login_required
def dokuman_merkezi():
    """Kurumsal dosya yönetimi sayfası"""
    try:
        # Kategorilere göre dosyaları getir
        from models import ProjectFile
        from sqlalchemy import or_, and_
        
        # Kurumsal dosyaları getir (scope='CORPORATE' veya project_id NULL)
        corporate_files = ProjectFile.query.filter(
            or_(
                ProjectFile.scope == 'CORPORATE',
                and_(ProjectFile.scope == 'PROJECT', ProjectFile.project_id.is_(None))
            ),
            ProjectFile.is_active == True
        ).order_by(ProjectFile.category, ProjectFile.file_name).all()
        
        # Kategorilere göre grupla
        files_by_category = {}
        for file in corporate_files:
            category = file.category or 'Diğer'
            if category not in files_by_category:
                files_by_category[category] = []
            files_by_category[category].append(file)
        
        # Kullanıcı yetkisi kontrolü (yönetici mi?)
        can_upload = current_user.sistem_rol in ['admin', 'kurum_yoneticisi']
        
        return render_template('corporate_files.html',
                             files_by_category=files_by_category,
                             can_upload=can_upload)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Doküman merkezi sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/dashboard/executive')
@login_required
def executive_dashboard():
    """Yönetim Kokpiti - Executive Dashboard"""
    try:
        # Yetki kontrolü - Sadece yönetici ve gözlemci erişebilir
        if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim', 'gözlemci']:
            return "Bu sayfaya erişim yetkiniz yok", 403
        
        from services.strategic_impact_service import get_strategic_impact_summary

        strategic_impact = get_strategic_impact_summary(current_user.kurum_id)

        return render_template('executive_dashboard.html',
                             strategic_impact=strategic_impact)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Executive dashboard sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/projeler/<int:project_id>/gantt')
@login_required
def proje_gantt(project_id):
    """Proje Gantt şeması görünümü"""
    try:
        project = Project.query.get_or_404(project_id)
        
        # Yetki kontrolü
        if project.kurum_id != current_user.kurum_id:
            flash('Bu projeye erişim yetkiniz yok.', 'danger')
            return redirect(url_for('main.projeler'))
        
        # Projenin görevlerini getir (predecessor ilişkileriyle birlikte)
        tasks = Task.query.filter_by(project_id=project_id).all()
        
        # Gantt verisi için görevleri hazırla
        gantt_data = []
        for task in tasks:
            start_date = task.start_date or task.due_date
            end_date = task.due_date or task.start_date
            gantt_item = {
                'id': task.id,
                'name': task.title,
                'start': start_date.strftime('%Y-%m-%d') if start_date else None,
                'end': end_date.strftime('%Y-%m-%d') if end_date else None,
                'progress': 0,
                'dependencies': []
            }
            
            # Tamamlanmış görevler için progress %100
            normalized_status = normalize_task_status(task.status) or task.status
            if normalized_status == 'Tamamlandı':
                gantt_item['progress'] = 100
            elif normalized_status == 'Devam Ediyor':
                gantt_item['progress'] = 50
            elif normalized_status == 'Beklemede':
                gantt_item['progress'] = 25
            
            # Predecessor'ları (bağımlılıkları) ekle
            if task.predecessors:
                for predecessor in task.predecessors:
                    gantt_item['dependencies'].append(str(predecessor.id))
            
            # Başlangıç tarihi yoksa varsayılan bir tarih ekle
            if not gantt_item['start']:
                gantt_item['start'] = (datetime.now().date() - timedelta(days=30)).strftime('%Y-%m-%d')
                gantt_item['end'] = datetime.now().date().strftime('%Y-%m-%d')
            
            gantt_data.append(gantt_item)
        
        return render_template('project_gantt.html',
                             project=project,
                             tasks=tasks,
                             gantt_data=gantt_data)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje Gantt sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"Sayfa yüklenirken hata oluştu: {str(e)}", 500


@main_bp.route('/proje-analitik')
@login_required
def proje_analitik():
    """Proje Analitik sayfası - Süreç sağlık skoru ve analitik raporlar"""
    try:
        from services.project_analytics import calculate_surec_saglik_skoru
        from models import Surec
        
        # Kullanıcının erişebileceği süreçleri getir
        if current_user.sistem_rol == 'admin':
            surecler = Surec.query.all()
        elif current_user.sistem_rol == 'kurum_yoneticisi':
            surecler = Surec.query.all()
        else:
            # Kullanıcının üye/lider olduğu süreçler
            surecler = Surec.query.join(surec_liderleri).filter(
                surec_liderleri.c.user_id == current_user.id
            ).union(
                Surec.query.join(surec_uyeleri).filter(
                    surec_uyeleri.c.user_id == current_user.id
                )
            ).all()
        
        # Her süreç için sağlık skorunu hesapla
        surec_skorlari = []
        for surec in surecler:
            skor = calculate_surec_saglik_skoru(surec.id)
            if skor:
                surec_skorlari.append({
                    'surec': surec,
                    'skor': skor
                })
        
        # Skorlara göre sırala (yüksekten düşüğe)
        surec_skorlari.sort(key=lambda x: x['skor']['skor'], reverse=True)
        
        return render_template('proje_analitik.html',
                             surec_skorlari=surec_skorlari)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Proje Analitik sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash('Analitik sayfası yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/zaman-takibi')
@login_required
def zaman_takibi():
    """Zaman Takibi (Timesheet) sayfası"""
    try:
        from models import TimeEntry, Task, Project
        from datetime import datetime, timedelta
        
        # Varsayılan tarih aralığı: Bu hafta
        bugun = datetime.now().date()
        hafta_basi = bugun - timedelta(days=bugun.weekday())
        hafta_sonu = hafta_basi + timedelta(days=6)
        
        # Kullanıcının time entry'lerini getir
        entries = TimeEntry.query.filter_by(user_id=current_user.id).filter(
            TimeEntry.start_time >= datetime.combine(hafta_basi, datetime.min.time()),
            TimeEntry.start_time <= datetime.combine(hafta_sonu, datetime.max.time())
        ).order_by(TimeEntry.start_time.desc()).all()
        
        # Aktif time entry var mı?
        active_entry = TimeEntry.query.filter_by(
            user_id=current_user.id,
            end_time=None
        ).first()
        
        # Toplam süre hesapla
        toplam_dakika = sum(entry.duration_minutes or 0 for entry in entries if entry.duration_minutes)
        toplam_saat = toplam_dakika / 60.0
        
        return render_template('zaman_takibi.html',
                             entries=entries,
                             active_entry=active_entry,
                             toplam_saat=toplam_saat,
                             hafta_basi=hafta_basi,
                             hafta_sonu=hafta_sonu)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Zaman Takibi sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash('Zaman takibi sayfası yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/gorev-aktivite-log')
@login_required
def gorev_aktivite_log():
    """Görev Aktivite Log sayfası - Görev değişiklik geçmişi"""
    try:
        from models import TaskActivity, Task, Project
        
        # Kullanıcının erişebileceği projeler
        if current_user.sistem_rol == 'admin':
            projects = Project.query.all()
        else:
            projects = Project.query.filter_by(kurum_id=current_user.kurum_id).all()
        
        project_ids = [p.id for p in projects]
        
        # Bu projelerdeki görevlerin aktivitelerini getir
        task_ids = [t.id for t in Task.query.filter(Task.project_id.in_(project_ids)).all()]
        
        activities = TaskActivity.query.filter(
            TaskActivity.task_id.in_(task_ids)
        ).order_by(TaskActivity.created_at.desc()).limit(100).all()
        
        return render_template('gorev_aktivite_log.html',
                             activities=activities)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Görev Aktivite Log sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash('Aktivite log sayfası yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/sistem-degisiklik-gunlugu')
@login_required
def sistem_degisiklik_gunlugu():
    """Sistem Değişiklik Günlüğü - Audit Trail UI"""
    # Sadece yöneticiler erişebilir
    if current_user.sistem_rol not in ['admin', 'kurum_yoneticisi', 'ust_yonetim']:
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('main.dashboard'))
    
    try:
        # Filtreleme parametreleri
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        islem_tipi = request.args.get('islem_tipi', '')
        user_id = request.args.get('user_id', 0, type=int)
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Base query
        query = PerformansGostergeVeriAudit.query
        
        # Filtreleme
        if islem_tipi:
            query = query.filter(PerformansGostergeVeriAudit.islem_tipi == islem_tipi)
        if user_id:
            query = query.filter(PerformansGostergeVeriAudit.islem_yapan_user_id == user_id)
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(PerformansGostergeVeriAudit.islem_tarihi >= start_dt)
            except:
                pass
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                end_dt = end_dt.replace(hour=23, minute=59, second=59)
                query = query.filter(PerformansGostergeVeriAudit.islem_tarihi <= end_dt)
            except:
                pass
        
        # Sıralama (en yeni önce)
        query = query.order_by(PerformansGostergeVeriAudit.islem_tarihi.desc())
        
        # Pagination
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        audit_logs = pagination.items
        
        # Kullanıcı listesi (filtre için)
        users = User.query.filter_by(kurum_id=current_user.kurum_id).order_by(User.first_name).all()
        
        # İşlem tipleri
        islem_tipleri = ['OLUSTUR', 'GUNCELLE', 'SIL']
        
        return render_template(
            'sistem_degisiklik_gunlugu.html',
            audit_logs=audit_logs,
            pagination=pagination,
            users=users,
            islem_tipleri=islem_tipleri,
            current_filters={
                'islem_tipi': islem_tipi,
                'user_id': user_id,
                'start_date': start_date,
                'end_date': end_date
            }
        )
    except Exception as e:
        current_app.logger.error(f'Audit log sayfası hatası: {e}', exc_info=True)
        flash('Sistem değişiklik günlüğü yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/akilli-planlama')
@login_required
def akilli_planlama():
    """Akıllı Planlama sayfası - Geciken görevler için otomatik tarih güncelleme"""
    try:
        from models import Task, Project
        from datetime import date
        
        # Kullanıcının erişebileceği projeler
        if current_user.sistem_rol == 'admin':
            projects = Project.query.all()
        else:
            projects = Project.query.filter_by(kurum_id=current_user.kurum_id).all()
        
        project_ids = [p.id for p in projects]
        
        # Geciken görevleri bul
        bugun = date.today()
        geciken_gorevler = Task.query.filter(
            Task.project_id.in_(project_ids),
            Task.status.notin_(COMPLETED_STATUSES),
            Task.due_date < bugun,
            Task.due_date.isnot(None)
        ).order_by(Task.due_date.asc()).all()
        
        return render_template('akilli_planlama.html',
                             geciken_gorevler=geciken_gorevler,
                             bugun=bugun)
    except Exception as e:
        import traceback
        current_app.logger.error(f'Akıllı Planlama sayfası hatası: {str(e)}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash('Akıllı planlama sayfası yüklenirken hata oluştu.', 'danger')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/setup_test_pg_automation')
@login_required
def setup_test_pg_automation():
    """Test için PG otomasyonu hazırlama route'u (Geçici)"""
    try:
        # 1. İlk görevi bul
        task = Task.query.first()
        
        if not task:
            return "❌ Görev bulunamadı. Önce bir görev oluşturun."
        
        # 2. İlk bireysel PG'yi bul (veya oluştur)
        pg = BireyselPerformansGostergesi.query.filter_by(user_id=current_user.id).first()
        
        if not pg:
            # Demo PG oluştur
            pg = BireyselPerformansGostergesi(
                user_id=current_user.id,
                ad="Test Performans Göstergesi (Otomasyon)",
                aciklama="Otomasyon testi için oluşturuldu",
                hedef_deger="100",
                olcum_birimi="Adet",
                periyot="Aylik",
                kaynak="Bireysel"
            )
            db.session.add(pg)
            db.session.flush()
        
        # 3. Görevi ölçülebilir yap ve PG'ye bağla
        task.is_measurable = True
        task.planned_output_value = 100.0
        task.related_indicator_id = pg.id
        
        db.session.commit()
        
        return f"""
        ✅ Test görevi hazırlandı!<br><br>
        <strong>Görev ID:</strong> {task.id}<br>
        <strong>Görev Adı:</strong> {task.title}<br>
        <strong>PG ID:</strong> {pg.id}<br>
        <strong>PG Adı:</strong> {pg.ad}<br>
        <strong>Planlanan Değer:</strong> {task.planned_output_value}<br><br>
        <strong>Şimdi yapılacaklar:</strong><br>
        1. Proje detay sayfasına gidin<br>
        2. Bu görevi tamamlayın (Bitir butonuna tıklayın)<br>
        3. Otomatik olarak PG verisi oluşturulacak!<br><br>
        <a href="/projeler/{task.project_id}">Proje Detay Sayfasına Git</a>
        """
        
    except Exception as e:
        db.session.rollback()
        import traceback
        current_app.logger.error(f'Test setup hatası: {e}')
        current_app.logger.error(f'Traceback: {traceback.format_exc()}')
        return f"❌ Hata: {str(e)}<br><br>Traceback:<br><pre>{traceback.format_exc()}</pre>"


@main_bp.route('/debug/schema_check')
@login_required
def debug_schema_check():
    """Veritabanı şema kontrolü - Task tablosu sütunlarını kontrol et"""
    try:
        from sqlalchemy import inspect
        
        # Task modelini inspect et
        inspector = inspect(db.engine)
        columns = [col['name'] for col in inspector.get_columns('task')]
        
        # Kontrol edilecek sütunlar
        required_columns = ['is_measurable', 'planned_output_value', 'related_indicator_id']
        
        result = []
        result.append("=== VERITABANI SEMA KONTROLU ===\n\n")
        result.append(f"Task tablosu toplam {len(columns)} sütun içeriyor.\n\n")
        
        all_present = True
        for col_name in required_columns:
            if col_name in columns:
                result.append(f"✅ {col_name}: MEVCUT\n")
            else:
                result.append(f"❌ {col_name}: EKSIK\n")
                all_present = False
        
        result.append("\n=== SONUC ===\n")
        if all_present:
            result.append("✅ Tüm sütunlar mevcut! Otomasyon için hazır.\n")
        else:
            result.append("❌ Bazı sütunlar eksik. Migration gerekli.\n")
        
        result.append("\n=== TUM SUTUNLAR ===\n")
        result.append(", ".join(columns))
        
        return "<pre>" + "".join(result) + "</pre>"
        
    except Exception as e:
        import traceback
        return f"<pre>❌ Hata: {str(e)}\n\nTraceback:\n{traceback.format_exc()}</pre>"


@main_bp.route('/debug/monitor')
@login_required
def debug_monitor():
    """Canlı izleme paneli - Son görevler ve PG verileri"""
    try:
        # Son 5 görev
        last_tasks = Task.query.order_by(Task.created_at.desc()).limit(5).all()
        
        # Son 5 PG verisi (otomasyonla oluşanlar)
        last_pg_veriler = PerformansGostergeVeri.query.order_by(PerformansGostergeVeri.created_at.desc()).limit(5).all()
        
        html = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Debug Monitor - Sistem İzleme</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
                .container { max-width: 1200px; margin: 0 auto; }
                h1 { color: #333; }
                table { width: 100%; border-collapse: collapse; background: white; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
                th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
                th { background-color: #4CAF50; color: white; }
                tr:hover { background-color: #f5f5f5; }
                .badge { padding: 4px 8px; border-radius: 4px; font-size: 12px; }
                .badge-success { background: #28a745; color: white; }
                .badge-danger { background: #dc3545; color: white; }
                .badge-warning { background: #ffc107; color: black; }
                .badge-info { background: #17a2b8; color: white; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>🔍 Sistem Tanı ve İzleme Paneli</h1>
                <p><strong>Son Güncelleme:</strong> """ + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + """</p>
                
                <h2>📋 Son 5 Görev</h2>
                <table>
                    <thead>
                        <tr>
                            <th>ID</th>
                            <th>Adı</th>
                            <th>Durum</th>
                            <th>Ölçülebilir?</th>
                            <th>Planlanan Değer</th>
                            <th>Bağlı PG ID</th>
                            <th>Tamamlanma Tarihi</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        for task in last_tasks:
            is_measurable_badge = '<span class="badge badge-success">Evet</span>' if task.is_measurable else '<span class="badge badge-danger">Hayır</span>'
            status_badge = f'<span class="badge badge-{"success" if task.status == "Tamamlandı" else "warning"}">{task.status}</span>'
            pg_id = task.related_indicator_id if task.related_indicator_id else '-'
            planned_value = task.planned_output_value if task.planned_output_value is not None else '-'
            completed = task.completed_at.strftime('%Y-%m-%d %H:%M') if task.completed_at else '-'
            
            html += f"""
                        <tr>
                            <td>{task.id}</td>
                            <td>{task.title}</td>
                            <td>{status_badge}</td>
                            <td>{is_measurable_badge}</td>
                            <td>{planned_value}</td>
                            <td>{pg_id}</td>
                            <td>{completed}</td>
                        </tr>
            """
        
        html += """
                    </tbody>
                </table>
                
                <h2>📊 Son 5 PG Verisi</h2>
                <table>
                    <thead>
                        <tr>
                            <th>ID</th>
                            <th>Tarih</th>
                            <th>Değer</th>
                            <th>Açıklama</th>
                            <th>PG ID</th>
                            <th>Kullanıcı</th>
                            <th>Oluşturulma</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        for pg_veri in last_pg_veriler:
            is_auto = 'Otomatik' in (pg_veri.aciklama or '')
            auto_badge = '<span class="badge badge-info">Otomatik</span>' if is_auto else '<span class="badge">Manuel</span>'
            user_name = pg_veri.user.first_name + ' ' + pg_veri.user.last_name if pg_veri.user else '-'
            created = pg_veri.created_at.strftime('%Y-%m-%d %H:%M') if pg_veri.created_at else '-'
            
            html += f"""
                        <tr>
                            <td>{pg_veri.id}</td>
                            <td>{pg_veri.veri_tarihi.strftime('%Y-%m-%d') if pg_veri.veri_tarihi else '-'}</td>
                            <td>{pg_veri.gerceklesen_deger}</td>
                            <td>{auto_badge} {pg_veri.aciklama[:50] if pg_veri.aciklama else '-'}</td>
                            <td>{pg_veri.bireysel_pg_id}</td>
                            <td>{user_name}</td>
                            <td>{created}</td>
                        </tr>
            """
        
        html += """
                    </tbody>
                </table>
                
                <p><a href="/debug/schema_check">🔍 Şema Kontrolü</a> | 
                   <a href="/dashboard">🏠 Dashboard</a></p>
            </div>
        </body>
        </html>
        """
        
        return html
        
    except Exception as e:
        import traceback
        return f"<pre>❌ Hata: {str(e)}\n\nTraceback:\n{traceback.format_exc()}</pre>"


@main_bp.route('/debug/force_trigger/<int:task_id>')
@login_required
def debug_force_trigger(task_id):
    """Manuel otomasyon tetikleme testi"""
    try:
        from datetime import date
        
        # Görevi bul
        task = Task.query.get_or_404(task_id)
        
        # Görevi ölçülebilir yap ve değerleri set et
        task.is_measurable = True
        task.planned_output_value = 50.0
        
        # PG ID'yi kontrol et - eğer yoksa ilk PG'yi bul veya oluştur
        if not task.related_indicator_id:
            pg = BireyselPerformansGostergesi.query.filter_by(user_id=current_user.id).first()
            if not pg:
                # Demo PG oluştur
                pg = BireyselPerformansGostergesi(
                    user_id=current_user.id,
                    ad="Test Performans Göstergesi (Manuel Tetikleme)",
                    aciklama="Manuel tetikleme testi için oluşturuldu",
                    hedef_deger="100",
                    olcum_birimi="Adet",
                    periyot="Aylik",
                    kaynak="Bireysel"
                )
                db.session.add(pg)
                db.session.flush()
            task.related_indicator_id = pg.id
        
        db.session.flush()
        
        # Otomasyon mantığını manuel çalıştır
        result = {
            'success': False,
            'task_id': task_id,
            'task_title': task.title,
            'is_measurable': task.is_measurable,
            'planned_output_value': task.planned_output_value,
            'related_indicator_id': task.related_indicator_id,
            'pg_created': False,
            'pg_veri_id': None,
            'error': None
        }
        
        if task.is_measurable and task.related_indicator_id:
            try:
                # İlişkili PG'yi kontrol et
                related_pg = BireyselPerformansGostergesi.query.get(task.related_indicator_id)
                if related_pg:
                    # Yeni performans değeri kaydı oluştur
                    today = date.today()
                    
                    # Değer hesapla
                    output_value = task.planned_output_value if task.planned_output_value is not None else 1.0
                    
                    # PerformansGostergeVeri kaydı oluştur
                    new_pg_veri = PerformansGostergeVeri(
                        bireysel_pg_id=task.related_indicator_id,
                        yil=today.year,
                        veri_tarihi=today,
                        giris_periyot_tipi='gunluk',
                        giris_periyot_no=today.day,
                        giris_periyot_ay=today.month,
                        ay=today.month,
                        gun=today.day,
                        gerceklesen_deger=str(output_value),
                        aciklama=f"Otomatik: {task.title} tamamlandı. (Manuel Tetikleme)",
                        user_id=current_user.id,
                        created_by=current_user.id,
                        updated_by=current_user.id
                    )
                    db.session.add(new_pg_veri)
                    db.session.flush()
                    
                    result['pg_created'] = True
                    result['pg_veri_id'] = new_pg_veri.id
                    result['success'] = True
                    result['message'] = f'PG verisi başarıyla oluşturuldu (ID: {new_pg_veri.id})'
                else:
                    result['error'] = f'İlişkili PG bulunamadı (ID: {task.related_indicator_id})'
            except Exception as pg_error:
                result['error'] = f'PG verisi oluşturulurken hata: {str(pg_error)}'
        else:
            result['error'] = 'Görev ölçülebilir değil veya PG ID yok'
        
        db.session.commit()
        
        return jsonify(result)
        
    except Exception as e:
        db.session.rollback()
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@main_bp.route('/debug/fix_and_reset')
@login_required
def debug_fix_and_reset():
    """Test ortamını sıfırla ve eksik referansları tamamla"""
    try:
        result_messages = []
        
        # Adım 1: İndikatör Kontrolü (BireyselPerformansGostergesi ID=1)
        pg = BireyselPerformansGostergesi.query.get(1)
        if not pg:
            # ID=1 olan PG yoksa oluştur
            # Önce mevcut kullanıcının PG'sini kontrol et
            existing_pg = BireyselPerformansGostergesi.query.filter_by(
                user_id=current_user.id,
                ad="Test KPI"
            ).first()
            
            if existing_pg:
                # Mevcut PG'yi kullan
                pg = existing_pg
                result_messages.append(f"ℹ️ Mevcut Test KPI kullanılıyor (ID: {pg.id})")
            else:
                # Yeni PG oluştur
                pg = BireyselPerformansGostergesi(
                    user_id=current_user.id,
                    ad="Test KPI",
                    aciklama="Test ortamı için otomatik oluşturuldu",
                    hedef_deger="100",
                    olcum_birimi="Adet",
                    periyot="Aylik",
                    kaynak="Bireysel",
                    durum="Devam Ediyor"
                )
                db.session.add(pg)
                db.session.flush()
                result_messages.append(f"✅ İndikatör oluşturuldu (ID: {pg.id}): 'Test KPI'")
        else:
            result_messages.append(f"ℹ️ İndikatör (ID:1) zaten mevcut: '{pg.ad}'")
        
        # Adım 2: Görevi Sıfırla (Task ID=1)
        task = Task.query.get(1)
        if not task:
            # Görev 1 yoksa, ilk görevi bul veya oluştur
            task = Task.query.first()
            if not task:
                # Hiç görev yoksa, bir proje bul ve görev oluştur
                project = Project.query.first()
                if project:
                    task = Task(
                        project_id=project.id,
                        title="Test Görevi (Otomasyon Testi)",
                        description="Otomasyon testi için oluşturuldu",
                        status="Yapılacak",
                        priority="Orta"
                    )
                    db.session.add(task)
                    db.session.flush()
                    result_messages.append("✅ Yeni test görevi oluşturuldu (ID: {})".format(task.id))
                else:
                    return "<pre>❌ Hata: Hiç proje bulunamadı. Önce bir proje oluşturun.</pre>"
            else:
                result_messages.append(f"ℹ️ Görev 1 bulunamadı, ilk görev kullanılıyor (ID: {task.id})")
        
        # Görevi sıfırla
        task.status = 'Yapılacak'
        task.completed_at = None
        task.is_measurable = True
        task.planned_output_value = 50.0
        task.related_indicator_id = pg.id  # ID=1 olan PG'ye bağla
        
        result_messages.append(f"✅ Görev (ID: {task.id}) sıfırlandı:")
        result_messages.append(f"   - Durum: 'Yapılacak'")
        result_messages.append(f"   - Tamamlanma Tarihi: None")
        result_messages.append(f"   - Ölçülebilir: True")
        result_messages.append(f"   - Planlanan Değer: 50.0")
        result_messages.append(f"   - Bağlı PG ID: {pg.id}")
        
        # Adım 3: Kaydet
        db.session.commit()
        
        result_messages.append("\n✅ Sistem sıfırlandı ve hazır!")
        result_messages.append(f"\n📝 Şimdi yapılacaklar:")
        result_messages.append(f"   1. Proje detay sayfasına gidin: /projeler/{task.project_id}")
        result_messages.append(f"   2. Görev (ID: {task.id}) 'Bitir' butonuna tıklayın")
        result_messages.append(f"   3. Otomatik olarak PG verisi oluşturulacak!")
        
        return "<pre>" + "\n".join(result_messages) + "</pre>"
        
    except Exception as e:
        db.session.rollback()
        import traceback
        return f"<pre>❌ Hata: {str(e)}\n\nTraceback:\n{traceback.format_exc()}</pre>"


@main_bp.route('/debug/init_strategy_db')
@login_required
def debug_init_strategy_db():
    """Stratejik Planlama V3.0 veritabanı migration - Yeni tabloları ve ilişkileri oluşturur"""
    try:
        from models import CorporateIdentity, AnaStrateji, AltStrateji, Surec, SurecPerformansGostergesi, Project
        
        result_messages = []
        result_messages.append("=== STRATEJİK PLANLAMA V3.0 VERİTABANI MİGRATİON ===\n")
        
        # Tüm tabloları oluştur
        db.create_all()
        result_messages.append("✅ Tüm tablolar oluşturuldu (veya zaten mevcut)")
        
        # Yeni tabloları kontrol et
        inspector = db.inspect(db.engine)
        existing_tables = inspector.get_table_names()
        
        # CorporateIdentity tablosu kontrolü
        if 'corporate_identity' in existing_tables:
            result_messages.append("✅ corporate_identity tablosu mevcut")
        else:
            result_messages.append("⚠️ corporate_identity tablosu bulunamadı (create_all çalıştırıldı)")
        
        # Association table'ları kontrol et
        if 'process_owners' in existing_tables:
            result_messages.append("✅ process_owners association table mevcut")
        else:
            result_messages.append("⚠️ process_owners association table bulunamadı")
        
        if 'strategy_process_matrix' in existing_tables:
            result_messages.append("✅ strategy_process_matrix association table mevcut")
        else:
            result_messages.append("⚠️ strategy_process_matrix association table bulunamadı")
        
        # Mevcut tablolardaki yeni kolonları kontrol et (SQLite için ALTER TABLE gerekebilir)
        result_messages.append("\n=== YENİ KOLONLAR KONTROLÜ ===")
        result_messages.append("⚠️ Not: SQLite'da ALTER TABLE ile kolon ekleme manuel yapılmalıdır.")
        result_messages.append("Aşağıdaki kolonlar eklenmeli:")
        result_messages.append("  - ana_strateji: code, name")
        result_messages.append("  - alt_strateji: code, name")
        result_messages.append("  - surec: code, name, weight")
        result_messages.append("  - surec_performans_gostergesi: calculation_method, target_method, unit, direction")
        
        result_messages.append("\n✅ Migration tamamlandı!")
        result_messages.append("\n📝 Sonraki Adımlar:")
        result_messages.append("  1. SQLite için ALTER TABLE komutlarını çalıştırın")
        result_messages.append("  2. Veya yeni bir migration scripti oluşturun")
        
        return "<pre>" + "\n".join(result_messages) + "</pre>"
        
    except Exception as e:
        db.session.rollback()
        import traceback
        return f"<pre>❌ Hata: {str(e)}\n\nTraceback:\n{traceback.format_exc()}</pre>"


@main_bp.route('/debug/init_strategy_v3')
@login_required
def debug_init_strategy_v3():
    """Stratejik Planlama V3.0 veritabanı initialization - Excel yapısına göre tabloları oluşturur"""
    try:
        from models import (
            CorporateIdentity, AnaStrateji, AltStrateji, Surec, 
            SurecPerformansGostergesi, Project
        )
        
        result_messages = []
        result_messages.append("=== STRATEJİK PLANLAMA V3.0 VERİTABANI INITIALIZATION ===\n")
        result_messages.append("Excel: SP VE SÜREÇ YAPISI dosyasına göre yapılandırılıyor...\n")
        
        # Tüm tabloları oluştur
        db.create_all()
        result_messages.append("✅ Tüm tablolar oluşturuldu (veya zaten mevcut)")
        
        # Yeni tabloları kontrol et
        inspector = db.inspect(db.engine)
        existing_tables = inspector.get_table_names()
        
        # CorporateIdentity tablosu kontrolü
        if 'corporate_identity' in existing_tables:
            result_messages.append("✅ corporate_identity tablosu mevcut (Excel: Misyon, Vizyon, Değerler)")
        else:
            result_messages.append("⚠️ corporate_identity tablosu bulunamadı (create_all çalıştırıldı)")
        
        # Association table'ları kontrol et
        if 'process_owners' in existing_tables:
            result_messages.append("✅ process_owners association table mevcut (Çoklu Süreç Sahipliği)")
        else:
            result_messages.append("⚠️ process_owners association table bulunamadı")
        
        if 'strategy_process_matrix' in existing_tables:
            result_messages.append("✅ strategy_process_matrix association table mevcut (Excel: SP - Süreç Matrisi)")
        else:
            result_messages.append("⚠️ strategy_process_matrix association table bulunamadı")
        
        # Mevcut tablolardaki yeni kolonları kontrol et
        result_messages.append("\n=== YENİ KOLONLAR KONTROLÜ ===")
        result_messages.append("⚠️ Not: SQLite'da ALTER TABLE ile kolon ekleme manuel yapılmalıdır.")
        result_messages.append("Aşağıdaki kolonlar eklenmeli:")
        result_messages.append("  - corporate_identity: YENİ TABLO (vizyon, misyon, kalite_politikasi, degerler)")
        result_messages.append("  - ana_strateji: code (UNIQUE), name")
        result_messages.append("  - alt_strateji: code, name, target_method")
        result_messages.append("  - surec: code, name, weight")
        result_messages.append("  - surec_performans_gostergesi: calculation_method, target_method, unit, direction")
        result_messages.append("  - strategy_process_matrix: relationship_score (A=9, B=3)")
        
        result_messages.append("\n=== EXCEL YAPISI EŞLEŞTİRMESİ ===")
        result_messages.append("✅ CorporateIdentity ↔ Excel: 'Misyon, Vizyon, Değerler' sayfası")
        result_messages.append("✅ AnaStrateji ↔ Excel: 'ST1, ST2' yapıları")
        result_messages.append("✅ AltStrateji ↔ Excel: 'ST1.1' yapıları (target_method ile)")
        result_messages.append("✅ Surec ↔ Excel: 'KMF Süreçler' sayfası (code, name, weight)")
        result_messages.append("✅ StrategyProcessMatrix ↔ Excel: 'SP - Süreç Matrisi' (A=9, B=3)")
        result_messages.append("✅ SurecPerformansGostergesi ↔ Excel: KPI yapısı (calculation_method, target_method)")
        
        result_messages.append("\n✅ V3.0 Mimari Kurulumu Tamamlandı!")
        result_messages.append("\n📝 Sonraki Adımlar:")
        result_messages.append("  1. SQLite için ALTER TABLE komutlarını çalıştırın")
        result_messages.append("  2. Excel verilerini import edin")
        result_messages.append("  3. CRUD endpoint'lerini kullanarak veri girişi yapın")
        
        return "<pre>" + "\n".join(result_messages) + "</pre>"
        
    except Exception as e:
        db.session.rollback()
        import traceback
        return f"<pre>❌ Hata: {str(e)}\n\nTraceback:\n{traceback.format_exc()}</pre>"


# ============================================
# FAZ 2: OPERASYONEL MÜKEMMELLİK ROUTE'LARI
# ============================================

# V59.0 - Hoshin Kanri Paketi
@main_bp.route('/okr/<int:objective_id>/comment', methods=['POST'])
@login_required
def okr_comment(objective_id):
    """OKR/Hedef yorum ekleme - Hoshin Catchball"""
    try:
        data = request.get_json() if request.is_json else request.form
        comment_text = data.get('comment_text', '').strip()
        comment_type = data.get('comment_type', 'feedback')
        
        if not comment_text:
            return jsonify({'success': False, 'message': 'Yorum metni boş olamaz'}), 400
        
        # AltStrateji kontrolü
        objective = AltStrateji.query.filter_by(id=objective_id).first_or_404()
        
        # Yetki kontrolü - kullanıcı aynı kurumda olmalı
        if objective.ana_strateji.kurum_id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Yetkisiz erişim'}), 403
        
        comment = ObjectiveComment(
            objective_id=objective_id,
            user_id=current_user.id,
            comment_text=comment_text,
            comment_type=comment_type,
            status='active'
        )
        
        db.session.add(comment)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Yorum eklendi',
            'comment': {
                'id': comment.id,
                'comment_text': comment.comment_text,
                'comment_type': comment.comment_type,
                'user_name': current_user.first_name or current_user.username,
                'created_at': comment.created_at.isoformat()
            }
        })
    
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'OKR comment hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/mtbp')
@login_required
def mtbp():
    """Mid-Term Business Plan (MTBP) sayfası"""
    try:
        kurum = current_user.kurum
        plans = StrategicPlan.query.filter_by(
            kurum_id=kurum.id
        ).order_by(StrategicPlan.start_date.desc()).all()
        
        return render_template('mtbp.html', plans=plans, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'MTBP sayfası hatası: {e}')
        flash('MTBP sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/mtbp/add', methods=['POST'])
@login_required
@csrf.exempt
def mtbp_add():
    """Yeni MTBP planı ekleme"""
    try:
        data = request.get_json() if request.is_json else request.form
        
        plan = StrategicPlan(
            kurum_id=current_user.kurum_id,
            plan_name=data.get('plan_name', '').strip(),
            plan_type='mtbp',
            start_date=datetime.strptime(data.get('start_date'), '%Y-%m-%d').date(),
            end_date=datetime.strptime(data.get('end_date'), '%Y-%m-%d').date(),
            description=data.get('description', ''),
            status='draft',
            created_by=current_user.id
        )
        
        db.session.add(plan)
        db.session.commit()
        
        if request.is_json:
            return jsonify({'success': True, 'message': 'Plan oluşturuldu', 'plan_id': plan.id})
        else:
            flash('Plan başarıyla oluşturuldu.', 'success')
            return redirect(url_for('main.mtbp'))
    
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'MTBP ekleme hatası: {e}')
        if request.is_json:
            return jsonify({'success': False, 'message': str(e)}), 500
        else:
            flash('Plan oluşturulurken bir hata oluştu.', 'error')
            return redirect(url_for('main.mtbp'))


@main_bp.route('/gemba')
@login_required
def gemba():
    """Gemba Walk sayfası - Dijital Gemba"""
    try:
        kurum = current_user.kurum
        walks = GembaWalk.query.filter_by(
            kurum_id=kurum.id
        ).order_by(GembaWalk.walk_date.desc()).limit(50).all()
        
        surecler = Surec.query.filter_by(kurum_id=kurum.id).all()
        
        return render_template('gemba.html', walks=walks, surecler=surecler, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Gemba sayfası hatası: {e}')
        flash('Gemba Walk sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/gemba/add', methods=['POST'])
@login_required
@csrf.exempt
def gemba_add():
    """Yeni Gemba Walk kaydı ekleme"""
    try:
        data = request.get_json() if request.is_json else request.form
        
        walk = GembaWalk(
            kurum_id=current_user.kurum_id,
            surec_id=int(data.get('surec_id')) if data.get('surec_id') else None,
            conducted_by=current_user.id,
            walk_date=datetime.strptime(data.get('walk_date', datetime.utcnow().strftime('%Y-%m-%d')), '%Y-%m-%d').date(),
            location=data.get('location', ''),
            observations=data.get('observations', '').strip(),
            findings=data.get('findings', ''),
            improvements=data.get('improvements', ''),
            status='completed'
        )
        
        db.session.add(walk)
        db.session.commit()
        
        if request.is_json:
            return jsonify({'success': True, 'message': 'Gemba Walk kaydı eklendi', 'walk_id': walk.id})
        else:
            flash('Gemba Walk kaydı başarıyla eklendi.', 'success')
            return redirect(url_for('main.gemba'))
    
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Gemba ekleme hatası: {e}')
        if request.is_json:
            return jsonify({'success': False, 'message': str(e)}), 500
        else:
            flash('Gemba Walk kaydı eklenirken bir hata oluştu.', 'error')
            return redirect(url_for('main.gemba'))


# V60.0 - Talent & Risk Paketi
@main_bp.route('/competencies')
@login_required
def competencies():
    """Yetkinlik Matrisi sayfası"""
    try:
        kurum = current_user.kurum
        competencies_list = Competency.query.filter_by(kurum_id=kurum.id).all()
        
        # Kullanıcı yetkinliklerini getir
        user_competencies = UserCompetency.query.filter_by(user_id=current_user.id).all()
        user_competency_map = {uc.competency_id: uc.level for uc in user_competencies}
        
        # Tüm kullanıcıların yetkinlikleri (heatmap için)
        all_user_competencies = UserCompetency.query.join(User).filter(
            User.kurum_id == kurum.id
        ).all()
        
        return render_template('competencies.html', 
                             competencies=competencies_list,
                             user_competency_map=user_competency_map,
                             all_user_competencies=all_user_competencies,
                             kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Competencies sayfası hatası: {e}')
        flash('Yetkinlik Matrisi sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/risks')
@login_required
def risks():
    """Stratejik Risk Yönetimi sayfası"""
    try:
        kurum = current_user.kurum
        risks_list = StrategicRisk.query.filter_by(kurum_id=kurum.id).order_by(
            StrategicRisk.risk_score.desc()
        ).all()
        
        # Risk seviyelerine göre grupla
        risk_levels = {
            'Kritik': [r for r in risks_list if r.risk_level == 'Kritik'],
            'Yüksek': [r for r in risks_list if r.risk_level == 'Yüksek'],
            'Orta': [r for r in risks_list if r.risk_level == 'Orta'],
            'Düşük': [r for r in risks_list if r.risk_level == 'Düşük']
        }
        
        return render_template('risks.html', risks=risks_list, risk_levels=risk_levels, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Risks sayfası hatası: {e}')
        flash('Risk Yönetimi sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/risks/add', methods=['POST'])
@login_required
def risks_add():
    """Yeni risk ekle"""
    try:
        kurum = current_user.kurum
        probability = int(request.form.get('probability', 3))
        impact = int(request.form.get('impact', 3))
        risk_score = probability * impact
        
        # Risk seviyesini belirle
        if risk_score >= 21:
            risk_level = 'Kritik'
        elif risk_score >= 13:
            risk_level = 'Yüksek'
        elif risk_score >= 7:
            risk_level = 'Orta'
        else:
            risk_level = 'Düşük'
        
        risk = StrategicRisk(
            kurum_id=kurum.id,
            risk_name=request.form.get('risk_name'),
            risk_category=request.form.get('risk_category', 'strategic'),
            description=request.form.get('description'),
            probability=probability,
            impact=impact,
            risk_score=risk_score,
            risk_level=risk_level,
            mitigation_strategy=request.form.get('mitigation_strategy'),
            created_by=current_user.id
        )
        db.session.add(risk)
        db.session.commit()
        flash('Risk başarıyla eklendi.', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Risk ekleme hatası: {e}')
        flash('Risk eklenirken bir hata oluştu.', 'error')
    return redirect(url_for('main.risks'))


@main_bp.route('/executive-report')
@login_required
def executive_report():
    """Yönetim Kurulu Özeti sayfası"""
    try:
        kurum = current_user.kurum
        
        # Üst yönetim rolü kontrolü
        if current_user.sistem_rol not in ['admin', 'ust_yonetim', 'kurum_yoneticisi']:
            flash('Bu sayfaya erişim yetkiniz bulunmamaktadır.', 'error')
            return redirect(url_for('main.dashboard'))
        
        # Özet verileri topla
        stats = {
            'total_processes': Surec.query.filter_by(kurum_id=kurum.id).count(),
            'active_risks': StrategicRisk.query.filter_by(kurum_id=kurum.id, status='open').count(),
            'critical_risks': StrategicRisk.query.filter_by(kurum_id=kurum.id).filter(
                StrategicRisk.risk_score >= 20
            ).count(),
            'total_users': User.query.filter_by(kurum_id=kurum.id).count()
        }
        
        # Son eklenen riskler
        recent_risks = StrategicRisk.query.filter_by(kurum_id=kurum.id).order_by(
            StrategicRisk.created_at.desc()
        ).limit(10).all()
        
        return render_template('executive_report.html', stats=stats, recent_risks=recent_risks, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Executive report hatası: {e}')
        flash('Yönetim Kurulu Özeti sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


# V65.0 - Muda Hunter
@main_bp.route('/muda-hunter')
@login_required
def muda_hunter():
    """Muda Hunter sayfası - Süreç Verimsizlik Analizi"""
    try:
        kurum = current_user.kurum
        surecler = Surec.query.filter_by(kurum_id=kurum.id).all()
        
        # Mevcut muda bulguları
        findings = MudaFinding.query.filter_by(kurum_id=kurum.id).order_by(
            MudaFinding.created_at.desc()
        ).all()
        
        # Verimlilik skoru
        from services.muda_analyzer import MudaAnalyzerService
        efficiency_score = MudaAnalyzerService.get_efficiency_score(kurum.id)
        
        return render_template('muda_hunter.html', 
                             surecler=surecler,
                             findings=findings,
                             efficiency_score=efficiency_score,
                             kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Muda Hunter sayfası hatası: {e}')
        flash('Muda Hunter sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/api/muda-hunter/analyze/<int:surec_id>', methods=['POST'])
@login_required
def muda_analyze(surec_id):
    """Süreç verimsizlik analizi API"""
    try:
        surec = Surec.query.filter_by(id=surec_id, kurum_id=current_user.kurum_id).first_or_404()
        
        from services.muda_analyzer import MudaAnalyzerService
        findings = MudaAnalyzerService.analyze_process_inefficiency(surec_id, current_user.kurum_id)
        
        return jsonify({
            'success': True,
            'findings': findings,
            'surec_name': surec.surec_adi
        })
    
    except Exception as e:
        current_app.logger.error(f'Muda analiz hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/api/muda-hunter/efficiency-score', methods=['GET'])
@login_required
def muda_efficiency_score():
    """Genel verimlilik skoru API"""
    try:
        from services.muda_analyzer import MudaAnalyzerService
        efficiency_score = MudaAnalyzerService.get_efficiency_score(current_user.kurum_id)
        
        return jsonify({
            'success': True,
            'efficiency_score': efficiency_score
        })
    
    except Exception as e:
        current_app.logger.error(f'Efficiency score hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================
# FAZ 3: İLERİ SEVİYE MODÜLLER ROUTE'LARI
# ============================================

# V61.0 - Titan & Zenith Paketi
@main_bp.route('/crisis')
@login_required
def crisis():
    """Kriz Komuta Merkezi sayfası"""
    try:
        kurum = current_user.kurum
        crises = CrisisMode.query.filter_by(kurum_id=kurum.id).order_by(
            CrisisMode.activated_at.desc()
        ).all()
        
        return render_template('crisis.html', crises=crises, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Crisis sayfası hatası: {e}')
        flash('Kriz Komuta Merkezi sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/crisis/add', methods=['POST'])
@login_required
def crisis_add():
    """Yeni kriz ekle"""
    try:
        kurum = current_user.kurum
        crisis = CrisisMode(
            kurum_id=kurum.id,
            crisis_name=request.form.get('crisis_name'),
            crisis_type=request.form.get('crisis_type', 'other'),
            description=request.form.get('description'),
            severity='high',
            status='active',
            activated_at=datetime.utcnow(),
            activated_by=current_user.id
        )
        db.session.add(crisis)
        db.session.commit()
        flash('Kriz başarıyla bildirildi.', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Kriz ekleme hatası: {e}')
        flash('Kriz eklenirken bir hata oluştu.', 'error')
    return redirect(url_for('main.crisis'))


@main_bp.route('/succession')
@login_required
def succession():
    """Halefiyet Planlaması sayfası"""
    try:
        kurum = current_user.kurum
        plans = SuccessionPlan.query.filter_by(kurum_id=kurum.id).order_by(
            SuccessionPlan.created_at.desc()
        ).all()
        
        return render_template('succession.html', plans=plans, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Succession sayfası hatası: {e}')
        flash('Halefiyet Planlaması sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/reorg')
@login_required
def reorg():
    """Dinamik Organizasyon Tasarımcısı sayfası"""
    try:
        kurum = current_user.kurum
        scenarios = OrgScenario.query.filter_by(kurum_id=kurum.id).order_by(
            OrgScenario.created_at.desc()
        ).all()
        
        return render_template('reorg.html', scenarios=scenarios, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Reorg sayfası hatası: {e}')
        flash('Organizasyon Tasarımcısı sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/ona')
@login_required
def ona():
    """ONA (Organizasyonel Ağ Analizi) sayfası"""
    try:
        kurum = current_user.kurum
        influence_scores = InfluenceScore.query.filter_by(kurum_id=kurum.id).order_by(
            InfluenceScore.score.desc()
        ).limit(50).all()
        
        return render_template('ona.html', influence_scores=influence_scores, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'ONA sayfası hatası: {e}')
        flash('ONA sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/market-watch')
@login_required
def market_watch():
    """Market Watcher sayfası"""
    try:
        kurum = current_user.kurum
        intels = MarketIntel.query.filter_by(kurum_id=kurum.id).order_by(
            MarketIntel.collected_at.desc()
        ).limit(100).all()
        
        return render_template('market_watch.html', intels=intels, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Market Watch sayfası hatası: {e}')
        flash('Market Watcher sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


# V62.0 - Corporate Consciousness
@main_bp.route('/wellbeing')
@login_required
def wellbeing():
    """Tükenmişlik Kalkanı sayfası"""
    try:
        kurum = current_user.kurum
        scores = WellbeingScore.query.filter_by(
            kurum_id=kurum.id
        ).order_by(WellbeingScore.score_date.desc()).limit(100).all()
        
        user_scores = WellbeingScore.query.filter_by(
            user_id=current_user.id
        ).order_by(WellbeingScore.score_date.desc()).limit(30).all()
        
        return render_template('wellbeing.html', scores=scores, user_scores=user_scores, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Wellbeing sayfası hatası: {e}')
        flash('Tükenmişlik Kalkanı sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/simulation')
@login_required
def simulation():
    """Monte Carlo Simülatörü sayfası"""
    try:
        kurum = current_user.kurum
        scenarios = SimulationScenario.query.filter_by(kurum_id=kurum.id).order_by(
            SimulationScenario.created_at.desc()
        ).all()
        
        return render_template('simulation.html', scenarios=scenarios, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Simulation sayfası hatası: {e}')
        flash('Monte Carlo Simülatörü sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/simulation/add', methods=['POST'])
@login_required
def simulation_add():
    """Yeni simülasyon senaryosu ekle"""
    try:
        kurum = current_user.kurum
        scenario = SimulationScenario(
            kurum_id=kurum.id,
            scenario_name=request.form.get('scenario_name'),
            scenario_type=request.form.get('scenario_type', 'financial'),
            description=request.form.get('description'),
            variables=request.form.get('variables', '{}'),
            iterations=int(request.form.get('iterations', 10000)),
            status='draft',
            created_by=current_user.id
        )
        db.session.add(scenario)
        db.session.commit()
        flash('Simülasyon senaryosu başarıyla eklendi.', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Simülasyon ekleme hatası: {e}')
        flash('Simülasyon eklenirken bir hata oluştu.', 'error')
    return redirect(url_for('main.simulation'))


@main_bp.route('/simulation/<int:scenario_id>/run', methods=['POST'])
@login_required
def simulation_run(scenario_id):
    """Simülasyon çalıştır"""
    try:
        scenario = SimulationScenario.query.get_or_404(scenario_id)
        if scenario.kurum_id != current_user.kurum_id:
            return jsonify({'error': 'Yetkisiz erişim'}), 403
        
        from services.monte_carlo_service import run_monte_carlo_simulation
        
        # Variables JSON formatında olmalı
        variables = json.loads(scenario.variables) if isinstance(scenario.variables, str) else scenario.variables
        
        result = run_monte_carlo_simulation(
            variables=variables,
            iterations=scenario.iterations,
            simulation_type=scenario.scenario_type
        )
        
        if result.get('success'):
            scenario.status = 'completed'
            scenario.results = json.dumps(result)
            db.session.commit()
            return jsonify({
                'success': True,
                'result': result
            })
        else:
            scenario.status = 'failed'
            db.session.commit()
            return jsonify({
                'success': False,
                'error': result.get('error', 'Bilinmeyen hata')
            }), 400
    except Exception as e:
        current_app.logger.error(f'Simülasyon çalıştırma hatası: {e}')
        return jsonify({'error': str(e)}), 500


@main_bp.route('/deep-work/toggle', methods=['POST'])
@login_required
@csrf.exempt
def deep_work_toggle():
    """Deep Work oturumu başlat/durdur"""
    try:
        data = request.get_json() if request.is_json else request.form
        action = data.get('action', 'start')  # start veya stop
        
        if action == 'start':
            session = DeepWorkSession(
                user_id=current_user.id,
                start_time=datetime.utcnow(),
                status='active'
            )
            db.session.add(session)
            db.session.commit()
            return jsonify({'success': True, 'message': 'Deep Work oturumu başlatıldı', 'session_id': session.id})
        else:
            # Son aktif oturumu bul ve durdur
            session = DeepWorkSession.query.filter_by(
                user_id=current_user.id,
                status='active'
            ).order_by(DeepWorkSession.start_time.desc()).first()
            
            if session:
                session.end_time = datetime.utcnow()
                session.duration_minutes = int((session.end_time - session.start_time).total_seconds() / 60)
                session.status = 'completed'
                db.session.commit()
                return jsonify({'success': True, 'message': 'Deep Work oturumu durduruldu', 'duration': session.duration_minutes})
            else:
                return jsonify({'success': False, 'message': 'Aktif oturum bulunamadı'}), 404
    
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Deep Work toggle hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


# V63.0 - Transcendence Pack
@main_bp.route('/synthetic-lab')
@login_required
def synthetic_lab():
    """Sentetik Müşteri Laboratuvarı sayfası"""
    try:
        kurum = current_user.kurum
        personas = Persona.query.filter_by(kurum_id=kurum.id).all()
        simulations = ProductSimulation.query.filter_by(kurum_id=kurum.id).order_by(
            ProductSimulation.created_at.desc()
        ).limit(50).all()
        
        return render_template('synthetic_lab.html', personas=personas, simulations=simulations, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Synthetic Lab sayfası hatası: {e}')
        flash('Sentetik Müşteri Laboratuvarı sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/governance')
@login_required
def governance():
    """DAO Yönetimi sayfası"""
    try:
        kurum = current_user.kurum
        contracts = SmartContract.query.filter_by(kurum_id=kurum.id).all()
        proposals = DaoProposal.query.filter_by(kurum_id=kurum.id).order_by(
            DaoProposal.created_at.desc()
        ).limit(50).all()
        
        return render_template('governance.html', contracts=contracts, proposals=proposals, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Governance sayfası hatası: {e}')
        flash('DAO Yönetimi sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/metaverse')
@login_required
def metaverse():
    """Metaverse Departman İkizi sayfası"""
    try:
        kurum = current_user.kurum
        departments = MetaverseDepartment.query.filter_by(kurum_id=kurum.id).all()
        
        return render_template('metaverse.html', departments=departments, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Metaverse sayfası hatası: {e}')
        flash('Metaverse Departman İkizi sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/legacy-chat')
@login_required
def legacy_chat():
    """Kurucu Miras AI sayfası"""
    try:
        kurum = current_user.kurum
        knowledge = LegacyKnowledge.query.filter_by(kurum_id=kurum.id).order_by(
            LegacyKnowledge.created_at.desc()
        ).limit(100).all()
        
        return render_template('legacy_chat.html', knowledge=knowledge, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Legacy Chat sayfası hatası: {e}')
        flash('Kurucu Miras AI sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


# ============================================
# FAZ 4: NIRVANA LEGACY ROUTE'LARI
# ============================================

# V66.0 - Nirvana Legacy Paketi
@main_bp.route('/game-theory')
@login_required
def game_theory():
    """Oyun Teorisi (Game Theory Grid) sayfası"""
    try:
        kurum = current_user.kurum
        competitors = Competitor.query.filter_by(kurum_id=kurum.id).all()
        scenarios = GameScenario.query.filter_by(kurum_id=kurum.id).order_by(
            GameScenario.created_at.desc()
        ).limit(50).all()
        
        # Nash dengesi hesaplamaları için servisi import et
        from services.game_theory_service import calculate_nash_equilibrium, get_strategy_recommendation
        
        # Senaryolar için Nash dengesi sonuçlarını hesapla
        scenario_results = []
        for scenario in scenarios:
            if scenario.payoffs and scenario.nash_equilibrium:
                try:
                    nash_result = calculate_nash_equilibrium(scenario.payoffs)
                    recommendation = get_strategy_recommendation(nash_result)
                    scenario_results.append({
                        'scenario': scenario,
                        'nash_result': nash_result,
                        'recommendation': recommendation
                    })
                except Exception as e:
                    current_app.logger.warning(f'Nash dengesi hesaplama hatası (scenario {scenario.id}): {e}')
        
        return render_template('game_theory.html', 
                             competitors=competitors, 
                             scenarios=scenarios, 
                             scenario_results=scenario_results,
                             kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Game Theory sayfası hatası: {e}')
        flash('Oyun Teorisi sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/game-theory/scenario/<int:scenario_id>/calculate-nash', methods=['POST'])
@login_required
def calculate_nash_for_scenario(scenario_id):
    """Belirli bir senaryo için Nash dengesi hesapla"""
    try:
        scenario = GameScenario.query.get_or_404(scenario_id)
        
        # Sadece kendi kurumunun senaryosunu hesaplayabilir
        if scenario.kurum_id != current_user.kurum_id:
            return jsonify({'error': 'Yetkisiz erişim'}), 403
        
        from services.game_theory_service import calculate_nash_equilibrium, get_strategy_recommendation
        
        if not scenario.payoffs:
            return jsonify({'error': 'Kazanç matrisi bulunamadı'}), 400
        
        nash_result = calculate_nash_equilibrium(scenario.payoffs)
        recommendation = get_strategy_recommendation(nash_result)
        
        # Sonucu senaryoya kaydet
        scenario.nash_equilibrium = json.dumps(nash_result)
        scenario.strategy_recommendation = recommendation
        db.session.commit()
        
        return jsonify({
            'success': True,
            'nash_result': nash_result,
            'recommendation': recommendation
        })
    except Exception as e:
        current_app.logger.error(f'Nash dengesi hesaplama hatası: {e}')
        return jsonify({'error': str(e)}), 500


@main_bp.route('/knowledge-graph')
@login_required
def knowledge_graph():
    """Bilgi Grafığı sayfası"""
    try:
        kurum = current_user.kurum
        # Bilgi grafığı için veri topla (stratejiler, süreçler, projeler, vb.)
        strategies = AltStrateji.query.join(AnaStrateji).filter(
            AnaStrateji.kurum_id == kurum.id
        ).all()
        processes = Surec.query.filter_by(kurum_id=kurum.id).all()
        projects = Project.query.filter_by(kurum_id=kurum.id).all()
        
        # Bilgi grafığı servisi ile veri yapısını oluştur
        from services.knowledge_graph_service import build_knowledge_graph_data, calculate_centrality_metrics
        
        graph_data = build_knowledge_graph_data(strategies, processes, projects)
        centrality_metrics = calculate_centrality_metrics(graph_data['nodes'], graph_data['edges'])
        
        return render_template('knowledge_graph.html', 
                             strategies=strategies, 
                             processes=processes, 
                             projects=projects,
                             graph_data=graph_data,
                             centrality_metrics=centrality_metrics,
                             kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Knowledge Graph sayfası hatası: {e}')
        flash('Bilgi Grafığı sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/black-swan')
@login_required
def black_swan():
    """Siyah Kuğu Simülatörü (Doomsday Prepper) sayfası"""
    try:
        kurum = current_user.kurum
        scenarios = DoomsdayScenario.query.filter_by(kurum_id=kurum.id).order_by(
            DoomsdayScenario.severity_level.desc()
        ).all()
        
        return render_template('black_swan.html', scenarios=scenarios, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Black Swan sayfası hatası: {e}')
        flash('Siyah Kuğu Simülatörü sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


@main_bp.route('/library')
@login_required
def library():
    """Omega'nın Kitabı (Auto-Biography) sayfası"""
    try:
        kurum = current_user.kurum
        chronicles = YearlyChronicle.query.filter_by(kurum_id=kurum.id).order_by(
            YearlyChronicle.year.desc()
        ).all()
        
        return render_template('library.html', chronicles=chronicles, kurum=kurum)
    except Exception as e:
        current_app.logger.error(f'Library sayfası hatası: {e}')
        flash('Omega\'nın Kitabı sayfası yüklenirken bir hata oluştu.', 'error')
        return redirect(url_for('main.dashboard'))


# ============================================
# ADMIN API ENDPOINTS
# ============================================

@main_bp.route('/admin/get-organization/<kisa_ad>')
@login_required
def admin_get_organization(kisa_ad):
    """Kurum bilgilerini getir"""
    try:
        allowed_roles = {'admin', 'kurum_yoneticisi', 'ust_yonetim'}
        if current_user.sistem_rol not in allowed_roles:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
        kurum = Kurum.query.filter_by(kisa_ad=kisa_ad).first()
        if not kurum:
            return jsonify({'success': False, 'message': 'Kurum bulunamadı'}), 404

        # Admin olmayan kullanıcılar sadece kendi kurumunu görüntüleyebilir
        if current_user.sistem_rol != 'admin' and kurum.id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
        return jsonify({
            'success': True,
            'organization': {
                'id': kurum.id,
                'kisa_ad': kurum.kisa_ad,
                'ticari_unvan': kurum.ticari_unvan,
                'faaliyet_alani': kurum.faaliyet_alani,
                'sektor': kurum.sektor,
                'calisan_sayisi': kurum.calisan_sayisi,
                'email': kurum.email,
                'telefon': kurum.telefon,
                'web_adresi': kurum.web_adresi,
                'vergi_dairesi': kurum.vergi_dairesi if hasattr(kurum, 'vergi_dairesi') else None,
                'vergi_numarasi': kurum.vergi_numarasi if hasattr(kurum, 'vergi_numarasi') else None,
                'logo_url': kurum.logo_url if hasattr(kurum, 'logo_url') else None
            }
        })
    except Exception as e:
        current_app.logger.error(f'Kurum getirme hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/add-organization', methods=['POST'])
@login_required
def admin_add_organization():
    """Yeni kurum ekle - Admin tüm kurumları oluşturabilir"""
    try:
        if current_user.sistem_rol != 'admin':
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

        data = request.get_json()
        kisa_ad = data.get('kisa_ad', '').strip()
        
        if not kisa_ad:
            return jsonify({'success': False, 'message': 'Kısa ad zorunludur'}), 400
        
        # Kurum zaten var mı kontrol et
        existing = Kurum.query.filter_by(kisa_ad=kisa_ad).first()
        if existing:
            return jsonify({'success': False, 'message': 'Bu kısa adla bir kurum zaten mevcut'}), 400
        
        kurum = Kurum(
            kisa_ad=kisa_ad,
            ticari_unvan=data.get('ticari_unvan', '').strip(),
            faaliyet_alani=data.get('faaliyet_alani', '').strip(),
            sektor=data.get('sektor', '').strip(),
            calisan_sayisi=data.get('calisan_sayisi'),
            email=data.get('email', '').strip(),
            telefon=data.get('telefon', '').strip(),
            web_adresi=data.get('web_adresi', '').strip()
        )
        
        # Opsiyonel alanlar
        if hasattr(Kurum, 'vergi_dairesi'):
            kurum.vergi_dairesi = data.get('vergi_dairesi', '').strip() or None
        if hasattr(Kurum, 'vergi_numarasi'):
            kurum.vergi_numarasi = data.get('vergi_numarasi', '').strip() or None
        if hasattr(Kurum, 'logo_url'):
            kurum.logo_url = data.get('logo_url', '').strip() or None
        
        db.session.add(kurum)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Kurum başarıyla oluşturuldu',
            'organization': {
                'id': kurum.id,
                'kisa_ad': kurum.kisa_ad
            }
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Kurum ekleme hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/update-organization', methods=['POST'])
@login_required
def admin_update_organization():
    """Kurum bilgilerini güncelle - Admin tüm kurumları düzenleyebilir"""
    try:
        allowed_roles = {'admin', 'kurum_yoneticisi', 'ust_yonetim'}
        if current_user.sistem_rol not in allowed_roles:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
        data = request.get_json()
        kurum_id = data.get('id')
        
        if not kurum_id:
            return jsonify({'success': False, 'message': 'Kurum ID gerekli'}), 400

        kurum = Kurum.query.filter_by(id=kurum_id).first()
        if not kurum:
            return jsonify({'success': False, 'message': 'Kurum bulunamadı'}), 404

        # Admin olmayan kullanıcılar sadece kendi kurumunu düzenleyebilir
        if current_user.sistem_rol != 'admin' and kurum.id != current_user.kurum_id:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403
        
        # Kısa ad değişiyorsa, yeni kısa ad zaten kullanılıyor mu kontrol et
        new_kisa_ad = data.get('kisa_ad', '').strip()
        # Admin olmayanlar kisa_ad değiştiremesin (URL/bağımlılıklar için güvenli)
        if current_user.sistem_rol != 'admin':
            new_kisa_ad = kurum.kisa_ad

        if new_kisa_ad and new_kisa_ad != kurum.kisa_ad:
            existing = Kurum.query.filter_by(kisa_ad=new_kisa_ad).first()
            if existing:
                return jsonify({'success': False, 'message': 'Bu kısa adla bir kurum zaten mevcut'}), 400
            kurum.kisa_ad = new_kisa_ad
        
        # Diğer alanları güncelle
        if 'ticari_unvan' in data:
            kurum.ticari_unvan = data.get('ticari_unvan', '').strip()
        if 'faaliyet_alani' in data:
            kurum.faaliyet_alani = data.get('faaliyet_alani', '').strip()
        if 'sektor' in data:
            kurum.sektor = data.get('sektor', '').strip()
        if 'calisan_sayisi' in data:
            kurum.calisan_sayisi = data.get('calisan_sayisi')
        if 'email' in data:
            kurum.email = data.get('email', '').strip()
        if 'telefon' in data:
            kurum.telefon = data.get('telefon', '').strip()
        if 'web_adresi' in data:
            kurum.web_adresi = data.get('web_adresi', '').strip()
        
        # Opsiyonel alanlar
        if hasattr(Kurum, 'vergi_dairesi') and 'vergi_dairesi' in data:
            kurum.vergi_dairesi = data.get('vergi_dairesi', '').strip() or None
        if hasattr(Kurum, 'vergi_numarasi') and 'vergi_numarasi' in data:
            kurum.vergi_numarasi = data.get('vergi_numarasi', '').strip() or None
        if hasattr(Kurum, 'logo_url') and 'logo_url' in data:
            kurum.logo_url = data.get('logo_url', '').strip() or None
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Kurum başarıyla güncellendi'
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Kurum güncelleme hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/delete-organization/<int:org_id>', methods=['DELETE'])
@login_required
def admin_delete_organization(org_id):
    """Kurum sil (soft delete) - Sadece sistem admini (kurum_id=1) kurumları silebilir"""
    try:
        # Sadece sistem admini kurumları silebilir
        is_system_admin = current_user.sistem_rol == 'admin' and current_user.kurum_id == 1
        
        if not is_system_admin:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok. Sadece sistem yöneticisi kurum silebilir.'}), 403

        kurum = Kurum.query.get(org_id)
        if not kurum:
            return jsonify({'success': False, 'message': 'Kurum bulunamadı'}), 404
        
        if kurum.silindi:
            return jsonify({'success': False, 'message': 'Bu kurum zaten silinmiş'}), 400

        # SOFT DELETE: Kurumu ve cascade ilişkilerini silindi=True yap
        kurum.silindi = True
        kurum.deleted_at = datetime.utcnow()
        kurum.deleted_by = current_user.id
        
        # Cascade: Kuruma bağlı tüm kullanıcıları soft delete yap
        User.query.filter_by(kurum_id=org_id, silindi=False).update({
            'silindi': True,
            'deleted_at': datetime.utcnow(),
            'deleted_by': current_user.id
        }, synchronize_session=False)
        
        # Cascade: Kuruma bağlı tüm süreçleri soft delete yap
        Surec.query.filter_by(kurum_id=org_id, silindi=False).update({
            'silindi': True,
            'deleted_at': datetime.utcnow(),
            'deleted_by': current_user.id
        }, synchronize_session=False)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Kurum ve ilgili tüm veriler arşivlendi. İsterseniz geri getirebilirsiniz.'
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Kurum silme hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/restore-organization/<int:org_id>', methods=['POST'])
@login_required
def admin_restore_organization(org_id):
    """Kurum geri getir (restore) - Sadece sistem admini restore edebilir"""
    try:
        # Sadece sistem admini restore edebilir
        is_system_admin = current_user.sistem_rol == 'admin' and current_user.kurum_id == 1
        
        if not is_system_admin:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok. Sadece sistem yöneticisi kurum geri getirebilir.'}), 403

        kurum = Kurum.query.get(org_id)
        if not kurum:
            return jsonify({'success': False, 'message': 'Kurum bulunamadı'}), 404
        
        if not kurum.silindi:
            return jsonify({'success': False, 'message': 'Bu kurum zaten aktif'}), 400

        # Restore seçeneklerini al
        data = request.get_json() or {}
        restore_users = data.get('restore_users', True)
        restore_processes = data.get('restore_processes', True)

        # Kurumu restore et
        kurum.silindi = False
        kurum.deleted_at = None
        kurum.deleted_by = None
        
        # Cascade: İlişkili verileri restore et
        if restore_users:
            User.query.filter_by(kurum_id=org_id, silindi=True).update({
                'silindi': False,
                'deleted_at': None,
                'deleted_by': None
            }, synchronize_session=False)
        
        if restore_processes:
            Surec.query.filter_by(kurum_id=org_id, silindi=True).update({
                'silindi': False,
                'deleted_at': None,
                'deleted_by': None
            }, synchronize_session=False)
        
        db.session.commit()
        
        restore_info = []
        if restore_users:
            restore_info.append('kullanıcılar')
        if restore_processes:
            restore_info.append('süreçler')
        
        message = f'Kurum başarıyla geri getirildi'
        if restore_info:
            message += f' ({", ".join(restore_info)} dahil)'
        
        return jsonify({
            'success': True,
            'message': message
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Kurum restore hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/admin/deleted-organizations', methods=['GET'])
@login_required
def admin_deleted_organizations():
    """Silinmiş kurumlar listesi - Sadece sistem admini görebilir"""
    try:
        # Sadece sistem admini görebilir
        is_system_admin = current_user.sistem_rol == 'admin' and current_user.kurum_id == 1
        
        if not is_system_admin:
            return jsonify({'success': False, 'message': 'Bu işlem için yetkiniz yok'}), 403

        deleted_kurumlar = Kurum.query.filter_by(silindi=True).all()
        
        kurumlar_data = []
        for kurum in deleted_kurumlar:
            user_count = User.query.filter_by(kurum_id=kurum.id, silindi=True).count()
            surec_count = Surec.query.filter_by(kurum_id=kurum.id, silindi=True).count()
            
            deleter = None
            if kurum.deleted_by:
                deleter_user = User.query.get(kurum.deleted_by)
                if deleter_user:
                    deleter = f"{deleter_user.first_name} {deleter_user.last_name}"
            
            kurumlar_data.append({
                'id': kurum.id,
                'kisa_ad': kurum.kisa_ad,
                'ticari_unvan': kurum.ticari_unvan,
                'deleted_at': kurum.deleted_at.strftime('%d.%m.%Y %H:%M') if kurum.deleted_at else None,
                'deleted_by_name': deleter,
                'user_count': user_count,
                'surec_count': surec_count
            })
        
        return jsonify({
            'success': True,
            'data': kurumlar_data
        })
    except Exception as e:
        current_app.logger.error(f'Silinmiş kurumlar listesi hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/api/guide/update-preferences', methods=['POST'])
@login_required
def update_guide_preferences():
    """Kullanıcının guide tercihlerini güncelle"""
    try:
        data = request.get_json()
        page_id = data.get('page_id')
        completed = data.get('completed', False)
        
        if not page_id:
            return jsonify({'success': False, 'message': 'page_id gerekli'}), 400
        
        # Completed walkthroughs JSON'ını güncelle
        completed_walkthroughs = {}
        if current_user.completed_walkthroughs:
            try:
                import json
                completed_walkthroughs = json.loads(current_user.completed_walkthroughs)
            except:
                pass
        
        if completed:
            completed_walkthroughs[page_id] = True
        
        import json
        current_user.completed_walkthroughs = json.dumps(completed_walkthroughs)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        current_app.logger.error(f'Guide preferences güncelleme hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/api/guide/update-settings', methods=['POST'])
@login_required
def update_guide_settings():
    """Kullanıcının guide ayarlarını güncelle (settings sayfasından)"""
    try:
        data = request.get_json()
        show_page_guides = data.get('show_page_guides', True)
        guide_character_style = data.get('guide_character_style', 'professional')
        
        # Validate character style
        if guide_character_style not in ['professional', 'friendly', 'minimal']:
            return jsonify({'success': False, 'message': 'Geçersiz karakter stili'}), 400
        
        # Update user settings
        current_user.show_page_guides = show_page_guides
        current_user.guide_character_style = guide_character_style
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        current_app.logger.error(f'Guide settings güncelleme hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/api/guide/reset-walkthroughs', methods=['POST'])
@login_required
def reset_walkthroughs():
    """Tüm completed walkthroughs'u sıfırla"""
    try:
        current_user.completed_walkthroughs = None
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        current_app.logger.error(f'Walkthrough sıfırlama hatası: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


@main_bp.route('/yardim-merkezi')
@login_required
def help_center():
    """İnteraktif Yardım Merkezi"""
    return render_template('help_center.html', title='Yardım Merkezi')
