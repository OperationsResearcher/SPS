# -*- coding: utf-8 -*-
"""Kurulum import sihirbazı servisi testleri (TASK-235).

Tasarım: docs/paketler/KURULUM-IMPORT-SIHIRBAZI.md — dry-run/apply/upsert/döngü.
"""
import io

import pytest
from openpyxl import Workbook

from app import create_app
from app.models import db
from app.models.core import Tenant, Strategy, SubStrategy
from app.models.process import Process, ProcessKpi
from app.models.plan_year import PlanYear
from app.services.setup_import_service import (
    SHEET_KPIS,
    SHEET_PROCESSES,
    SHEET_STRATEGY,
    apply_workbook,
    make_setup_template_excel,
    parse_workbook,
)
from config import TestingConfig


@pytest.fixture
def app():
    app = create_app(TestingConfig)
    ctx = app.app_context()
    ctx.push()
    db.create_all()
    yield app
    db.session.remove()
    db.drop_all()
    ctx.pop()


@pytest.fixture
def tenant(app):
    t = Tenant(name="T", short_name="t", is_active=True)
    db.session.add(t)
    db.session.commit()
    return t


def _wb_bytes(procs=None, kpis=None, strategies=None):
    """Test çalışma kitabı üret. Satırlar: başlıksız tuple listeleri."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_PROCESSES
    ws.append(["Süreç Kodu", "Süreç Adı", "Üst Süreç Kodu", "Ağırlık", "Doküman No", "Revizyon No"])
    for r in procs or []:
        ws.append(list(r))
    ws2 = wb.create_sheet(SHEET_KPIS)
    ws2.append(["Süreç Kodu", "PG Kodu", "PG Adı", "Hedef", "Birim", "Periyot",
                "Toplama Yöntemi", "Gösterge Türü", "Ağırlık"])
    for r in kpis or []:
        ws2.append(list(r))
    ws3 = wb.create_sheet(SHEET_STRATEGY)
    ws3.append(["Ana Strateji Adı", "Alt Strateji Adı", "Bağlı Süreç Kodları"])
    for r in strategies or []:
        ws3.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_dry_run_add_ve_hata(tenant):
    data = _wb_bytes(
        procs=[("SR1", "Satış", None, 30, None, None),
               ("SR2", "", None, None, None, None)],          # ad boş → hata
        kpis=[("SR1", "PG-01", "Memnuniyet", "90", "%", "Aylık", "Ortalama", None, 40),
              ("SR1", "PG-02", "Süre", "3", "gün", "Haftalık", None, None, None)],  # periyot geçersiz
    )
    plan = parse_workbook(data, tenant.id)
    assert plan["success"]
    assert plan["summary"]["add"] == 2  # SR1 + PG-01
    assert plan["summary"]["error"] == 2
    sheets = {e["sheet"] for e in plan["errors"]}
    assert sheets == {SHEET_PROCESSES, SHEET_KPIS}


def test_apply_tek_transaction_hata_varsa_yazmaz(tenant):
    data = _wb_bytes(procs=[("SR1", "Satış", None, None, None, None),
                            ("", "Adsız", None, None, None, None)])  # kod boş → hata
    result = apply_workbook(data, tenant.id, user_id=1, skip_errors=False)
    assert result.get("applied") is False
    assert Process.query.filter_by(tenant_id=tenant.id).count() == 0  # hiçbir şey yazılmadı


def test_apply_skip_errors_gecerli_satirlari_yazar(tenant):
    data = _wb_bytes(procs=[("SR1", "Satış", None, None, None, None),
                            ("", "Adsız", None, None, None, None)])
    result = apply_workbook(data, tenant.id, user_id=1, skip_errors=True)
    assert result["applied"] is True
    assert result["result"]["processes"]["add"] == 1


def test_upsert_gunceller(tenant):
    db.session.add(Process(tenant_id=tenant.id, code="SR1", name="Eski Ad", is_active=True))
    db.session.commit()
    data = _wb_bytes(procs=[("SR1", "Yeni Ad", None, 55, None, None)])
    result = apply_workbook(data, tenant.id, user_id=1)
    assert result["applied"] and result["result"]["processes"]["update"] == 1
    p = Process.query.filter_by(tenant_id=tenant.id, code="SR1").one()
    assert p.name == "Yeni Ad" and p.weight == 55


def test_parent_dongusu_yakalanir(tenant):
    data = _wb_bytes(procs=[("A", "A Süreci", "B", None, None, None),
                            ("B", "B Süreci", "A", None, None, None)])
    plan = parse_workbook(data, tenant.id)
    assert any("döngü" in e["error"] for e in plan["errors"])


def test_parent_bagi_kurulur(tenant):
    data = _wb_bytes(procs=[("SR1", "Ana", None, None, None, None),
                            ("SR1.1", "Alt", "SR1", None, None, None)])
    result = apply_workbook(data, tenant.id, user_id=1)
    assert result["applied"]
    child = Process.query.filter_by(tenant_id=tenant.id, code="SR1.1").one()
    parent = Process.query.filter_by(tenant_id=tenant.id, code="SR1").one()
    assert child.parent_id == parent.id


def test_strateji_plan_yili_yoksa_atlanir(tenant):
    data = _wb_bytes(procs=[("SR1", "Satış", None, None, None, None)],
                     strategies=[("Büyüme", "Kanal genişletme", "SR1")])
    result = apply_workbook(data, tenant.id, user_id=1, plan_year_id=None)
    assert result["applied"]
    assert any("plan yılı" in w.lower() for w in result["warnings"])
    assert Strategy.query.count() == 0


def test_strateji_tam_zincir(tenant):
    py = PlanYear(tenant_id=tenant.id, year=2026)
    db.session.add(py)
    db.session.commit()
    data = _wb_bytes(procs=[("SR1", "Satış", None, None, None, None)],
                     strategies=[("Büyüme", "Kanal genişletme", "SR1")])
    result = apply_workbook(data, tenant.id, user_id=1, plan_year_id=py.id)
    assert result["applied"]
    st = result["result"]["strategy"]
    assert st == {"strategy": 1, "sub": 1, "link": 1}
    s = Strategy.query.one()
    sub = SubStrategy.query.one()
    assert s.plan_year_id == py.id and sub.strategy_id == s.id


def test_sablon_uretimi():
    data = make_setup_template_excel()
    assert data[:2] == b"PK"  # xlsx = zip
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    assert set(wb.sheetnames) == {SHEET_PROCESSES, SHEET_KPIS, SHEET_STRATEGY}
