# -*- coding: utf-8 -*-
"""
docs/kokpitim_yedek_eski.db -> docs/eski_pgv_raporu.xlsx
performans_gosterge_veri + kurum / user / surec / surec_performans_gostergesi (LEFT JOIN)
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
import sqlite3

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "docs" / "kokpitim_yedek_eski.db"
OUT_PATH = ROOT / "docs" / "eski_pgv_raporu.xlsx"

# PG: surec PG varsa onun ad/kodu; yoksa bireysel PG satiri (kaynak baglantisi olmayanlar)
SQL = """
SELECT
    k.ticari_unvan AS kurum_adi,
    TRIM(COALESCE(u.first_name, '') || ' ' || COALESCE(u.last_name, '')) AS kullanici,
    s.ad AS surec_ad,
    COALESCE(spg.ad, bpg.ad) AS pg_ad,
    COALESCE(spg.kodu, bpg.kodu) AS pg_kod,
    pgv.yil,
    pgv.ay,
    pgv.ceyrek,
    pgv.hedef_deger,
    pgv.gerceklesen_deger,
    pgv.durum,
    pgv.created_at
FROM performans_gosterge_veri pgv
INNER JOIN bireysel_performans_gostergesi bpg ON bpg.id = pgv.bireysel_pg_id
INNER JOIN user u ON u.id = pgv.user_id
INNER JOIN kurum k ON k.id = u.kurum_id
LEFT JOIN surec s ON s.id = bpg.kaynak_surec_id
LEFT JOIN surec_performans_gostergesi spg ON spg.id = bpg.kaynak_surec_pg_id
ORDER BY
    datetime(COALESCE(pgv.created_at, pgv.veri_tarihi)) DESC,
    pgv.id DESC
"""


def main() -> None:
    if not DB_PATH.is_file():
        raise SystemExit(f"DB bulunamadi: {DB_PATH}")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(SQL).fetchall()
    conn.close()

    headers = [
        "Kurum adı",
        "Kullanıcı adı soyadı",
        "Süreç adı",
        "PG adı",
        "PG kodu",
        "Yıl",
        "Ay",
        "Çeyrek",
        "Hedef değer",
        "Gerçekleşen değer",
        "Durum",
        "Oluşturma tarihi",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "performans_gosterge_veri"

    bold = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = bold

    for r, row in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=row["kurum_adi"] or "")
        ws.cell(row=r, column=2, value=row["kullanici"] or "")
        ws.cell(row=r, column=3, value=row["surec_ad"] or "")
        ws.cell(row=r, column=4, value=row["pg_ad"] or "")
        ws.cell(row=r, column=5, value=row["pg_kod"] or "")
        ws.cell(row=r, column=6, value=row["yil"])
        ws.cell(row=r, column=7, value=row["ay"] if row["ay"] is not None else "")
        ws.cell(row=r, column=8, value=row["ceyrek"] if row["ceyrek"] is not None else "")
        ws.cell(row=r, column=9, value=row["hedef_deger"])
        ws.cell(row=r, column=10, value=row["gerceklesen_deger"])
        ws.cell(row=r, column=11, value=row["durum"] or "")
        ws.cell(row=r, column=12, value=row["created_at"])

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Yazildi: {OUT_PATH} ({len(rows)} satir)")


if __name__ == "__main__":
    main()
