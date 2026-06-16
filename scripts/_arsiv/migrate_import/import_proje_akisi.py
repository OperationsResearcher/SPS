# -*- coding: utf-8 -*-
"""Import PROJE AKIŞI.xlsx into Project + Task.

- Reads task rows: ADIM / SORUMLU / DURUM + dynamic HAFTA columns (1..N)
- Computes task start/due dates from project start date (1 hafta = 7 gün)
- Stores optional project notification settings if present as columns

Usage (example):
  python import_proje_akisi.py --excel "belge/PROJE AKIŞI.xlsx" --kurum-id 2 --manager-id 1 --name "Danışmanlık Projesi" --start-date 2026-01-13

If assignee cannot be matched to a user, task will be left unassigned and the raw assignee text is appended to description.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from difflib import SequenceMatcher
from collections import Counter
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook


def _parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def _is_marked(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    s = str(value).strip()
    if not s:
        return False
    return s.upper() in {"X", "✓", "OK", "1", "VAR"}


def _normalize_status(raw: Any) -> str:
    if raw is None:
        return "Yapılacak"
    s = str(raw).strip().lower()
    if not s:
        return "Yapılacak"

    mapping = {
        "yapılacak": "Yapılacak",
        "yapilacak": "Yapılacak",
        "todo": "Yapılacak",
        "to do": "Yapılacak",
        "devam": "Devam Ediyor",
        "devam ediyor": "Devam Ediyor",
        "in progress": "Devam Ediyor",
        "beklemede": "Beklemede",
        "blocked": "Beklemede",
        "tamamlandı": "Tamamlandı",
        "tamamlandi": "Tamamlandı",
        "done": "Tamamlandı",
    }

    for k, v in mapping.items():
        if s == k:
            return v

    # best-effort fuzzy
    if "tamam" in s:
        return "Tamamlandı"
    if "devam" in s:
        return "Devam Ediyor"
    if "bekle" in s or "blok" in s:
        return "Beklemede"
    return "Yapılacak"


def _extract_week_number(header_value: Any) -> int | None:
    if header_value is None:
        return None
    text = str(header_value)
    m = re.search(r"(\d+)", text)
    if not m:
        return None
    try:
        n = int(m.group(1))
    except Exception:
        return None
    if n <= 0:
        return None
    # If header doesn't mention 'hafta' but is a large number, ignore.
    if "haft" in text.lower() or n <= 120:
        return n
    return None


def _get_default_notification_settings() -> dict:
    return {
        "reminder_days": [7, 3, 1],
        "overdue_frequency": "daily",
        "channels": {"in_app": True, "email": False},
        "notify_manager": True,
        "notify_observers": False,
    }


def _read_project_notification_settings(headers: list[Any], row_values: list[Any]) -> dict:
    """Reads optional per-project notification settings from the sheet.

    Convention: these are project-level columns (same value repeated or only first row filled).
    If not present, returns defaults.
    """
    defaults = _get_default_notification_settings()
    header_to_index = {str(h).strip().lower(): i for i, h in enumerate(headers) if h is not None}

    def pick(*names: str) -> Any:
        for name in names:
            idx = header_to_index.get(name)
            if idx is not None and idx < len(row_values):
                v = row_values[idx]
                if v is not None and str(v).strip() != "":
                    return v
        return None

    reminder_raw = pick("reminder_days", "uyari_gunleri", "uyarı_günleri", "yaklasan_uyari_gunleri")
    if reminder_raw is not None:
        s = str(reminder_raw)
        parts = [p.strip() for p in re.split(r"[;,\s]+", s) if p.strip()]
        days: list[int] = []
        for p in parts:
            if p.isdigit():
                days.append(int(p))
        days = sorted(list(set([d for d in days if d >= 0])), reverse=True)
        if days:
            defaults["reminder_days"] = days

    overdue_freq = pick("overdue_frequency", "gecikme_sikligi", "gecikme_sıklığı", "gecikme_uyarisi")
    if overdue_freq is not None:
        val = str(overdue_freq).strip().lower()
        if val in {"off", "kapali", "kapalı", "0", "hayir", "false"}:
            defaults["overdue_frequency"] = "off"
        elif val in {"daily", "gunluk", "günlük"}:
            defaults["overdue_frequency"] = "daily"

    notify_manager = pick("notify_manager", "yonetici_bildir", "yönetici_bildir")
    if notify_manager is not None:
        val = str(notify_manager).strip().lower()
        defaults["notify_manager"] = val not in {"0", "hayir", "false", "no"}

    notify_observers = pick("notify_observers", "gozlemci_bildir", "gözlemci_bildir")
    if notify_observers is not None:
        val = str(notify_observers).strip().lower()
        defaults["notify_observers"] = val in {"1", "evet", "true", "yes"}

    email_enabled = pick("channel_email", "email", "e_posta", "e-posta")
    if email_enabled is not None:
        val = str(email_enabled).strip().lower()
        defaults["channels"]["email"] = val in {"1", "evet", "true", "yes"}

    return defaults


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, help="Excel path")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active)")
    parser.add_argument(
        "--personel-excel",
        default="belge/PERSONEL LİSTESİ.xlsx",
        help="Optional PERSONEL LİSTESİ.xlsx path (used to validate names / external assignee capture)",
    )
    parser.add_argument("--kurum-id", type=int, required=True)
    parser.add_argument("--manager-id", type=int, default=None, help="Project manager user id (optional)")
    parser.add_argument("--name", required=True, help="Project name")
    parser.add_argument("--start-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--end-date", default=None, help="YYYY-MM-DD (optional)")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--report-limit",
        type=int,
        default=10,
        help="Print top N unmatched assignee names (default: 10)",
    )

    args = parser.parse_args()

    from __init__ import create_app
    from models import db, Project, Task, User

    app = create_app()

    project_start = _parse_date(args.start_date)
    project_end = _parse_date(args.end_date) if args.end_date else None

    wb = load_workbook(args.excel, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    def norm(s: Any) -> str:
        if s is None:
            return ""
        t = str(s).strip().lower()
        t = (
            t.replace("ı", "i")
            .replace("ş", "s")
            .replace("ğ", "g")
            .replace("ü", "u")
            .replace("ö", "o")
            .replace("ç", "c")
            .replace("İ", "i")
        )
        t = unicodedata.normalize("NFKD", t)
        t = "".join(ch for ch in t if not unicodedata.combining(ch))
        # keep alnum and spaces only
        t = re.sub(r"[^a-z0-9\s._@-]+", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        return t

    def _similarity(a: str, b: str) -> float:
        if not a or not b:
            return 0.0
        return SequenceMatcher(None, a, b).ratio()

    def _read_personel_names(path: str) -> list[str]:
        try:
            wb_p = load_workbook(path, data_only=True)
            ws_p = wb_p.active
        except Exception:
            return []

        header_row = None
        for r in range(1, min(25, ws_p.max_row) + 1):
            vals = [ws_p.cell(row=r, column=c).value for c in range(1, min(15, ws_p.max_column) + 1)]
            if any(norm(v) in {"ad soyad", "adsoyad", "ad_soyad"} for v in vals if v is not None):
                header_row = r
                break
        if header_row is None:
            header_row = 1

        headers_p = [ws_p.cell(row=header_row, column=c).value for c in range(1, ws_p.max_column + 1)]
        header_lower_p = [norm(h) for h in headers_p]
        try:
            name_col = header_lower_p.index("ad soyad") + 1
        except ValueError:
            # fallback: first non-empty column
            name_col = 1

        names: list[str] = []
        for r in range(header_row + 1, ws_p.max_row + 1):
            v = ws_p.cell(row=r, column=name_col).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            names.append(s)
        return names

    # Find header row (some files have a title row above)
    header_row_idx = None
    for r in range(1, min(25, ws.max_row) + 1):
        row_vals = [c.value for c in ws[r]]
        row_norm = [norm(v) for v in row_vals]
        if any(v == "adim" for v in row_norm) and any(v == "sorumlu" for v in row_norm) and any(v == "durum" for v in row_norm):
            header_row_idx = r
            break

    if header_row_idx is None:
        raise RuntimeError("Header satırı bulunamadı (ADIM/SORUMLU/DURUM)")

    headers = [cell.value for cell in ws[header_row_idx]]
    header_lower = [norm(h) for h in headers]

    def find_col(*names: str) -> int | None:
        for name in names:
            name_l = norm(name)
            for i, h in enumerate(header_lower):
                if h == name_l:
                    return i
        return None

    col_adim = find_col("adim", "adım", "görev", "task")
    col_sorumlu = find_col("sorumlu", "assigned_to", "atanan", "assignee")
    col_durum = find_col("durum", "status")

    if col_adim is None:
        raise RuntimeError("Excel header içinde 'ADIM' kolonu bulunamadı")

    week_cols: list[tuple[int, int]] = []  # (col_index, week_no)
    for i, h in enumerate(headers):
        week_no = _extract_week_number(h)
        if week_no is None:
            continue
        # Heuristic: must look like a week column
        if "haft" in str(h).lower() or week_no <= 120:
            # avoid colliding with id-like columns
            if i not in {col_adim, col_sorumlu, col_durum}:
                week_cols.append((i, week_no))

    if not week_cols:
        raise RuntimeError("Hafta kolonları bulunamadı (başlıklarda 'Hafta 1' gibi olmalı)")

    week_cols.sort(key=lambda t: t[1])

    # Read first data row to derive optional project settings
    first_data_row = header_row_idx + 1
    first_values = [cell.value for cell in ws[first_data_row]]
    notification_settings = _read_project_notification_settings(headers, first_values)

    tasks_to_create: list[dict] = []

    for r in range(header_row_idx + 1, ws.max_row + 1):
        row = ws[r]
        values = [c.value for c in row]
        title = values[col_adim] if col_adim < len(values) else None
        if title is None or str(title).strip() == "":
            continue

        title_str = str(title).strip()
        assignee_raw = values[col_sorumlu] if col_sorumlu is not None and col_sorumlu < len(values) else None
        status_raw = values[col_durum] if col_durum is not None and col_durum < len(values) else None

        marked_weeks = [week_no for (ci, week_no) in week_cols if ci < len(values) and _is_marked(values[ci])]
        if marked_weeks:
            min_w = min(marked_weeks)
            max_w = max(marked_weeks)
            start = project_start + timedelta(days=(min_w - 1) * 7)
            due = project_start + timedelta(days=max_w * 7 - 1)
        else:
            start = None
            due = None

        tasks_to_create.append(
            {
                "title": title_str,
                "assignee_raw": str(assignee_raw).strip() if assignee_raw is not None else "",
                "status": _normalize_status(status_raw),
                "start_date": start,
                "due_date": due,
            }
        )

    with app.app_context():
        # Resolve manager
        manager = None
        if args.manager_id is not None:
            manager = User.query.get(args.manager_id)
            if not manager or manager.kurum_id != args.kurum_id:
                raise RuntimeError("manager_id geçersiz veya farklı kurumda")
        else:
            # Pick a sensible default manager from the kurum
            manager = (
                User.query.filter_by(kurum_id=args.kurum_id)
                .filter(User.sistem_rol.in_(['kurum_yoneticisi', 'ust_yonetim', 'admin']))
                .order_by(User.id.asc())
                .first()
            )
            if manager is None:
                manager = User.query.filter_by(kurum_id=args.kurum_id).order_by(User.id.asc()).first()
            if manager is None:
                raise RuntimeError("Bu kurum için kullanıcı bulunamadı (manager seçilemedi)")

        if args.dry_run:
            print(f"[DRY RUN] Project: {args.name} (tasks={len(tasks_to_create)})")
            print(f"manager: id={manager.id} username={manager.username}")
            print("notification_settings:", notification_settings)
            # Assignee match report (dry run)
            matched_internal = 0
            matched_external = 0
            empty_assignee = 0
            unmatched_counter: Counter[str] = Counter()

            # Build quick lookup for users in kurum
            users = User.query.filter_by(kurum_id=args.kurum_id).all()
            by_username = {u.username.lower(): u for u in users if u.username}
            personel_names = _read_personel_names(args.personel_excel) if args.personel_excel else []
            personel_norm = [(n, norm(n)) for n in personel_names]

            def resolve_assignee(text: str):
                if not text:
                    return None
                raw = str(text).strip()
                if not raw:
                    return None

                key_raw = raw.strip().lower()
                if key_raw in by_username:
                    return by_username[key_raw]

                key = norm(raw)

                if "@" in key:
                    for u in users:
                        if hasattr(u, "email") and u.email and norm(u.email) == key:
                            return u

                best = None
                best_score = 0.0
                for u in users:
                    uname = (u.username or "").strip().lower()
                    full = f"{u.first_name or ''} {u.last_name or ''}".strip()
                    score = max(_similarity(key, norm(full)), _similarity(key, norm(uname)))
                    if score > best_score:
                        best_score = score
                        best = u
                if best is not None and best_score >= 0.86:
                    return best
                return None

            def resolve_external_assignee(raw: str) -> str | None:
                if not raw:
                    return None
                raw_s = str(raw).strip()
                if not raw_s:
                    return None
                raw_n = norm(raw_s)
                if not raw_n:
                    return raw_s
                if personel_norm:
                    best_name = None
                    best_score = 0.0
                    for original, normalized in personel_norm:
                        score = _similarity(raw_n, normalized)
                        if score > best_score:
                            best_score = score
                            best_name = original
                    if best_name is not None and best_score >= 0.90:
                        return str(best_name).strip()
                return raw_s

            for t in tasks_to_create:
                u = resolve_assignee(t["assignee_raw"])
                ext = None
                if t["assignee_raw"] and not u:
                    ext = resolve_external_assignee(t["assignee_raw"])
                if u:
                    matched_internal += 1
                elif ext:
                    matched_external += 1
                    unmatched_counter[ext] += 1
                else:
                    empty_assignee += 1

            print(
                f"Assignee report: internal={matched_internal}, external={matched_external}, empty={empty_assignee}"
            )
            if unmatched_counter and args.report_limit > 0:
                top = unmatched_counter.most_common(args.report_limit)
                print("Top external assignees:")
                for name, count in top:
                    print(f" - {name} ({count})")

            for t in tasks_to_create[:5]:
                print(" -", t["title"], t["start_date"], t["due_date"], t["assignee_raw"], t["status"])
            return 0

        project = Project(
            kurum_id=args.kurum_id,
            name=args.name,
            description=None,
            manager_id=manager.id,
            start_date=project_start,
            end_date=project_end,
            priority="Orta",
            notification_settings=json.dumps(notification_settings, ensure_ascii=False),
        )
        db.session.add(project)
        db.session.flush()

        # Build quick lookup for users in kurum
        users = User.query.filter_by(kurum_id=args.kurum_id).all()
        by_username = {u.username.lower(): u for u in users if u.username}

        personel_names = _read_personel_names(args.personel_excel) if args.personel_excel else []
        personel_norm = [(n, norm(n)) for n in personel_names]

        def resolve_assignee(text: str):
            if not text:
                return None
            raw = str(text).strip()
            if not raw:
                return None

            key_raw = raw.strip().lower()
            if key_raw in by_username:
                return by_username[key_raw]

            key = norm(raw)

            # Email match if the User model has email
            if "@" in key:
                for u in users:
                    if hasattr(u, "email") and u.email and norm(u.email) == key:
                        return u

            # Fuzzy match against users (full name and username)
            best = None
            best_score = 0.0
            for u in users:
                uname = (u.username or "").strip().lower()
                full = f"{u.first_name or ''} {u.last_name or ''}".strip()
                score = max(_similarity(key, norm(full)), _similarity(key, norm(uname)))
                if score > best_score:
                    best_score = score
                    best = u

            if best is not None and best_score >= 0.86:
                return best
            return None

        def resolve_external_assignee(raw: str) -> str | None:
            if not raw:
                return None
            raw_s = str(raw).strip()
            if not raw_s:
                return None
            raw_n = norm(raw_s)
            if not raw_n:
                return raw_s

            # If PERSONEL list exists, prefer its canonical formatting
            if personel_norm:
                best_name = None
                best_score = 0.0
                for original, normalized in personel_norm:
                    score = _similarity(raw_n, normalized)
                    if score > best_score:
                        best_score = score
                        best_name = original
                if best_name is not None and best_score >= 0.90:
                    return str(best_name).strip()

            return raw_s

        created = 0
        matched_internal = 0
        matched_external = 0
        empty_assignee = 0
        unmatched_counter: Counter[str] = Counter()
        for t in tasks_to_create:
            user = resolve_assignee(t["assignee_raw"])
            external_name = None
            if t["assignee_raw"] and not user:
                external_name = resolve_external_assignee(t["assignee_raw"])

            if user:
                matched_internal += 1
            elif external_name:
                matched_external += 1
                unmatched_counter[external_name] += 1
            else:
                empty_assignee += 1

            desc = None
            if external_name:
                desc = f"Sorumlu (Excel): {external_name}"

            task = Task(
                project_id=project.id,
                title=t["title"],
                description=desc,
                assigned_to_id=user.id if user else None,
                external_assignee_name=external_name,
                reporter_id=manager.id,
                status=t["status"],
                priority="Orta",
                start_date=t["start_date"],
                due_date=t["due_date"],
            )
            db.session.add(task)
            created += 1

        db.session.commit()

        print(f"OK: Project created id={project.id}, tasks={created}")
        print(
            f"Assignee report: internal={matched_internal}, external={matched_external}, empty={empty_assignee}"
        )
        if unmatched_counter and args.report_limit > 0:
            top = unmatched_counter.most_common(args.report_limit)
            print("Top external assignees:")
            for name, count in top:
                print(f" - {name} ({count})")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
