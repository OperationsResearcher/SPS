
import os
import sys
from openpyxl import load_workbook

def inspect():
    # Construct path
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(base_dir, 'belge', 'SP_V2.xlsx')
    
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    try:
        wb = load_workbook(file_path, data_only=True)
        print("Sheets found:", wb.sheetnames)
        
        target_sheet = None
        for name in wb.sheetnames:
            if "vizyon" in name.lower() or "misyon" in name.lower() or "swot" in name.lower():
                target_sheet = name
                break
        
        if target_sheet:
            print(f"\n--- Content of '{target_sheet}' ---")
            ws = wb[target_sheet]
            # Print first 20 rows and 5 columns to get an idea
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=5, values_only=True):
                # Filter out completely None rows for cleaner output
                if any(row):
                    print(row)
        else:
            print("\n No sheet matching 'vizyon/misyon/swot' found.")
            
    except Exception as e:
        print(f"Error reading excel: {e}")

if __name__ == "__main__":
    inspect()
