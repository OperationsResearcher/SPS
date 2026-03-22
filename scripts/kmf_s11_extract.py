#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KMF / BOUN tipi süreç karnesi Excel (ör. s11.xlsx) — ham veriyi yapılandırılmış JSON'a döker.
Veritabanına yazmaz; eşleme ve import ayrı adımdır.

Kullanım:
  py scripts/kmf_s11_extract.py docs/KMF/s11.xlsx
  py scripts/kmf_s11_extract.py docs/KMF/s11.xlsx -o docs/KMF/s11-extracted.json

Sayfalar: "Veriler" ve dört haneli yıl adlı sayfalar (2021, 2022, ...).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl gerekli: pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

HEADER_ANCHOR = "Ana Strateji"
COL_FIILI_HEDEF = "Fiili/ Hedef"  # Excel başlığı (boşluklu)
QUARTER_KEYS = ("q1", "q2", "q3", "q4", "yil_sonu")

# Dosyada görülen kısaltma → Kokpitim PG formu ile olası eşleşme (doğrulama sizde)
GOSTERGE_TURU_KISA = {
    "iylş.": "İyileştirme",
    "iylş": "İyileştirme",
}

# Kokpitim target_method kodları (Excel sütunu "Hedef Belirl. Yön.")
# RG, HKY, HK, SH, DH, SGH — dosyada DH, HKY, SH görülüyor


def _cell_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    return s if s else None


def _cell_raw(v: Any) -> Any:
    """JSON için sayı/string koru."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def _normalize_doc_no(raw: str | None) -> str | None:
    """Hücrede yalnızca 'DOKÜMAN NO :' etiketi varsa None döner (dosyada boş olabiliyor)."""
    if not raw:
        return None
    t = raw.strip()
    if re.match(r"^DOK[ÜU]MAN\s*NO\s*:?\s*$", t, re.I):
        return None
    return t


def _find_header_row(ws) -> int | None:
    for r in range(1, min(ws.max_row or 0, 30) + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() == HEADER_ANCHOR:
            return r
    return None


def _build_header_map(ws, header_row: int) -> dict[str, int]:
    """Başlık metni (normalize) -> 1-based sütun indeksi."""
    mapping: dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        key = str(v).strip()
        if not key:
            continue
        norm = re.sub(r"\s+", " ", key)
        mapping[norm] = c
        mapping[key] = c
    return mapping


def _get_col(hmap: dict[str, int], *candidates: str) -> int | None:
    for cand in candidates:
        if cand in hmap:
            return hmap[cand]
        n = re.sub(r"\s+", " ", cand.strip())
        if n in hmap:
            return hmap[n]
    return None


def _parse_meta(ws, header_row: int) -> dict[str, Any]:
    meta: dict[str, Any] = {
        "sheet_title": _cell_str(ws.cell(1, 1).value),
        "revision_date": _cell_raw(ws.cell(2, 3).value),
        "publish_date": _cell_raw(ws.cell(2, 7).value),
        "document_no": _normalize_doc_no(
            _cell_str(ws.cell(2, 16).value) or _cell_str(ws.cell(2, 17).value) or _cell_str(ws.cell(2, 15).value),
        ),
        "process_leader": _cell_str(ws.cell(3, 3).value),
        "process_team": _cell_str(ws.cell(4, 3).value),
        "parametre_basari_baslik": _cell_str(ws.cell(4, 19).value),
    }
    # Satır 5: genel 1-5 puan bandı başlıkları (sütun 19-23)
    meta["basari_band_genel"] = [
        _cell_str(ws.cell(5, 19 + i).value) for i in range(5)
    ]
    return meta


def _read_indicator_row(
    ws,
    r: int,
    hmap: dict[str, int],
) -> dict[str, Any] | None:
    c_ana = hmap.get(HEADER_ANCHOR) or 1
    c_alt = _get_col(hmap, "Alt Strateji") or 2
    c_gost = _get_col(hmap, "Gösterge") or 3
    c_tur = _get_col(hmap, "Göst. Türü")
    c_hby = _get_col(hmap, "Hedef Belirl. Yön.")
    c_ag = _get_col(hmap, "Göst. Ağırlığı (%)")
    c_birim = _get_col(hmap, "Birim")
    c_per = _get_col(hmap, "Ölçüm Per.")
    c_oya = _get_col(hmap, "Önceki Yıl Ort.")
    c_fh = _get_col(hmap, COL_FIILI_HEDEF, "Fiili/ Hedef", "Fiili/Hedef")
    c_q = [_get_col(hmap, f"{i}.Ç") for i in range(1, 5)]
    c_ys = _get_col(hmap, "Yıl Sonu")
    c_bp = _get_col(hmap, "Başarı Puanı")
    c_abp = _get_col(hmap, "Ağırlıklı Başarı Puanı")

    if c_fh is None:
        return None

    fh = ws.cell(r, c_fh).value
    if fh is None:
        return None
    fh_s = str(fh).strip()
    if fh_s not in ("Fiili", "Hedef"):
        return None

    quarters = {}
    for i, cc in enumerate(c_q, start=1):
        if cc is not None:
            quarters[f"q{i}"] = _cell_raw(ws.cell(r, cc).value)
    yil = _cell_raw(ws.cell(r, c_ys).value) if c_ys else None

    bands = [_cell_raw(ws.cell(r, 19 + i).value) for i in range(5)]

    return {
        "row": r,
        "satir_tipi": fh_s,
        "ana_strateji_kod": _cell_str(ws.cell(r, c_ana).value),
        "alt_strateji_kod": _cell_str(ws.cell(r, c_alt).value),
        "gosterge_adi": _cell_str(ws.cell(r, c_gost).value),
        "gosterge_turu_excel": _cell_str(ws.cell(r, c_tur).value) if c_tur else None,
        "hedef_belirleme_yontemi": _cell_str(ws.cell(r, c_hby).value) if c_hby else None,
        "agirlik_yuzde": _cell_raw(ws.cell(r, c_ag).value) if c_ag else None,
        "birim": _cell_str(ws.cell(r, c_birim).value) if c_birim else None,
        "olcum_periyodu_excel": _cell_str(ws.cell(r, c_per).value) if c_per else None,
        "onceki_yil_ortalamasi": _cell_raw(ws.cell(r, c_oya).value) if c_oya else None,
        "ceyrekler": quarters,
        "yil_sonu": yil,
        "basari_puani": _cell_raw(ws.cell(r, c_bp).value) if c_bp else None,
        "agirlikli_basari_puani": _cell_raw(ws.cell(r, c_abp).value) if c_abp else None,
        "basari_esikleri_satir": bands,
    }


def _merge_fiili_hedef(pairs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Aynı gösterge için Fiili + Hedef satırlarını tek kayıtta birleştir."""
    out: list[dict[str, Any]] = []
    i = 0
    while i < len(pairs):
        a = pairs[i]
        if a.get("satir_tipi") != "Fiili":
            out.append(
                {
                    "parse_note": "Beklenmeyen: ilk satır Fiili değil",
                    "raw": a,
                }
            )
            i += 1
            continue
        b = pairs[i + 1] if i + 1 < len(pairs) else None
        merged = {
            "ana_strateji_kod": a.get("ana_strateji_kod"),
            "alt_strateji_kod": a.get("alt_strateji_kod"),
            "gosterge_adi": a.get("gosterge_adi"),
            "gosterge_turu_excel": a.get("gosterge_turu_excel"),
            "gosterge_turu_kokpitim_oneri": GOSTERGE_TURU_KISA.get(
                (a.get("gosterge_turu_excel") or "").strip().lower(),
            ),
            "hedef_belirleme_yontemi": a.get("hedef_belirleme_yontemi"),
            "agirlik_yuzde": a.get("agirlik_yuzde"),
            "birim": a.get("birim"),
            "olcum_periyodu_excel": a.get("olcum_periyodu_excel"),
            "onceki_yil_ortalamasi": a.get("onceki_yil_ortalamasi"),
            "fiili": {
                "ceyrekler": a.get("ceyrekler") or {},
                "yil_sonu": a.get("yil_sonu"),
                "basari_puani": a.get("basari_puani"),
                "agirlikli_basari_puani": a.get("agirlikli_basari_puani"),
            },
            "basari_esikleri": a.get("basari_esikleri_satir"),
            "excel_fiili_row": a.get("row"),
        }
        if b and b.get("satir_tipi") == "Hedef":
            merged["hedef"] = {
                "ceyrekler": b.get("ceyrekler") or {},
                "yil_sonu": b.get("yil_sonu"),
            }
            merged["excel_hedef_row"] = b.get("row")
            i += 2
        else:
            merged["hedef"] = None
            merged["parse_note"] = "Hedef satırı yok veya hemen sonra değil"
            i += 1
        out.append(merged)
    return out


def parse_sheet(ws, sheet_name: str) -> dict[str, Any]:
    hr = _find_header_row(ws)
    if hr is None:
        return {
            "sheet": sheet_name,
            "error": f"'{HEADER_ANCHOR}' başlığı bulunamadı (ilk 30 satır).",
        }

    hmap = _build_header_map(ws, hr)
    meta = _parse_meta(ws, hr)

    raw_rows: list[dict[str, Any]] = []
    warnings: list[str] = []

    for r in range(hr + 1, (ws.max_row or hr) + 1):
        rowd = _read_indicator_row(ws, r, hmap)
        if rowd is None:
            continue
        raw_rows.append(rowd)

    # Ardışık Fiili-Hedef çiftlerini doğrula
    for j in range(len(raw_rows) - 1):
        if raw_rows[j]["satir_tipi"] == "Fiili" and raw_rows[j + 1]["satir_tipi"] != "Hedef":
            warnings.append(
                f"Satır {raw_rows[j]['row']} Fiili sonrası beklenen Hedef yok veya satır {raw_rows[j+1]['row']} farklı tip.",
            )

    indicators = _merge_fiili_hedef(raw_rows)

    yil = sheet_name if re.fullmatch(r"\d{4}", sheet_name) else None

    return {
        "sheet": sheet_name,
        "yil": yil,
        "meta": meta,
        "header_row": hr,
        "uyarilar": warnings,
        "gostergeler": indicators,
        "ham_fiili_hedef_satirlari": raw_rows,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="KMF s11 benzeri karneden JSON çıkarımı")
    ap.add_argument("xlsx", type=Path, help="Excel dosyası")
    ap.add_argument("-o", "--out", type=Path, help="UTF-8 JSON çıktı (yoksa stdout)")
    ap.add_argument(
        "--no-raw-rows",
        action="store_true",
        help="Çıktıdan ham_fiili_hedef_satirlari alanını kaldır (daha küçük dosya)",
    )
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"Dosya yok: {args.xlsx}", file=sys.stderr)
        return 1

    wb = load_workbook(args.xlsx, data_only=True)
    sheets_out: list[dict[str, Any]] = []
    for name in wb.sheetnames:
        if name == "Veriler" or re.fullmatch(r"\d{4}", name):
            sheets_out.append(parse_sheet(wb[name], name))
    wb.close()

    doc = {
        "kaynak_dosya": str(args.xlsx.resolve()),
        "sayfalar": sheets_out,
        "kokpitim_notlari": {
            "KpiData": "Her çeyrek fiili değeri için ayrı kayıt: year, period_type=ceyrek, period_no=1..4, data_date=çeyrek son günü, actual_value.",
            "ProcessKpi.target_value": "Import: çeyrek+yıl sonu hedeflerinden sayı çıkarılıp ortalaması tek target_value olarak yazılır (scripts/kmf_s11_import.py).",
            "import": "py scripts/kmf_s11_import.py --process-id ID --actor-user-id UID --xlsx docs/KMF/s11.xlsx [--wipe-kpis] [--dry-run] [--seed N]",
        },
        "sizden_beklenen_netlestirmeler": SORULAR_TR,
    }

    if args.no_raw_rows:
        for s in doc["sayfalar"]:
            s.pop("ham_fiili_hedef_satirlari", None)

    text = json.dumps(doc, ensure_ascii=False, indent=2)

    if args.out:
        args.out.write_text(text, encoding="utf-8")
        print(f"Yazıldı: {args.out}", file=sys.stderr)
    else:
        sys.stdout.reconfigure(encoding="utf-8") if hasattr(sys.stdout, "reconfigure") else None
        print(text)
    return 0


# Kullanıcı kararları (import betiği kmf_s11_import.py ile uygulanır):
# 1) Alt strateji: H1.1  2) Gösterge türü: İyileştirme  3) Periyot: 3 ay→Çeyreklik, 1 Yıl→Yıllık, 6 ay→"6 ay"
# 4) Fiili = gerçekleşen  5) Hedef metin aralığı → sayıların ortalaması tek değer
# 6) 2 lider + 6 üye: tenant aktif kullanıcılarından rastgele (actor hariç en az 8 kişi)
SORULAR_TR: list[str] = []


if __name__ == "__main__":
    raise SystemExit(main())
