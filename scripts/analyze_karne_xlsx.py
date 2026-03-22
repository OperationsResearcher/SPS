#!/usr/bin/env python3
"""
BOUN / KMF tipi süreç karnesi Excel dosyasının yapısını metin olarak döker.
Kullanım:
  py scripts/analyze_karne_xlsx.py "docs/KMF/SR4 Pazarlama Stratejileri Yönetimi Süreç Karnesi.xlsx"

openpyxl gerekir (requirements.txt).
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path


def main() -> int:
    p = argparse.ArgumentParser(description="Karne xlsx özet çıktısı")
    p.add_argument("xlsx", type=Path, help="xlsx dosya yolu")
    p.add_argument("--rows", type=int, default=8, help="Her sayfadan kaç satır örnek")
    args = p.parse_args()
    path: Path = args.xlsx
    if not path.exists():
        print(f"Dosya yok: {path}", file=sys.stderr)
        return 1
    if path.name.startswith("~$"):
        print(
            "Uyarı: Bu Excel kilit dosyası (~$...) olabilir. Asıl .xlsx dosyasını seçin.",
            file=sys.stderr,
        )

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl yüklü değil: pip install openpyxl", file=sys.stderr)
        return 1

    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"Dosya: {path.name}")
    print(f"Sayfa sayısı: {len(wb.sheetnames)}")
    print(f"Sayfalar: {wb.sheetnames}")
    print("-" * 60)

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n## Sayfa: {name!r}  (max_row~{ws.max_row}, max_column~{ws.max_column})")
        n = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, args.rows), values_only=True):
            cells = ["" if c is None else str(c).replace("\n", " ")[:80] for c in row]
            # Sondaki boşları kırp
            while cells and cells[-1] == "":
                cells.pop()
            if not any(cells):
                continue
            print(f"  R{n+1}: {' | '.join(cells)}")
            n += 1
        if n == 0:
            print("  (ilk satırlar boş veya okunamadı)")

    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
