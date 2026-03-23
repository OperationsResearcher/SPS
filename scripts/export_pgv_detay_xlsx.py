# -*- coding: utf-8 -*-
"""kpi_data detaylarini Excel'e aktarir (docs/pgv_detay_raporu.xlsx)."""

import sqlite3
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "instance" / "kokpitim.db"
OUT_PATH = ROOT / "docs" / "pgv_detay_raporu.xlsx"

SQL = """
SELECT
    t.name AS tenant_name,
    TRIM(COALESCE(u.first_name, '') || ' ' || COALESCE(u.last_name, '')) AS user_fullname,
    p.name AS process_name,
    pk.name AS pg_name,
    pk.code AS pg_code,
    kd.data_date,
    kd.created_at,
    kd.year,
    kd.period_month,
    kd.period_type,
    kd.period_no,
    kd.target_value,
    kd.actual_value,
    kd.status,
    kd.id AS kpi_data_id
FROM kpi_data kd
INNER JOIN process_kpis pk ON pk.id = kd.process_kpi_id AND pk.is_active = 1
INNER JOIN processes p ON p.id = pk.process_id AND p.is_active = 1
INNER JOIN tenants t ON t.id = p.tenant_id AND t.is_active = 1
LEFT JOIN users u ON u.id = kd.user_id
WHERE kd.is_active = 1
  AND kd.deleted_at IS NULL
ORDER BY
    COALESCE(kd.data_date, date(kd.created_at)) DESC,
    kd.created_at DESC,
    kd.id DESC
"""


def _quarter_from_row(data_date: str | None, period_type: str | None, period_no: int | None) -> str | int | None:
    pt = (period_type or "").lower()
    if period_no is not None and ("ceyrek" in pt or "quarter" in pt or pt in ("q", "ceyrek")):
        return period_no
    if data_date:
        try:
            s = str(data_date)[:10]
            y, m, _ = int(s[0:4]), int(s[5:7]), int(s[8:10])
            return (m - 1) // 3 + 1
        except (ValueError, IndexError):
            pass
    return ""


def main() -> None:
    if not DB_PATH.is_file():
        raise SystemExit(f"DB bulunamadi: {DB_PATH}")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(SQL).fetchall()
    conn.close()

    headers = [
        "Kurum adı",
        "Kullanıcı adı soyadı",
        "Süreç adı",
        "PG adı",
        "PG kodu",
        "Veri tarihi",
        "Yıl",
        "Ay",
        "Çeyrek",
        "Hedef değer",
        "Gerçekleşen değer",
        "Durum",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "kpi_data"

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for r, row in enumerate(rows, start=2):
        data_date = row["data_date"]
        created = row["created_at"]
        veri_tarihi = data_date or (str(created)[:10] if created else "")
        ay = row["period_month"]
        quarter = _quarter_from_row(data_date, row["period_type"], row["period_no"])

        ws.cell(row=r, column=1, value=row["tenant_name"])
        ws.cell(row=r, column=2, value=row["user_fullname"] or "")
        ws.cell(row=r, column=3, value=row["process_name"])
        ws.cell(row=r, column=4, value=row["pg_name"])
        ws.cell(row=r, column=5, value=row["pg_code"] or "")
        ws.cell(row=r, column=6, value=veri_tarihi)
        ws.cell(row=r, column=7, value=row["year"])
        ws.cell(row=r, column=8, value=ay if ay is not None else "")
        ws.cell(row=r, column=9, value=quarter if quarter != "" else "")
        ws.cell(row=r, column=10, value=row["target_value"])
        ws.cell(row=r, column=11, value=row["actual_value"])
        ws.cell(row=r, column=12, value=row["status"] or "")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Yazildi: {OUT_PATH} ({len(rows)} satir)")


if __name__ == "__main__":
    main()
