# -*- coding: utf-8 -*-
"""KMF PGV verilerini Yayin ve Yerel icin AYRI AYRI Excel'e dokar — SALT OKUNUR.

Amac: kullanici iki tabloyu yan yana koyup "hangi veri gidecek, hangisi
kalacak" kararini versin; sonra ayni dosyayi geri verip uygulatsin.

BU SCRIPT HICBIR SEY YAZMAZ. Yayin'da yalnizca SELECT calisir.

SUTUNLAR
    Kullanicinin istedigi 6 sutun + uygulama adimi icin ZORUNLU olanlar.

    Neden ek sutun var: "geri verip uygula" derken hedef satiri TEK ANLAMLI
    bulabilmem gerekiyor. Surec adi + PG adi + tarih bunu saglamiyor —
    ayni PG'nin ayni tarihte birden fazla kaydi olabiliyor (olculdu: yerelde
    385 satir, 331 benzersiz is anahtari → 54 tekrar).
    Bu yuzden `kd_id` (satirin kimligi) ve `is_anahtari` (iki ortam arasinda
    eslestirme) tasiniyor. KARAR sutununu siz doldurursunuz.

Kullanim:
    python scripts/ops/pgv_excel_uret.py                      # KMF
    python scripts/ops/pgv_excel_uret.py --kurum "Eskişehir"
    python scripts/ops/pgv_excel_uret.py --yerel-only
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SSH_KEY = "C:/crt/ssh-key-2026-04-18_v4.key"
SSH_HOST = "ubuntu@129.159.30.175"
YAYIN_CONTAINER = "kokpitim-web"

# Iki tarafta AYNI kod kosar ki elma-elma olsun.
OLCUM_KODU = r'''
import json, sys
from app import create_app
from extensions import db
import sqlalchemy as sa

KURUM = sys.argv[1] if len(sys.argv) > 1 else "Kayseri Model Fabrika"

a = create_app()
with a.app_context():
    rows = db.session.execute(sa.text("""
        SELECT kd.id                AS kd_id,
               p.code               AS surec_kod,
               p.name               AS surec_ad,
               pk.id                AS pg_id,
               pk.code              AS pg_kod,
               pk.name              AS pg_ad,
               py.year              AS pg_plan_yili,
               kd.year              AS pgv_yili,
               kd.period_type, kd.period_no, kd.period_month,
               kd.data_date, kd.target_value, kd.actual_value,
               kd.status, kd.status_percentage,
               kd.is_active, kd.description,
               kd.created_at, kd.updated_at,
               u.email              AS giren_eposta,
               COALESCE(NULLIF(TRIM(CONCAT(u.first_name,' ',u.last_name)),''), u.email) AS giren_ad
        FROM kpi_data kd
        JOIN process_kpis pk ON pk.id = kd.process_kpi_id
        JOIN processes    p  ON p.id  = pk.process_id
        JOIN tenants      t  ON t.id  = p.tenant_id
        LEFT JOIN plan_years py ON py.id = p.plan_year_id
        LEFT JOIN users      u  ON u.id  = kd.user_id
        WHERE t.name ILIKE :k
        ORDER BY p.code, pk.name, kd.year, kd.period_type, kd.period_no
    """), {"k": "%" + KURUM + "%"}).fetchall()

    out = []
    for r in rows:
        out.append({
            "kd_id": r.kd_id,
            "surec_kod": r.surec_kod, "surec_ad": r.surec_ad,
            "pg_id": r.pg_id, "pg_kod": r.pg_kod, "pg_ad": r.pg_ad,
            "pg_plan_yili": r.pg_plan_yili, "pgv_yili": r.pgv_yili,
            "period_type": r.period_type, "period_no": r.period_no,
            "period_month": r.period_month,
            "data_date": str(r.data_date) if r.data_date else None,
            "hedef": r.target_value, "gerceklesen": r.actual_value,
            "durum": r.status,
            "yuzde": float(r.status_percentage) if r.status_percentage is not None else None,
            "aktif": bool(r.is_active) if r.is_active is not None else None,
            "aciklama": r.description,
            "created_at": str(r.created_at) if r.created_at else None,
            "updated_at": str(r.updated_at) if r.updated_at else None,
            "giren_eposta": r.giren_eposta, "giren_ad": r.giren_ad,
        })
    print("@@JSON@@" + json.dumps(out, default=str))
'''


def _ayikla(cikti: str) -> list:
    for satir in cikti.splitlines():
        if satir.startswith("@@JSON@@"):
            return json.loads(satir[len("@@JSON@@"):])
    raise RuntimeError("JSON satiri bulunamadi:\n" + cikti[-800:])


def yerel(kurum: str) -> list:
    kod = "import sys; sys.path.insert(0, '.')\n" + OLCUM_KODU
    p = subprocess.run([sys.executable, "-c", kod, kurum],
                       capture_output=True, text=True, cwd=str(ROOT), timeout=300)
    if p.returncode != 0:
        raise RuntimeError("Yerel olcum basarisiz:\n" + p.stderr[-800:])
    return _ayikla(p.stdout)


def yayin(kurum: str) -> list:
    uzak = f"sudo docker exec -i {YAYIN_CONTAINER} python - {kurum!r}"
    p = subprocess.run(
        ["ssh", "-i", SSH_KEY, "-o", "StrictHostKeyChecking=no",
         "-o", "ConnectTimeout=25", SSH_HOST, uzak],
        input=OLCUM_KODU, capture_output=True, text=True, timeout=400)
    if p.returncode != 0:
        raise RuntimeError(f"Yayin olcumu basarisiz (rc={p.returncode}):\n{p.stderr[-800:]}")
    return _ayikla(p.stdout)


def is_anahtari(k: dict) -> str:
    """Iki ortam arasinda ayni PGV'yi eslestiren anahtar (id'ler farkli olabilir)."""
    return "|".join(str(x) for x in (
        k.get("surec_kod") or "", k.get("pg_kod") or "", k.get("pg_ad") or "",
        k.get("pgv_yili"), k.get("period_type") or "",
        k.get("period_no"), k.get("period_month") or ""))


# (baslik, anahtar, genislik)  — ilk 6'si kullanicinin istedigi sutunlar
SUTUNLAR = [
    ("Süreç Adı",          "surec_ad",     34),
    ("Kullanıcı Adı",      "giren_ad",     26),
    ("PGV Adı",            "pg_ad",        42),
    ("Girilen Değer",      "gerceklesen",  14),
    ("Veri Tarihi",        "data_date",    12),
    ("Veri Giriş Tarihi",  "created_at",   19),
    # ── uygulama adimi icin gerekli ──
    ("Yıl",                "pgv_yili",      6),
    ("Dönem",              "_donem",       10),
    ("Hedef",              "hedef",        14),
    ("Durum",              "durum",         9),
    ("Skor %",             "yuzde",         8),
    ("Aktif",              "_aktif",        7),
    ("Diğer Ortamda Var?", "_karsi",       17),
    ("Fark",               "_fark",        24),
    ("KARAR",              "_karar",       14),
    ("Süreç Kodu",         "surec_kod",    12),
    ("PG Kodu",            "pg_kod",       10),
    ("PG Planı",           "pg_plan_yili",  9),
    ("Güncelleme",         "updated_at",   19),
    ("E-posta",            "giren_eposta", 30),
    ("kd_id",              "kd_id",        10),
    ("pg_id",              "pg_id",        10),
    ("İş Anahtarı",        "_anahtar",     46),
]


def yaz(kayitlar: list, karsi: dict, ortam: str, yol: Path, kurum: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "PGV"

    baslik_dolgu = PatternFill("solid", fgColor="1F4E79")
    baslik_font = Font(color="FFFFFF", bold=True, size=10)
    ek_dolgu = PatternFill("solid", fgColor="4472A8")   # ek sütunlar
    karar_dolgu = PatternFill("solid", fgColor="C55A11")
    yok_dolgu = PatternFill("solid", fgColor="FCE4D6")  # karşı ortamda yok
    fark_dolgu = PatternFill("solid", fgColor="FFF2CC")  # değer farkı

    ws.cell(1, 1, f"{kurum} — PGV · {ortam}").font = Font(bold=True, size=13)
    ws.cell(2, 1, "Üretim: %s · SALT OKUNUR · KARAR sütununu siz doldurun "
                  "(KALSIN / SİL / GÜNCELLE)" % datetime.now().strftime("%Y-%m-%d %H:%M")
            ).font = Font(size=9, italic=True, color="808080")

    BAS = 4
    for j, (bas, _, gen) in enumerate(SUTUNLAR, 1):
        c = ws.cell(BAS, j, bas)
        c.font = baslik_font
        c.fill = karar_dolgu if bas == "KARAR" else (ek_dolgu if j > 6 else baslik_dolgu)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = gen
    ws.row_dimensions[BAS].height = 30

    for i, k in enumerate(kayitlar, BAS + 1):
        ank = is_anahtari(k)
        oteki = karsi.get(ank)
        k["_anahtar"] = ank
        k["_donem"] = "%s%s" % (k.get("period_type") or "", k.get("period_no") or "")
        k["_aktif"] = "aktif" if k.get("aktif") else "SİLİNMİŞ"
        k["_karsi"] = "VAR" if oteki else "YOK"
        farklar = []
        if oteki:
            for alan, etiket in (("hedef", "hedef"), ("gerceklesen", "değer"),
                                 ("data_date", "tarih"), ("aktif", "aktif")):
                if k.get(alan) != oteki.get(alan):
                    farklar.append("%s: %s↔%s" % (etiket, k.get(alan), oteki.get(alan)))
        k["_fark"] = " ; ".join(farklar)
        k["_karar"] = ""

        for j, (bas, anahtar, _) in enumerate(SUTUNLAR, 1):
            c = ws.cell(i, j, k.get(anahtar))
            c.alignment = Alignment(vertical="top", wrap_text=(anahtar in ("pg_ad", "surec_ad", "_fark")))
            if bas == "KARAR":
                c.fill = karar_dolgu
                c.font = Font(color="FFFFFF")
        if not oteki:
            for j in range(1, len(SUTUNLAR) + 1):
                ws.cell(i, j).fill = yok_dolgu
        elif farklar:
            for j in range(1, len(SUTUNLAR) + 1):
                ws.cell(i, j).fill = fark_dolgu

    ws.freeze_panes = ws.cell(BAS + 1, 4)
    ws.auto_filter.ref = "A%d:%s%d" % (BAS, get_column_letter(len(SUTUNLAR)), BAS + len(kayitlar))

    # Özet sayfası
    oz = wb.create_sheet("Özet")
    yok = sum(1 for k in kayitlar if k["_karsi"] == "YOK")
    farkli = sum(1 for k in kayitlar if k["_fark"])
    silinmis = sum(1 for k in kayitlar if not k.get("aktif"))
    yil = {}
    for k in kayitlar:
        yil[k["pgv_yili"]] = yil.get(k["pgv_yili"], 0) + 1
    satirlar = [
        ("Ortam", ortam), ("Kurum", kurum), ("Toplam PGV", len(kayitlar)),
        ("Diğer ortamda YOK", yok), ("Değer farkı olan", farkli),
        ("Silinmiş (is_active=false)", silinmis), ("", ""), ("YIL DAĞILIMI", ""),
    ] + [(str(y), yil[y]) for y in sorted(yil)]
    for i, (a, b) in enumerate(satirlar, 1):
        oz.cell(i, 1, a).font = Font(bold=(b == "" or a in ("Ortam", "Kurum")))
        oz.cell(i, 2, b)
    oz.column_dimensions["A"].width = 30
    oz.column_dimensions["B"].width = 40

    wb.save(yol)
    return len(kayitlar), yok, farkli


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--kurum", default="Kayseri Model Fabrika")
    ap.add_argument("--yerel-only", action="store_true")
    ap.add_argument("--dizin", default="docs/buyukkesif")
    args = ap.parse_args()

    print("\nPGV Excel üretimi: %r  (SALT OKUNUR)\n" % args.kurum)
    print("  yerel okunuyor…")
    y = yerel(args.kurum)
    a = []
    if not args.yerel_only:
        print("  yayın okunuyor…")
        a = yayin(args.kurum)

    y_idx = {is_anahtari(k): k for k in y}
    a_idx = {is_anahtari(k): k for k in a}

    dizin = ROOT / args.dizin
    dizin.mkdir(parents=True, exist_ok=True)
    ad = args.kurum.split()[0]

    ys = dizin / f"PGV_YEREL_{ad}.xlsx"
    n1, yok1, f1 = yaz(y, a_idx, "YEREL", ys, args.kurum)
    print(f"\n  ✅ {ys}")
    print(f"     {n1} satır · Yayın'da yok: {yok1} · değer farkı: {f1}")

    if a:
        as_ = dizin / f"PGV_YAYIN_{ad}.xlsx"
        n2, yok2, f2 = yaz(a, y_idx, "YAYIN", as_, args.kurum)
        print(f"\n  ✅ {as_}")
        print(f"     {n2} satır · Yerelde yok: {yok2} · değer farkı: {f2}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
