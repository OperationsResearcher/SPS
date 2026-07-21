# -*- coding: utf-8 -*-
"""Excel'deki KARAR sutunlarini uygular — Yayin ve Yerel.

GIRDI (3 dosya, docs/buyukkesif/):
    PGV_YEREL_<kurum>.xlsx     KARAR=SİL olanlar yerelde silinir
    PGV_YAYIN_<kurum>.xlsx     KARAR=SİL olanlar Yayin'da silinir
    PGV_SORUNLU_<kurum>.xlsx   EKSİK / DEĞER-FARKI / ÇELİŞKİLİ-KARAR

KARAR SOZLUGU (esnek eslesme — kullanicinin yazdigi ifadeler):
    "yayin*"  iceren  -> Yayin referans alinir
    "sil"     iceren  -> silinir
    "yerel"   iceren  -> yerel korunur (islem yok)
    "yok say" iceren  -> islem yok

⚠ HARD DELETE (kullanici karari 2026-07-21). KURALLAR §3 soft delete
  istiyor; burada acik talep uzerine gercek DELETE kullaniliyor.
  Bu yuzden script YEDEK OLMADAN CALISMAZ (--yedek-alindi zorunlu).

Kullanim:
    python scripts/ops/pgv_karar_uygula.py                     # KONTROL
    python scripts/ops/pgv_karar_uygula.py --calistir --yedek-alindi
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook  # noqa: E402

SSH_KEY = "C:/crt/ssh-key-2026-04-18_v4.key"
SSH_HOST = "ubuntu@129.159.30.175"
YAYIN_CONTAINER = "kokpitim-web"
BAS = 4


def _oku(yol: Path, sayfa: str) -> list[dict]:
    ws = load_workbook(yol, data_only=True)[sayfa]
    basliklar = [ws.cell(BAS, j).value for j in range(1, ws.max_column + 1)]
    out = []
    for i in range(BAS + 1, ws.max_row + 1):
        d = dict(zip(basliklar, [ws.cell(i, j).value for j in range(1, ws.max_column + 1)]))
        if d.get("İş Anahtarı"):
            out.append(d)
    return out


def _k(v) -> str:
    return str(v).strip().upper() if v is not None else ""


def yayin_referans(karar: str) -> bool:
    k = _k(karar)
    return "YAYIN" in k or "YAYIN" in k.replace("I", "I")


def sil_mi(karar: str) -> bool:
    return "SİL" in _k(karar) or "SIL" in _k(karar)


# ── Uzak/yerel calistirma ────────────────────────────────────────────────────
# ⚠ ARGUMAN KODUN ICINE GOMULUR — komut satiri/stdin DEGIL.
#
# Denenen ve BASARISIZ olan iki yol:
#   1. `docker exec … python - '<json>'` → JSON'daki tirnak/kose parantez uzak
#      bash'te "unexpected EOF while looking for matching `''" ile patladi.
#   2. Kod + arguman ayni stdin akisinda (`head -n1` ile ayirma) → uzak taraf
#      hic cikti dondurmedi (stdin iki tuketici arasinda bolunemedi).
#
# Calisan yol: argumani `repr()` ile kodun basina sabit olarak yazmak. Tek
# stdin akisi, kabuk yorumu yok.
def _sarmala(kod: str, arg: str) -> str:
    return (
        "ARG = " + repr(arg) + "\n"
        + kod.replace("sys.argv[1]", "ARG").replace("len(sys.argv) > 1", "bool(ARG)")
    )


def _yerelde(kod: str, arg: str = "") -> dict:
    tam = "import sys; sys.path.insert(0,'.')\n" + _sarmala(kod, arg)
    p = subprocess.run([sys.executable, "-c", tam],
                       capture_output=True, text=True, cwd=str(ROOT), timeout=600)
    if p.returncode != 0:
        raise RuntimeError("Yerel islem basarisiz:\n" + p.stderr[-1500:])
    for s in p.stdout.splitlines():
        if s.startswith("@@JSON@@"):
            return json.loads(s[8:])
    raise RuntimeError("JSON yok:\n" + p.stdout[-800:])


def _yayinda(kod: str, arg: str = "") -> dict:
    p = subprocess.run(["ssh", "-i", SSH_KEY, "-o", "StrictHostKeyChecking=no",
                        "-o", "ConnectTimeout=25", SSH_HOST,
                        f"sudo docker exec -i {YAYIN_CONTAINER} python -"],
                       input=_sarmala(kod, arg),
                       capture_output=True, text=True, timeout=600)
    if p.returncode != 0:
        raise RuntimeError(f"Yayin islemi basarisiz (rc={p.returncode}):\n{p.stderr[-1500:]}")
    for s in p.stdout.splitlines():
        if s.startswith("@@JSON@@"):
            return json.loads(s[8:])
    raise RuntimeError("JSON yok:\n" + p.stdout[-800:] + "\nSTDERR:\n" + p.stderr[-600:])


SIL_KODU = r'''
import json, sys
from app import create_app
from extensions import db
import sqlalchemy as sa
idler = json.loads(sys.argv[1]) if len(sys.argv) > 1 else []
UYGULA = idler and idler[0] == "UYGULA"
if UYGULA: idler = idler[1:]
a = create_app()
with a.app_context():
    S = db.session
    var = [r[0] for r in S.execute(sa.text(
        "SELECT id FROM kpi_data WHERE id = ANY(:i)"), {"i": idler}).fetchall()]
    n = 0
    if UYGULA and var:
        n = S.execute(sa.text("DELETE FROM kpi_data WHERE id = ANY(:i)"), {"i": var}).rowcount
        S.commit()
    print("@@JSON@@" + json.dumps({"istenen": len(idler), "bulunan": len(var), "silinen": n}))
'''

CEK_KODU = r'''
import json, sys
from app import create_app
from extensions import db
import sqlalchemy as sa
a = create_app()
with a.app_context():
    anahtarlar = json.loads(sys.argv[1]) if len(sys.argv) > 1 else []
    rows = db.session.execute(sa.text("""
        SELECT kd.id, p.code surec_kod, p.name surec_ad, pk.code pg_kod, pk.name pg_ad,
               kd.year, kd.period_type, kd.period_no, kd.period_month,
               kd.data_date, kd.target_value, kd.actual_value, kd.status,
               kd.status_percentage, kd.description, kd.is_active,
               kd.created_at, kd.updated_at, u.email giren
        FROM kpi_data kd
        JOIN process_kpis pk ON pk.id = kd.process_kpi_id
        JOIN processes p ON p.id = pk.process_id
        JOIN tenants t ON t.id = p.tenant_id
        LEFT JOIN users u ON u.id = kd.user_id
        WHERE t.name ILIKE '%Kayseri Model Fabrika%'
    """)).fetchall()
    def ank(r):
        return "|".join(str(x) for x in (r.surec_kod or "", r.pg_kod or "", r.pg_ad or "",
                                         r.year, r.period_type or "", r.period_no,
                                         r.period_month or ""))
    ist = set(anahtarlar)
    out = []
    for r in rows:
        k = ank(r)
        if k in ist:
            out.append({"id": r.id, "anahtar": k, "pg_ad": r.pg_ad, "pg_kod": r.pg_kod,
                        "surec_kod": r.surec_kod, "surec_ad": r.surec_ad, "year": r.year,
                        "period_type": r.period_type, "period_no": r.period_no,
                        "period_month": r.period_month,
                        "data_date": str(r.data_date) if r.data_date else None,
                        "target_value": r.target_value, "actual_value": r.actual_value,
                        "status": r.status,
                        "status_percentage": float(r.status_percentage) if r.status_percentage is not None else None,
                        "description": r.description, "is_active": bool(r.is_active),
                        "giren": r.giren})
    print("@@JSON@@" + json.dumps(out, default=str))
'''


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--kurum", default="Kayseri")
    ap.add_argument("--calistir", action="store_true")
    ap.add_argument("--yedek-alindi", action="store_true",
                    help="HARD DELETE oncesi yedek aldiginizi onaylar")
    args = ap.parse_args()

    if args.calistir and not args.yedek_alindi:
        print("\n⛔ HARD DELETE yapilacak. Once yedek alin, sonra --yedek-alindi ekleyin.")
        return 1

    d = ROOT / "docs" / "buyukkesif"
    ye = _oku(d / f"PGV_YEREL_{args.kurum}.xlsx", "PGV")
    ya = _oku(d / f"PGV_YAYIN_{args.kurum}.xlsx", "PGV")
    so = _oku(d / f"PGV_SORUNLU_{args.kurum}.xlsx", "SORUNLU")

    yi = {k["İş Anahtarı"]: k for k in ye}
    ai = {k["İş Anahtarı"]: k for k in ya}

    # ── PLAN ──
    yerel_sil, yayin_sil, cek, guncelle, atlanan = [], [], [], [], []

    for k in ye:
        if sil_mi(k["KARAR"]):
            yerel_sil.append(k)
    for k in ya:
        if sil_mi(k["KARAR"]):
            yayin_sil.append(k)

    for k in so:
        a, sorun, karar = k["İş Anahtarı"], k["SORUN"], k["KARAR"]
        if not _k(karar):
            atlanan.append((a, sorun, "(boş)")); continue
        if sorun == "EKSİK":
            if yayin_referans(karar):
                cek.append(a)
            elif "YOK SAY" in _k(karar):
                pass
            else:
                atlanan.append((a, sorun, karar))
        elif sorun == "DEĞER-FARKI":
            if yayin_referans(karar):
                guncelle.append(k)
            elif "YEREL" in _k(karar):
                pass
            else:
                atlanan.append((a, sorun, karar))
        elif sorun == "ÇELİŞKİLİ-KARAR":
            if sil_mi(karar):
                if a in yi and yi[a] not in yerel_sil:
                    yerel_sil.append(yi[a])
                if a in ai and ai[a] not in yayin_sil:
                    yayin_sil.append(ai[a])
            elif "KALSIN" in _k(karar):
                yerel_sil = [x for x in yerel_sil if x["İş Anahtarı"] != a]
                yayin_sil = [x for x in yayin_sil if x["İş Anahtarı"] != a]
            else:
                atlanan.append((a, sorun, karar))

    print("\n" + "=" * 72)
    print("PGV KARAR UYGULAMA " + ("(ÇALIŞTIR)" if args.calistir else "(KONTROL — yazma yok)"))
    print("=" * 72)
    print(f"\n  YEREL'de silinecek   : {len(yerel_sil)}")
    print(f"  YAYIN'da silinecek   : {len(yayin_sil)}")
    print(f"  Yayın'dan çekilecek  : {len(cek)}")
    print(f"  Yayın değeriyle güncellenecek : {len(guncelle)}")
    print(f"  Karar tanınmayan/boş : {len(atlanan)}")
    for a, s_, kk in atlanan[:10]:
        print(f"      {s_:<17} {kk!r:<26} {a[:40]}")

    if not args.calistir:
        print("\n  Uygulamak için: --calistir --yedek-alindi")
        return 0

    # ── UYGULA ──
    print("\n  1) Yayın'dan eksik kayıtlar okunuyor…")
    veri = _yayinda(CEK_KODU, json.dumps(cek + [k["İş Anahtarı"] for k in guncelle]))
    print(f"     {len(veri)} kayıt alındı (istenen {len(cek) + len(guncelle)})")
    (ROOT / "backups" / "pgv_islem").mkdir(parents=True, exist_ok=True)
    (ROOT / "backups" / "pgv_islem" / "yayin_kaynak.json").write_text(
        json.dumps(veri, ensure_ascii=False, indent=1), encoding="utf-8")

    print("\n  2) YEREL silme…")
    r1 = _yerelde(SIL_KODU, json.dumps(["UYGULA"] + [int(k["kd_id"]) for k in yerel_sil]))
    print(f"     istenen={r1['istenen']} bulunan={r1['bulunan']} SİLİNEN={r1['silinen']}")

    print("\n  3) YAYIN silme…")
    r2 = _yayinda(SIL_KODU, json.dumps(["UYGULA"] + [int(k["kd_id"]) for k in yayin_sil]))
    print(f"     istenen={r2['istenen']} bulunan={r2['bulunan']} SİLİNEN={r2['silinen']}")

    print("\n  ⚠ 4) Çekme/güncelleme AYRI ADIM — pgv_cek_uygula.py ile yapılacak.")
    print("     Kaynak veri kaydedildi: backups/pgv_islem/yayin_kaynak.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
