# -*- coding: utf-8 -*-
"""KARAR sonrasi HALA SORUNLU olan PGV'leri TEK Excel'de toplar.

Girdi : docs/buyukkesif/PGV_{YEREL,YAYIN}_<kurum>.xlsx  (KARAR sutunu dolu)
Cikti : docs/buyukkesif/PGV_SORUNLU_<kurum>.xlsx

Tek liste, tek KARAR sutunu. Her satir bir SORUN; sorun turu ilk sutunda.

SORUN TURLERI
    EKSIK          Yayin'da var, yerelde yok  -> ne yapilacak?
    DEGER-FARKI    Iki tarafta farkli deger   -> hangisi dogru?
    CELISKILI-KARAR Bir tarafta SIL, digerinde bos
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

BAS = 4


def oku(yol: Path) -> list[dict]:
    ws = load_workbook(yol, data_only=True)["PGV"]
    basliklar = [ws.cell(BAS, j).value for j in range(1, ws.max_column + 1)]
    out = []
    for i in range(BAS + 1, ws.max_row + 1):
        satir = [ws.cell(i, j).value for j in range(1, ws.max_column + 1)]
        d = dict(zip(basliklar, satir))
        if d.get("kd_id") is not None:
            out.append(d)
    return out


def norm(v) -> str:
    return str(v).strip().upper() if v is not None else ""


def s(v) -> str:
    return "" if v is None else str(v)


# (baslik, genislik)
SUTUNLAR = [
    ("SORUN",              16),
    ("Süreç Adı",          32),
    ("PGV Adı",            40),
    ("Yıl",                 6),
    ("Dönem",              10),
    ("Veri Tarihi",        12),
    # ── yan yana karsilastirma ──
    ("YEREL Değer",        16),
    ("YAYIN Değer",        16),
    ("YEREL Hedef",        14),
    ("YAYIN Hedef",        14),
    ("YEREL Aktif",        12),
    ("YAYIN Aktif",        12),
    ("YEREL Tarih",        12),
    ("YAYIN Tarih",        12),
    ("Kullanıcı",          22),
    ("Veri Giriş Tarihi",  19),
    # ── karar ──
    ("KARAR",              20),
    ("AÇIKLAMA",           30),
    # ── iz surme ──
    ("Mevcut KARAR (yerel)", 14),
    ("Mevcut KARAR (yayın)", 14),
    ("yerel kd_id",        12),
    ("yayın kd_id",        12),
    ("İş Anahtarı",        46),
]

KARAR_SECENEK = {
    "EKSIK": "YAYINDAN ÇEK / YOK SAY",
    "DEGER-FARKI": "YEREL DOĞRU / YAYIN DOĞRU",
    "CELISKILI-KARAR": "İKİSİNİ DE SİL / İKİSİ DE KALSIN",
}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--kurum", default="Kayseri")
    args = ap.parse_args()
    d = ROOT / "docs" / "buyukkesif"

    ye = oku(d / f"PGV_YEREL_{args.kurum}.xlsx")
    ya = oku(d / f"PGV_YAYIN_{args.kurum}.xlsx")
    yi = {k["İş Anahtarı"]: k for k in ye}
    ai = {k["İş Anahtarı"]: k for k in ya}

    sil_y = {a for a, k in yi.items() if norm(k["KARAR"]) == "SİL"}
    sil_a = {a for a, k in ai.items() if norm(k["KARAR"]) == "SİL"}
    kalan_y = set(yi) - sil_y
    kalan_a = set(ai) - sil_a

    satirlar = []

    # 1) CELISKILI KARAR — once, cunku digerlerini etkiler
    for a in sorted((sil_y & kalan_a) | (sil_a & kalan_y)):
        y, x = yi.get(a), ai.get(a)
        k = y or x
        satirlar.append({
            "SORUN": "ÇELİŞKİLİ-KARAR", "_a": a, "y": y, "x": x, "k": k,
            "AÇIKLAMA": "bir ortamda SİL, diğerinde işaretsiz",
        })

    # 2) EKSIK — Yayin'da kaliyor, yerelde yok
    for a in sorted(kalan_a - kalan_y):
        if any(r["_a"] == a for r in satirlar):
            continue
        satirlar.append({
            "SORUN": "EKSİK", "_a": a, "y": None, "x": ai[a], "k": ai[a],
            "AÇIKLAMA": "Yayın'da var, yerelde yok",
        })

    # 3) DEGER FARKI — ikisi de kaliyor ama degerler farkli
    for a in sorted(kalan_y & kalan_a):
        y, x = yi[a], ai[a]
        f = [alan for alan in ("Girilen Değer", "Hedef", "Veri Tarihi", "Aktif")
             if s(y[alan]) != s(x[alan])]
        if f:
            satirlar.append({
                "SORUN": "DEĞER-FARKI", "_a": a, "y": y, "x": x, "k": y,
                "AÇIKLAMA": "farklı alan: " + ", ".join(f),
            })

    # ── Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "SORUNLU"

    ws.cell(1, 1, f"{args.kurum} — KARAR SONRASI HÂLÂ SORUNLU PGV'ler").font = Font(bold=True, size=13)
    ws.cell(2, 1, "Üretim: %s · Tek liste · KARAR sütununu doldurun. "
                  "Boş bırakılan satıra DOKUNULMAZ."
                  % datetime.now().strftime("%Y-%m-%d %H:%M")
            ).font = Font(size=9, italic=True, color="808080")

    bd = PatternFill("solid", fgColor="1F4E79")
    kd = PatternFill("solid", fgColor="C55A11")
    bf = Font(color="FFFFFF", bold=True, size=10)
    for j, (bas, gen) in enumerate(SUTUNLAR, 1):
        c = ws.cell(BAS, j, bas)
        c.font = bf
        c.fill = kd if bas in ("KARAR", "AÇIKLAMA") else bd
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = gen
    ws.row_dimensions[BAS].height = 30

    RENK = {
        "EKSİK":           PatternFill("solid", fgColor="FCE4D6"),
        "DEĞER-FARKI":     PatternFill("solid", fgColor="FFF2CC"),
        "ÇELİŞKİLİ-KARAR": PatternFill("solid", fgColor="F8CBAD"),
    }

    for i, r in enumerate(satirlar, BAS + 1):
        y, x, k = r["y"], r["x"], r["k"]
        deger = {
            "SORUN": r["SORUN"],
            "Süreç Adı": k["Süreç Adı"], "PGV Adı": k["PGV Adı"],
            "Yıl": k["Yıl"], "Dönem": k["Dönem"], "Veri Tarihi": k["Veri Tarihi"],
            "YEREL Değer": y["Girilen Değer"] if y else "— yok —",
            "YAYIN Değer": x["Girilen Değer"] if x else "— yok —",
            "YEREL Hedef": y["Hedef"] if y else "",
            "YAYIN Hedef": x["Hedef"] if x else "",
            "YEREL Aktif": y["Aktif"] if y else "",
            "YAYIN Aktif": x["Aktif"] if x else "",
            "YEREL Tarih": y["Veri Tarihi"] if y else "",
            "YAYIN Tarih": x["Veri Tarihi"] if x else "",
            "Kullanıcı": k["Kullanıcı Adı"],
            "Veri Giriş Tarihi": k["Veri Giriş Tarihi"],
            "KARAR": "",
            "AÇIKLAMA": r["AÇIKLAMA"] + "  →  seçenek: " + KARAR_SECENEK.get(
                r["SORUN"].replace("İ", "I").replace("Ğ", "G").replace("Ç", "C")
                          .replace("Ş", "S").replace("Ü", "U").replace("Ö", "O"), ""),
            "Mevcut KARAR (yerel)": y["KARAR"] if y else "—",
            "Mevcut KARAR (yayın)": x["KARAR"] if x else "—",
            "yerel kd_id": y["kd_id"] if y else "",
            "yayın kd_id": x["kd_id"] if x else "",
            "İş Anahtarı": r["_a"],
        }
        for j, (bas, _) in enumerate(SUTUNLAR, 1):
            c = ws.cell(i, j, deger.get(bas))
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(bas in ("PGV Adı", "Süreç Adı", "AÇIKLAMA")))
            if bas == "KARAR":
                c.fill = kd
                c.font = Font(color="FFFFFF")
            elif r["SORUN"] in RENK:
                c.fill = RENK[r["SORUN"]]

    ws.freeze_panes = ws.cell(BAS + 1, 4)
    ws.auto_filter.ref = "A%d:%s%d" % (BAS, get_column_letter(len(SUTUNLAR)), BAS + len(satirlar))

    # Özet
    oz = wb.create_sheet("Özet")
    from collections import Counter
    c = Counter(r["SORUN"] for r in satirlar)
    cy = Counter(r["k"]["Yıl"] for r in satirlar)
    bilgi = [("TOPLAM SORUNLU", len(satirlar)), ("", "")]
    bilgi += [(t, n) for t, n in c.most_common()]
    bilgi += [("", ""), ("YIL DAĞILIMI", "")]
    bilgi += [(str(y), cy[y]) for y in sorted(cy)]
    bilgi += [("", ""), ("SEÇENEKLER", "")]
    bilgi += [("EKSİK", "YAYINDAN ÇEK  /  YOK SAY"),
              ("DEĞER-FARKI", "YEREL DOĞRU  /  YAYIN DOĞRU"),
              ("ÇELİŞKİLİ-KARAR", "İKİSİNİ DE SİL  /  İKİSİ DE KALSIN")]
    for i, (a, b) in enumerate(bilgi, 1):
        oz.cell(i, 1, a).font = Font(bold=(b == "" or a.startswith("TOPLAM")))
        oz.cell(i, 2, b)
    oz.column_dimensions["A"].width = 26
    oz.column_dimensions["B"].width = 42

    yol = d / f"PGV_SORUNLU_{args.kurum}.xlsx"
    wb.save(yol)

    print(f"\n  ✅ {yol}")
    print(f"     {len(satirlar)} sorunlu satır\n")
    for t, n in c.most_common():
        print(f"       {t:<18} {n}")
    print("\n     yıl: " + "  ".join(f"{y}={cy[y]}" for y in sorted(cy)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
