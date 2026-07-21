# -*- coding: utf-8 -*-
"""Yayin'dan CEKILECEK 225 PGV satirini Excel'e doker — karar icin.

NEDEN VAR (2026-07-21):
    Onceki Excel'lerde satirlar "is anahtari" ile eslestiriliyordu:
        surec_kod | pg_kod | pg_ad | yil | period_type | period_no | period_month
    Bu anahtar TEKIL DEGILMIS. Olcum: 126 anahtar -> 225 gercek satir.
    43 anahtarda birden fazla olcum var; ornegin
        "Hizmet Verilen Kurumlardaki Getiri Miktari | 2025 | Aylik | ay=6"
    altinda 6 ayri tarihte 6 ayri deger bulunuyor (period_no bos).

    Yani kullanicinin "117 eksik" diye gordugu sey aslinda 225 satir.
    Bu Excel her SATIRI ayri gosterir; karar satir bazinda verilir.

Girdi : backups/pgv_islem/yayin_kaynak.json   (Yayin'dan okunan ham veri)
Cikti : docs/buyukkesif/PGV_CEKILECEK_<kurum>.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

BAS = 4

SUTUNLAR = [
    ("SORUN",              15),
    ("Süreç Adı",          30),
    ("PGV Adı",            42),
    ("Yıl",                 6),
    ("Dönem",              12),
    ("Ay",                  5),
    ("Veri Tarihi",        12),
    ("Girilen Değer",      15),
    ("Hedef",              14),
    ("Aktif",              10),
    ("Kullanıcı",          24),
    ("Veri Giriş Tarihi",  19),
    ("Bu grupta kaç kayıt", 18),
    ("Grup sırası",        12),
    ("KARAR",              18),
    ("AÇIKLAMA",           34),
    ("yayın kd_id",        12),
    ("İş Anahtarı",        46),
]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--kurum", default="Kayseri")
    args = ap.parse_args()

    kaynak = ROOT / "backups" / "pgv_islem" / "yayin_kaynak.json"
    veri = json.loads(kaynak.read_text(encoding="utf-8"))

    # SORUNLU excelinden her anahtarin sorun turunu al
    d = ROOT / "docs" / "buyukkesif"
    ws_s = load_workbook(d / f"PGV_SORUNLU_{args.kurum}.xlsx", data_only=True)["SORUNLU"]
    bs = [ws_s.cell(BAS, j).value for j in range(1, ws_s.max_column + 1)]
    sorun_turu, aciklama = {}, {}
    for i in range(BAS + 1, ws_s.max_row + 1):
        r = dict(zip(bs, [ws_s.cell(i, j).value for j in range(1, ws_s.max_column + 1)]))
        if r.get("İş Anahtarı"):
            sorun_turu[r["İş Anahtarı"]] = r["SORUN"]
            aciklama[r["İş Anahtarı"]] = r["AÇIKLAMA"]

    grup = Counter(x["anahtar"] for x in veri)
    # Grup icinde tarihe gore sirala ki "1/6, 2/6…" anlamli olsun
    veri.sort(key=lambda x: (x["anahtar"], x["data_date"] or "", x["id"]))
    sira: Counter = Counter()

    wb = Workbook()
    ws = wb.active
    ws.title = "ÇEKİLECEK"
    ws.cell(1, 1, f"{args.kurum} — Yayın'dan ÇEKİLECEK PGV satırları").font = Font(bold=True, size=13)
    ws.cell(2, 1, "Üretim: %s · %d satır · %d benzersiz iş anahtarı · "
                  "KARAR: ÇEK / ATLA · boş bırakılan satır ÇEKİLMEZ"
                  % (datetime.now().strftime("%Y-%m-%d %H:%M"), len(veri), len(grup))
            ).font = Font(size=9, italic=True, color="808080")

    bd = PatternFill("solid", fgColor="1F4E79")
    kd = PatternFill("solid", fgColor="C55A11")
    coklu = PatternFill("solid", fgColor="FFF2CC")   # grupta >1 kayıt
    bf = Font(color="FFFFFF", bold=True, size=10)
    for j, (bas, gen) in enumerate(SUTUNLAR, 1):
        c = ws.cell(BAS, j, bas)
        c.font = bf
        c.fill = kd if bas in ("KARAR", "AÇIKLAMA") else bd
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = gen
    ws.row_dimensions[BAS].height = 30

    for i, x in enumerate(veri, BAS + 1):
        a = x["anahtar"]
        sira[a] += 1
        n = grup[a]
        deger = {
            "SORUN": sorun_turu.get(a, "—"),
            # Süreç ADI gösterilir. Önceki sürümde `surec_kod` yazılıyordu ve
            # KMF'de süreç kodları DB'de NULL olduğu için sütun boş görünüyordu.
            "Süreç Adı": x.get("surec_ad") or x.get("surec_kod") or "—",
            "PGV Adı": x.get("pg_ad"),
            "Yıl": x.get("year"),
            "Dönem": x.get("period_type"),
            "Ay": x.get("period_month"),
            "Veri Tarihi": x.get("data_date"),
            "Girilen Değer": x.get("actual_value"),
            "Hedef": x.get("target_value"),
            "Aktif": "aktif" if x.get("is_active") else "SİLİNMİŞ",
            "Kullanıcı": (x.get("giren") or "").split("@")[0] or "—",
            "Veri Giriş Tarihi": None,
            "Bu grupta kaç kayıt": n,
            "Grup sırası": "%d / %d" % (sira[a], n),
            "KARAR": "",
            "AÇIKLAMA": ("⚠ aynı dönemde %d ayrı ölçüm" % n) if n > 1 else (aciklama.get(a) or ""),
            "yayın kd_id": x.get("id"),
            "İş Anahtarı": a,
        }
        for j, (bas, _) in enumerate(SUTUNLAR, 1):
            c = ws.cell(i, j, deger.get(bas))
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(bas in ("PGV Adı", "AÇIKLAMA")))
            if bas == "KARAR":
                c.fill = kd
                c.font = Font(color="FFFFFF")
            elif n > 1:
                c.fill = coklu

    ws.freeze_panes = ws.cell(BAS + 1, 4)
    ws.auto_filter.ref = "A%d:%s%d" % (BAS, get_column_letter(len(SUTUNLAR)), BAS + len(veri))

    # Özet
    oz = wb.create_sheet("Özet")
    cy = Counter(x["year"] for x in veri)
    coklu_grup = {a: n for a, n in grup.items() if n > 1}
    bilgi = [
        ("TOPLAM SATIR", len(veri)),
        ("Benzersiz iş anahtarı", len(grup)),
        ("Tek kayıtlı anahtar", len(grup) - len(coklu_grup)),
        ("ÇOK KAYITLI anahtar", len(coklu_grup)),
        ("  → bunlardaki fazla satır", sum(coklu_grup.values()) - len(coklu_grup)),
        ("", ""), ("YIL DAĞILIMI", ""),
    ] + [(str(y), cy[y]) for y in sorted(cy)] + [
        ("", ""), ("KARAR SEÇENEKLERİ", ""),
        ("ÇEK", "satır yerele kopyalanır"),
        ("ATLA", "satır çekilmez"),
        ("(boş)", "ÇEKİLMEZ — dokunulmaz"),
    ]
    for i, (a, b) in enumerate(bilgi, 1):
        oz.cell(i, 1, a).font = Font(bold=(b == "" or a.startswith("TOPLAM")))
        oz.cell(i, 2, b)
    oz.column_dimensions["A"].width = 30
    oz.column_dimensions["B"].width = 40

    # En çok tekrar eden gruplar
    oz.cell(len(bilgi) + 2, 1, "EN ÇOK KAYITLI GRUPLAR").font = Font(bold=True)
    for k, (a, n) in enumerate(sorted(coklu_grup.items(), key=lambda x: -x[1])[:20], 1):
        p = a.split("|")
        oz.cell(len(bilgi) + 2 + k, 1, "%s · %s · %s ay=%s" % (p[2][:34], p[3], p[4], p[6]))
        oz.cell(len(bilgi) + 2 + k, 2, n)

    yol = d / f"PGV_CEKILECEK_{args.kurum}.xlsx"
    wb.save(yol)
    print(f"\n  ✅ {yol}")
    print(f"     {len(veri)} satır · {len(grup)} anahtar · çok kayıtlı: {len(coklu_grup)}")
    print("\n     yıl: " + "  ".join(f"{y}={cy[y]}" for y in sorted(cy)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
