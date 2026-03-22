#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
s11.xlsx / s11-extracted.json → Kokpitim ProcessKpi + KpiData + süreç lider/üye ataması.

Kurallar (TASK kullanıcı yanıtları):
  - Tüm PG'ler alt strateji kodu H1.1 ile bağlanır (süreçte link yoksa eklenir).
  - gosterge_turu = İyileştirme
  - Ölçüm periyodu: Excel metnine göre (3 ay→Çeyreklik, 1 Yıl→Yıllık, 6 ay→6 ay)
  - Hedef hücresi metin aralığıysa sayıları çıkarılıp ortalaması tek hedef olarak yazılır.
  - 2 lider + 6 üye: tenant aktif kullanıcılarından rastgele (en az 8 kullanıcı gerekir).

Örnek:
  py scripts/kmf_s11_import.py --process-id 12 --actor-user-id 1 --xlsx docs/KMF/s11.xlsx --dry-run
  py scripts/kmf_s11_import.py --process-id 12 --actor-user-id 1 --xlsx docs/KMF/s11.xlsx --wipe-kpis --seed 42

Ortam: proje kökünden çalıştırın (.env yüklü app factory).
"""
from __future__ import annotations

import argparse
import importlib.util
import json
import math
import random
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _load_kmf_extract_module():
    p = Path(__file__).resolve().parent / "kmf_s11_extract.py"
    spec = importlib.util.spec_from_file_location("kmf_s11_extract", p)
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader
    spec.loader.exec_module(mod)
    return mod


def excel_olcum_to_pg_period(raw: str | None) -> str:
    if not raw:
        return "Çeyreklik"
    s = raw.strip().lower().replace("ı", "i").replace("ü", "u").replace("ö", "o")
    if re.search(r"(^|[^0-9])6\s*ay", s):
        return "6 ay"
    if "1" in s and ("yil" in s or "yıl" in raw.lower()):
        return "Yıllık"
    if "3" in s and "ay" in s:
        return "Çeyreklik"
    if "6" in s and "ay" in s:
        return "6 ay"
    return "Çeyreklik"


def tr_cell_to_float(chunk: str) -> float | None:
    chunk = chunk.strip().replace("%", "").replace(" ", "")
    if not chunk or chunk == "-":
        return None
    neg = chunk.startswith("-")
    if neg:
        chunk = chunk[1:].strip()
    if not chunk:
        return None
    if re.fullmatch(r"-?\d+\.\d{3}(?:\.\d{3})*(?:,\d+)?", chunk.replace(" ", "")):
        pass
    if "," in chunk and chunk.rsplit(",", 1)[-1].isdigit() and len(chunk.rsplit(",", 1)[-1]) <= 2:
        main, dec = chunk.rsplit(",", 1)
        main = main.replace(".", "")
        chunk = f"{main}.{dec}"
    else:
        chunk = chunk.replace(".", "").replace(",", ".")
    try:
        v = float(chunk)
        return -v if neg else v
    except ValueError:
        return None


def cell_to_scalar(val: object) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
            return None
        return float(val)
    s = str(val).strip()
    if not s or s == "-":
        return None
    nums: list[float] = []
    for m in re.finditer(r"-?[\d]+(?:[.,]\d+)*", s.replace(" ", "")):
        span = m.group(0)
        n = tr_cell_to_float(span)
        if n is not None:
            nums.append(n)
    if len(nums) >= 2:
        return sum(nums) / len(nums)
    if len(nums) == 1:
        return nums[0]
    return tr_cell_to_float(s)


def average_hedef_indicator(g: dict) -> float | None:
    parts: list[float] = []
    h = g.get("hedef") or {}
    for k in ("q1", "q2", "q3", "q4"):
        v = cell_to_scalar((h.get("ceyrekler") or {}).get(k))
        if v is not None:
            parts.append(v)
    ys = cell_to_scalar(h.get("yil_sonu"))
    if ys is not None:
        parts.append(ys)
    if not parts:
        return None
    return sum(parts) / len(parts)


def basari_esikleri_to_json(g: dict) -> str | None:
    bands = g.get("basari_esikleri")
    if not bands:
        return None
    obj = {}
    for i, b in enumerate(bands[:5], start=1):
        if b is None or (isinstance(b, str) and not str(b).strip()):
            continue
        obj[str(i)] = str(b).strip() if not isinstance(b, (int, float)) else str(b)
    if not obj:
        return None
    return json.dumps(obj, ensure_ascii=False)


def norm_gosterge_name(name: str | None) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--process-id", type=int, required=True)
    ap.add_argument("--actor-user-id", type=int, required=True, help="KpiData.user_id (veri giren)")
    ap.add_argument("--xlsx", type=Path, help="Kaynak Excel")
    ap.add_argument("--json", type=Path, help="kmf_s11_extract çıktısı (tercih edilir)")
    ap.add_argument("--sub-strategy-code", default="H1.1", help="Alt strateji kodu")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--wipe-kpis", action="store_true", help="Mevcut aktif PG+veriyi pasifleştir")
    ap.add_argument("--seed", type=int, default=None, help="Rastgele lider/üye için")
    args = ap.parse_args()

    if not args.json and not args.xlsx:
        print("Either --json or --xlsx gerekli.", file=sys.stderr)
        return 1

    if args.seed is not None:
        random.seed(args.seed)

    if args.json and args.json.exists():
        doc = json.loads(args.json.read_text(encoding="utf-8"))
    else:
        assert args.xlsx
        kmf = _load_kmf_extract_module()
        from openpyxl import load_workbook

        wb = load_workbook(args.xlsx, data_only=True)
        doc = {"sayfalar": []}
        for name in wb.sheetnames:
            if name == "Veriler" or re.fullmatch(r"\d{4}", name):
                doc["sayfalar"].append(kmf.parse_sheet(wb[name], name))
        wb.close()

    from app import create_app
    from app.models import db
    from app.models.core import Strategy, SubStrategy, User
    from app.models.process import (
        Process,
        ProcessKpi,
        KpiData,
        KpiDataAudit,
        ProcessSubStrategyLink,
    )
    from app.utils.process_utils import last_day_of_period

    app = create_app()
    with app.app_context():
        proc = Process.query.filter_by(id=args.process_id, is_active=True).first()
        if not proc:
            print(f"Süreç bulunamadı veya pasif: id={args.process_id}", file=sys.stderr)
            return 1

        actor = User.query.filter_by(id=args.actor_user_id, is_active=True).first()
        if not actor or actor.tenant_id != proc.tenant_id:
            print("actor-user-id tenant ile uyuşan aktif kullanıcı olmalı.", file=sys.stderr)
            return 1

        code_lc = args.sub_strategy_code.strip().lower()
        ss = (
            SubStrategy.query.join(Strategy)
            .filter(Strategy.tenant_id == proc.tenant_id)
            .filter(db.func.lower(SubStrategy.code) == code_lc)
            .filter(SubStrategy.is_active.is_(True))
            .order_by(SubStrategy.id.desc())
            .first()
        )
        if not ss:
            print(
                f"Alt strateji kodu bulunamadı: {args.sub_strategy_code!r} (tenant stratejilerinde tanımlı mı?)",
                file=sys.stderr,
            )
            return 1

        users = User.query.filter_by(tenant_id=proc.tenant_id, is_active=True).all()
        if len(users) < 8:
            print(
                f"Aktif kullanıcı sayısı yetersiz (en az 8 gerekli: 2 lider + 6 üye), var={len(users)}.",
                file=sys.stderr,
            )
            return 1
        picked = random.sample(users, 8)
        leaders = picked[:2]
        members = picked[2:10]

        all_pages = [p for p in doc["sayfalar"] if not p.get("error")]
        year_pages = sorted(
            [p for p in all_pages if p.get("yil")],
            key=lambda x: int(x["yil"]),
            reverse=True,
        )
        veriler_page = next((p for p in all_pages if p.get("sheet") == "Veriler"), None)

        by_name: dict[str, dict] = {}
        for page in year_pages:
            for g in page.get("gostergeler") or []:
                if g.get("parse_note"):
                    continue
                key = norm_gosterge_name(g.get("gosterge_adi"))
                if not key:
                    continue
                if key not in by_name:
                    by_name[key] = {"template": g, "years": {}}
                by_name[key]["years"][int(page["yil"])] = g

        if veriler_page:
            for g in veriler_page.get("gostergeler") or []:
                if g.get("parse_note"):
                    continue
                key = norm_gosterge_name(g.get("gosterge_adi"))
                if not key:
                    continue
                if key not in by_name:
                    by_name[key] = {"template": g, "years": {}}

        print(
            f"PG şablonları: {len(by_name)} | Yıl sayfaları: {[p['yil'] for p in year_pages]} | dry_run={args.dry_run}",
            file=sys.stderr,
        )

        if args.dry_run:
            print(json.dumps({k: v["template"].get("gosterge_adi") for k, v in by_name.items()}, ensure_ascii=False, indent=2))
            print("Liderler:", [u.email for u in leaders], file=sys.stderr)
            print("Üyeler:", [u.email for u in members], file=sys.stderr)
            return 0

        proc.leaders = leaders
        proc.members = members

        if not ProcessSubStrategyLink.query.filter_by(
            process_id=proc.id, sub_strategy_id=ss.id
        ).first():
            db.session.add(ProcessSubStrategyLink(process_id=proc.id, sub_strategy_id=ss.id))

        if args.wipe_kpis:
            for k in ProcessKpi.query.filter_by(process_id=proc.id, is_active=True).all():
                k.is_active = False
            for kd in (
                KpiData.query.join(ProcessKpi)
                .filter(ProcessKpi.process_id == proc.id, KpiData.is_active.is_(True))
                .all()
            ):
                kd.is_active = False

        db.session.flush()

        kpi_ids: dict[str, int] = {}
        idx = 0
        for key, bundle in by_name.items():
            g = bundle["template"]
            idx += 1
            avg_tgt = average_hedef_indicator(g)
            target_str = "" if avg_tgt is None else str(round(avg_tgt, 6)).rstrip("0").rstrip(".")
            bp_json = basari_esikleri_to_json(g)
            wraw = g.get("agirlik_yuzde")
            if wraw is None or (isinstance(wraw, str) and str(wraw).strip() in ("", "-")):
                w_db = 0.0
            else:
                try:
                    wf = float(wraw)
                except (TypeError, ValueError):
                    w_db = 0.0
                else:
                    w_db = wf * 100 if wf <= 1.0 + 1e-9 else wf
            oya = cell_to_scalar(g.get("onceki_yil_ortalamasi"))
            kpi = ProcessKpi(
                process_id=proc.id,
                name=(g.get("gosterge_adi") or key)[:200],
                code=f"KMF-{idx:03d}",
                description=None,
                target_value=target_str or None,
                unit=(g.get("birim") or "")[:50] or None,
                period=excel_olcum_to_pg_period(g.get("olcum_periyodu_excel")),
                data_collection_method="Ortalama",
                gosterge_turu="İyileştirme",
                target_method=(g.get("hedef_belirleme_yontemi") or "")[:10] or None,
                basari_puani_araliklari=bp_json,
                onceki_yil_ortalamasi=oya,
                weight=w_db,
                direction="Increasing",
                sub_strategy_id=ss.id,
                is_active=True,
            )

            db.session.add(kpi)
            db.session.flush()
            kpi_ids[key] = kpi.id

        for page in year_pages:
            y = int(page["yil"])
            for g in page.get("gostergeler") or []:
                key = norm_gosterge_name(g.get("gosterge_adi"))
                if not key or key not in kpi_ids:
                    continue
                kid = kpi_ids[key]
                fi = g.get("fiili") or {}
                qmap = fi.get("ceyrekler") or {}
                for qn in range(1, 5):
                    raw = qmap.get(f"q{qn}")
                    if raw is None or str(raw).strip() == "" or str(raw).strip() == "-":
                        continue
                    actual = str(raw).strip()
                    exists = KpiData.query.filter_by(
                        process_kpi_id=kid,
                        year=y,
                        period_type="ceyrek",
                        period_no=qn,
                        is_active=True,
                    ).first()
                    if exists:
                        continue
                    dd = last_day_of_period(y, "ceyrek", qn, None)
                    if not dd:
                        continue
                    ht = (g.get("hedef") or {}).get("ceyrekler") or {}
                    hv = ht.get(f"q{qn}")
                    tgt_scalar = cell_to_scalar(hv)
                    entry = KpiData(
                        process_kpi_id=kid,
                        year=y,
                        data_date=dd,
                        period_type="ceyrek",
                        period_no=qn,
                        period_month=None,
                        target_value=None if tgt_scalar is None else str(tgt_scalar),
                        actual_value=actual,
                        user_id=args.actor_user_id,
                        is_active=True,
                    )
                    db.session.add(entry)
                    db.session.flush()
                    db.session.add(
                        KpiDataAudit(
                            kpi_data_id=entry.id,
                            action_type="CREATE",
                            new_value=entry.actual_value,
                            action_detail="kmf_s11_import",
                            user_id=args.actor_user_id,
                        )
                    )

        db.session.commit()
        print(f"Tamamlandı. PG sayısı={len(kpi_ids)}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
