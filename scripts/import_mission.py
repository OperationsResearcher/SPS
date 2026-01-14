
import sys
import os
from openpyxl import load_workbook

# Add project root to sys.path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app
from extensions import db
from models.user import Kurum

def import_mission():
    try:
        app = create_app()
    except Exception as e:
        print(f"Failed to create app: {e}")
        return

    with app.app_context():
        # Load Excel
        file_path = os.path.join(app.root_path, 'belge', 'SP_V2.xlsx')
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return

        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            print(f"Failed to load excel: {e}")
            return

        # Find sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "misyon" in name.lower() and "vizyon" in name.lower():
                sheet_name = name
                break
        
        if not sheet_name:
            print("Sheet 'Misyon Vizyon SWOT' not found.")
            return

        ws = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        mission_text = None
        vision_text = None

        # Scan column B (col index 2)
        # We assume structure: Header -> Next Row Value
        for row in ws.iter_rows(min_col=2, max_col=2, values_only=False):
            cell = row[0]
            if not cell.value:
                continue
            
            val = str(cell.value).strip().lower()
            
            if "misyonumuz" in val or val == "misyon":
                # Get next row, same column
                next_cell = ws.cell(row=cell.row + 1, column=cell.column)
                if next_cell.value:
                    mission_text = str(next_cell.value).strip()
            
            if "vizyonumuz" in val or val == "vizyon":
                next_cell = ws.cell(row=cell.row + 1, column=cell.column)
                if next_cell.value:
                    vision_text = str(next_cell.value).strip()

        if not mission_text:
            print("Mission (Misyonumuz) not found in excel.")
        if not vision_text:
            print("Vision (Vizyonumuz) not found in excel.")

        if not mission_text and not vision_text:
            print("Nothing to update.")
            return

        print(f"Found Mission: {mission_text[:50]}..." if mission_text else "Mission empty")
        print(f"Found Vision: {vision_text[:50]}..." if vision_text else "Vision empty")

        # Update Database
        # Try to find KMF
        kurum = Kurum.query.filter(Kurum.kisa_ad.ilike('KMF')).first()
        
        if not kurum:
            # Check count of all institutions
            all_kurums = Kurum.query.all()
            if len(all_kurums) == 1:
                kurum = all_kurums[0]
                print(f"Kurum 'KMF' not found, applying to the only existing kurum: {kurum.kisa_ad}")
            elif len(all_kurums) == 0:
                print("No Kurum found in DB. Creating 'KMF'...")
                kurum = Kurum(kisa_ad='KMF', ticari_unvan='KMF Danışmanlık') # Default placeholder
                db.session.add(kurum)
            else:
                print(f"Multiple kurums found ({len(all_kurums)}). Cannot decide which one to update for 'KMF'.")
                # List them
                for k in all_kurums:
                    print(f"- {k.kisa_ad}")
                return

        # Update fields
        changes = []
        if mission_text:
            kurum.amac = mission_text
            changes.append("Mission->Amac")
        
        if vision_text:
            kurum.vizyon = vision_text
            changes.append("Vision->Vizyon")

        if changes:
            try:
                db.session.commit()
                print(f"Successfully updated {kurum.kisa_ad}: {', '.join(changes)}")
            except Exception as e:
                db.session.rollback()
                print(f"Error saving to DB: {e}")
        else:
            print("No changes needed.")

if __name__ == '__main__':
    import_mission()
